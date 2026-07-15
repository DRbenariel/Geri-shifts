import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import calendar
import random
import io
import os
import uuid
import streamlit_shadcn_ui as ui
import streamlit_antd_components as sac  # Added for Chips/Menu
import hashlib
import gspread
from streamlit_gsheets import GSheetsConnection
import time

calendar.setfirstweekday(calendar.SUNDAY)

# ── Daily-schedule departments (Phase 1+, single source of truth) ──
DAILY_DEPTS_ALL = ["שיקום גריאטרי א'", "שיקום גריאטרי ב'", "פנימית גריאטרית", "זה״ב", "בדיקה"]

# ── Hebrew month names (single source of truth) ───────────────────────────────
_HEB_MONTHS = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני",
               "יולי","אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

# ── פנימית גריאטרית — two sub-ward constants ─────────────────────────────────
PNIM_DEPT   = "פנימית גריאטרית"
PNIM_SIDES  = ["ורוד", "כחול"]
PNIM_COLORS = {"ורוד": "#fce7f3", "כחול": "#dbeafe"}
PNIM_ICONS  = {"ורוד": "🌸", "כחול": "🔵"}

import ui_components # Modular UI components

# --- 1. עיצוב ו-CSS ---
st.set_page_config(page_title="מערכת סידור עבודה", layout="wide")
ui_components.setup_style()

# RTL global fix — must run on every render (ungated).
# Uses st.html() which injects inline (not sandboxed) so <style> applies globally.
st.html("""<style>
.stMarkdown, .stMarkdownContainer, .stHeadingWithActionElements,
.stAlert, .stCaptionContainer {
    direction: rtl !important;
    unicode-bidi: bidi-override;
    text-align: right !important;
}
.stTextInput, .stTextArea, .stSelectbox, .stMultiSelect {
    direction: rtl !important;
    text-align: right !important;
}
input, label, textarea {
    direction: rtl !important;
    text-align: right !important;
}
</style>""")

import hashlib
from streamlit_gsheets import GSheetsConnection

# --- 2. ניהול מסד נתונים (Google Sheets) ---

def _strip_cell(s):
    """
    Strip only MATCHING surrounding quote pairs from a cell value.
    '...'  →  ...    (removes surrounding single quotes)
    "..."  →  ...    (removes surrounding double quotes)
    שיקום גריאטרי א'  →  unchanged  (trailing apostrophe is part of the name)
    """
    s = str(s).strip()
    if len(s) >= 2 and s[0] == s[-1] and s[0] in ("'", '"'):
        return s[1:-1]
    return s

# Apostrophe/geresh normalisation map — any of these variants → standard ASCII apostrophe.
# Google Sheets may store Hebrew dept names with geresh (׳ U+05F3), right-single-quote
# (' U+2019), or modifier-letter-apostrophe (ʼ U+02BC) instead of plain ' (U+0027).
_APOS_VARIANTS = str.maketrans({
    '׳': "'",   # ׳ HEBREW PUNCTUATION GERESH
    '’': "'",   # ' RIGHT SINGLE QUOTATION MARK
    'ʼ': "'",   # ʼ MODIFIER LETTER APOSTROPHE
    '‘': "'",   # ' LEFT SINGLE QUOTATION MARK
    '`': "'",   # ` GRAVE ACCENT
})

def _norm_dept(v: str) -> str:
    """Normalise a daily_dept string so apostrophe variants all compare equal."""
    return str(v).strip().translate(_APOS_VARIANTS)

def _norm_dr(df: "pd.DataFrame") -> "pd.DataFrame":
    """Normalise the daily_dept and side columns of a dept_rotation DataFrame."""
    if not df.empty and 'daily_dept' in df.columns:
        df = df.copy()
        df['daily_dept'] = df['daily_dept'].astype(str).apply(_norm_dept)
    # Ensure the `side` column always exists (added for פנימית two-ward display)
    if 'side' not in df.columns:
        df['side'] = ''
    df['side'] = df['side'].fillna('').astype(str).str.strip()
    return df

def _clean_sheet_values(vals):
    """
    Convert raw get_all_values() output to a clean DataFrame.
    Strips matching surrounding quote pairs from headers and cell values —
    these are baked in when data was imported via CSV or written with
    USER_ENTERED mode that serialised Python string reprs into cells.
    """
    if not vals:
        return pd.DataFrame()
    headers = [_strip_cell(h) for h in vals[0]]
    rows = []
    for row in vals[1:]:
        cleaned = [_strip_cell(c) for c in row]
        # Pad short rows to match header length
        while len(cleaned) < len(headers):
            cleaned.append('')
        rows.append(cleaned[:len(headers)])
    return pd.DataFrame(rows, columns=headers)

@st.cache_data(ttl=600, show_spinner=False)
def _fetch_sheet_data_silently(worksheet_name):
    """
    Cached sheet read. Uses get_all_values() so a sheet with a header row but
    no data rows still returns a DataFrame with correct column names.
    Strips surrounding quotes from all headers and cell values.
    """
    gc = get_gspread_client()
    url = st.secrets["connections"]["gsheets"]["spreadsheet"]
    sh = gc.open_by_url(url)
    try:
        ws = sh.worksheet(worksheet_name)
        vals = ws.get_all_values()
        return _clean_sheet_values(vals)
    except gspread.exceptions.WorksheetNotFound:
        return pd.DataFrame()

def _fetch_live(worksheet_name):
    """
    Direct Sheets read that bypasses the cache — use only immediately before a write.
    Uses get_all_values() + _clean_sheet_values() to preserve column names even on
    header-only sheets and to strip surrounding quotes baked in by CSV imports.
    """
    for attempt in range(3):
        try:
            gc = get_gspread_client()
            url = st.secrets["connections"]["gsheets"]["spreadsheet"]
            sh = gc.open_by_url(url)
            ws = sh.worksheet(worksheet_name)
            return _clean_sheet_values(ws.get_all_values())
        except gspread.exceptions.WorksheetNotFound:
            return pd.DataFrame()
        except gspread.exceptions.APIError as e:
            if e.response.status_code == 429:
                time.sleep(2 ** attempt)
            else:
                raise
    return pd.DataFrame()

def get_db_data(worksheet_name):
    try:
        df = _fetch_sheet_data_silently(worksheet_name)
        return df
    except Exception as e:
        err_msg = str(e).lower()
        if "worksheet" in err_msg and "not found" in err_msg:
             return pd.DataFrame()
        raise e

def get_gspread_client():
    # יצירת קליינט gspread מתוך ה-secrets הקיימים
    # נניח שהם במבנה סטנדרטי תחת connections.gsheets
    creds = st.secrets["connections"]["gsheets"]
    
    # gspread מצפה למילון או קובץ, st.secrets מחזיר מילון-כמו-אובייקט, נמיר למילון רגיל
    creds_dict = dict(creds)
    
    # במידה ויש צורך ב-scopes ספציפיים, gspread מטפל בזה לרוב אוטומטית עם service_account_from_dict
    gc = gspread.service_account_from_dict(creds_dict)
    return gc

def send_notification_email(to_address, subject, body_html):
    """
    Send an HTML email via Gmail SMTP.
    Credentials from [email] section in .streamlit/secrets.toml.
    Entire body wrapped in except — NEVER crashes the app.
    """
    try:
        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText

        cfg = st.secrets.get("email", {})
        smtp_server = str(cfg.get("smtp_server", "smtp.gmail.com"))
        smtp_port = int(cfg.get("smtp_port", 587))
        sender = str(cfg.get("sender_address", ""))
        password = str(cfg.get("sender_password", ""))
        if not sender or not password or not to_address:
            return  # not configured — silently skip

        msg = MIMEMultipart("alternative")
        msg["Subject"] = subject
        msg["From"] = sender
        msg["To"] = to_address
        msg.attach(MIMEText(body_html, "html", "utf-8"))

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender, password)
            server.sendmail(sender, to_address, msg.as_string())
    except Exception:
        pass  # Email failure must never affect the main app

def _update_absence_status(req_id, new_status, responder_name):
    """Update an absence_requests row's status + responded_at + approved_by, persist + email user."""
    df = st.session_state.absence_requests.copy()
    if df.empty or 'id' not in df.columns:
        return False
    mask = df['id'].astype(str) == str(req_id)
    if not mask.any():
        return False
    df.loc[mask, 'status']       = new_status
    df.loc[mask, 'approved_by']  = responder_name
    df.loc[mask, 'responded_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    st.session_state.absence_requests = df
    # Fire-and-forget write so admin can approve/reject in rapid succession
    # without each click freezing on the gspread round-trip. The coalescing
    # queue in _save_async guarantees the latest snapshot wins.
    _save_async("absence_requests", df.copy())

    # Notify the requester by email if we have one
    try:
        row = df[mask].iloc[0]
        emp_name = str(row.get('employee', '')).strip()
        # Look up requester's email
        staff_df = st.session_state.staff
        emp_match = staff_df[staff_df['name'].astype(str).str.strip() == emp_name]
        emp_email = str(emp_match.iloc[0].get('email', '')).strip() if not emp_match.empty else ''
        if emp_email:
            verb = "אושרה" if new_status == 'approved' else "נדחתה"
            send_notification_email(
                emp_email,
                f"בקשת היעדרות {verb}",
                f"<div dir='rtl'><p>שלום {emp_name},</p>"
                f"<p>בקשת ההיעדרות שלך לתאריכים "
                f"<b>{row.get('start_date','')} – {row.get('end_date','')}</b> "
                f"({row.get('type','')}) <b>{verb}</b> על ידי {responder_name}.</p></div>"
            )
    except Exception:
        pass
    return True

def _materialize_absence_to_wsd(employee, atype, start_date, end_date):
    """Write one work_schedule_daily row per day in [start_date, end_date] for an
    approved absence, so the request appears directly in the WSD sheet — not only
    as a runtime overlay via _derive_auto_status. Normalizes type via
    _ABSENCE_TYPE_TO_STATUS so 'חופש עתידי' becomes 'חופש', etc."""
    try:
        emp = str(employee).strip()
        status = _ABSENCE_TYPE_TO_STATUS.get(str(atype).strip(), str(atype).strip()) or "חופש"
        sd = datetime.strptime(str(start_date)[:10], '%Y-%m-%d').date()
        ed = datetime.strptime(str(end_date)[:10],   '%Y-%m-%d').date()
        if ed < sd:
            sd, ed = ed, sd
        cur = sd
        while cur <= ed:
            date_str = cur.strftime('%Y-%m-%d')
            dept = _emp_dept_for_date(emp, date_str) or ''
            _wsd_upsert(date_str, emp, dept, status, is_manual=True, note='')
            cur = cur + timedelta(days=1)
    except Exception:
        pass


def _approve_request(req_id, responder_name):
    result = _update_absence_status(req_id, 'approved', responder_name)
    if result:
        try:
            df = st.session_state.absence_requests
            row = df[df['id'].astype(str) == str(req_id)].iloc[0]
            _materialize_absence_to_wsd(
                row.get('employee', ''), row.get('type', ''),
                row.get('start_date', ''), row.get('end_date', ''))
        except Exception:
            pass
    _build_approved_map()   # keep cache in sync immediately
    return result

def _reject_request(req_id, responder_name):
    result = _update_absence_status(req_id, 'rejected', responder_name)
    _build_approved_map()   # keep cache in sync immediately
    return result


def _delete_absence_request(req_id):
    """Permanently delete one row from `absence_requests` by id.
    Used for hard-removing an APPROVED request the admin/manager decides was
    a mistake — distinct from _reject_request which keeps a paper trail."""
    try:
        df = st.session_state.absence_requests
        if df.empty or 'id' not in df.columns:
            return False
        m = df['id'].astype(str) == str(req_id)
        if not m.any():
            return False
        new_df = df[~m].reset_index(drop=True)
        st.session_state.absence_requests = new_df
        save_to_db("absence_requests", new_df)
        _build_approved_map()
        return True
    except Exception:
        return False


def _migrate_requests_on_dept_change(emp_name: str, old_dept: str, new_dept: str, year_month: str) -> int:
    """
    Called when the Gantt moves emp_name from old_dept → new_dept for year_month (YYYY-MM).
    Updates:
      • absence_requests  — dept_at_request + manager_email  (pending & approved rows only)
      • work_schedule_daily — daily_dept on manual חופש/202 rows in old_dept for this month
    Returns the number of absence_request rows migrated.
    """
    emp_n = str(emp_name).strip()
    ym    = str(year_month)   # e.g. "2026-07"

    # Resolve new manager email (first מנהל מחלקה whose manage_depts includes new_dept)
    new_mgr_email = ""
    try:
        sf = st.session_state.staff
        for _, sr in sf[sf['type'].astype(str).str.strip() == 'מנהל מחלקה'].iterrows():
            if new_dept in _parse_manage_depts(sr.get('manage_depts', '')):
                new_mgr_email = str(sr.get('email', '')).strip()
                break
    except Exception:
        pass

    migrated = 0

    # ── 1. absence_requests ──────────────────────────────────────────────────
    ar = st.session_state.absence_requests.copy()
    if not ar.empty and 'id' in ar.columns:
        mask = (
            (ar['employee'].astype(str).str.strip() == emp_n) &
            (ar['start_date'].astype(str).str.startswith(ym)) &
            (ar['dept_at_request'].astype(str).str.strip() == old_dept) &
            (ar['status'].astype(str).isin(['pending', 'approved']))
        )
        migrated = int(mask.sum())
        if migrated:
            ar.loc[mask, 'dept_at_request'] = new_dept
            ar.loc[mask, 'manager_email']   = new_mgr_email
            st.session_state.absence_requests = ar
            save_to_db("absence_requests", ar)
            _build_approved_map()

    # ── 2. work_schedule_daily — manual absence rows in old dept ─────────────
    wsd = st.session_state.work_schedule_daily.copy()
    if not wsd.empty and 'date' in wsd.columns:
        wsd_mask = (
            (wsd['employee'].astype(str).str.strip() == emp_n) &
            (wsd['date'].astype(str).str.startswith(ym)) &
            (wsd['daily_dept'].astype(str).str.strip() == old_dept) &
            (wsd['is_manual'].astype(str).str.lower() == 'true') &
            (wsd['status'].astype(str).isin(['חופש', '202']))
        )
        if wsd_mask.any():
            wsd.loc[wsd_mask, 'daily_dept'] = new_dept
            st.session_state.work_schedule_daily = wsd
            _rebuild_wsd_index()
            _save_async("work_schedule_daily", wsd.copy())

    return migrated


def _generate_work_schedule(year_month, view_month):
    """
    Generate work_schedule_daily for one month.
    Reads dept_rotation, absence_requests (approved), and schedule (night shifts) from session_state.
    Preserves rows with is_manual=True.
    Returns: {'written': N, 'kept_manual': K, 'employees': E, 'absences_applied': A, 'post_shifts': P, 'error': err}
    """
    try:
        year = 2026
        # Always reload absence_requests, work_schedule_daily, AND dept_rotation fresh from Sheets
        # so data edited directly in Google Sheets is picked up immediately.
        _ar_fresh = get_db_data("absence_requests")
        if not _ar_fresh.empty and 'employee' in _ar_fresh.columns:
            st.session_state.absence_requests = _ar_fresh
        _wsd_fresh = get_db_data("work_schedule_daily")
        if not _wsd_fresh.empty and 'date' in _wsd_fresh.columns:
            _wsd_fresh = _wsd_fresh[_wsd_fresh['date'].astype(str).str.startswith('2026')]
            st.session_state.work_schedule_daily = _wsd_fresh
            _rebuild_wsd_index()
        _dr_fresh = get_db_data("dept_rotation")
        if not _dr_fresh.empty and 'employee' in _dr_fresh.columns:
            st.session_state.dept_rotation = _norm_dr(_dr_fresh)

        # Inputs
        dr = st.session_state.dept_rotation.copy()
        ar = st.session_state.absence_requests.copy()
        sched = st.session_state.schedule.copy()
        wsd  = st.session_state.work_schedule_daily.copy()

        if dr.empty or 'employee' not in dr.columns:
            return {'error': "אין נתוני dept_rotation. אנא הגדר תחילה את השיבוץ החודשי."}

        dr['employee']   = dr['employee'].astype(str).str.strip()
        dr['year_month'] = dr['year_month'].astype(str)
        dr['daily_dept'] = dr['daily_dept'].astype(str).apply(_norm_dept)
        month_rotation = dr[dr['year_month'] == year_month]
        if month_rotation.empty:
            return {'error': f"אין שיבוצים לחודש {year_month}. אנא שבץ עובדים תחילה."}

        # Recurring weekly absences map: employee → set of Hebrew weekday indexes (Sun=0..Sat=6)
        HEB_DAY_TO_IDX = {"א": 0, "ב": 1, "ג": 2, "ד": 3, "ה": 4, "ו": 5, "ש": 6}
        recurring_map = {}
        staff_for_rec = st.session_state.staff
        if not staff_for_rec.empty and 'name' in staff_for_rec.columns:
            for _, sr in staff_for_rec.iterrows():
                rec_raw = str(sr.get('recurring_absent_days', '') or '').strip()
                if not rec_raw:
                    continue
                idxs = set()
                for tok in rec_raw.split(','):
                    t = tok.strip()
                    if t in HEB_DAY_TO_IDX:
                        idxs.add(HEB_DAY_TO_IDX[t])
                if idxs:
                    recurring_map[str(sr['name']).strip()] = idxs

        # Approved absences map: employee → list of (start_d, end_d, type)
        approved_map = {}
        if not ar.empty and 'status' in ar.columns:
            ar['status']     = ar['status'].astype(str).str.lower()
            ar['employee']   = ar['employee'].astype(str).str.strip()
            ar['start_date'] = ar['start_date'].astype(str)
            ar['end_date']   = ar['end_date'].astype(str)
            ar['type']       = ar['type'].astype(str)
            for _, r in ar[ar['status'] == 'approved'].iterrows():
                try:
                    sd = datetime.strptime(r['start_date'], '%Y-%m-%d').date()
                    ed = datetime.strptime(r['end_date'],   '%Y-%m-%d').date()
                except Exception:
                    continue
                approved_map.setdefault(r['employee'], []).append((sd, ed, r['type']))

        # Night shifts map: (employee, date_str) → True (for D+1 lookup)
        night_map = set()
        if not sched.empty and 'employee' in sched.columns and 'date' in sched.columns:
            sched['employee'] = sched['employee'].astype(str).str.strip()
            sched['date']     = sched['date'].astype(str)
            sched['dept']     = sched['dept'].astype(str)
            # Only main dept shifts count as "night shift" for אחרי תורנות (not שישי בוקר)
            real_shifts = sched[~sched['dept'].str.contains('שישי בוקר', na=False)]
            real_shifts = real_shifts[real_shifts['employee'].str.len() > 0]
            real_shifts = real_shifts[~real_shifts['employee'].isin(['', '---'])]
            for _, r in real_shifts.iterrows():
                night_map.add((r['employee'], r['date']))

        # Existing manual rows for this month (preserve untouched)
        if wsd.empty or 'date' not in wsd.columns:
            wsd = pd.DataFrame(columns=['date','employee','daily_dept','status','note','is_manual'])
        wsd['date']     = wsd['date'].astype(str)
        wsd['employee'] = wsd['employee'].astype(str).str.strip()
        wsd['is_manual'] = wsd['is_manual'].apply(
            lambda v: v if isinstance(v, bool) else str(v).strip().lower() == 'true'
        )
        in_month = wsd['date'].str.startswith(year_month)
        manual_rows  = wsd[in_month & wsd['is_manual']].copy()
        # Preserved manual rows (we keep these as-is)
        manual_keys  = set(zip(manual_rows['date'], manual_rows['employee']))

        # Other-month rows (untouched)
        other_rows = wsd[~in_month].copy()

        # Build new rows for this month
        new_rows = []
        absences_applied = 0
        post_shifts = 0
        num_days = calendar.monthrange(year, view_month)[1]

        _staff_mso = {
            str(r['name']).strip(): (r['manual_schedule_only']
                                     if isinstance(r['manual_schedule_only'], bool)
                                     else str(r.get('manual_schedule_only', '')).strip().lower() == 'true')
            for _, r in st.session_state.staff.iterrows()
            if 'manual_schedule_only' in r.index
        }

        for _, rot in month_rotation.iterrows():
            emp_n = rot['employee']
            dept_n = rot['daily_dept']
            if dept_n == "— לא שובץ —" or not dept_n:
                continue
            if _staff_mso.get(emp_n, False):
                continue  # only appear when manually planted
            for d in range(1, num_days + 1):
                date_obj = date(year, view_month, d)
                date_str = date_obj.strftime('%Y-%m-%d')

                # Skip: manual override exists for this (date, employee) → keep manual_rows version
                if (date_str, emp_n) in manual_keys:
                    continue

                status = "עובד"  # default
                note   = ""

                wd_idx_gen = (date_obj.weekday() + 1) % 7  # Sun=0, Fri=5, Sat=6

                # Priority 0a: Saturday — department closed
                if wd_idx_gen == 6:
                    status = "חופש"
                    note   = "שבת"

                # Priority 0b: Friday — working employees from שישי בוקר schedule
                elif wd_idx_gen == 5:
                    fri_shifts = _DAILY_DEPT_TO_FRIDAY_SHIFTS.get(str(dept_n), [])
                    if fri_shifts:
                        # Check if this employee is assigned to the mapped שישי בוקר shift
                        fri_assigned = {
                            str(r['employee']).strip()
                            for _, r in sched.iterrows()
                            if str(r['date']) == date_str and str(r['dept']) in fri_shifts
                        }
                        if emp_n not in fri_assigned:
                            status = "חופש"
                            note   = "שישי — לא משובץ"

                # Priority 1: approved absence covers this day?
                if status == "עובד":
                    for sd, ed, atype in approved_map.get(emp_n, []):
                        if sd <= date_obj <= ed:
                            status = _ABSENCE_TYPE_TO_STATUS.get(atype, atype) if atype else "חופש"
                            break

                # Priority 2: recurring weekly absence? (weekdays only — Fri/Sat handled above)
                if status == "עובד":
                    wd_heb_idx = wd_idx_gen
                    if wd_heb_idx in recurring_map.get(emp_n, set()):
                        status = "חופש"
                        note = "היעדרות קבועה"

                # Priority 3: night shift the day before?
                if status == "עובד":
                    prev_str = (date_obj - timedelta(days=1)).strftime('%Y-%m-%d')
                    if (emp_n, prev_str) in night_map:
                        status = "אחרי תורנות"
                        post_shifts += 1
                else:
                    absences_applied += 1

                new_rows.append({
                    'date': date_str,
                    'employee': emp_n,
                    'daily_dept': dept_n,
                    'status': status,
                    'note': note,
                    'is_manual': False,
                })

        # Merge: other_rows + manual_rows (this month) + new_rows (auto for this month)
        new_df = pd.concat(
            [other_rows, manual_rows, pd.DataFrame(new_rows)],
            ignore_index=True
        )
        st.session_state.work_schedule_daily = new_df
        save_to_db("work_schedule_daily", new_df)

        return {
            'written': len(new_rows),
            'kept_manual': len(manual_rows),
            'employees': len(month_rotation),
            'absences_applied': absences_applied,
            'post_shifts': post_shifts,
            'error': None,
        }
    except Exception as e:
        return {'error': f"שגיאה ביצירת הסידור: {e}"}

def _export_schedule_wide(view_month):
    """
    Export the schedule sheet (night shifts) to 'Schedule_Export' in wide format.
    One row per date, columns: תאריך, יום, פנימית גריאטרית, שיקום, שישי בוקר *4.
    Returns (ok, error_msg).
    """
    try:
        days_in_month = calendar.monthrange(2026, view_month)[1]
        dates_list = [date(2026, view_month, d) for d in range(1, days_in_month + 1)]
        export_rows = []
        schedule_data = st.session_state.schedule

        for d_obj in dates_list:
            d_str = str(d_obj)
            day_name = ["ב'", "ג'", "ד'", "ה'", "ו'", "ש'", "א'"][d_obj.weekday()]
            row_data = {
                'תאריך': d_str, 'יום': day_name,
                'פנימית גריאטרית': '', 'שיקום': '',
                'שישי בוקר - פנימית (1)': '', 'שישי בוקר - פנימית (2)': '',
                'שישי בוקר - שיקום (1)':   '', 'שישי בוקר - שיקום (2)':   '',
            }
            daily_shifts = schedule_data[schedule_data['date'] == d_str]
            if not daily_shifts.empty:
                for _, shift in daily_shifts.iterrows():
                    dept = shift['dept']; emp = shift['employee']
                    if emp == '---': continue
                    if dept in row_data:
                        row_data[dept] = emp
            export_rows.append(row_data)

        df_export = pd.DataFrame(export_rows)
        col_order = [
            'תאריך', 'יום',
            'פנימית גריאטרית', 'שיקום',
            'שישי בוקר - פנימית (1)', 'שישי בוקר - פנימית (2)',
            'שישי בוקר - שיקום (1)',   'שישי בוקר - שיקום (2)',
        ]
        for col in col_order:
            if col not in df_export.columns:
                df_export[col] = ''
        df_export = df_export[col_order]
        save_to_db("Schedule_Export", df_export, is_rtl=True)
        return True, None
    except Exception as e:
        return False, str(e)

def _export_dept_grid(dept_name, year_month, view_month, year=None):
    """
    Export the dept grid to a Google Sheets worksheet `WSD_<dept>_<year_month>`,
    matching the same 3-section batched layout as the Excel export
    (uses `_build_batched_day_data`):
      Row 1 : day-number + weekday-letter header
      Row 2 : עובדים section label
      Row 3+: worker slots
      Row N : תורן row (label + names)
      Row N+1: לא נמצאים section label
      Row N+2+: absent slots
    Friday: workers = שישי בוקר only; absences hidden.
    Saturday: only תורן (workers + absences hidden).
    """
    try:
        # Feature 4: derive year from caller (or year_month prefix), defaulting to 2026.
        if year is None:
            try:
                year = int(str(year_month).split('-')[0])
            except Exception:
                year = 2026
        num_days = calendar.monthrange(year, view_month)[1]
        days = list(range(1, num_days + 1))
        WD = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

        result = _build_batched_day_data(dept_name, year_month, view_month, year=year)
        if not result or result[0] is None:
            return False
        (_employees, per_day_workers, per_day_toranet, per_day_absent,
         per_day_workers_by_side) = result

        toranet_label = "תורן פנגר" if dept_name == "פנימית גריאטרית" else "תורן שיקום"

        max_work_rows = max((len(v) for v in per_day_workers.values()), default=1)
        max_abs_rows  = max((len(v) for v in per_day_absent.values()),  default=0)

        col_names = ['#'] + [str(d) for d in days]
        rows = []

        # Row 1: day-number + weekday letter (one cell, e.g. "15 / ה")
        rows.append({
            '#': f"{dept_name} — {year_month}",
            **{str(d): f"{d} / {WD[(date(year, view_month, d).weekday() + 1) % 7]}"
               for d in days}
        })

        if per_day_workers_by_side is not None:
            # פנימית: one block per side (🌸 ורוד then 🔵 כחול), each רופא בכיר→מתמחים
            for _s in PNIM_SIDES:
                _max_s = max((len(v.get(_s, [])) for v in per_day_workers_by_side.values()),
                             default=0)
                rows.append({'#': f"{PNIM_ICONS[_s]} צד {_s}",
                             **{str(d): '' for d in days}})
                for i in range(max(_max_s, 1)):
                    row = {'#': ''}
                    for d in days:
                        lst = per_day_workers_by_side.get(d, {}).get(_s, [])
                        row[str(d)] = lst[i] if i < len(lst) else ''
                    rows.append(row)
        else:
            # Section עובדים
            rows.append({'#': '— עובדים —', **{str(d): '' for d in days}})
            for i in range(max_work_rows):
                row = {'#': ''}
                for d in days:
                    lst = per_day_workers.get(d, [])
                    row[str(d)] = lst[i] if i < len(lst) else ''
                rows.append(row)

        # תורן row (single data row)
        row_t = {'#': toranet_label}
        for d in days:
            row_t[str(d)] = per_day_toranet.get(d) or ''
        rows.append(row_t)

        # Section לא נמצאים (only Sun–Thu have entries; Fri/Sat have empty per the builder)
        if max_abs_rows > 0:
            rows.append({'#': '— לא נמצאים —', **{str(d): '' for d in days}})
            for i in range(max_abs_rows):
                row = {'#': ''}
                for d in days:
                    lst = per_day_absent.get(d, [])
                    row[str(d)] = lst[i] if i < len(lst) else ''
                rows.append(row)

        export_df = pd.DataFrame(rows, columns=col_names)

        safe = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:25]
        sheet_name = f"WSD_{safe}_{year_month}"
        save_to_db(sheet_name, export_df)

        # Set worksheet to RTL
        try:
            gc  = get_gspread_client()
            url = st.secrets["connections"]["gsheets"]["spreadsheet"]
            sh  = gc.open_by_url(url)
            ws  = sh.worksheet(sheet_name)
            sh.batch_update({"requests": [{
                "updateSheetProperties": {
                    "properties": {"sheetId": ws.id, "rightToLeft": True},
                    "fields": "rightToLeft"
                }
            }]})
        except Exception:
            pass  # RTL is cosmetic — don't fail the export if it errors

        return True
    except Exception:
        return False

# ── Excel colour map (PatternFill hex) ──────────────────────────────────────
_XL_STATUS_FILL = {
    "עובד":         "DCFCE7",
    "חופש":         "DBEAFE",
    "202":          "FDE047",   # bright yellow — matches #eab308 on screen
    "אחרי תורנות":  "FFEDD5",
    "תורנות":       "EDE9FE",
    "אחר":          "F1F5F9",
}

def _get_user_email(user_name):
    """Return email for user_name from staff sheet, or '' if missing."""
    try:
        sf = st.session_state.staff
        if sf.empty or 'email' not in sf.columns:
            return ""
        row = sf[sf['name'].astype(str).str.strip() == str(user_name).strip()]
        if row.empty:
            return ""
        return str(row.iloc[0].get('email', '') or '').strip()
    except Exception:
        return ""

def _export_dept_grid_excel(dept_name, year_month, view_month):
    """
    Build an in-memory Excel workbook for the dept grid and return (bytes, filename).
    Returns (None, None) on failure.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        year = 2026
        num_days = calendar.monthrange(year, view_month)[1]
        days = list(range(1, num_days + 1))
        WD = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

        dr = st.session_state.dept_rotation
        if dr.empty or 'employee' not in dr.columns:
            return None, None
        mask = ((dr['year_month'].astype(str) == year_month) &
                (dr['daily_dept'].astype(str) == dept_name))
        employees = dr[mask]['employee'].astype(str).str.strip().tolist()
        if not employees:
            return None, None

        wb = openpyxl.Workbook()
        ws_xl = wb.active
        ws_xl.title = dept_name[:31]
        ws_xl.sheet_view.rightToLeft = True

        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Row 1: dept + month title
        ws_xl.cell(1, 1, f"{dept_name} — {year_month}").font = Font(bold=True, size=12)
        ws_xl.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=num_days + 1)

        # Row 2: header — col A = "עובד", then day numbers
        ws_xl.cell(2, 1, "עובד/ת").font = Font(bold=True)
        for i, d in enumerate(days):
            wd_letter = WD[(date(year, view_month, d).weekday() + 1) % 7]
            cell = ws_xl.cell(2, i + 2, f"{d}\n{wd_letter}")
            cell.font = Font(bold=True, size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border = border
        ws_xl.row_dimensions[2].height = 28
        ws_xl.column_dimensions['A'].width = 16

        # Data rows
        for ri, emp in enumerate(employees):
            row_num = ri + 3
            ws_xl.cell(row_num, 1, emp).font = Font(bold=False, size=10)
            ws_xl.cell(row_num, 1).alignment = Alignment(horizontal='right')
            for i, d in enumerate(days):
                date_str = f"{year}-{view_month:02d}-{d:02d}"
                raw = _wsd_get_status(date_str, emp, default=None)
                status = raw if raw is not None else _derive_auto_status(date_str, emp, daily_dept=dept_name)
                note   = _wsd_get_note(date_str, emp)
                col_num = i + 2
                cell = ws_xl.cell(row_num, col_num)
                lbl = _GRID_STATUS_LABEL_SHORT.get(status, status)
                cell.value = lbl
                cell.alignment = Alignment(horizontal='center', vertical='center',
                                           wrap_text=True, shrink_to_fit=False)
                fill_hex = _XL_STATUS_FILL.get(status, "F8FAFC")
                cell.fill = PatternFill("solid", fgColor=fill_hex)
                cell.border = border
                cell.font   = Font(size=9)
            ws_xl.column_dimensions[
                openpyxl.utils.get_column_letter(i + 2)].width = 6

        # Freeze header rows
        ws_xl.freeze_panes = "B3"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:25]
        return buf.getvalue(), f"WSD_{safe}_{year_month}.xlsx"
    except Exception:
        return None, None

def _export_personal_schedule_excel(user_name, view_month, daily_dept=None):
    """
    Build an Excel file with the personal day schedule for one employee.
    Returns (bytes, filename) or (None, None).
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        year = 2026
        num_days = calendar.monthrange(year, view_month)[1]
        WD_FULL = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]

        wb = openpyxl.Workbook()
        ws_xl = wb.active
        ws_xl.title = "לוח עבודה אישי"
        ws_xl.sheet_view.rightToLeft = True

        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Header
        headers = ["תאריך", "יום", "סטטוס", "הערה"]
        header_fills = ["334155"] * 4
        for ci, h in enumerate(headers):
            cell = ws_xl.cell(1, ci + 1, h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = PatternFill("solid", fgColor="334155")
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        ws_xl.column_dimensions['A'].width = 12
        ws_xl.column_dimensions['B'].width = 10
        ws_xl.column_dimensions['C'].width = 14
        ws_xl.column_dimensions['D'].width = 22

        for d in range(1, num_days + 1):
            date_str = f"{year}-{view_month:02d}-{d:02d}"
            wd_name  = WD_FULL[(date(year, view_month, d).weekday() + 1) % 7]
            raw = _wsd_get_status(date_str, user_name, default=None)
            manual = _wsd_is_manual(date_str, user_name)
            if raw is None or (raw == "עובד" and not manual):
                status = _derive_auto_status(date_str, user_name,
                                             daily_dept=daily_dept)
            else:
                status = raw
            note   = _wsd_get_note(date_str, user_name)
            row_num = d + 1
            ws_xl.cell(row_num, 1, date_str).border = border
            ws_xl.cell(row_num, 2, wd_name).border = border
            status_cell = ws_xl.cell(row_num, 3, status)
            status_cell.fill = PatternFill("solid", fgColor=_XL_STATUS_FILL.get(status, "F8FAFC"))
            status_cell.alignment = Alignment(horizontal='center')
            status_cell.border = border
            ws_xl.cell(row_num, 4, note).border = border

        ws_xl.freeze_panes = "A2"
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = str(user_name).replace(' ', '_')[:20]
        return buf.getvalue(), f"סידור_{safe_name}_{year}-{view_month:02d}.xlsx"
    except Exception:
        return None, None

def _export_dept_to_new_gsheet(dept_name, year_month, view_month):
    """
    Export the dept grid into the EXISTING Shifts_scheduler spreadsheet
    (same as _export_dept_grid) and return a direct URL to that specific
    worksheet tab so it can be opened immediately in the browser.
    Returns (True, tab_url, '') or (False, '', error_message).
    No new files are created — avoids Drive quota issues.
    """
    try:
        # Run the standard export into the main spreadsheet
        ok = _export_dept_grid(dept_name, year_month, view_month)
        if not ok:
            return False, '', "שגיאה בייצוא — וודא שיש שיבוצים למחלקה זו"

        # Build the worksheet name (same logic as _export_dept_grid)
        safe = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:25]
        sheet_name = f"WSD_{safe}_{year_month}"

        # Open the spreadsheet and get the worksheet's gid for a direct tab URL
        gc  = get_gspread_client()
        url = st.secrets["connections"]["gsheets"]["spreadsheet"]
        sh  = gc.open_by_url(url)
        ws  = sh.worksheet(sheet_name)
        tab_url = f"{sh.url}#gid={ws.id}"

        # Make the spreadsheet viewable by anyone with the link (read-only)
        try:
            sh.share('', perm_type='anyone', role='reader')
        except Exception:
            pass  # non-critical — link still works for existing editors

        return True, tab_url, ''
    except Exception as e:
        return False, '', str(e)


# ── Batched export helpers ────────────────────────────────────────────────────
_BATCHED_ABSENT_STATUSES = {"חופש", "202", "אחרי תורנות", "אחר"}
_BATCHED_SHIKUM_DEPTS    = {"שיקום גריאטרי א'", "שיקום גריאטרי ב'"}


def _build_batched_day_data(dept_name, year_month, view_month, year=None):
    """
    Build per-day batches for the batched export format.
    Returns (employees, per_day_workers, per_day_toranet, per_day_absent,
             per_day_workers_by_side) or a 5-tuple of None if no data.

    `per_day_workers_by_side` is None for non-פנימית depts; for פנימית it is
    {d: {'ורוד': [...], 'כחול': [...]}} with each side ordered רופא בכיר→מתמחים.

    Batch rules:
      - Regular day : workers = dept_rotation employees not absent (sorted by role)
      - Friday      : workers = שישי בוקר workers for this dept
      - Saturday    : workers = [] (nothing for שיקום; only תורן for פנימית)
      - תורן        : from _get_night_duty()
      - absent      : employees with ABSENT status
      - day-specific transfers (Feature 3): folded into that day's workers only
    """
    # Feature 4: derive year from caller (or year_month prefix), defaulting to 2026.
    if year is None:
        try:
            year = int(str(year_month).split('-')[0])
        except Exception:
            year = 2026
    num_days = calendar.monthrange(year, view_month)[1]
    days = list(range(1, num_days + 1))
    is_pnim = (dept_name == PNIM_DEPT)

    dr = st.session_state.dept_rotation
    if dr.empty or 'employee' not in dr.columns:
        return None, None, None, None, None
    mask = ((dr['year_month'].astype(str) == year_month) &
            (dr['daily_dept'].astype(str) == dept_name))
    employees = dr[mask]['employee'].astype(str).str.strip().tolist()

    # מנהל מחלקה: managers of this dept are auto-planted even without a
    # dept_rotation row (their dept comes from manage_depts). Merge them in so
    # exports match the on-screen grid.
    employees = employees + [m for m in _get_dept_managers(dept_name) if m not in employees]

    # Side map (פנימית) + day-specific incoming transfers
    side_map = {}
    if is_pnim and 'side' in dr.columns:
        side_map = (dr[dr['year_month'].astype(str) == year_month]
                    .set_index('employee')['side']
                    .fillna('').astype(str).str.strip().to_dict())
    incoming = _incoming_transfers(dept_name, year_month)   # {emp: {'days':set,'side':str}}
    transfer_side = {e: r.get('side', '') for e, r in incoming.items()}

    if not employees and not incoming:
        return None, None, None, None, None
    employees = _sort_employees_by_role(employees)

    def _emp_side(emp):
        s = side_map.get(emp, '')
        if not s:
            s = transfer_side.get(emp, '')
        return s

    per_day_workers = {}
    per_day_toranet = {}
    per_day_absent  = {}
    per_day_workers_by_side = {} if is_pnim else None

    for d in days:
        date_str = f"{year}-{view_month:02d}-{d:02d}"
        date_obj = date(year, view_month, d)
        weekday  = (date_obj.weekday() + 1) % 7  # 0=Sun 5=Fri 6=Sat
        is_friday   = weekday == 5
        is_saturday = weekday == 6

        workers, absent = [], []

        if is_saturday:
            # Saturday: only תורן row populated (workers + absences hidden — it's a given)
            pass
        elif is_friday:
            # Friday: workers = שישי בוקר workers for this dept.
            # Absences hidden (it's a given).
            workers = list(_get_fri_shift_workers(date_str, dept_name))
        else:
            # Regular weekday (Sun–Thu)
            for emp in employees:
                status = _derive_auto_status(date_str, emp, daily_dept=dept_name)
                if status == "":
                    # מנהל מחלקה not planted (or transferred out) this day → skip
                    continue
                if status in _BATCHED_ABSENT_STATUSES:
                    if status == "אחרי תורנות":
                        prev_str = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
                        _ns_lbl = _night_shift_dept_label(prev_str, emp)
                        absent.append(f"{emp} - אחרי תורנות ({_ns_lbl})" if _ns_lbl else f"{emp} - אחרי תורנות")
                    else:
                        absent.append(f"{emp} - {status}")
                else:
                    workers.append(emp)

        # Fold day-specific incoming transfers into this day's workers
        for emp, rec in incoming.items():
            if d in rec['days'] and emp not in workers:
                status = _derive_auto_status(date_str, emp, daily_dept=dept_name)
                if status and status not in _BATCHED_ABSENT_STATUSES:
                    workers.append(emp)

        workers = _sort_employees_by_role(workers)
        per_day_workers[d] = workers
        per_day_toranet[d] = _get_night_duty(date_str, dept_name)
        per_day_absent[d]  = absent

        if is_pnim:
            by_side = {s: [] for s in PNIM_SIDES}
            for emp in workers:
                s = _emp_side(emp)
                if s in by_side:
                    by_side[s].append(emp)   # drop side-less workers
            per_day_workers_by_side[d] = by_side

    return (employees, per_day_workers, per_day_toranet, per_day_absent,
            per_day_workers_by_side)


def _write_batched_sheet(ws, dept_name, year_month, view_month,
                         per_day_workers, per_day_toranet, per_day_absent,
                         per_day_workers_by_side=None, year=None):
    """
    Write batched schedule data into an existing openpyxl Worksheet.
    Layout (rows × cols-per-day):
      Row 1  : dept title header
      Row 2  : day-number + weekday-letter header
      Row 3  : 'עובדים' section label
      Rows 4+ : worker slots (max across all days)
      Next   : תורן row (single data row, label in col A)
      Next   : 'לא נמצאים' section label
      Next+  : absent slots (max across all days)
    No merged cells.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    # Feature 4: derive year from caller (or year_month prefix), defaulting to 2026.
    if year is None:
        try:
            year = int(str(year_month).split('-')[0])
        except Exception:
            year = 2026
    num_days = calendar.monthrange(year, view_month)[1]
    days = list(range(1, num_days + 1))
    WD = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

    toranet_label = "תורן פנגר" if dept_name == "פנימית גריאטרית" else "תורן שיקום"

    max_work_rows = max((len(v) for v in per_day_workers.values()), default=1)
    max_abs_rows  = max((len(v) for v in per_day_absent.values()),  default=0)

    # ── Styles ──────────────────────────────────────────────────────────
    thin = Side(style='thin', color='CBD5E1')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    f_hdr_dept   = PatternFill("solid", fgColor="1E293B")
    f_hdr_day    = PatternFill("solid", fgColor="334155")
    f_sec_work   = PatternFill("solid", fgColor="DCFCE7")
    f_sec_toran  = PatternFill("solid", fgColor="EDE9FE")
    f_sec_absent = PatternFill("solid", fgColor="FEF2F2")
    f_work_cell  = PatternFill("solid", fgColor="F0FDF4")
    f_fri_cell   = PatternFill("solid", fgColor="FFFBEB")
    f_sat_cell   = PatternFill("solid", fgColor="F1F5F9")
    f_abs_cell   = PatternFill("solid", fgColor="FFF5F5")

    ws.sheet_view.rightToLeft = True

    def _set(row, col, val="", **kw):
        c = ws.cell(row, col, val)
        c.border = bdr
        for attr, v in kw.items():
            setattr(c, attr, v)
        return c

    # ── Row 1: dept title ────────────────────────────────────────────────
    _set(1, 1, f"{dept_name} — {year_month}",
         font=Font(bold=True, size=12, color="FFFFFF"),
         fill=f_hdr_dept,
         alignment=Alignment(horizontal='right'))
    for d in days:
        _set(1, d + 1, fill=f_hdr_dept)

    # ── Row 2: day numbers + weekday letters ─────────────────────────────
    _set(2, 1, "יום / תאריך",
         font=Font(bold=True, color="FFFFFF", size=9),
         fill=f_hdr_day,
         alignment=Alignment(horizontal='center'))
    for d in days:
        date_obj = date(year, view_month, d)
        wd_idx   = (date_obj.weekday() + 1) % 7
        _set(2, d + 1, f"{d}\n{WD[wd_idx]}",
             font=Font(bold=True, color="FFFFFF", size=9),
             fill=f_hdr_day,
             alignment=Alignment(horizontal='center', vertical='center', wrap_text=True))
    ws.row_dimensions[2].height = 26

    # ── Worker section ───────────────────────────────────────────────────
    r = 3

    def _write_worker_rows(getter, n_rows):
        """Write n_rows worker-slot rows; getter(d) → that day's name list."""
        nonlocal r
        for i in range(n_rows):
            _set(r, 1)
            for d in days:
                wd_idx = (date(year, view_month, d).weekday() + 1) % 7
                lst = getter(d)
                val = lst[i] if i < len(lst) else ""
                if wd_idx == 6:
                    bg = f_sat_cell
                elif wd_idx == 5:
                    bg = f_fri_cell
                else:
                    bg = f_work_cell
                _set(r, d + 1, val,
                     font=Font(size=9),
                     fill=bg,
                     alignment=Alignment(horizontal='right', vertical='center'))
            r += 1

    if per_day_workers_by_side is not None:
        # פנימית: one labelled block per side (🌸 ורוד then 🔵 כחול)
        f_side = {"ורוד": PatternFill("solid", fgColor="FCE7F3"),
                  "כחול": PatternFill("solid", fgColor="DBEAFE")}
        for _s in PNIM_SIDES:
            _max_s = max((len(v.get(_s, [])) for v in per_day_workers_by_side.values()),
                         default=0)
            _set(r, 1, f"{PNIM_ICONS[_s]} צד {_s}",
                 font=Font(bold=True, size=9, color="1E293B"),
                 fill=f_side.get(_s, f_sec_work),
                 alignment=Alignment(horizontal='right'))
            for d in days:
                _set(r, d + 1, fill=f_side.get(_s, f_sec_work))
            r += 1
            _write_worker_rows(
                lambda d, _s=_s: per_day_workers_by_side.get(d, {}).get(_s, []),
                max(_max_s, 1))
    else:
        _set(r, 1, "עובדים",
             font=Font(bold=True, size=9, color="166534"),
             fill=f_sec_work,
             alignment=Alignment(horizontal='right'))
        for d in days:
            _set(r, d + 1, fill=f_sec_work)
        r += 1
        _write_worker_rows(lambda d: per_day_workers.get(d, []), max_work_rows)

    # ── תורן row (label + data in same row) ──────────────────────────────
    _set(r, 1, toranet_label,
         font=Font(bold=True, size=9, color="5B21B6"),
         fill=f_sec_toran,
         alignment=Alignment(horizontal='right'))
    for d in days:
        nd = per_day_toranet.get(d) or ""
        _set(r, d + 1, nd,
             font=Font(size=9, color="5B21B6"),
             fill=f_sec_toran,
             alignment=Alignment(horizontal='right', vertical='center'))
    r += 1

    # ── לא נמצאים section label ──────────────────────────────────────────
    _set(r, 1, "לא נמצאים",
         font=Font(bold=True, size=9, color="991B1B"),
         fill=f_sec_absent,
         alignment=Alignment(horizontal='right'))
    for d in days:
        _set(r, d + 1, fill=f_sec_absent)
    r += 1

    # Absent rows
    for i in range(max_abs_rows):
        _set(r, 1)
        for d in days:
            lst = per_day_absent.get(d, [])
            val = lst[i] if i < len(lst) else ""
            _set(r, d + 1, val,
                 font=Font(size=9, color="991B1B" if val else "9CA3AF"),
                 fill=f_abs_cell,
                 alignment=Alignment(horizontal='right', vertical='center'))
        r += 1

    # ── Column widths + freeze ────────────────────────────────────────────
    ws.column_dimensions['A'].width = 14
    for d in days:
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(d + 1)].width = 11
    ws.freeze_panes = "B3"


def _export_dept_grid_batched(dept_name, year_month, view_month):
    """
    Export single dept in batched format. Returns (bytes, filename) or (None, None).
    """
    try:
        import openpyxl
        employees, pw, pt, pa, pws = _build_batched_day_data(
            dept_name, year_month, view_month)
        if employees is None:
            return None, None

        wb = openpyxl.Workbook()
        ws = wb.active
        safe_title = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:31]
        ws.title = safe_title

        _write_batched_sheet(ws, dept_name, year_month, view_month, pw, pt, pa, pws)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:25]
        return buf.getvalue(), f"סידור_{safe}_{year_month}.xlsx"
    except Exception:
        return None, None


def _export_all_depts_batched(year_month, view_month):
    """
    Export all 3 dept grids in a single Excel workbook (one sheet per dept).
    Returns (bytes, filename) or (None, None).
    """
    try:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet
        any_added = False

        for dept_name in DAILY_DEPTS_ALL:
            employees, pw, pt, pa, pws = _build_batched_day_data(
                dept_name, year_month, view_month)
            if employees is None:
                continue
            safe_title = dept_name.replace("'", "").replace('"', '').replace('/', '_')[:31]
            ws = wb.create_sheet(title=safe_title)
            _write_batched_sheet(ws, dept_name, year_month, view_month, pw, pt, pa, pws)
            any_added = True

        if not any_added:
            return None, None

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue(), f"סידור_כל_המחלקות_{year_month}.xlsx"
    except Exception:
        return None, None


def _get_fri_shift_workers(date_str, daily_dept):
    """
    Return list of employee names assigned to the department's Friday-morning
    shifts on date_str (Fridays only). Returns [] for any other weekday.
    For פנימית גריאטרית returns both slot-1 and slot-2 names;
    for each שיקום dept returns the single matching slot name.
    """
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        if (date_obj.weekday() + 1) % 7 != 5:   # not Friday
            return []
        fri_shifts = _DAILY_DEPT_TO_FRIDAY_SHIFTS.get(str(daily_dept), [])
        if not fri_shifts:
            return []
        sch = st.session_state.schedule
        if sch.empty or 'date' not in sch.columns:
            return []
        mask = (
            (sch['date'].astype(str) == date_str) &
            (sch['dept'].astype(str).isin(fri_shifts))
        )
        workers = sch[mask]['employee'].astype(str).str.strip().tolist()
        return [w for w in workers if w and w != '---']
    except Exception:
        return []


def _rebuild_wsd_index():
    """Build a (date_str, employee) → dict index from work_schedule_daily.
    Called at session init and after every _wsd_upsert write.
    """
    idx = {}
    wsd = st.session_state.get('work_schedule_daily', pd.DataFrame())
    if not wsd.empty and 'date' in wsd.columns:
        for _, r in wsd.iterrows():
            key = (str(r['date']), str(r.get('employee', '')).strip())
            raw_manual = r.get('is_manual', False)
            idx[key] = {
                'status':     str(r.get('status', 'עובד')),
                'note':       str(r.get('note', '') or ''),
                'is_manual':  raw_manual if isinstance(raw_manual, bool)
                              else str(raw_manual).strip().lower() == 'true',
                'daily_dept': str(r.get('daily_dept', '') or ''),
                'side':       str(r.get('side', '') or ''),
            }
    st.session_state.wsd_index = idx


def _build_approved_map():
    """
    Build {employee: [(start_date, end_date, type), ...]} from approved absence_requests.
    Called at module level on every Streamlit rerun so it is always current.
    Stored in st.session_state._approved_map for O(1) access inside _derive_auto_status.
    Also called explicitly after any absence_requests mutation (approve/reject/admin-add).
    """
    result = {}
    try:
        ar = st.session_state.get('absence_requests', pd.DataFrame())
        if not ar.empty and 'status' in ar.columns:
            ar2 = ar.copy()
            ar2['status']     = ar2['status'].astype(str).str.lower()
            ar2['employee']   = ar2['employee'].astype(str).str.strip()
            ar2['start_date'] = ar2['start_date'].astype(str)
            ar2['end_date']   = ar2['end_date'].astype(str)
            ar2['type']       = ar2['type'].astype(str).str.strip()
            for _, r in ar2[ar2['status'] == 'approved'].iterrows():
                try:
                    sd = datetime.strptime(r['start_date'], '%Y-%m-%d').date()
                    ed = datetime.strptime(r['end_date'],   '%Y-%m-%d').date()
                except Exception:
                    continue
                _atype = _ABSENCE_TYPE_TO_STATUS.get(r['type'], r['type'])
                result.setdefault(r['employee'], []).append((sd, ed, _atype))
    except Exception:
        pass
    st.session_state._approved_map = result


def _wsd_get_status(date_str, employee, default="עובד"):
    """O(1) status lookup via pre-built wsd_index."""
    idx = st.session_state.get('wsd_index')
    if idx is None:
        _rebuild_wsd_index()
        idx = st.session_state.wsd_index
    return idx.get((str(date_str), str(employee).strip()), {}).get('status', default)

def _wsd_is_manual(date_str, employee):
    """O(1) is_manual lookup via pre-built wsd_index."""
    idx = st.session_state.get('wsd_index')
    if idx is None:
        _rebuild_wsd_index()
        idx = st.session_state.wsd_index
    return idx.get((str(date_str), str(employee).strip()), {}).get('is_manual', False)

def _wsd_upsert(date_str, employee, daily_dept, status, is_manual=True, note="", side=""):
    """Insert/update one row in work_schedule_daily; persist to sheet."""
    wsd = st.session_state.work_schedule_daily.copy()
    if wsd.empty or 'date' not in wsd.columns:
        wsd = pd.DataFrame(columns=['date','employee','daily_dept','status','note','is_manual','side'])
    if 'side' not in wsd.columns:
        wsd['side'] = ''
    emp_n = str(employee).strip()
    is_manual_str = str(is_manual)   # ArrowString columns reject bool; store "True"/"False"
    m = (wsd['date'].astype(str) == date_str) & (wsd['employee'].astype(str).str.strip() == emp_n)
    if m.any():
        wsd.loc[m, 'status']     = status
        wsd.loc[m, 'is_manual']  = is_manual_str
        wsd.loc[m, 'daily_dept'] = daily_dept
        wsd.loc[m, 'note']       = note  # always update — empty string clears the note
        # side is sticky: an empty side (e.g. an in-grid status click) keeps the
        # existing value so a transfer's side isn't wiped by routine edits.
        if side:
            wsd.loc[m, 'side'] = side
    else:
        wsd = pd.concat([wsd, pd.DataFrame([{
            'date': date_str, 'employee': emp_n, 'daily_dept': daily_dept,
            'status': status, 'note': note, 'is_manual': is_manual_str, 'side': side,
        }])], ignore_index=True)
    st.session_state.work_schedule_daily = wsd
    _rebuild_wsd_index()   # keep O(1) index in sync
    # Fire-and-forget write: UI updates instantly; gspread write happens in
    # a background thread so the click doesn't freeze waiting for Google Sheets.
    _save_async("work_schedule_daily", wsd.copy())


def _wsd_delete(date_str, employee):
    """Remove one row from work_schedule_daily; persist. Used to undo a temporary transfer."""
    wsd = st.session_state.work_schedule_daily.copy()
    if wsd.empty or 'date' not in wsd.columns:
        return
    emp_n = str(employee).strip()
    m = (wsd['date'].astype(str) == date_str) & (wsd['employee'].astype(str).str.strip() == emp_n)
    if m.any():
        wsd = wsd[~m].reset_index(drop=True)
        st.session_state.work_schedule_daily = wsd
        _rebuild_wsd_index()
        _save_async("work_schedule_daily", wsd.copy())


def _incoming_transfers(dept, year_month):
    """
    Day-specific temporary transfers INTO `dept` for the given month.
    A transfer = a manual WSD row whose daily_dept == dept, on a date in this month,
    for an employee who is NOT a regular dept_rotation member of this dept.
    Returns {employee: {'days': {day_ints}, 'side': '<side>'}}.
    """
    result = {}
    idx = st.session_state.get('wsd_index')
    if idx is None:
        _rebuild_wsd_index()
        idx = st.session_state.wsd_index
    # Regular members of this dept (excluded — their manual rows aren't transfers)
    members = set()
    try:
        dr = st.session_state.dept_rotation
        if not dr.empty and 'employee' in dr.columns:
            dmask = ((dr['year_month'].astype(str) == year_month) &
                     (dr['daily_dept'].astype(str) == dept))
            members = set(dr[dmask]['employee'].astype(str).str.strip().tolist())
    except Exception:
        pass
    for (date_str, emp), entry in idx.items():
        if not entry.get('is_manual'):
            continue
        if str(entry.get('daily_dept', '') or '') != dept:
            continue
        if not str(date_str).startswith(year_month):
            continue
        if emp in members:
            continue
        try:
            day_int = int(str(date_str).split('-')[2])
        except Exception:
            continue
        rec = result.setdefault(emp, {'days': set(), 'side': ''})
        rec['days'].add(day_int)
        if entry.get('side'):
            rec['side'] = entry['side']
    return result


# ── Dept resolution single source of truth (Gantt-first) ───────────────────
def _emp_dept_for_date(employee, date_str_or_obj) -> str:
    """
    Return the Gantt-canonical daily_dept for (employee, the year-month of date).
    Used wherever `dept_at_request` is written/displayed so the same logical dept
    renders consistently (e.g. 'שיקום גריאטרי א'' instead of generic 'שיקום').
    Fallback to staff.dept only when no dept_rotation row exists for that month.
    """
    try:
        emp_n = str(employee).strip()
        if not emp_n:
            return ""
        if hasattr(date_str_or_obj, 'strftime'):
            ym = date_str_or_obj.strftime('%Y-%m')
        else:
            ym = str(date_str_or_obj)[:7]
        dr = st.session_state.get('dept_rotation', pd.DataFrame())
        if not dr.empty and 'employee' in dr.columns:
            mask = ((dr['year_month'].astype(str) == ym) &
                    (dr['employee'].astype(str).str.strip() == emp_n))
            if mask.any():
                dd = str(dr[mask].iloc[0].get('daily_dept', '') or '').strip()
                if dd:
                    return dd
        # Fallback: legacy staff.dept (generic 'שיקום' / 'פנימית גריאטרית')
        sf = st.session_state.get('staff', pd.DataFrame())
        if not sf.empty and 'name' in sf.columns and 'dept' in sf.columns:
            srow = sf[sf['name'].astype(str).str.strip() == emp_n]
            if not srow.empty:
                return str(srow.iloc[0].get('dept', '') or '').strip()
    except Exception:
        pass
    return ""


def _emp_night_dept(employee, year_month) -> str:
    """
    Return the NIGHT-SHIFT dept ('שיקום' or 'פנימית גריאטרית') for an employee
    in the given year-month. Gantt-first, with intern-default fallback.

      • dept_rotation.daily_dept contains 'שיקום' → 'שיקום'
      • dept_rotation.daily_dept == 'פנימית גריאטרית' → 'פנימית גריאטרית'
      • no rotation row + role == 'מתמחה'           → 'פנימית גריאטרית'
      • no rotation row + any other role            → staff.dept (legacy fallback)
    """
    try:
        emp_n = str(employee).strip()
        if not emp_n:
            return ""
        ym = str(year_month)
        dr = st.session_state.get('dept_rotation', pd.DataFrame())
        if not dr.empty and 'employee' in dr.columns:
            mask = ((dr['year_month'].astype(str) == ym) &
                    (dr['employee'].astype(str).str.strip() == emp_n))
            if mask.any():
                dd = str(dr[mask].iloc[0].get('daily_dept', '') or '').strip()
                if 'שיקום' in dd:
                    return 'שיקום'
                if dd == 'פנימית גריאטרית':
                    return 'פנימית גריאטרית'
        # No Gantt row → intern default to פנימית; others fall back to staff.dept
        sf = st.session_state.get('staff', pd.DataFrame())
        if not sf.empty and 'name' in sf.columns:
            srow = sf[sf['name'].astype(str).str.strip() == emp_n]
            if not srow.empty:
                role = str(srow.iloc[0].get('type', '') or '').strip()
                if role == 'מתמחה':
                    return 'פנימית גריאטרית'
                if 'dept' in sf.columns:
                    return str(srow.iloc[0].get('dept', '') or '').strip()
    except Exception:
        pass
    return ""


def _absence_conflicts(employee, dept, start_date, end_date,
                       absence_df=None, exclude_id=None):
    """
    Return a list of conflict dicts for OTHER employees whose APPROVED absence
    overlaps [start_date, end_date] AND whose normalized dept matches.

    `dept` is compared *as the Gantt-canonical value* — use the value returned
    by `_emp_dept_for_date(employee, start_date)` on the caller side, and we
    re-normalize the OTHER employees' dept here so e.g. a stale 'שיקום' row and
    a fresh 'שיקום גריאטרי א'' row are treated as the same logical dept ONLY
    when both resolve to the same Gantt value.

    Each item: {employee, start, end, type, overlap_start, overlap_end}.
    `exclude_id` skips one absence_requests row by id (used on self-edit).
    """
    out = []
    try:
        emp_n = str(employee).strip()
        dept_n = str(dept or '').strip()
        _filter_by_dept = bool(dept_n)
        # Coerce inputs to date
        def _to_date(v):
            if hasattr(v, 'strftime') and not isinstance(v, str):
                return v if isinstance(v, date) else v.date()
            return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
        sd = _to_date(start_date); ed = _to_date(end_date)
        if ed < sd:
            sd, ed = ed, sd
        df = absence_df if absence_df is not None else st.session_state.get(
            'absence_requests', pd.DataFrame())
        if df.empty or 'status' not in df.columns:
            return out
        df2 = df.copy()
        df2['_status'] = df2['status'].astype(str).str.lower()
        df2['_emp']    = df2['employee'].astype(str).str.strip()
        df2 = df2[(df2['_status'] == 'approved') & (df2['_emp'] != emp_n)]
        for _, r in df2.iterrows():
            try:
                r_sd = _to_date(r['start_date'])
                r_ed = _to_date(r['end_date'])
            except Exception:
                continue
            if r_ed < r_sd:
                r_sd, r_ed = r_ed, r_sd
            ov_s = max(sd, r_sd); ov_e = min(ed, r_ed)
            if ov_s > ov_e:
                continue
            if exclude_id is not None and str(r.get('id', '')) == str(exclude_id):
                continue
            if _filter_by_dept:
                other_dept = str(r.get('dept_at_request', '') or '').strip()
                if not other_dept:
                    other_dept = _emp_dept_for_date(r['_emp'], r_sd)
                if other_dept and str(other_dept).strip() != dept_n:
                    continue
            out.append({
                'employee': r['_emp'],
                'start':    r_sd,
                'end':      r_ed,
                'type':     str(r.get('type', '') or ''),
                'overlap_start': ov_s,
                'overlap_end':   ov_e,
            })
    except Exception:
        pass
    return out


def _format_absence_conflict_warning(conflicts) -> str:
    """Render a Hebrew warning string for one or more absence conflicts."""
    if not conflicts:
        return ""
    parts = []
    for c in conflicts:
        if c['overlap_start'] == c['overlap_end']:
            rng = c['overlap_start'].strftime('%d/%m')
        else:
            rng = (f"{c['overlap_start'].strftime('%d/%m')}–"
                   f"{c['overlap_end'].strftime('%d/%m')}")
        parts.append(f"{c['employee']} ({rng})")
    return "⚠️ חופש כבר אושר מאותה מחלקה בתאריכים אלו: " + " · ".join(parts)


def _format_absence_conflict_question(conflicts) -> str:
    """Approver-facing variant: warning + explicit yes/no question."""
    if not conflicts:
        return ""
    return (_format_absence_conflict_warning(conflicts)
            + "  \nהאם לאשר בקשת היעדרות בכל זאת?")


# Status cycle for in-grid editing (clicks rotate through these)
_GRID_STATUS_CYCLE = {
    "":             "עובד",    # empty (manager-unplanted) → first click plants עובד
    "עובד":         "חופש",
    "חופש":         "202",
    "202":          "אחרי תורנות",
    "אחרי תורנות":  "אחר",
    "אחר":          "",          # clears back to empty (un-plant)
    "תורנות":       "חופש",   # clicking overrides auto-derived night-shift status
}
_GRID_STATUS_LABEL_SHORT = {
    "":             "·",   # plant-me dot for empty
    "עובד":         "ע",
    "חופש":         "ח",
    "202":          "202",
    "אחרי תורנות":  "א",
    "אחר":          "+",
    "תורנות":       "ת",
}
# Maps absence_requests.type → valid work_schedule_daily status.
# "חופש עתידי" is a request type, not a display status; normalize to "חופש".
# "היעדרות אחרת" normalizes to "אחר" (the grid's catch-all status).
_ABSENCE_TYPE_TO_STATUS = {
    "חופש עתידי":   "חופש",
    "היעדרות אחרת": "אחר",
}

def _make_initials(name: str) -> str:
    """Return first letter of each word: 'סלאמה קאסם' → 'ס ק'"""
    parts = str(name).strip().split()
    return " ".join(p[0] for p in parts if p)
_GRID_STATUS_PFX = {
    "":             "wsdcell_e",   # empty/unplanted manager cell
    "עובד":         "wsdcell_w",
    "חופש":         "wsdcell_h",
    "202":          "wsdcell_2",
    "אחרי תורנות":  "wsdcell_p",
    "אחר":          "wsdcell_a",
    "תורנות":       "wsdcell_t",
}

def _wsd_get_note(date_str, employee):
    """O(1) note lookup via pre-built wsd_index."""
    idx = st.session_state.get('wsd_index')
    if idx is None:
        _rebuild_wsd_index()
        idx = st.session_state.wsd_index
    return idx.get((str(date_str), str(employee).strip()), {}).get('note', '')

_FRIDAY_SHIFT_DEPTS = {
    "שישי בוקר - שיקום (1)", "שישי בוקר - שיקום (2)",
    "שישי בוקר - פנימית (1)", "שישי בוקר - פנימית (2)",
}

_HEB_DAY_TO_IDX = {"א": 0, "ב": 1, "ג": 2, "ד": 3, "ה": 4, "ו": 5, "ש": 6}

# ── Role-specific dept-grid cycles ────────────────────────────────────────────
# מנהל מחלקה: toggle empty ↔ עובד only (they don't take vacation through the grid)
# רופא בכיר: skip 202 (which is comp-day-off for מתמחה weekend night-shifts only)
# Everyone else: full cycle (uses _GRID_STATUS_CYCLE)
_GRID_STATUS_CYCLE_MANAGER = {"": "עובד", "עובד": ""}
_GRID_STATUS_CYCLE_SENIOR = {
    "":             "עובד",
    "עובד":         "חופש",
    "חופש":         "אחרי תורנות",   # skip 202
    "אחרי תורנות":  "אחר",
    "אחר":          "",
    "תורנות":       "חופש",
}

def _cycle_for_role(emp: str) -> dict:
    """Pick the click cycle for the dept grid based on the employee's staff.type."""
    try:
        sf = st.session_state.staff
        row = sf[sf['name'].astype(str).str.strip() == str(emp).strip()]
        if not row.empty:
            t = str(row.iloc[0].get('type', '')).strip()
            if t == 'מנהל מחלקה':
                return _GRID_STATUS_CYCLE_MANAGER
            if t == 'רופא בכיר':
                return _GRID_STATUS_CYCLE_SENIOR
    except Exception:
        pass
    return _GRID_STATUS_CYCLE

def _manual_statuses_for_role(emp: str) -> list:
    """Pick the popover selectbox options based on the employee's staff.type."""
    try:
        sf = st.session_state.staff
        row = sf[sf['name'].astype(str).str.strip() == str(emp).strip()]
        if not row.empty:
            t = str(row.iloc[0].get('type', '')).strip()
            if t == 'מנהל מחלקה':
                return ["", "עובד"]
            if t == 'רופא בכיר':
                return ["עובד", "חופש", "אחרי תורנות", "אחר"]
    except Exception:
        pass
    return _MANUAL_STATUSES

# Friday shift → daily dept mapping (for deriving who works which daily dept on Fridays)
_DAILY_DEPT_TO_FRIDAY_SHIFTS = {
    "פנימית גריאטרית":    ["שישי בוקר - פנימית (1)", "שישי בוקר - פנימית (2)"],
    "שיקום גריאטרי א'":   ["שישי בוקר - שיקום (1)"],
    "שיקום גריאטרי ב'":   ["שישי בוקר - שיקום (2)"],
}

# Daily dept → night-shift dept name (for תורנ/ית display row)
_DAILY_DEPT_TO_NIGHT_DEPT = {
    "פנימית גריאטרית":    "פנימית גריאטרית",
    "שיקום גריאטרי א'":   "שיקום",
    "שיקום גריאטרי ב'":   "שיקום",
}

def _derive_auto_status(date_str, employee, daily_dept=None):
    """
    SINGLE SOURCE OF TRUTH for day-schedule status.
    Priority order:
      0a. Saturday → "חופש"
      0b. Friday + daily_dept known → check שישי בוקר assignment
      1.  is_manual=True override in WSD → return stored status (human decision)
      2.  Approved absence in absence_requests covers date → return type
      3.  Recurring weekly absence (staff.recurring_absent_days) → "חופש"
      4.  Night shift today → "תורנות"
      5.  Night shift yesterday → "אחרי תורנות"
      else → "עובד"
    """
    try:
        emp = str(employee).strip()
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
        wd_idx = (date_obj.weekday() + 1) % 7  # Sun=0 … Sat=6

        # 0c. manual_schedule_only: never auto-planted, blank unless explicit manual override.
        #     מנהל מחלקה: auto-planted as worker in managed depts, blank elsewhere; manual wins.
        try:
            sf = st.session_state.staff
            _erow = sf[sf['name'].astype(str).str.strip() == emp]
            if not _erow.empty:
                _er = _erow.iloc[0]
                _mso_raw = _er.get('manual_schedule_only', '')
                _manual_only = _mso_raw if isinstance(_mso_raw, bool) else str(_mso_raw).strip().lower() == 'true'
                if _manual_only:
                    _idx = st.session_state.get('wsd_index', {})
                    _ent = _idx.get((date_str, emp), {})
                    if _ent.get('is_manual'):
                        return _ent.get('status', '')
                    return ""
                if str(_er.get('type', '')).strip() == 'מנהל מחלקה':
                    _idx = st.session_state.get('wsd_index', {})
                    _ent = _idx.get((date_str, emp), {})
                    if _ent.get('is_manual'):
                        return _ent.get('status', '')
                    _managed = _parse_manage_depts(_er.get('manage_depts', ''))
                    if daily_dept and str(daily_dept).strip() not in _managed:
                        return ""   # not their dept → blank
                    # else: fall through to normal auto logic → planted as עובד
        except Exception:
            pass

        # 0a. Saturday: department closed
        if wd_idx == 6:
            return "חופש"

        # 0b. Friday: working employees come from שישי בוקר assignments
        if wd_idx == 5 and daily_dept:
            fri_shifts = _DAILY_DEPT_TO_FRIDAY_SHIFTS.get(str(daily_dept), [])
            if fri_shifts:
                try:
                    sch = st.session_state.schedule
                    if not sch.empty and 'date' in sch.columns:
                        fri_mask = (
                            (sch['date'].astype(str) == date_str) &
                            (sch['dept'].astype(str).isin(fri_shifts))
                        )
                        fri_emps = sch[fri_mask]['employee'].astype(str).str.strip().tolist()
                        fri_emps = [e for e in fri_emps if e and e != '---']
                        return "עובד" if emp in fri_emps else "חופש"
                except Exception:
                    pass

        # 1. Manual override in WSD (is_manual=True) — human decision, beats all auto logic.
        #    Dept-aware: a manual row whose daily_dept differs from the queried dept means the
        #    employee was transferred elsewhere that day → blank here (pulled out of home dept).
        try:
            idx = st.session_state.get('wsd_index', {})
            entry = idx.get((date_str, emp), {})
            if entry.get('is_manual'):
                ent_dept = str(entry.get('daily_dept', '') or '')
                if daily_dept and ent_dept and ent_dept != str(daily_dept):
                    return ""
                return entry['status']
        except Exception:
            pass

        # 2. Approved absence covers this date
        try:
            approved_map = st.session_state.get('_approved_map', {})
            for sd, ed, atype in approved_map.get(emp, []):
                if sd <= date_obj <= ed:
                    atype = _ABSENCE_TYPE_TO_STATUS.get(str(atype).strip(), str(atype).strip())
                    return atype if atype else "חופש"
        except Exception:
            pass

        # 3. Recurring weekly absence
        try:
            sf = st.session_state.staff
            if not sf.empty and 'recurring_absent_days' in sf.columns:
                emp_row = sf[sf['name'].astype(str).str.strip() == emp]
                if not emp_row.empty:
                    rec_raw = str(emp_row.iloc[0].get('recurring_absent_days', '') or '').strip()
                    if rec_raw:
                        for tok in rec_raw.split(','):
                            if _HEB_DAY_TO_IDX.get(tok.strip()) == wd_idx:
                                return "חופש"
        except Exception:
            pass

        # 4 & 5. Night shift schedule
        try:
            sch = st.session_state.schedule
            if not sch.empty and 'date' in sch.columns:
                night_sch = sch[~sch['dept'].astype(str).isin(_FRIDAY_SHIFT_DEPTS)]
                names = night_sch['employee'].astype(str).str.strip()
                dates = night_sch['date'].astype(str)
                if ((dates == date_str) & (names == emp)).any():
                    return "תורנות"
                prev = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
                if ((dates == prev) & (names == emp)).any():
                    return "אחרי תורנות"
        except Exception:
            pass
    except Exception:
        pass
    return "עובד"


def _get_night_duty(date_str, daily_dept):
    """Return the name of the on-call (night-shift) employee for a given date and daily dept."""
    night_dept = _DAILY_DEPT_TO_NIGHT_DEPT.get(str(daily_dept))
    if not night_dept:
        return None
    try:
        sch = st.session_state.schedule
        if sch.empty or 'date' not in sch.columns:
            return None
        mask = (sch['date'].astype(str) == date_str) & (sch['dept'].astype(str) == night_dept)
        rows = sch[mask]
        if rows.empty:
            return None
        emp = str(rows.iloc[0]['employee']).strip()
        return emp if emp and emp != '---' else None
    except Exception:
        return None

def _night_shift_dept_label(prev_date_str: str, emp: str) -> str:
    """Return 'פנג"ר', 'שיקום', or '' — the dept of the employee's night shift on prev_date_str."""
    try:
        sch = st.session_state.schedule
        if sch.empty or 'date' not in sch.columns:
            return ''
        night_sch = sch[~sch['dept'].astype(str).isin(_FRIDAY_SHIFT_DEPTS)]
        mask = (
            (night_sch['date'].astype(str) == prev_date_str) &
            (night_sch['employee'].astype(str).str.strip() == emp)
        )
        rows = night_sch[mask]
        if rows.empty:
            return ''
        dept = str(rows.iloc[0].get('dept', '')).strip()
        if 'פנימית' in dept:
            return 'פנג"ר'
        if 'שיקום' in dept:
            return 'שיקום'
        return dept
    except Exception:
        return ''


_MANUAL_STATUSES = ["עובד", "חופש", "202", "אחרי תורנות", "אחר"]
# ── Inclusive display labels for role types (stored values unchanged) ─────────
_TYPE_DISPLAY = {
    'מתמחה':      'מתמחה',
    'תורן חוץ':   'תורנ/ית חוץ',
    'מנהל/ת':     'מנהל/ת',
    'רופא בכיר':  'רופא/ה בכירה',
    'מנהל מחלקה': 'מנהל/ת מחלקה',
}


def _render_export_buttons(dept_name, year_month, view_month, key_ns, user_name=""):
    """
    Render export buttons for a dept:
      1. 📥 Excel (batched format) — download_button
      2. 🔗 פתח ב-Sheets — export to Google Sheets + open in browser
    """
    import streamlit.components.v1 as _components
    c1, c2 = st.columns(2)

    # 1. Batched Excel download
    with c1:
        xl_bytes, xl_fname = _export_dept_grid_batched(dept_name, year_month, view_month)
        if xl_bytes:
            st.download_button(
                "📥 הורד Excel",
                data=xl_bytes,
                file_name=xl_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"xl_dl_{key_ns}",
                use_container_width=True,
            )
        else:
            st.button("📥 Excel", key=f"xl_dl_{key_ns}_na", disabled=True,
                      use_container_width=True)

    # 2. Google Sheets export — auto-open in browser
    with c2:
        if st.button("🔗 פתח ב-Sheets", key=f"exp_new_{key_ns}",
                     use_container_width=True):
            ok, url, err = _export_dept_to_new_gsheet(dept_name, year_month, view_month)
            if ok:
                _components.html(
                    f'<script>window.open("{url}", "_blank");</script>',
                    height=0)
                st.success(f"✅ גיליון נוצר!")
                st.markdown(f"[🔗 פתח גיליון]({url})")
            else:
                st.error(f"שגיאה: {err}")

def _render_dept_grid(dept_name, year_month, view_month, key_ns,
                      employees=None, max_days=None,
                      readonly=False, highlight_user=None,
                      allow_temp_add=False, side_groups=None,
                      transfer_days=None, year=None):
    """
    Render a dept × month grid.

    readonly=False (default): editable — buttons cycle status, popover notes.
    readonly=True            : same visual layout but cells are styled <div>s
                               (no clicking, no notes popover). Used in the
                               employee's סידור יומי view.
    highlight_user           : if set, that employee's name gets ▶ + bold.
    side_groups              : פנימית only — dict {'ורוד':[emps], 'כחול':[emps]}.
                               When set, each week's employee rows are grouped
                               under a coloured side label (per-week interleave).
    transfer_days            : dict {employee: {day_ints}} for day-specific
                               temporary transfers. Such an employee is active
                               ONLY on the listed days; all other day cells in
                               their row render as a muted locked '—'.
    """
    # Inline gradient styles for readonly mode (must match the CSS keyed by prefix)
    _RO_CELL_STYLES = {
        "wsdcell_w": "background:linear-gradient(135deg,#6ee7b7,#059669);color:#fff;",
        "wsdcell_h": "background:linear-gradient(135deg,#bfdbfe,#3b82f6);color:#fff;",
        "wsdcell_2": "background:linear-gradient(135deg,#fef08a,#ca8a04);color:#713f12;",
        "wsdcell_p": "background:linear-gradient(135deg,#fed7aa,#ea580c);color:#fff;",
        "wsdcell_a": "background:linear-gradient(135deg,#e2e8f0,#64748b);color:#fff;",
        "wsdcell_t": "background:linear-gradient(135deg,#ddd6fe,#7c3aed);color:#fff;",
        "wsdcell_e": "background:#f8fafc;color:#cbd5e1;border:1.5px dashed #cbd5e1;",
    }

    def _ro_cell(lbl, pfx, cur_note=""):
        """Render a readonly-mode cell that visually matches the editable button."""
        sty = _RO_CELL_STYLES.get(pfx, "background:#f1f5f9;color:#475569;")
        _tt = f" title=\"{str(cur_note).replace(chr(34), '')}\"" if cur_note else ""
        st.markdown(
            f"<div{_tt} style='{sty}min-height:36px;border-radius:8px;"
            f"text-align:center;font-weight:700;font-size:1rem;display:flex;"
            f"align-items:center;justify-content:center;padding:2px'>"
            f"{lbl}</div>",
            unsafe_allow_html=True)

    def _emp_name_html(emp, font_size="0.82rem"):
        """Render an employee-name cell, with ▶ + bold when emp matches highlight_user."""
        is_me = (highlight_user is not None and
                 str(emp).strip() == str(highlight_user).strip())
        if is_me:
            return (f"<div style='font-weight:800;font-size:{font_size};"
                    f"color:#0f172a;padding:4px 8px;background:#fef3c7;"
                    f"border-radius:6px;border:1px solid #fde68a;"
                    f"text-align:right'>▶ {emp}</div>")
        return (f"<div style='font-weight:600;font-size:{font_size};color:#0f172a;"
                f"padding:4px 8px;background:white;border-radius:6px;"
                f"border:1px solid #e2e8f0;text-align:right'>{emp}</div>")

    # Feature 4: derive year from caller (or year_month prefix), defaulting to 2026.
    if year is None:
        try:
            year = int(str(year_month).split('-')[0])
        except Exception:
            year = 2026
    num_days = calendar.monthrange(year, view_month)[1]
    seg1 = list(range(1, 11))
    seg2 = list(range(11, 21))
    seg3 = list(range(21, num_days + 1))
    halves = [s for s in [seg1, seg2, seg3] if s]

    # Derive employees from dept_rotation if not provided
    if employees is None:
        dr = st.session_state.dept_rotation
        if dr.empty or 'employee' not in dr.columns:
            employees = []
        else:
            mask = ((dr['year_month'].astype(str) == year_month) &
                    (dr['daily_dept'].astype(str) == dept_name))
            employees = dr[mask]['employee'].astype(str).str.strip().tolist()

    # Merge day-specific temporary transfers (persisted in work_schedule_daily).
    # When transfer_days is supplied by the caller (פנימית path), it has already
    # folded transfers into `employees` + `side_groups`; don't double-merge here.
    if transfer_days is None:
        _inc = _incoming_transfers(dept_name, year_month)
        if _inc:
            _emp_set = {str(e).strip() for e in (employees or [])}
            employees = (employees or []) + [e for e in _inc if e not in _emp_set]
            transfer_days = {e: rec['days'] for e, rec in _inc.items()}

    if not employees:
        st.info(f"אין עובדים משובצים ל-{dept_name} בחודש זה.")
        if not allow_temp_add:
            return

    # ── Side grouping + day-specific transfer helpers ───────────────────
    _transfer_days = transfer_days or {}

    def _week_row_items():
        """Flat row sequence for the week loop: ('__label__', side) | ('emp', name).
        Lets us inject per-side labels without re-indenting the row body."""
        items = []
        if side_groups:
            for _s in PNIM_SIDES:
                _g = side_groups.get(_s, [])
                if _g:
                    items.append(('__label__', _s))
                    items.extend(('emp', e) for e in _g)
        else:
            items.extend(('emp', e) for e in employees)
        return items

    def _side_label_row(side):
        color = PNIM_COLORS.get(side, "#f1f5f9")
        icon  = PNIM_ICONS.get(side, "")
        st.markdown(
            f"<div style='background:{color};padding:5px 12px;border-radius:7px;"
            f"margin:6px 0 2px;font-weight:700;font-size:0.9rem;color:#1e293b;"
            f"text-align:right'>{icon}&nbsp;צד {side}</div>",
            unsafe_allow_html=True)

    def _is_locked_day(emp, d):
        """True when emp is a day-specific transfer and d is NOT an active day."""
        days = _transfer_days.get(str(emp).strip())
        return days is not None and d not in days


    # ── Shared setup ────────────────────────────────────────────────────
    _WD_HDRS  = ["ש'", "ו'", "ה'", "ד'", "ג'", "ב'", "א'"]
    _WD_IS_WK = [True, True, False, False, False, False, False]  # Sat, Fri = weekend
    cal_weeks = [list(reversed(w))
                 for w in calendar.Calendar(firstweekday=6).monthdayscalendar(year, view_month)]

    # Pre-load konenut on-call doctor per day for this dept/month
    _KONENUT_DEPT_SET = {"פנימית גריאטרית", "שיקום גריאטרי א'", "שיקום גריאטרי ב'"}
    _konenut_day_map = {}
    if dept_name in _KONENUT_DEPT_SET:
        try:
            _kdf = st.session_state.get('konenut', pd.DataFrame())
            if not _kdf.empty and 'date' in _kdf.columns:
                _km = _kdf[_kdf['date'].str.startswith(f"{year}-{view_month:02d}")]
                for _, _kr in _km.iterrows():
                    _ds  = str(_kr['date'])
                    _r1  = str(_kr.get('rehab_dr1', '') or '').strip()
                    _r2  = str(_kr.get('rehab_dr2', '') or '').strip()
                    if dept_name == "פנימית גריאטרית":
                        _konenut_day_map[_ds] = str(_kr.get('pnim_dr', '') or '').strip()
                    elif dept_name == "שיקום גריאטרי א'":
                        _konenut_day_map[_ds] = _r1
                    else:  # שיקום גריאטרי ב'
                        _konenut_day_map[_ds] = _r2 if _r2 else _r1
        except Exception:
            pass

    # Mobile auto-detection: set the toggle's session-state key directly
    # so it flips ON as soon as JS resolves the device type (render 2).
    # Once detected, we stop overriding so manual user flips are preserved.
    _tog_key      = f"mob_toggle_{key_ns}"
    _detected_key = f"mob_det_{key_ns}"
    if not st.session_state.get(_detected_key, False):
        _dev = st.session_state.get('analytics_device_type', 'unknown')
        _vp  = 0
        try:
            _vp = int(st.session_state.get('analytics_vp_width', 0) or 0)
        except (ValueError, TypeError):
            pass
        _ua      = str(st.session_state.get('analytics_ua', '') or '')
        _mob_ua  = any(x in _ua for x in ['Android', 'iPhone', 'iPad', 'Mobile', 'webOS'])
        _mob_any = (_dev == 'mobile' or
                    st.session_state.get('mobile_detected_persistent', False) or
                    _mob_ua or (0 < _vp < 768))
        if _dev != 'unknown':
            # Device confirmed — lock the toggle and stop re-detecting
            st.session_state[_tog_key]      = _mob_any
            st.session_state[_detected_key] = True
        # If still 'unknown': leave toggle at its default (False) until next render
    is_mobile = st.toggle("📱 תצוגת מובייל", value=False, key=_tog_key)

    # ── CSS: colour every status button by key prefix (both modes) ───────
    _kn = key_ns
    st.markdown(f"""<style>
div[class*="st-key-wsdcell_w_{_kn}"] button {{
    background:linear-gradient(135deg,#6ee7b7,#059669)!important;
    color:white!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_h_{_kn}"] button {{
    background:linear-gradient(135deg,#bfdbfe,#3b82f6)!important;
    color:white!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_2_{_kn}"] button {{
    background:linear-gradient(135deg,#fef08a,#ca8a04)!important;
    color:#713f12!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_p_{_kn}"] button {{
    background:linear-gradient(135deg,#fed7aa,#ea580c)!important;
    color:white!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_a_{_kn}"] button {{
    background:linear-gradient(135deg,#e2e8f0,#64748b)!important;
    color:white!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_t_{_kn}"] button {{
    background:linear-gradient(135deg,#ddd6fe,#7c3aed)!important;
    color:white!important;border:none!important;
    min-height:36px!important;font-weight:700!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
div[class*="st-key-wsdcell_e_{_kn}"] button {{
    background:#f8fafc!important;
    color:#cbd5e1!important;border:1.5px dashed #cbd5e1!important;
    min-height:36px!important;font-weight:400!important;
    font-size:1rem!important;border-radius:8px!important;padding:2px!important;}}
</style>""", unsafe_allow_html=True)

    if is_mobile:
        # ── MOBILE: exactly 7 cols (proven to work — same as constraint calendar)
        # Employee name shown as full-width RTL banner above each button row.
        # One st.button() per cell, no popover inside cell.

        # Day-name header (7 cols)
        hdr_cols = st.columns(7)
        for _ci, (_h, _wk) in enumerate(zip(_WD_HDRS, _WD_IS_WK)):
            _hbg = "#fef2f2" if _wk else "#f1f5f9"
            _hfg = "#b91c1c" if _wk else "#334155"
            hdr_cols[_ci].markdown(
                f"<div style='background:{_hbg};font-weight:700;text-align:center;"
                f"padding:5px 1px;border-radius:6px;font-size:0.85rem;color:{_hfg}'>{_h}</div>",
                unsafe_allow_html=True)

        for week_rtl in cal_weeks:
            if all(d == 0 for d in week_rtl):
                continue
            st.markdown("<div style='height:3px'></div>", unsafe_allow_html=True)

            # Date-number row (7 cols)
            num_cols = st.columns(7)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    num_cols[_ci].write("")
                else:
                    _nbg = "#fef2f2" if _WD_IS_WK[_ci] else "#f8fafc"
                    _nfg = "#b91c1c" if _WD_IS_WK[_ci] else "#334155"
                    num_cols[_ci].markdown(
                        f"<div style='background:{_nbg};font-weight:700;text-align:center;"
                        f"padding:3px 1px;border-radius:6px;font-size:0.75rem;color:{_nfg}'>{d}</div>",
                        unsafe_allow_html=True)

            # Employee rows: full-name banner (full width) + 7-col button row
            for _kind, emp in _week_row_items():
                if _kind == '__label__':
                    _side_label_row(emp)
                    continue
                _is_me_m = (highlight_user is not None and
                            str(emp).strip() == str(highlight_user).strip())
                _name_bg = "#fef3c7" if _is_me_m else "#f1f5f9"
                _name_bd = "#fde68a" if _is_me_m else "#e2e8f0"
                _name_pre = "▶ " if _is_me_m else ""
                _name_fw  = 800 if _is_me_m else 700
                st.markdown(
                    f"<div style='font-weight:{_name_fw};font-size:0.82rem;color:#0f172a;"
                    f"padding:2px 8px;background:{_name_bg};border-radius:5px;"
                    f"border:1px solid {_name_bd};text-align:right;margin:2px 0 1px'>"
                    f"{_name_pre}{emp}</div>",
                    unsafe_allow_html=True)
                row_cols = st.columns(7)
                for _ci, d in enumerate(week_rtl):
                    with row_cols[_ci]:
                        if d == 0:
                            st.write("")
                            continue
                        if _is_locked_day(emp, d):
                            _ro_cell("—", "wsdcell_e")
                            continue
                        date_str   = f"{year}-{view_month:02d}-{d:02d}"
                        cur_status = _derive_auto_status(date_str, emp, daily_dept=dept_name)
                        cur_note   = _wsd_get_note(date_str, emp)
                        lbl        = _GRID_STATUS_LABEL_SHORT.get(cur_status, "?")
                        pfx        = _GRID_STATUS_PFX.get(cur_status, "wsdcell_w")
                        cell_key   = f"{pfx}_{key_ns}_{emp}_{d}"
                        if _WD_IS_WK[_ci] and cur_status == "חופש" and not _wsd_is_manual(date_str, emp):
                            st.markdown(
                                "<div style='height:36px;background:#f8fafc;"
                                "border-radius:8px;border:1px solid #e2e8f0'></div>",
                                unsafe_allow_html=True)
                            continue
                        if readonly:
                            _ro_cell(lbl, pfx, cur_note)
                            continue
                        if st.button(lbl, key=cell_key, use_container_width=True,
                                     help=cur_note if cur_note else None):
                            _wsd_upsert(date_str, emp, dept_name,
                                        _cycle_for_role(emp).get(cur_status, ""),
                                        is_manual=True, note=cur_note)
                            st.rerun()

            # Night-duty row (full-width label + 7 cols)
            # Initials shown; clicking opens popover with full name
            st.markdown(
                "<div style='background:#ede9fe;font-weight:700;text-align:right;"
                "padding:2px 8px;border-radius:5px;font-size:0.75rem;"
                "color:#5b21b6;margin:2px 0 1px'>🌙 תורנ/ית</div>",
                unsafe_allow_html=True)
            nd_cols = st.columns(7)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    nd_cols[_ci].write("")
                    continue
                nd = _get_night_duty(f"{year}-{view_month:02d}-{d:02d}", dept_name)
                with nd_cols[_ci]:
                    if nd:
                        try:
                            with st.popover(_make_initials(nd),
                                            key=f"ndpop_{key_ns}_{d}_{_ci}",
                                            use_container_width=True):
                                st.markdown(f"**{nd}**")
                        except Exception:
                            _nbg = "#fef2f2" if _WD_IS_WK[_ci] else "#f5f3ff"
                            st.markdown(
                                f"<div style='background:{_nbg};text-align:center;"
                                f"padding:4px 1px;border-radius:6px;font-size:0.68rem;"
                                f"color:#6d28d9'>{_make_initials(nd)}</div>",
                                unsafe_allow_html=True)
                    else:
                        st.markdown(
                            "<div style='text-align:center;padding:4px 1px;"
                            "font-size:0.68rem;color:#94a3b8'>—</div>",
                            unsafe_allow_html=True)

            # Friday-morning row (full-width label + 7 cols)
            # Initials shown; clicking opens popover with full name(s)
            st.markdown(
                "<div style='background:#fef9c3;font-weight:700;text-align:right;"
                "padding:2px 8px;border-radius:5px;font-size:0.75rem;"
                "color:#854d0e;margin:2px 0 1px'>☀️ שישי בוקר</div>",
                unsafe_allow_html=True)
            fri_cols = st.columns(7)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    fri_cols[_ci].write("")
                    continue
                date_str_fri = f"{year}-{view_month:02d}-{d:02d}"
                with fri_cols[_ci]:
                    if _ci == 1:
                        fw = _get_fri_shift_workers(date_str_fri, dept_name)
                        if fw:
                            fw_initials = " / ".join(_make_initials(n) for n in fw)
                            try:
                                with st.popover(fw_initials,
                                                key=f"fwpop_{key_ns}_{d}_{_ci}",
                                                use_container_width=True):
                                    for _fn in fw:
                                        st.markdown(f"**{_fn}**")
                            except Exception:
                                st.markdown(
                                    f"<div style='background:#fef3c7;text-align:center;"
                                    f"padding:4px 1px;border-radius:6px;font-size:0.68rem;"
                                    f"color:#92400e'>{fw_initials}</div>",
                                    unsafe_allow_html=True)
                        else:
                            st.markdown(
                                "<div style='text-align:center;padding:4px 1px;"
                                "font-size:0.68rem;color:#94a3b8'>—</div>",
                                unsafe_allow_html=True)
                    else:
                        st.write("")

            # Konenut (on-call doctor) row — mobile
            if _konenut_day_map:
                st.markdown(
                    "<div style='background:#dcfce7;font-weight:700;text-align:right;"
                    "padding:2px 8px;border-radius:5px;font-size:0.75rem;"
                    "color:#166534;margin:2px 0 1px'>🔔 כונן/ית</div>",
                    unsafe_allow_html=True)
                _kn_mob_cols = st.columns(7)
                for _ci, d in enumerate(week_rtl):
                    if d == 0:
                        _kn_mob_cols[_ci].write("")
                        continue
                    _kds  = f"{year}-{view_month:02d}-{d:02d}"
                    _kval = _konenut_day_map.get(_kds, '')
                    with _kn_mob_cols[_ci]:
                        st.markdown(
                            f"<div style='background:#{'bbf7d0' if _kval else 'f0fdf4'};"
                            f"text-align:center;padding:4px 1px;border-radius:6px;"
                            f"font-size:0.68rem;color:#166534;"
                            f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>"
                            f"{_kval or '—'}</div>",
                            unsafe_allow_html=True)

    else:
        # ════════════════════════════════════════════════════════════════
        # DESKTOP — 8 cols: full name on right, button + popover per cell
        # ════════════════════════════════════════════════════════════════
        _COL_W_D = [2] * 7 + [3]   # day cols narrow, name col wider

        # Day-name header
        hdr_cols = st.columns(_COL_W_D)
        for _ci, (_h, _wk) in enumerate(zip(_WD_HDRS, _WD_IS_WK)):
            _hbg = "#fef2f2" if _wk else "#f1f5f9"
            _hfg = "#b91c1c" if _wk else "#334155"
            hdr_cols[_ci].markdown(
                f"<div style='background:{_hbg};font-weight:700;text-align:center;"
                f"padding:6px 2px;border-radius:6px;font-size:0.8rem;color:{_hfg}'>{_h}</div>",
                unsafe_allow_html=True)
        hdr_cols[7].markdown(
            "<div style='background:#f1f5f9;font-weight:700;text-align:center;"
            "padding:6px 2px;border-radius:6px;font-size:0.75rem;color:#334155'>עובד/ת</div>",
            unsafe_allow_html=True)

        for week_rtl in cal_weeks:
            if all(d == 0 for d in week_rtl):
                continue
            st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

            # Date-number row
            num_cols = st.columns(_COL_W_D)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    num_cols[_ci].write("")
                else:
                    _nbg = "#fef2f2" if _WD_IS_WK[_ci] else "#f8fafc"
                    _nfg = "#b91c1c" if _WD_IS_WK[_ci] else "#334155"
                    num_cols[_ci].markdown(
                        f"<div style='background:{_nbg};font-weight:700;text-align:center;"
                        f"padding:4px 2px;border-radius:6px;font-size:0.75rem;color:{_nfg}'>{d}</div>",
                        unsafe_allow_html=True)
            num_cols[7].write("")

            # Employee rows
            for _kind, emp in _week_row_items():
                if _kind == '__label__':
                    _side_label_row(emp)
                    continue
                row_cols = st.columns(_COL_W_D)
                row_cols[7].markdown(_emp_name_html(emp), unsafe_allow_html=True)
                for _ci, d in enumerate(week_rtl):
                    with row_cols[_ci]:
                        if d == 0:
                            st.write("")
                            continue
                        if _is_locked_day(emp, d):
                            _ro_cell("—", "wsdcell_e")
                            continue
                        date_str   = f"{year}-{view_month:02d}-{d:02d}"
                        cur_status = _derive_auto_status(date_str, emp, daily_dept=dept_name)
                        cur_note   = _wsd_get_note(date_str, emp)
                        lbl        = _GRID_STATUS_LABEL_SHORT.get(cur_status, "?")
                        pfx        = _GRID_STATUS_PFX.get(cur_status, "wsdcell_w")
                        cell_key   = f"{pfx}_{key_ns}_{emp}_{d}"
                        _is_wknd   = _WD_IS_WK[_ci]
                        if _is_wknd and cur_status == "חופש" and not _wsd_is_manual(date_str, emp):
                            st.markdown(
                                "<div style='height:32px;background:#f8fafc;"
                                "border-radius:6px;border:1px solid #e2e8f0'></div>",
                                unsafe_allow_html=True)
                            continue
                        if readonly:
                            _ro_cell(lbl, pfx, cur_note)
                            continue
                        if st.button(lbl, key=cell_key, use_container_width=True,
                                     help=cur_note if cur_note else None):
                            _wsd_upsert(date_str, emp, dept_name,
                                        _cycle_for_role(emp).get(cur_status, ""),
                                        is_manual=True, note=cur_note)
                            st.rerun()
                        try:
                            with st.popover("💬" if cur_note else "+",
                                            key=f"notepop_{key_ns}_{emp}_{d}",
                                            use_container_width=True):
                                st.markdown(f"<b>{emp}</b> — {d}/{view_month}",
                                            unsafe_allow_html=True)
                                _pop_opts = _manual_statuses_for_role(emp)
                                ps = st.selectbox(
                                    "סטטוס:", _pop_opts,
                                    index=_pop_opts.index(cur_status)
                                           if cur_status in _pop_opts else 0,
                                    key=f"popst_{key_ns}_{emp}_{d}")
                                new_note = st.text_input(
                                    "הערה:", value=cur_note,
                                    placeholder="הוסף הערה",
                                    key=f"popnote_{key_ns}_{emp}_{d}")
                                if st.button("💾 שמור",
                                             key=f"popsave_{key_ns}_{emp}_{d}",
                                             use_container_width=True):
                                    _wsd_upsert(date_str, emp, dept_name, ps,
                                                is_manual=True, note=new_note.strip())
                                    st.rerun()
                        except Exception:
                            pass

            # Night-duty row
            nd_cols = st.columns(_COL_W_D)
            nd_cols[7].markdown(
                "<div style='background:#ede9fe;font-weight:700;text-align:center;"
                "padding:6px 2px;border-radius:6px;font-size:0.75rem;color:#5b21b6'>🌙 תורנ/ית</div>",
                unsafe_allow_html=True)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    nd_cols[_ci].write("")
                    continue
                nd = _get_night_duty(f"{year}-{view_month:02d}-{d:02d}", dept_name)
                _nbg = "#fef2f2" if _WD_IS_WK[_ci] else "#f5f3ff"
                nd_cols[_ci].markdown(
                    f"<div style='background:{_nbg};text-align:center;padding:4px 2px;"
                    f"border-radius:6px;font-size:0.7rem;color:#6d28d9;"
                    f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>"
                    f"{nd or '—'}</div>",
                    unsafe_allow_html=True)

            # Friday-morning row (col 1 = Friday)
            fri_cols = st.columns(_COL_W_D)
            fri_cols[7].markdown(
                "<div style='background:#fef9c3;font-weight:700;text-align:center;"
                "padding:6px 2px;border-radius:6px;font-size:0.75rem;color:#854d0e'>☀️ שישי בוקר</div>",
                unsafe_allow_html=True)
            for _ci, d in enumerate(week_rtl):
                if d == 0:
                    fri_cols[_ci].write("")
                    continue
                date_str_fri = f"{year}-{view_month:02d}-{d:02d}"
                if _ci == 1:
                    fw = _get_fri_shift_workers(date_str_fri, dept_name)
                    cell_txt    = " / ".join(fw) if fw else "—"
                    cell_bg_fri = "#fef3c7"
                else:
                    cell_txt    = ""
                    cell_bg_fri = "#f8fafc"
                fri_cols[_ci].markdown(
                    f"<div style='background:{cell_bg_fri};text-align:center;padding:4px 2px;"
                    f"border-radius:6px;font-size:0.68rem;color:#92400e;"
                    f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>"
                    f"{cell_txt}</div>",
                    unsafe_allow_html=True)

            # Konenut (on-call doctor) row — green
            if _konenut_day_map:
                _krow_cols = st.columns(_COL_W_D)
                _krow_cols[7].markdown(
                    "<div style='background:#dcfce7;font-weight:700;text-align:center;"
                    "padding:6px 2px;border-radius:6px;font-size:0.75rem;color:#166534'>"
                    "🔔 כונן/ית</div>",
                    unsafe_allow_html=True)
                for _ci, d in enumerate(week_rtl):
                    if d == 0:
                        _krow_cols[_ci].write("")
                        continue
                    _kds  = f"{year}-{view_month:02d}-{d:02d}"
                    _kval = _konenut_day_map.get(_kds, '')
                    _kbg  = "#bbf7d0" if _kval else ("#f0fdf4" if not _WD_IS_WK[_ci] else "#f0fdf4")
                    _krow_cols[_ci].markdown(
                        f"<div style='background:{_kbg};min-height:36px;border-radius:6px;"
                        f"text-align:center;font-weight:600;font-size:0.72rem;color:#166534;"
                        f"display:flex;align-items:center;justify-content:center;padding:2px;"
                        f"white-space:nowrap;overflow:hidden;text-overflow:ellipsis'>"
                        f"{_kval or '—'}</div>",
                        unsafe_allow_html=True)

    # ── Day-specific temporary transfer form (persisted to work_schedule_daily) ──
    if allow_temp_add and not readonly:
        _is_pnim = (dept_name == PNIM_DEPT)
        st.divider()
        st.markdown(
            "<div style='background:#f0fdf4;border:1px solid #86efac;border-radius:8px;"
            "padding:10px 14px;text-align:right;margin-bottom:6px'>"
            "<b>➕ העברה זמנית — הוסף עובד/ת ממחלקה אחרת ליום מסוים</b></div>",
            unsafe_allow_html=True)

        # Employees already permanently rostered to this dept this month — not addable.
        # Exception: in sides mode (פנימית ורוד/כחול), don't exclude same-dept employees
        # so admins can transfer between sides; the target side is chosen explicitly below.
        _roster_set = set()
        if side_groups is None:
            try:
                _dr = st.session_state.dept_rotation
                if not _dr.empty and 'employee' in _dr.columns:
                    _rmask = ((_dr['year_month'].astype(str) == year_month) &
                              (_dr['daily_dept'].astype(str) == dept_name))
                    _roster_set = set(_dr[_rmask]['employee'].astype(str).str.strip().tolist())
            except Exception:
                pass
        _all_staff_names = []
        if not st.session_state.staff.empty and 'name' in st.session_state.staff.columns:
            _all_staff_names = (st.session_state.staff['name']
                                .astype(str).str.strip().tolist())
        _available_to_add = [n for n in _all_staff_names
                             if n and n not in _roster_set]

        _month_start = date(year, view_month, 1)
        _month_end   = date(year, view_month, num_days)
        if _is_pnim:
            _ta_c1, _ta_c2, _ta_c3, _ta_c4 = st.columns([3, 2, 2, 1])
        else:
            _ta_c1, _ta_c2, _ta_c4 = st.columns([4, 2, 1])
            _ta_c3 = None
        with _ta_c1:
            _sel_ta = st.selectbox(
                "עובד/ת:", ["—"] + _available_to_add,
                key=f"temp_emp_sel_{key_ns}")
        with _ta_c2:
            _sel_date = st.date_input(
                "תאריך:", value=_month_start,
                min_value=_month_start, max_value=_month_end,
                format="DD/MM/YYYY", key=f"temp_emp_date_{key_ns}")
        _sel_side = ""
        if _is_pnim and _ta_c3 is not None:
            with _ta_c3:
                _sel_side = st.selectbox(
                    "צד:", PNIM_SIDES, key=f"temp_emp_side_{key_ns}")
        with _ta_c4:
            st.write("")
            if st.button("➕ הוסף", key=f"temp_emp_add_{key_ns}",
                         use_container_width=True):
                if _sel_ta and _sel_ta != "—" and _sel_date:
                    _ds = _sel_date.strftime("%Y-%m-%d")
                    _wsd_upsert(_ds, _sel_ta, dept_name, "עובד",
                                is_manual=True, side=_sel_side)
                    st.rerun()

        # List existing day-specific transfers into this dept this month
        _existing = _incoming_transfers(dept_name, year_month)
        if _existing:
            st.markdown("**העברות זמניות בחודש זה:**")
            for _te in sorted(_existing.keys()):
                _rec = _existing[_te]
                for _dday in sorted(_rec['days']):
                    _ds = f"{year}-{view_month:02d}-{_dday:02d}"
                    _side_txt = (f" · {PNIM_ICONS.get(_rec['side'], '')} {_rec['side']}"
                                 if _is_pnim and _rec.get('side') else "")
                    _tr1, _tr2 = st.columns([5, 1])
                    _tr1.markdown(f"• {_te} — {_dday}/{view_month}{_side_txt}")
                    if _tr2.button("✕ הסר", key=f"temp_emp_rm_{key_ns}_{_te}_{_dday}",
                                   use_container_width=True):
                        _wsd_delete(_ds, _te)
                        st.rerun()


def _render_pnim_sided(year_month, view_month, key_ns,
                       employees, readonly=False,
                       highlight_user=None, allow_temp_add=False,
                       year=None):
    """Render פנימית גריאטרית with per-week 🌸 ורוד / 🔵 כחול interleaving.

    Employees are split by their `side` field in dept_rotation, then sorted
    within each group: רופא בכיר first, then מתמחה alphabetically. Day-specific
    temporary transfers are merged into their stored side. Employees with no
    side assigned are surfaced via a non-blocking caption (no grid section).
    A single `_render_dept_grid` call drives the whole layout (one mobile toggle
    + one CSS block) with the weekly side labels emitted inside the week loop.
    """
    # Build side map from dept_rotation
    dr = st.session_state.dept_rotation
    _side_map: dict = {}
    if not dr.empty and 'employee' in dr.columns and 'side' in dr.columns:
        _cur_dr = dr[dr['year_month'].astype(str) == year_month]
        _side_map = (
            _cur_dr.set_index('employee')['side']
            .fillna('').astype(str).str.strip()
            .to_dict()
        )

    staff_df = st.session_state.staff

    def _role_order(emp_name):
        row = staff_df[staff_df['name'].astype(str).str.strip() == str(emp_name).strip()]
        t = str(row.iloc[0]['type']).strip() if not row.empty else ''
        return 0 if t == 'רופא בכיר' else 1

    def _sorted_group(emps):
        return sorted(emps, key=lambda e: (_role_order(e), str(e)))

    emp_strs = [str(e).strip() for e in employees]

    # Roster members per side
    pink = _sorted_group([e for e in emp_strs if _side_map.get(e, '') == 'ורוד'])
    blue = _sorted_group([e for e in emp_strs if _side_map.get(e, '') == 'כחול'])
    unassigned = _sorted_group([e for e in emp_strs if _side_map.get(e, '') == ''])

    # Day-specific temporary transfers INTO פנימית — fold into the matching side
    _inc = _incoming_transfers(PNIM_DEPT, year_month)
    _transfer_days = {}
    for _emp, _rec in _inc.items():
        _transfer_days[_emp] = _rec['days']
        _tside = _rec.get('side') or ''
        if _tside == 'כחול':
            if _emp not in blue:
                blue.append(_emp)
        else:  # default any sideless / ורוד transfer to ורוד
            if _emp not in pink:
                pink.append(_emp)

    if unassigned:
        st.caption(f"⚠️ {len(unassigned)} עובדים ללא צד — שבץ/י צד בגאנט: "
                   + ", ".join(unassigned))

    side_groups = {'ורוד': pink, 'כחול': blue}
    _render_dept_grid(
        PNIM_DEPT, year_month, view_month, key_ns,
        employees=pink + blue,
        side_groups=side_groups,
        transfer_days=_transfer_days,
        readonly=readonly,
        highlight_user=highlight_user,
        allow_temp_add=allow_temp_add,
        year=year,
    )


# ── Feature 3: approved-absence Gantt view ────────────────────────────────
_ABSENCE_GANTT_COLORS = {
    "חופש":          ("#dbeafe", "#1e3a8a"),
    "חופש עתידי":    ("#dbeafe", "#1e3a8a"),
    "202":           ("#fef08a", "#854d0e"),
    "היעדרות אחרת":  ("#e2e8f0", "#475569"),
}

def _render_absence_gantt(year: int, month: int, dept_filter=None):
    """
    Render approved future absences as a Gantt-style grid:
      • Rows grouped by Gantt-canonical dept (via _emp_dept_for_date), then employee.
      • Columns = days 1..N of (year, month).
      • Cells coloured by absence type; overlap days (≥2 employees same dept) get
        a red inset border + tooltip listing the other absent employee(s).
    `dept_filter` is an optional iterable of normalized dept names; when given,
    only employees whose canonical dept is in this set are shown.
    """
    num_days = calendar.monthrange(year, month)[1]
    days = list(range(1, num_days + 1))
    month_start = date(year, month, 1)
    month_end   = date(year, month, num_days)

    ar = st.session_state.get('absence_requests', pd.DataFrame())
    if ar.empty or 'status' not in ar.columns:
        st.info("אין בקשות שאושרו בחודש זה.")
        return

    ar2 = ar.copy()
    ar2['_status'] = ar2['status'].astype(str).str.lower()
    ar2 = ar2[ar2['_status'] == 'approved']
    if ar2.empty:
        st.info("אין בקשות שאושרו בחודש זה.")
        return

    def _to_date(v):
        try:
            if hasattr(v, 'strftime') and not isinstance(v, str):
                return v if isinstance(v, date) else v.date()
            return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
        except Exception:
            return None

    # Collect (emp, dept_canon, sd, ed, type) intersecting this month
    rows = []
    for _, r in ar2.iterrows():
        sd = _to_date(r.get('start_date'))
        ed = _to_date(r.get('end_date'))
        if not sd or not ed:
            continue
        if ed < sd:
            sd, ed = ed, sd
        # Intersect with [month_start, month_end]
        i_s = max(sd, month_start); i_e = min(ed, month_end)
        if i_s > i_e:
            continue
        emp = str(r.get('employee', '')).strip()
        dept_canon = _emp_dept_for_date(emp, sd) or '—'
        if dept_filter and dept_canon not in dept_filter:
            continue
        rows.append({
            'emp':  emp,
            'dept': dept_canon,
            'sd':   i_s,
            'ed':   i_e,
            'type': str(r.get('type', '') or '').strip() or 'חופש',
        })

    if not rows:
        st.info("אין בקשות שאושרו בחודש זה.")
        return

    # Group by (dept, employee) → list of (day_int, type)
    by_dept = {}
    for row in rows:
        cur = row['sd']
        while cur <= row['ed']:
            d = cur.day
            by_dept.setdefault(row['dept'], {}).setdefault(row['emp'], []).append((d, row['type']))
            cur = cur + timedelta(days=1)

    # Compute conflict days per (dept, day) — ≥2 distinct employees absent that day
    conflicts = {}   # {(dept, day_int): [emp1, emp2, ...]}
    for dept_n, emps in by_dept.items():
        day_to_emps = {}
        for emp, items in emps.items():
            for d, _t in items:
                day_to_emps.setdefault(d, set()).add(emp)
        for d, eset in day_to_emps.items():
            if len(eset) >= 2:
                conflicts[(dept_n, d)] = sorted(eset)

    # Legend
    st.markdown(
        "<div style='direction:rtl;font-size:0.78rem;color:#475569;margin:6px 0 4px;line-height:2'>"
        "<span style='background:#dbeafe;border-radius:4px;padding:1px 10px'>&nbsp;</span> חופש &nbsp;|&nbsp; "
        "<span style='background:#fef08a;border-radius:4px;padding:1px 10px'>&nbsp;</span> 202 &nbsp;|&nbsp; "
        "<span style='background:#e2e8f0;border-radius:4px;padding:1px 10px'>&nbsp;</span> היעדרות אחרת &nbsp;|&nbsp; "
        "<span style='background:#dbeafe;border:2px solid #ef4444;border-radius:4px;padding:1px 10px'>&nbsp;</span> "
        "חפיפה במחלקה</div>",
        unsafe_allow_html=True)

    # RTL day-number header. Use ~26 cols: 1 emp-name col + 25 days. For 28-31
    # day months, split into segments to keep cells legible.
    WD = ["א", "ב", "ג", "ד", "ה", "ו", "ש"]

    def _wd_label(d):
        wi = (date(year, month, d).weekday() + 1) % 7
        return WD[wi]

    def _cell_for(emp, d, items_by_day):
        if d in items_by_day:
            typ = items_by_day[d]
            bg, fg = _ABSENCE_GANTT_COLORS.get(typ, ("#cbd5e1", "#0f172a"))
            return bg, fg, typ[:3] or "·"
        return "#f8fafc", "#cbd5e1", ""

    # RTL ordering: name column is LAST (rightmost in display), days laid out
    # in reverse so day 1 sits at the right (next to the name) and day N at the
    # left. Mirrors the _render_dept_grid desktop layout.
    days_rtl = list(reversed(days))
    name_col = num_days   # last column index

    for dept_n in sorted(by_dept.keys()):
        st.markdown(
            f"<div style='background:#f1f5f9;padding:6px 12px;border-radius:7px;"
            f"margin:14px 0 4px;font-weight:700;font-size:0.95rem;color:#1e293b;"
            f"text-align:right'>📌 {dept_n}</div>",
            unsafe_allow_html=True)

        col_widths = [1] * num_days + [3]   # name col LAST + wider
        hdr_cols = st.columns(col_widths)
        hdr_cols[name_col].markdown(
            "<div style='font-weight:700;font-size:0.78rem;text-align:right;"
            "padding:4px 8px'>עובד/ת / יום</div>",
            unsafe_allow_html=True)
        for i, d in enumerate(days_rtl):
            wi = (date(year, month, d).weekday() + 1) % 7
            is_wk = wi in (5, 6)
            hbg = "#fef2f2" if is_wk else "#eef2ff"
            hfg = "#b91c1c" if is_wk else "#3730a3"
            hdr_cols[i].markdown(
                f"<div style='background:{hbg};color:{hfg};text-align:center;"
                f"padding:2px 0;border-radius:5px;font-size:0.65rem;line-height:1.2'>"
                f"{d}<br>{_wd_label(d)}</div>",
                unsafe_allow_html=True)

        for emp in sorted(by_dept[dept_n].keys()):
            items_by_day = {d: t for d, t in by_dept[dept_n][emp]}
            row_cols = st.columns(col_widths)
            row_cols[name_col].markdown(
                f"<div style='font-weight:600;font-size:0.78rem;color:#0f172a;"
                f"padding:4px 8px;background:white;border-radius:5px;"
                f"border:1px solid #e2e8f0;text-align:right'>{emp}</div>",
                unsafe_allow_html=True)
            for i, d in enumerate(days_rtl):
                bg, fg, lbl = _cell_for(emp, d, items_by_day)
                # Conflict day overlay
                _ttl = ""
                _border = "1px solid #e2e8f0"
                if (dept_n, d) in conflicts and emp in items_by_day:
                    _others = [n for n in conflicts[(dept_n, d)] if n != emp]
                    if _others:
                        _border = "2px solid #ef4444"
                        _ttl = f" title='חפיפה: {', '.join(_others)}'"
                row_cols[i].markdown(
                    f"<div{_ttl} style='background:{bg};color:{fg};text-align:center;"
                    f"padding:3px 0;border-radius:5px;border:{_border};"
                    f"font-size:0.7rem;font-weight:700;min-height:22px;line-height:1.2'>"
                    f"{lbl}</div>",
                    unsafe_allow_html=True)


def log_event(event_type, detail_1='', detail_2=''):
    """
    Append a single analytics event row to the 'analytics_log' Google Sheet.
    Never clears the sheet — append-only. Entire body wrapped in bare except so it
    NEVER crashes the main app. Dropped events are acceptable.
    """
    try:
        now = datetime.now()
        row = [
            str(uuid.uuid4()),                                          # event_id
            str(st.session_state.get('analytics_session_id', '')),     # session_id
            now.strftime('%Y-%m-%d %H:%M:%S'),                         # timestamp
            str(st.session_state.get('user_name', '')),                # user_name
            str(st.session_state.get('user_role', '')),                # user_role
            event_type,                                                 # event_type
            str(detail_1),                                             # detail_1
            str(detail_2),                                             # detail_2
            str(st.session_state.get('analytics_device_type', 'unknown')),  # device_type
            str(st.session_state.get('analytics_ua', ''))[:200],      # ua_string
            str(st.session_state.get('analytics_vp_width', '')),      # viewport_width
            str(st.session_state.get('active_month_int', '')),        # active_month
            str(now.day),                                              # day_of_month
        ]
        gc = get_gspread_client()
        url = st.secrets["connections"]["gsheets"]["spreadsheet"]
        sh = gc.open_by_url(url)
        # Create sheet with header on first call if missing
        try:
            ws = sh.worksheet('analytics_log')
        except gspread.exceptions.WorksheetNotFound:
            ws = sh.add_worksheet('analytics_log', rows=5000, cols=13)
            ws.append_row([
                'event_id', 'session_id', 'timestamp', 'user_name', 'user_role',
                'event_type', 'detail_1', 'detail_2', 'device_type', 'ua_string',
                'viewport_width', 'active_month', 'day_of_month'
            ], value_input_option='USER_ENTERED')
        ws.append_row(row, value_input_option='USER_ENTERED')
    except Exception:
        pass  # Analytics failure must never affect the main app


import threading as _threading

def _log_async(event_type, detail_1='', detail_2=''):
    """Fire-and-forget wrapper for log_event — never blocks the UI."""
    _threading.Thread(
        target=log_event,
        args=(event_type, detail_1, detail_2),
        daemon=True,
    ).start()


import queue as _queue
import sys as _sys
import traceback as _traceback

# Single background queue + drainer thread for all async writes.
# Coalesces consecutive writes to the same worksheet so rapid clicks
# don't race (the latest snapshot always wins). This fixes the bug
# where two daemon threads writing the same sheet could complete out
# of order and silently drop the newer click.
_save_queue: "_queue.Queue[tuple[str, object, bool]]" = _queue.Queue()
_save_worker_started = False
_save_worker_lock = _threading.Lock()

def _save_worker():
    while True:
        worksheet, df, is_rtl = _save_queue.get()   # blocks
        # Coalesce: drain any queued writes for the same worksheet
        # and keep only the latest snapshot — older ones are stale.
        try:
            held = []
            while not _save_queue.empty():
                try:
                    w2, d2, r2 = _save_queue.get_nowait()
                except _queue.Empty:
                    break
                if w2 == worksheet:
                    df, is_rtl = d2, r2   # newer snapshot wins
                else:
                    held.append((w2, d2, r2))
            # Re-queue items for other worksheets (preserve order)
            for item in held:
                _save_queue.put(item)
            # Perform the (coalesced) write
            try:
                save_to_db(worksheet, df, is_rtl=is_rtl)
            except Exception:
                # Surface to the Streamlit server log so silent loss
                # doesn't go completely unnoticed.
                _traceback.print_exc(file=_sys.stderr)
        finally:
            _save_queue.task_done()


def _save_async(worksheet_name, df, is_rtl=False):
    """Queue a non-blocking write to Google Sheets — never freezes the UI.

    Use ONLY when the calling code has just updated the in-memory
    st.session_state copy. The DataFrame is .copy()ed so background
    mutations don't corrupt the write.

    Writes are serialized through a single daemon worker thread that
    coalesces consecutive writes to the same worksheet (latest wins).
    This prevents the race where two daemon threads could complete out
    of order and silently drop the newer click's data.
    """
    global _save_worker_started
    if not _save_worker_started:
        with _save_worker_lock:
            if not _save_worker_started:
                _threading.Thread(target=_save_worker, daemon=True).start()
                _save_worker_started = True
    _save_queue.put((worksheet_name, df, is_rtl))


def save_to_db(worksheet_name, df, is_rtl=False):
    """
    Write df to a Google Sheets worksheet. Retries up to 4 times on 429 rate-limit errors.

    SAFE write order — prevents data loss on transient API failures:
      1. ws.update()  ← write new data first (old data still present in stale rows below)
      2. ws.resize()  ← trim stale rows/cols after a successful write (soft failure OK)
    Never calls ws.clear() before writing, so a mid-write failure never leaves the sheet empty.
    """
    url = st.secrets["connections"]["gsheets"]["spreadsheet"]
    df_str = df.astype(str).replace('nan', '', regex=True).replace('None', '', regex=True)
    data = [df_str.columns.tolist()] + df_str.values.tolist()
    n_rows = len(data)
    n_cols = max((len(r) for r in data), default=1)

    last_err = None
    for attempt in range(4):
        try:
            gc = get_gspread_client()
            sh = gc.open_by_url(url)
            try:
                ws = sh.worksheet(worksheet_name)
            except gspread.exceptions.WorksheetNotFound:
                ws = sh.add_worksheet(title=worksheet_name,
                                      rows=max(n_rows, 1), cols=max(n_cols, 1))

            # Write first — data is safe even if the trim below fails
            ws.update(range_name='A1', values=data, value_input_option='RAW')

            # Trim stale rows/cols that exceed the new data (soft failure is OK)
            try:
                ws.resize(rows=max(n_rows, 1), cols=max(n_cols, 1))
            except Exception:
                pass

            if is_rtl:
                try:
                    sh.batch_update({"requests": [{
                        "updateSheetProperties": {
                            "properties": {"sheetId": ws.id, "rightToLeft": True,
                                           "gridProperties": {"columnCount": 8}},
                            "fields": "rightToLeft,gridProperties.columnCount"
                        }
                    }]})
                    ws.format('A1:H1', {'textFormat': {'bold': True}, 'horizontalAlignment': 'CENTER'})
                    ws.format('A2:H100', {'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE'})
                except Exception as e:
                    print(f"RTL/Format warning: {e}")
            return  # success

        except gspread.exceptions.APIError as e:
            last_err = e
            if e.response.status_code == 429:
                wait = 2 ** attempt  # 1s, 2s, 4s, 8s
                time.sleep(wait)
                continue
            break  # non-429 API error — don't retry
        except Exception as e:
            last_err = e
            break

    st.error(f"שגיאה קריטית בשמירה לגיליון '{worksheet_name}': {last_err}")

def init_db():
    """
    Safe, additive initialisation.
    ONLY creates worksheets that do not yet exist — NEVER overwrites or clears
    any worksheet that is already present (even if it has zero data rows).
    This prevents accidental data loss on API hiccups or cold-start empty reads.
    """
    # Required sheets and their header columns
    REQUIRED_SHEETS = {
        "staff": ['name', 'type', 'dept', 'monthly_quota', 'weekend_quota',
                  'password', 'only_home_dept', 'email', 'manage_depts',
                  'recurring_absent_days'],
        "schedule":     ['date', 'dept', 'employee', 'is_manual', 'empty_reason'],
        "requests":     ['employee', 'date', 'status'],
        "special_days": ['date', 'description', 'day_type'],
        "dept_rotation":        ['employee', 'year_month', 'daily_dept'],
        "absence_requests":     ['id', 'employee', 'start_date', 'end_date', 'type',
                                 'status', 'dept_at_request', 'manager_email',
                                 'approved_by', 'notes', 'created_at', 'responded_at'],
        "work_schedule_daily":  ['date', 'employee', 'daily_dept', 'status', 'note', 'is_manual', 'side'],
    }
    SETTINGS_DEFAULTS = [
        ('active_month',        str((date.today().replace(day=1) + timedelta(days=32)).month)),
        ('daily_active_month',  str((date.today().replace(day=1) + timedelta(days=32)).month)),
        ('daily_requests_open', 'True'),
    ]

    try:
        gc  = get_gspread_client()
        url = st.secrets["connections"]["gsheets"]["spreadsheet"]
        sh  = gc.open_by_url(url)

        existing_titles = {ws.title for ws in sh.worksheets()}
        created_any = False

        for sheet_name, columns in REQUIRED_SHEETS.items():
            if sheet_name not in existing_titles:
                # Create the worksheet with just the header row — no data rows
                empty_df = pd.DataFrame(columns=columns)
                save_to_db(sheet_name, empty_df)
                created_any = True

        # settings sheet: create if missing, else only ADD missing keys (never overwrite)
        if "settings" not in existing_titles:
            next_m = (date.today().replace(day=1) + timedelta(days=32)).month
            settings_df = pd.DataFrame(
                [{'key': k, 'value': v} for k, v in SETTINGS_DEFAULTS])
            save_to_db("settings", settings_df)
            created_any = True
        else:
            # Add any settings keys that do not yet exist
            cur = get_db_data("settings")
            existing_keys = set(cur['key'].astype(str)) if not cur.empty and 'key' in cur.columns else set()
            for key, val in SETTINGS_DEFAULTS:
                if key not in existing_keys:
                    cur = pd.concat([cur, pd.DataFrame([{'key': key, 'value': val}])],
                                    ignore_index=True)
            if set(cur['key'].astype(str)) != existing_keys:
                save_to_db("settings", cur)

        if created_any:
            pass  # silent — no st.info/success shown to the user

    except Exception as e:
        if "403" in str(e) or "Public Spreadsheet" in str(e):
            st.error("שגיאת הרשאות Google Sheets — וודא ש-Secrets הוגדרו נכון.")
        # All other errors: silent — don't block the app from loading

# --- 3. ניהול התחברות (Login) ---
def login_screen():
    st.markdown("""
        <style>
            /* Center the entire screen content for login */
            .main .block-container {
                max-width: 450px !important;
                padding-top: 5rem !important;
                background-color: white;
                border-radius: 20px;
                box-shadow: 0 10px 25px rgba(0,0,0,0.05);
                margin-top: 40px;
                padding-bottom: 2rem !important;
            }
            
            div[data-testid="stTextInput"] input {
                border: 2px solid #e2e8f0 !important;
                background-color: #f8fafc;
                border-radius: 8px;
                padding: 10px;
                transition: all 0.3s;
                color: #1e293b !important; /* Ensure input text is dark */
            }
            div[data-testid="stTextInput"] input:focus {
                border-color: #3b82f6 !important;
                background-color: #ffffff;
                box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2);
            }
        </style>
    """, unsafe_allow_html=True)

    st.markdown("<h1 style='text-align: center; color: #1e293b; margin-bottom: 24px; font-size: 1.8rem;'>מערכת סידור עבודה המערך הגריאטרי</h1>", unsafe_allow_html=True)
    
    username = st.text_input("שם משתמש:").strip()
    password = st.text_input("סיסמה:", type="password")
    
    st.markdown("<br>", unsafe_allow_html=True) # Spacing before button
    
    if st.button("כניסה", use_container_width=True):
        # Use _fetch_live so login always reads fresh data, bypassing the cache
        # (a stale cached empty-DataFrame would otherwise block every login attempt)
        staff_df = _fetch_live("staff")
        hashed_pass = hashlib.sha256(password.encode()).hexdigest()

        # Guard: 'name' column missing means the sheet could not be read at all
        # (staff_df.empty is valid — sheet may have headers but no employees yet)
        if 'name' not in staff_df.columns:
            st.error("לא ניתן לטעון את רשימת הצוות — נסה שוב")
            st.stop()

        # בדיקה אם המשתמש קיים (strip whitespace from sheet names to avoid invisible-character mismatches)
        staff_df['name'] = staff_df['name'].astype(str).str.strip()
        user_match = staff_df[staff_df['name'] == username]
        
        if not user_match.empty:
            user = user_match[user_match['password'] == hashed_pass]
            if not user.empty:
                st.session_state.logged_in = True
                st.session_state.user_name = username
                st.session_state.user_role = user.iloc[0]['type']
                st.rerun()
            else:
                st.error("סיסמה שגויה")
                log_event('login_fail', username, 'סיסמה שגויה')
        else:
            st.error("שם המשתמש לא נמצא במערכת")
            log_event('login_fail', username, 'משתמש לא קיים')

if 'logged_in' not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    login_screen()
    st.stop()
else:
    # אתחול מסד הנתונים (רק פעם אחת, אחרי כניסה — לא מוצג על מסך הכניסה)
    if 'db_initialized' not in st.session_state:
        init_db()
        st.session_state.db_initialized = True

    # --- אתחול סשן אנליטיקה (פעם אחת לכל כניסה) ---
    if 'analytics_session_id' not in st.session_state:
        st.session_state.analytics_session_id = str(uuid.uuid4())
        st.session_state.analytics_login_time = datetime.now()
        st.session_state.analytics_last_tab = None
        st.session_state.analytics_tab_enter = datetime.now()
        st.session_state.analytics_device_captured = False
        st.session_state.analytics_device_type = 'unknown'
        st.session_state.analytics_ua = ''
        st.session_state.analytics_vp_width = ''

    # --- זיהוי מכשיר ב-JS (frame 2: ה-JS מחזיר None בפריים 1, ערך בפריים 2) ---
    try:
        from streamlit_javascript import st_javascript
        _ua_cap  = st_javascript("window.navigator.userAgent", key="analytics_ua_cap")
        _vp_cap  = st_javascript("window.innerWidth",          key="analytics_vp_cap")
        if _ua_cap and isinstance(_ua_cap, str) and not st.session_state.analytics_device_captured:
            st.session_state.analytics_ua = _ua_cap
            st.session_state.analytics_vp_width = str(_vp_cap) if _vp_cap else ''
            _is_mobile = any(x in _ua_cap for x in ["Android", "iPhone", "iPad", "Mobile", "webOS"])
            _is_mobile = _is_mobile or (isinstance(_vp_cap, (int, float)) and 0 < _vp_cap < 768)
            st.session_state.analytics_device_type = 'mobile' if _is_mobile else 'desktop'
            st.session_state.analytics_device_captured = True
            _log_async('login_success', st.session_state.analytics_device_type, str(_vp_cap))
    except Exception:
        pass

    # --- הזרקת CSS ייעודי למובייל (רק למשתמש מחובר) ---
    # מונע את עיוות מסך הכניסה
    st.markdown("""
        <style>
        /* 1. Aggressive Sidebar Hiding on Mobile */
        @media (max-width: 768px) {
            section[data-testid="stSidebar"][aria-expanded="false"], 
            section[data-testid="stSidebar"][aria-hidden="true"] {
                display: none !important;
                width: 0 !important;
                flex: 0 !important;
            }
            div[data-testid="stSidebarCollapsedControl"] {
                display: none !important;
            }
            
            /* 2. Mini-Calendar Grid for Mobile (Target specific classes instead of all columns) */
            /* Wrap the calendar in a container and target that, or target the specific 7-column child count if possible */
            /* To avoid breaking Manager UI (Team/Settings/Swap), we only apply this 7-column grid to the specific Calendar container */
            div.calendar-grid-container > div[data-testid="stHorizontalBlock"],
            div[data-testid="stHorizontalBlock"]:has(> div:nth-child(7):last-child) {
                display: grid !important;
                grid-template-columns: repeat(7, 1fr) !important;
                gap: 1px !important;
                padding: 0 !important;
                width: 100% !important;
            }
            div.calendar-grid-container div[data-testid="column"],
            div[data-testid="stHorizontalBlock"]:has(> div:nth-child(7):last-child) > div[data-testid="column"] {
                width: auto !important;
                min-width: 0 !important;
                flex: 1 1 0 !important;
                padding: 1px !important;
            }
            div.calendar-grid-container div[data-testid="column"] > div,
            div[data-testid="stHorizontalBlock"]:has(> div:nth-child(7):last-child) > div[data-testid="column"] > div {
                display: flex;
                flex-direction: column;
                align-items: center;
                justify-content: center;
            }
            div.calendar-grid-container div[data-testid="column"] div[data-testid="stMarkdownContainer"] p,
            div[data-testid="stHorizontalBlock"]:has(> div:nth-child(7):last-child) div[data-testid="column"] div[data-testid="stMarkdownContainer"] p {
                 font-size: 11px !important;
                 text-align: center !important;
                 margin: 0 !important;
                 line-height: 1.2 !important;
            }
            div[data-testid="stCheckbox"] {
                margin-top: -2px !important;
                justify-content: center !important;
            } 
            /* Label hiding is removed as it hides the checkbox itself in some structures. 
               We rely on label_visibility="collapsed" in Python. */
            
            div[data-testid="stCheckbox"] div[role="checkbox"] {
                transform: scale(0.75) !important;
            }

        }
        </style>
    """, unsafe_allow_html=True)

# --- Helper Function for Validity Checks (Shared by Auto-Scheduler and Swap Tool) ---
# --- Helper Function for Validity Checks (Shared by Auto-Scheduler and Swap Tool) ---
# --- Helper Function for Validity Checks (Shared by Auto-Scheduler and Swap Tool) ---
def get_functional_day_type(date_obj, special_days_df):
    """Returns the functional day type ('רגיל', 'כמו שישי (ערב חג)', 'כמו שבת (חג)')"""
    date_str = date_obj.strftime('%Y-%m-%d') if isinstance(date_obj, date) else date_obj
    if not special_days_df.empty and 'day_type' in special_days_df.columns:
        match = special_days_df[special_days_df['date'] == str(date_str)]
        if not match.empty:
            return match.iloc[0]['day_type']
    return 'רגיל'

def is_functional_weekend(date_obj, special_days_df):
    """Returns True if the date acts as a weekend, including days functioning as Friday or Saturday."""
    if isinstance(date_obj, str):
        date_obj = datetime.strptime(date_obj, '%Y-%m-%d').date()
    # 4 = Friday, 5 = Saturday
    if date_obj.weekday() in [4, 5]:
        return True
    
    day_type = get_functional_day_type(date_obj, special_days_df)
    if day_type in ['כמו שישי (ערב חג)', 'כמו שבת (חג)']:
        return True
        
    # The user specifically requested that the day BEFORE "Like Friday" behaves like a weekend day too
    tomorrow = date_obj + timedelta(days=1)
    if get_functional_day_type(tomorrow, special_days_df) == 'כמו שישי (ערב חג)':
        return True
        
    return False

def check_assignment_validity(schedule_data, person_name, check_date, target_dept, staff_df, requests_df, ignore_quota=False, ignore_home_restrict=False, ignore_rest=False):
    """
    Checks if assigning person_name to target_dept on check_date is valid.
    schedule_data: List of dicts OR DataFrame
    ignore_quota: If True, skips the max shifts check (useful for Swaps where count doesn't increase)
    ignore_home_restrict: If True, skips "Restricted to Home Dept" check
    ignore_rest: If True, skips "Rest Violation" (nearby days) check
    """
    # Normalize to list of dicts if DataFrame
    if isinstance(schedule_data, pd.DataFrame):
        schedule_data = schedule_data.to_dict('records')
        
    if person_name in ['ADMIN', '---']:
        return False, "Invalid Person"
        
    p_row = staff_df[staff_df['name'] == person_name]
    if p_row.empty: return False, "Unknown Employee"
    person = p_row.iloc[0]
    
    # --- 1. Static Constraints (Critical - Must Fail First) ---
    # Type Check (External vs Internal)
    p_type = str(person.get('type', '')).strip()
    if 'חוץ' in p_type and 'פנימית' in target_dept:
        return False, "External cannot work Internal"

    # Home Dept Check — Feature 5: derive the night-shift dept from the Gantt
    # (`dept_rotation`) for the year-month of check_date instead of reading the
    # legacy staff.dept column. Falls back to staff.dept when no rotation row.
    if not ignore_home_restrict:
        only_home = person.get('only_home_dept', False)
        if only_home:
             target_context = target_dept
             if "שישי בוקר" in target_dept:
                 target_context = "שיקום" if "שיקום" in target_dept else "פנימית גריאטרית"
             _person_home_dept = _emp_night_dept(person_name, str(check_date)[:7]) \
                 or str(person.get('dept', '') or '')
             if _person_home_dept != 'כללי' and _person_home_dept != target_context:
                 return False, f"Restricted to Home Dept ({_person_home_dept})"

    # --- 2. Hard User Constraints ---
    req_date_str = requests_df['date'].astype(str)
    if not requests_df[(requests_df['employee'] == person_name) & (req_date_str == check_date) & (requests_df['status'] == "אילוץ")].empty:
        return False, "User Restriction (Blocked)"
        
    # --- 3. Quota Check ---
    # Default quota to 6 if missing
    try:
        max_quota = int(person.get('monthly_quota', 6))
    except:
        max_quota = 6

    if max_quota == 0:
        return False, "Quota is 0 (Inactive)"

    if not ignore_quota:
        # Filter for current month only
        # check_date is 'YYYY-MM-DD'
        target_month_prefix = check_date[:7] # 'YYYY-MM'
        
        current_shifts = len([
            s for s in schedule_data 
            if s['employee'] == person_name 
            and str(s['date']).startswith(target_month_prefix)
        ])
        
        if current_shifts >= max_quota:
             return False, f"Quota Exceeded ({current_shifts}/{max_quota})"

    # --- 4. Situational Constraints ---
    check_d_obj = datetime.strptime(check_date, '%Y-%m-%d').date()
    
    # Rest Check (2 days gap)
    if not ignore_rest:
        for offset in [-2, -1, 1, 2]:
            nearby_date = str(check_d_obj + timedelta(days=offset))
            if any(s for s in schedule_data if str(s['date']) == nearby_date and s['employee'] == person_name):
                return False, "Rest Violation (Worked nearby)"
            
    # Duplicate Check (Same Day)
    if any(s for s in schedule_data if str(s['date']) == check_date and s['employee'] == person_name):
        return False, "Already working this day"

    return True, "Valid"

def find_swap_candidates(schedule_df, requests_df, staff_df, user_name, swap_date, swap_dept, sel_month, year=2026):
    """
    For an employee wanting to swap their shift on (swap_date, swap_dept), find:
    - 'full': candidates who can cover the user's shift AND have a shift the user can take in return
    - 'partial': candidates who can cover the user's shift but no mutual shift was found

    Hard constraints enforced via check_assignment_validity (ignore_quota=True, since a swap
    doesn't change anyone's monthly count). The user's own shift is removed from the simulated
    schedule before validating, so the 2-day rest gap recalculates correctly.
    """
    user_name_n = str(user_name).strip()

    sched_minus_user = schedule_df[
        ~((schedule_df['date'].astype(str) == swap_date) &
          (schedule_df['employee'].astype(str).str.strip() == user_name_n) &
          (schedule_df['dept'] == swap_dept))
    ].copy()

    def _safe_q(v):
        try: return int(v)
        except: return 6

    active_staff = staff_df[
        (staff_df['name'].astype(str).str.strip() != user_name_n) &
        (staff_df['type'].astype(str).str.strip() != 'מנהל/ת')
    ].copy()
    active_staff = active_staff[active_staff['monthly_quota'].apply(_safe_q) > 0]

    req_dates = requests_df['date'].astype(str)
    req_emps = requests_df['employee'].astype(str).str.strip()
    wished_set = set(req_emps[(req_dates == swap_date) & (requests_df['status'] == 'בקשה')])

    # Search candidate's mutual shifts in the same month as the swap (derived from swap_date,
    # not the active month — handles swaps in any future month).
    month_prefix = swap_date[:7]
    full, partial = [], []

    for _, cand_row in active_staff.iterrows():
        cand_name = str(cand_row['name']).strip()

        ok, _ = check_assignment_validity(
            sched_minus_user, cand_name, swap_date, swap_dept,
            staff_df, requests_df,
            ignore_quota=True, ignore_home_restrict=False, ignore_rest=False
        )
        if not ok:
            continue

        _today_str = date.today().strftime('%Y-%m-%d')
        their_shifts = schedule_df[
            (schedule_df['employee'].astype(str).str.strip() == cand_name) &
            (schedule_df['date'].astype(str).str.startswith(month_prefix)) &
            (schedule_df['date'].astype(str) >= _today_str)   # exclude past shifts
        ]

        mutual = None
        for _, ts in their_shifts.iterrows():
            ts_date = str(ts['date'])
            ts_dept = ts['dept']

            sched_sim = sched_minus_user[
                ~((sched_minus_user['date'].astype(str) == ts_date) &
                  (sched_minus_user['employee'].astype(str).str.strip() == cand_name) &
                  (sched_minus_user['dept'] == ts_dept))
            ]
            sched_sim = pd.concat([sched_sim, pd.DataFrame([{
                'date': swap_date, 'dept': swap_dept, 'employee': cand_name,
                'is_manual': False, 'empty_reason': ''
            }])], ignore_index=True)

            ok_me, _ = check_assignment_validity(
                sched_sim, user_name_n, ts_date, ts_dept,
                staff_df, requests_df,
                ignore_quota=True, ignore_home_restrict=False, ignore_rest=False
            )
            if ok_me:
                mutual = {'date': ts_date, 'dept': ts_dept}
                break

        info = {
            'name': cand_name,
            'dept': str(cand_row.get('dept', '')),
            'type': str(cand_row.get('type', '')),
            'wished': cand_name in wished_set,
            'their_shift': mutual,
        }
        if mutual:
            full.append(info)
        else:
            partial.append(info)

    full.sort(key=lambda x: (not x['wished'], x['name']))
    partial.sort(key=lambda x: (not x['wished'], x['name']))

    # ── 3-way chain swap (only relevant when user's shift is in פנימית גריאטרית) ─
    # Logic: find a מתמחה currently on שיקום on swap_date who CAN work פנימית,
    # then simulate them moving to פנימית and check if any תורן חוץ can fill the
    # now-vacant שיקום slot. Result is a list of chain dicts.
    chain = []
    if swap_dept == 'פנימית גריאטרית':
        # Candidates for facilitator role: מתמחה assigned to שיקום on swap_date
        _facilitators_sched = sched_minus_user[
            (sched_minus_user['date'].astype(str) == swap_date) &
            (sched_minus_user['dept'] == 'שיקום') &
            (sched_minus_user['employee'].astype(str).str.strip() != user_name_n)
        ]
        _externals = staff_df[
            (staff_df['type'].astype(str).str.strip() == 'תורן חוץ') &
            (staff_df['name'].astype(str).str.strip() != user_name_n)
        ].copy()

        for _, _fac_row in _facilitators_sched.iterrows():
            _fac_name = str(_fac_row['employee']).strip()
            _fac_staff = staff_df[staff_df['name'].astype(str).str.strip() == _fac_name]
            if _fac_staff.empty:
                continue
            if str(_fac_staff.iloc[0].get('type', '')).strip() != 'מתמחה':
                continue  # must be מתמחה to work פנימית

            # Check if facilitator can work פנימית on swap_date
            _ok_fac, _ = check_assignment_validity(
                sched_minus_user, _fac_name, swap_date, 'פנימית גריאטרית',
                staff_df, requests_df, ignore_quota=True
            )
            if not _ok_fac:
                continue

            # Simulate: remove facilitator from שיקום, add them to פנימית
            _sched_sim = sched_minus_user[
                ~((sched_minus_user['date'].astype(str) == swap_date) &
                  (sched_minus_user['employee'].astype(str).str.strip() == _fac_name) &
                  (sched_minus_user['dept'] == 'שיקום'))
            ].copy()
            _sched_sim = pd.concat([_sched_sim, pd.DataFrame([{
                'date': swap_date, 'dept': 'פנימית גריאטרית', 'employee': _fac_name,
                'is_manual': False, 'empty_reason': ''
            }])], ignore_index=True)

            # Find a תורן חוץ who can now cover the vacant שיקום slot
            _fac_wished = _fac_name in wished_set
            for _, _ext_row in _externals.iterrows():
                _ext_name = str(_ext_row['name']).strip()
                _ok_ext, _ = check_assignment_validity(
                    _sched_sim, _ext_name, swap_date, 'שיקום',
                    staff_df, requests_df, ignore_quota=True
                )
                if _ok_ext:
                    chain.append({
                        'facilitator_name': _fac_name,
                        'facilitator_dept': str(_fac_staff.iloc[0].get('dept', '')),
                        'facilitator_wished': _fac_wished,
                        'external_name': _ext_name,
                        'external_wished': _ext_name in wished_set,
                    })
                    break  # one valid external per facilitator is enough

    chain.sort(key=lambda x: (not x['facilitator_wished'], not x['external_wished'], x['facilitator_name']))
    return {'full': full, 'partial': partial, 'chain': chain}

# טעינת נתונים ממסד הנתונים ל-Session State
if 'staff' not in st.session_state:
    st.session_state.staff = get_db_data("staff")
    # Initialize only_home_dept if missing
    if 'only_home_dept' not in st.session_state.staff.columns:
        st.session_state.staff['only_home_dept'] = False
    else:
        # Google Sheets returns booleans as strings "True"/"False".
        # astype(bool) on a non-empty string always returns True, so we parse manually.
        st.session_state.staff['only_home_dept'] = st.session_state.staff['only_home_dept'].apply(
            lambda v: v if isinstance(v, bool) else str(v).strip().lower() == 'true'
        )
    # Daily-schedule feature columns (added after initial DB creation — handle missing gracefully)
    if 'email' not in st.session_state.staff.columns:
        st.session_state.staff['email'] = ''
    if 'manage_depts' not in st.session_state.staff.columns:
        st.session_state.staff['manage_depts'] = ''
    if 'recurring_absent_days' not in st.session_state.staff.columns:
        st.session_state.staff['recurring_absent_days'] = ''
    if 'manual_schedule_only' not in st.session_state.staff.columns:
        st.session_state.staff['manual_schedule_only'] = False
    else:
        st.session_state.staff['manual_schedule_only'] = st.session_state.staff['manual_schedule_only'].apply(
            lambda v: v if isinstance(v, bool) else str(v).strip().lower() == 'true'
        )

if 'schedule' not in st.session_state:
    st.session_state.schedule = get_db_data("schedule")
if 'requests' not in st.session_state:
    st.session_state.requests = get_db_data("requests")
if 'special_days' not in st.session_state:
    try:
        st.session_state.special_days = get_db_data("special_days")
        if 'day_type' not in st.session_state.special_days.columns:
            st.session_state.special_days['day_type'] = 'רגיל'
    except:
        st.session_state.special_days = pd.DataFrame(columns=['date', 'description', 'day_type'])
        save_to_db("special_days", st.session_state.special_days)

if 'swap_requests' not in st.session_state:
    st.session_state.swap_requests = get_db_data("swap_requests")

if 'konenut' not in st.session_state:
    try:
        _kdf = get_db_data("konenut")
        if _kdf.empty or 'date' not in _kdf.columns:
            _kdf = pd.DataFrame(columns=['date', 'pnim_dr', 'rehab_dr1', 'rehab_dr2'])
    except Exception:
        _kdf = pd.DataFrame(columns=['date', 'pnim_dr', 'rehab_dr1', 'rehab_dr2'])
    for _c in ['pnim_dr', 'rehab_dr1', 'rehab_dr2']:
        if _c not in _kdf.columns:
            _kdf[_c] = ''
    _kdf['date'] = _kdf['date'].astype(str)
    for _c in ['pnim_dr', 'rehab_dr1', 'rehab_dr2']:
        _kdf[_c] = _kdf[_c].fillna('').astype(str)
    st.session_state.konenut = _kdf

# ── Daily schedule feature data (load lazily, create empty if missing) ──
if 'dept_rotation' not in st.session_state:
    df = get_db_data("dept_rotation")
    if df.empty or 'employee' not in df.columns:
        df = pd.DataFrame(columns=['employee', 'year_month', 'daily_dept'])
    st.session_state.dept_rotation = _norm_dr(df)

if 'absence_requests' not in st.session_state:
    df = get_db_data("absence_requests")
    if df.empty or 'employee' not in df.columns:
        df = pd.DataFrame(columns=[
            'id', 'employee', 'start_date', 'end_date', 'type', 'status',
            'dept_at_request', 'manager_email', 'approved_by', 'notes',
            'created_at', 'responded_at'])
    st.session_state.absence_requests = df

if 'work_schedule_daily' not in st.session_state:
    df = get_db_data("work_schedule_daily")
    if df.empty or 'date' not in df.columns:
        df = pd.DataFrame(columns=[
            'date', 'employee', 'daily_dept', 'status', 'note', 'is_manual', 'side'])
    else:
        # Scope to 2026 only — prevents multi-year data from bloating session state
        df = df[df['date'].astype(str).str.startswith('2026')]
    if 'side' not in df.columns:
        df['side'] = ''
    st.session_state.work_schedule_daily = df
    _rebuild_wsd_index()   # build O(1) lookup index at session start
elif 'wsd_index' not in st.session_state:
    _rebuild_wsd_index()   # rebuild if session restored without index

# כלי שיבוץ ידני
if 'manual_date' not in st.session_state:
    st.session_state.manual_date = date(2026, 1, 1)
if 'manual_dept' not in st.session_state:
    st.session_state.manual_dept = "שיקום"
if 'manual_emp' not in st.session_state:
    st.session_state.manual_emp = st.session_state.staff['name'].iloc[0] if not st.session_state.staff.empty else ""

# --- 3. לוגיקת שיבוץ עם אבחון ---
def render_modern_calendar(year, month, default_constraint_days, default_wish_days,
                           special_days_df=None, key_prefix='cal', show_validation=True):
    """
    Renders a unified modern calendar for constraint/wish selection using colored buttons.
    Returns (constraint_day_nums, wish_day_nums) as lists of int day numbers.
    """
    const_key = f"{key_prefix}_c_{month}"
    wish_key  = f"{key_prefix}_w_{month}"
    mode_key  = f"{key_prefix}_m_{month}"
    init_key  = f"{key_prefix}_init_{month}"
    hash_key  = f"{key_prefix}_hash_{month}"

    # Fingerprint the DB-derived defaults. If they differ from what was loaded last time
    # (e.g. employee switched, or DB updated after save), reinitialize the calendar.
    # If defaults haven't changed, session state is preserved so unsaved edits survive reruns.
    defaults_hash = (
        tuple(sorted(int(d) for d in default_constraint_days)),
        tuple(sorted(int(d) for d in default_wish_days)),
    )
    if init_key not in st.session_state or st.session_state.get(hash_key) != defaults_hash:
        st.session_state[const_key] = sorted(list(set(int(d) for d in default_constraint_days)))
        st.session_state[wish_key]  = sorted(list(set(int(d) for d in default_wish_days)))
        st.session_state[init_key]  = True
        st.session_state[hash_key]  = defaults_hash

    sel_c    = st.session_state[const_key]
    sel_w    = st.session_state[wish_key]
    pfx      = key_prefix

    # ── Legend ────────────────────────────────────────────────────
    st.markdown(
        "<div style='display:flex;gap:12px;justify-content:center;margin-bottom:8px;flex-wrap:wrap;direction:rtl;'>"
        "<span style='background:linear-gradient(135deg,#6366f1,#4f46e5);color:white;border-radius:8px;padding:4px 12px;font-size:0.8rem;font-weight:600;'>רגיל</span>"
        "<span style='background:linear-gradient(135deg,#fecaca,#f87171);color:white;border-radius:8px;padding:4px 12px;font-size:0.8rem;font-weight:600;'>🔒 חסימה</span>"
        "<span style='background:linear-gradient(135deg,#86efac,#22c55e);color:white;border-radius:8px;padding:4px 12px;font-size:0.8rem;font-weight:600;'>⭐ בקשה</span>"
        "<span style='color:#64748b;font-size:0.8rem;padding-top:4px;'>לחץ יום כדי לעבור בין המצבים</span>"
        "</div>",
        unsafe_allow_html=True
    )

    # ── Dynamic CSS: color each day cell based on its state ────────
    css_parts = [f"""
    div[class*="st-key-{pfx}_d{month}_"] button {{
        min-height: 42px !important; font-size: 0.95rem !important;
        font-weight: 600 !important; border-radius: 10px !important;
        padding: 4px 2px !important;
        background: linear-gradient(135deg, #6366f1, #4f46e5) !important;
        color: white !important; border: none !important;
        box-shadow: 0 1px 3px rgba(79,70,229,0.3) !important; transform: none !important; letter-spacing: 0 !important;
    }}
    div[class*="st-key-{pfx}_d{month}_"] button:hover {{
        background: linear-gradient(135deg, #4f46e5, #4338ca) !important;
        transform: none !important; box-shadow: 0 2px 6px rgba(79,70,229,0.4) !important;
    }}"""]
    for d in sel_c:
        css_parts.append(f"""
    div[class*="st-key-{pfx}_d{month}_{d}x"] button {{
        background: linear-gradient(135deg,#fecaca,#f87171) !important;
        color: white !important; border-color: #f87171 !important;
    }}
    div[class*="st-key-{pfx}_d{month}_{d}x"] button:hover {{
        background: linear-gradient(135deg,#f87171,#ef4444) !important;
    }}""")
    for d in sel_w:
        css_parts.append(f"""
    div[class*="st-key-{pfx}_d{month}_{d}x"] button {{
        background: linear-gradient(135deg,#86efac,#22c55e) !important;
        color: white !important; border-color: #16a34a !important;
    }}
    div[class*="st-key-{pfx}_d{month}_{d}x"] button:hover {{
        background: linear-gradient(135deg,#22c55e,#16a34a) !important;
    }}""")
    st.markdown(f"<style>{''.join(css_parts)}</style>", unsafe_allow_html=True)

    # ── Special days lookup ────────────────────────────────────────
    special_map = {}
    if special_days_df is not None and not special_days_df.empty:
        for _, row in special_days_df.iterrows():
            try:
                d_obj = datetime.strptime(row['date'], '%Y-%m-%d').date()
                if d_obj.month == month and d_obj.year == year:
                    special_map[d_obj.day] = row['description']
            except:
                pass

    # ── Day headers (RTL: Saturday on left col 0, Sunday on right col 6) ──
    # DOM order left→right: ש' ו' ה' ד' ג' ב' א'  → reads right→left: א' ב' ג' ד' ה' ו' ש'
    day_headers = ["ש'", "ו'", "ה'", "ד'", "ג'", "ב'", "א'"]
    hcols = st.columns(7)
    for idx, h in enumerate(day_headers):
        is_wk_col = idx in [0, 1]   # col 0=Saturday, col 1=Friday
        hcols[idx].markdown(
            f"<div style='text-align:center;font-weight:700;"
            f"color:{'#7c3aed' if is_wk_col else '#64748b'};"
            f"font-size:0.8rem;padding:4px 0 2px;'>{h}</div>",
            unsafe_allow_html=True
        )

    # ── Calendar grid ──────────────────────────────────────────────
    # Use Calendar(6) so weeks start on Sunday (Python weekday 6).
    # Then reverse each week so Saturday ends up in col 0 (left) and Sunday in col 6 (right),
    # giving correct RTL visual order: read right→left = Sun … Fri Sat.
    cal_weeks = calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)
    for week in cal_weeks:
        week_rtl = list(reversed(week))   # Saturday→col 0, Sunday→col 6
        wcols = st.columns(7)
        for col_idx, day_num in enumerate(week_rtl):
            with wcols[col_idx]:
                if day_num == 0:
                    st.write("")
                else:
                    is_c       = day_num in sel_c
                    is_w       = day_num in sel_w
                    is_special = day_num in special_map
                    icon  = "🔒" if is_c else ("⭐" if is_w else ("📌" if is_special else ""))
                    label = f"{icon}{day_num}" if icon else str(day_num)

                    if st.button(label, key=f"{pfx}_d{month}_{day_num}x", use_container_width=True):
                        # Cycle: regular → חסימה → בקשה → regular
                        if day_num in st.session_state[const_key]:
                            # חסימה → בקשה
                            st.session_state[const_key].remove(day_num)
                            st.session_state[wish_key].append(day_num)
                        elif day_num in st.session_state[wish_key]:
                            # בקשה → regular
                            st.session_state[wish_key].remove(day_num)
                        else:
                            # regular → חסימה
                            st.session_state[const_key].append(day_num)
                        st.rerun()

                    if is_special:
                        desc = special_map[day_num]
                        st.markdown(
                            f"<div title='{desc}' style='font-size:9px;color:#b91c1c;"
                            f"text-align:center;margin-top:-6px;line-height:1.2;"
                            f"white-space:normal;word-break:break-word;overflow:visible;'>"
                            f"{desc}</div>",
                            unsafe_allow_html=True
                        )

    # ── Real-time status counter ───────────────────────────────────
    days_in_month = calendar.monthrange(year, month)[1]
    if show_validation:
        all_dates_m = [date(year, month, d) for d in range(1, days_in_month + 1)]
        total_thu = sum(1 for d in all_dates_m if d.weekday() == 3)
        total_wk  = sum(1 for d in all_dates_m if d.weekday() in [4, 5])
        blocked   = [date(year, month, d) for d in sel_c]
        av_thu = total_thu - sum(1 for d in blocked if d.weekday() == 3)
        av_wk  = total_wk  - sum(1 for d in blocked if d.weekday() in [4, 5])
        thu_ok = av_thu >= 2
        wk_ok  = av_wk  >= 4

        def _chip(label, ok=None):
            if ok is None:
                bg, bc, tc = "#f1f5f9", "#e2e8f0", "#334155"
            elif ok:
                bg, bc, tc = "#dcfce7", "#22c55e", "#166534"
            else:
                bg, bc, tc = "#fee2e2", "#ef4444", "#991b1b"
            icon = (" ✅" if ok else " ⚠️") if ok is not None else ""
            return (f"<span style='background:{bg};border:1px solid {bc};border-radius:8px;"
                    f"padding:5px 12px;font-weight:600;color:{tc};"
                    f"font-size:0.85rem;white-space:nowrap;'>{label}{icon}</span>")

        st.markdown(
            f"<div style='display:flex;gap:8px;justify-content:center;margin-top:14px;"
            f"flex-wrap:wrap;direction:rtl;'>"
            f"{_chip(f'🔒 חסימות: {len(sel_c)}')}"
            f"{_chip(f'⭐ בקשות: {len(sel_w)}')}"
            f"{_chip(f'ימי ה׳ פנויים: {av_thu}/{total_thu}', thu_ok)}"
            f"{_chip(f'סופ״ש פנויים: {av_wk}/{total_wk}', wk_ok)}"
            f"</div>",
            unsafe_allow_html=True
        )

    return list(st.session_state[const_key]), list(st.session_state[wish_key])

def run_smart_scheduling(year, month, only_weekends=False):
    num_days = calendar.monthrange(year, month)[1]
    staff_df = st.session_state.staff.copy()
    
    # טעינת כל השיבוצים הקיימים (לא רק ידניים) כדי לשמר מצב קיים
    all_current_records = st.session_state.schedule.to_dict('records')
    
    # סינון מקדים:
    # 1. שומרים את כל השיבוצים של חודשים אחרים.
    # 2. בחודש הנוכחי: שומרים כל שיבוץ שיש בו עובד אמיתי (לא '---').
    #    נמחקים רק שיבוצים של '---' בחודש הנוכחי כדי לאפשר למערכת לנסות לשבץ אותם מחדש.
    new_schedule = []
    current_month_prefix = f"{year}-{month:02d}"
    
    for r in all_current_records:
        # אם התאריך לא בחודש הנוכחי - שומרים
        if not str(r['date']).startswith(current_month_prefix):
            new_schedule.append(r)
        else:
            # אם בחודש הנוכחי - שומרים רק אם זה לא כישלון ('---') 
            # (כך שאם היה כישלון, המשבצת "מתפנה" לניסיון חוזר)
            if r['employee'] != '---':
                new_schedule.append(r)
    
    work_load = {row['name']: 0 for _, row in staff_df.iterrows()}
    weekends_worked = {row['name']: set() for _, row in staff_df.iterrows()}
    last_assignment = {row['name']: -999 for _, row in staff_df.iterrows()}
    wed_counts = {row['name']: 0 for _, row in staff_df.iterrows()}
    thu_counts = {row['name']: 0 for _, row in staff_df.iterrows()}

    # איסוף נתונים לדיווח הוגנות
    initial_wed_stats = {}
    initial_thu_stats = {}

    # עדכון מונים: הפרדה בין מונים גלובליים (הוגנות) למונים חודשיים (מכסות)
    for s in new_schedule:
        if s['employee'] not in work_load or s['employee'] == '---': continue

        dt = datetime.strptime(s['date'], '%Y-%m-%d')

        # 1. מונים גלובליים (היסטוריה מלאה) - לאיזון ימי כוח
        if dt.weekday() == 2: wed_counts[s['employee']] += 1
        if dt.weekday() == 3: thu_counts[s['employee']] += 1

        # 2. last_assignment מעודכן מכל ההיסטוריה (כולל חודשים קודמים)
        # כדי שמי שעבד ב-31 למרץ לא יקבל בונוס ריווח מקסימלי ב-1 לאפריל
        if dt.toordinal() > last_assignment[s['employee']]:
            last_assignment[s['employee']] = dt.toordinal()

        # 3. מונים חודשיים בלבד - לבדיקת מכסות
        if str(s['date']).startswith(current_month_prefix):
            work_load[s['employee']] += 1

            # תיקון: החרגת שישי בוקר ממכסת הסופ"ש + שילוב "סופ"שים פונקציונליים"
            if is_functional_weekend(dt, st.session_state.special_days) and "שישי בוקר" not in s.get('dept', ''):
                weekends_worked[s['employee']].add(dt.isocalendar()[1])

    # חישוב סטטיסטיקה לפני השיבוץ הנוכחי
    avg_wed = sum(wed_counts.values()) / len(wed_counts) if wed_counts else 0
    avg_thu = sum(thu_counts.values()) / len(thu_counts) if thu_counts else 0
    
    # זיהוי עובדים שזקוקים לאיזון (מתחת לממוצע)
    priority_wed = [k for k, v in wed_counts.items() if v < avg_wed]
    priority_thu = [k for k, v in thu_counts.items() if v < avg_thu]
    
    balancing_msg = f"**דוח איזון הוגנות (רב-חודשי):**\n"
    balancing_msg += f"- **רביעי:** ממוצע {avg_wed:.1f}. תועדפו: {len(priority_wed)} עובדים.\n"
    balancing_msg += f"- **חמישי:** ממוצע {avg_thu:.1f}. תועדפו: {len(priority_thu)} עובדים.\n"
    balancing_msg += f"- **שישי בוקר:** האיזון מבוצע אוטומטית על סמך כל ההיסטוריה."

    all_dates = [date(year, month, d) for d in range(1, num_days + 1)]
    
    # סינון תאריכים לפי דרישה
    if only_weekends:
        # אם ביקשו רק סופ"שים, ניקח רק שישי-שבת או ימים מיוחדים שמוגדרים סופ"ש
        sorted_dates = [d for d in all_dates if is_functional_weekend(d, st.session_state.special_days)]
    else:
        # תעדוף סופי שבוע פונקציונליים ואז אמצע שבוע
        sorted_dates = [d for d in all_dates if is_functional_weekend(d, st.session_state.special_days)] + [d for d in all_dates if not is_functional_weekend(d, st.session_state.special_days)]
    
    # פונקציית עזר להמרה בטוחה למספר
    def safe_int(val, default=0):
        try:
            if pd.isna(val) or val == "":
                return default
            return int(float(val))
        except (ValueError, TypeError):
            return default

    for d in sorted_dates:
        d_str = str(d)
        week_num = d.isocalendar()[1]
        
        # תעדוף פנימית גריאטרית (כי רק מתמחים יכולים) ואז שיקום (שיש בו תורני חוץ)
        for dept in ["פנימית גריאטרית", "שיקום"]:
            if any(s for s in new_schedule if s['date'] == d_str and s['dept'] == dept): continue
            
            candidates = []
            failure_reasons = []
            
            for _, person in staff_df.iterrows():
                name = person['name']
                if person['type'] == 'תורן חוץ' and dept == 'פנימית גריאטרית': continue
                
                # BUG FIX: Ensure employee is not already assigned to another department on the same day
                if any(s for s in new_schedule if s['date'] == d_str and s['employee'] == name):
                    continue

                # --- Department Restriction Check (Feature 5: Gantt-canonical) ---
                only_home = person.get('only_home_dept', False)
                if only_home:
                     # Determine target shift's "Real" Department context
                     target_context = dept
                     if "שישי בוקר" in dept:
                         target_context = "שיקום" if "שיקום" in dept else "פנימית גריאטרית"

                     # Resolve night-shift home dept from the Gantt for this month,
                     # falling back to staff.dept when no rotation row exists.
                     _person_home_dept = _emp_night_dept(name, f"{year}-{month:02d}") \
                         or str(person.get('dept', '') or '')
                     if _person_home_dept != 'כללי' and _person_home_dept != target_context:
                         # failure_reasons.append(f"{name}: מוגבל למחלקת אם")
                         continue
                # ------------------------------------

                # בדיקת מכסה קשיחה (חודשית)
                monthly_quota = safe_int(person['monthly_quota'], 0)
                if work_load[name] >= monthly_quota:
                    failure_reasons.append(f"{name}: מכסה מלאה ({monthly_quota})")
                    continue

                # שמירת מכסה רכה (Soft Quota Reservation)
                # אם אנחנו בחצי הראשון של החודש (לפני יום 15) ועובד כבר ניצל 50%+ מהמכסה שלו —
                # נדחה אותו מהמאגר הראשי כדי לשמור לו משמרות לחצי השני.
                # ה-Fallback עדיין יוכל לשלוף אותו כמוצא אחרון אם אין אף אחד אחר.
                if d.day < 15 and monthly_quota > 0 and work_load[name] >= monthly_quota * 0.5:
                    failure_reasons.append(f"{name}: שמירת מכסה (חצי ראשון)")
                    continue

                # בדיקת סופ"ש (כולל סופ"ש פונקציונלי)
                weekend_quota = safe_int(person['weekend_quota'], 0)
                if is_functional_weekend(d, st.session_state.special_days) and len(weekends_worked[name]) >= weekend_quota and week_num not in weekends_worked[name]:
                    failure_reasons.append(f"{name}: מכסת סופ\"ש")
                    continue
                
                # מנוחה - מרווח של שנתיים (2 ימים)
                # בדיקה של יומיים אחורה ויומיים קדימה
                gap_days = [-2, -1, 1, 2]
                has_rest_conflict = False
                for offset in gap_days:
                    if any(s for s in new_schedule if s['date'] == str(d + timedelta(days=offset)) and s['employee'] == name):
                        has_rest_conflict = True
                        break
                
                if has_rest_conflict:
                    failure_reasons.append(f"{name}: מרווח מנוחה")
                    continue

                # כלל שבת-רביעי למתמחים
                if person['type'] == 'מתמחה':
                    # אם היום שבת, בדוק אם שובץ ברביעי הקודם
                    if d.weekday() == 5: # שבת
                        if any(s for s in new_schedule if s['date'] == str(d - timedelta(days=3)) and s['employee'] == name):
                            failure_reasons.append(f"{name}: שובץ ברביעי")
                            continue
                    # אם היום רביעי, בדוק אם שובץ בשבת הבאה
                    if d.weekday() == 2: # רביעי
                        # בדיקה אם משובץ לשבת הבאה
                        if any(s for s in new_schedule if s['date'] == str(d + timedelta(days=3)) and s['employee'] == name):
                            failure_reasons.append(f"{name}: משובץ בשבת")
                            continue
                        # בדיקה אם משובץ לשישי הבא (כולל שישי בוקר) - לוגיקה הפוכה חדשה
                        fri_check = str(d + timedelta(days=2))
                        if any(s for s in new_schedule if s['date'] == fri_check and s['employee'] == name):
                            failure_reasons.append(f"{name}: משובץ בשישי הקרוב")
                            continue
                
                # אילוצים (חסמים)
                if not st.session_state.requests[(st.session_state.requests['employee'] == name) & (st.session_state.requests['date'] == d_str) & (st.session_state.requests['status'] == "אילוץ")].empty:
                    failure_reasons.append(f"{name}: אילוץ")
                    continue

                candidates.append(person)

            if candidates:
                # --- לוגיקה חדשה: תעדוף בקשות (Wishes - Option 3) ---
                # נבדוק אם יש מועמדים שביקשו במיוחד את המשמרת הזו (וברור שהם עומדים בכל שאר הכללים כי הם עברו את הסינון למעלה)
                requesters = st.session_state.requests[(st.session_state.requests['date'] == d_str) & (st.session_state.requests['status'] == "בקשה")]['employee'].tolist()
                
                wish_candidates = [c for c in candidates if c['name'] in requesters]
                
                # אם יש כאלו שביקשו, נצמצם את הרשימה רק אליהם
                final_pool = wish_candidates if wish_candidates else candidates
                
                # לוגיקת תיעדוף משופרת:
                # שימוש בשיטת ניקוד כדי:
                # 1. לפזר את התורנויות (מי שמילא אחוז נמוך יותר מהמכסה מקבל עדיפות)
                # 2. לתת עדיפות ברורה לתורני חוץ בימי חמישי, שישי ושבת במחלקת שיקום
                
                def calculate_score(cand):
                    name = cand['name']
                    # חישוב אחוז ניצול מכסה (נרצה לתת עדיפות למי שניצל פחות)
                    quota = safe_int(cand['monthly_quota'], 1)
                    usage_ratio = work_load[name] / quota if quota > 0 else 1.0
                    
                    # ציון בסיס: אחוז ניצול מכסה (שלילי, כי אנו רוצים יחס נמוך)
                    # מכפילים ב-100 כדי שיהיה משקל משמעותי
                    score = -usage_ratio * 100
                    
                    # פקטור פיזור: כמה ימים עברו מאז המשמרת האחרונה?
                    last_day = last_assignment.get(name, -999)
                    days_diff = d.toordinal() - last_day
                    score += days_diff * 2  # בונוס על כל יום שעבר
                    
                    # --- פקטור ריווח (Spacing/Pacing) ---
                    # המטרה: לפזר את המשמרות לאורך החודש ולמנוע דחיסה בהתחלה
                    # נחשב "קצב צפוי": באיזה יום בחודש אנו נמצאים, וכמה משמרות היה אמור לעשות עד עכשיו באופן לינארי.
                    current_day_in_month = d.day
                    month_progress = current_day_in_month / num_days # 0.0 to 1.0
                    expected_shifts = quota * month_progress
                    actual_shifts = work_load[name]
                    
                    # הציון הוא ההפרש:
                    # אם expected (2.5) > actual (1) -> נקבל 1.5 חיובי (דחוף לשבץ)
                    # אם expected (1.0) < actual (3) -> נקבל -2.0 שלילי (כבר עשה יותר מדי, להרגיע)
                    pacing_score = (expected_shifts - actual_shifts) * 500 
                    score += pacing_score
                    # ------------------------------------

                    # פקטור סופ"ש לתורני חוץ בשיקום
                    if dept == "שיקום" and cand['type'] == 'תורן חוץ':
                        # ימי חמישי (3), שישי (4), שבת (5) או סופ"ש פונקציונלי
                        if d.weekday() in [3, 4, 5] or is_functional_weekend(d, st.session_state.special_days):
                            score += 2000 # בונוס ענק שמבטיח בחירה
                    
                    # פקטור איזון ימי רביעי/חמישי למתמחים (שלא תורני חוץ)
                    if cand['type'] == 'מתמחה':
                        # יום רביעי (2) - מבוקש, ננסה לתת למי שעשה הכי מעט
                        if d.weekday() == 2:
                            score -= wed_counts[name] * 200
                        # יום חמישי (3) - לא מבוקש/קשה, ננסה לתת למי שעשה הכי מעט
                        if d.weekday() == 3:
                            score -= thu_counts[name] * 200
                            
                    # פקטור מחלקה - העדפה למחלקת האם! (Feature 5: Gantt-canonical)
                    # אם המועמד שייך למחלקה הנוכחית או ל'כללי' - מקבל בונוס
                    # אם המועמד ממחלקה אחרת - נמצא רק בעדיפות אחרונה (ענישה)
                    cand_dept = (_emp_night_dept(name, d.strftime('%Y-%m'))
                                 or str(cand.get('dept', '') or ''))
                    
                    # בדיקת התאמה מלאה - אם מוגבל למחלקת אם, הציון לא רלוונטי כי הוא נפסל למעלה,
                    # אבל כאן זה נותן בונוס למי שנמצא במחלקה הנכונה
                    if cand_dept == dept or cand_dept == 'כללי':
                         score += 500
                    else:
                         score -= 5000  # קנס משמעותי מאוד (היה 500, הגברנו ל-5000 ליתר ביטחון)

                    return score

                final_choice = max(final_pool, key=calculate_score)['name']
                
                new_schedule.append({'date': d_str, 'dept': dept, 'employee': final_choice, 'is_manual': False, 'empty_reason': ''})
                work_load[final_choice] += 1
                last_assignment[final_choice] = d.toordinal()
                if d.weekday() == 2: wed_counts[final_choice] += 1
                if d.weekday() == 3: thu_counts[final_choice] += 1
                if is_functional_weekend(d, st.session_state.special_days): weekends_worked[final_choice].add(week_num)
            else:
                # --- NEW STARVATION FALLBACK ---
                fallback_pool = []
                for _, cand in staff_df.iterrows():
                    c_name = cand['name']
                    if c_name == '---' or str(c_name).upper() == 'ADMIN': continue
                    if cand['type'] == 'תורן חוץ' and dept == 'פנימית גריאטרית': continue
                    if any(s for s in new_schedule if s['date'] == d_str and s['employee'] == c_name): continue
                    req = st.session_state.requests[(st.session_state.requests['employee'] == c_name) & (st.session_state.requests['date'] == d_str) & (st.session_state.requests['status'] == 'אילוץ')]
                    if not req.empty: continue
                    
                    # HARD QUOTA CHECK: Do not bypass quotas even in fallback
                    monthly_quota = safe_int(cand['monthly_quota'], 0)
                    if work_load.get(c_name, 0) >= monthly_quota: continue
                    
                    # Weekend quota
                    weekend_quota = safe_int(cand['weekend_quota'], 0)
                    if is_functional_weekend(d, st.session_state.special_days) and len(weekends_worked.get(c_name, set())) >= weekend_quota and week_num not in weekends_worked.get(c_name, set()):
                        continue
                        
                    fallback_pool.append(c_name)
                
                if fallback_pool:
                    fallback_choice = min(fallback_pool, key=lambda x: work_load.get(x, 0))
                    new_schedule.append({'date': d_str, 'dept': dept, 'employee': fallback_choice, 'is_manual': False, 'empty_reason': 'הושלם מחוסר ברירה (הגמשת חוקים)'})
                    work_load[fallback_choice] += 1
                    last_assignment[fallback_choice] = d.toordinal()
                    if d.weekday() == 2: wed_counts[fallback_choice] += 1
                    if d.weekday() == 3: thu_counts[fallback_choice] += 1
                    if is_functional_weekend(d, st.session_state.special_days): weekends_worked[fallback_choice].add(week_num)
                    continue
                # --- END FALLBACK ---
                
                # --- לוגיקת הצעת החלפות ועזרה (Swap & Suggest) ---
                suggestions = []
                # אתחול מילון הצעות מובנה אם לא קיים (נמחק בתחילת הריצה)
                if 'swap_suggestions' not in st.session_state: st.session_state.swap_suggestions = {}
                
                # כלי עזר לבדיקת תקינות מלאה (כולל מנוחה, רצף, וכו') להחלפה
                def is_valid_assignment_for_swap(person_name, check_date, target_dept):
                    # Use the shared global validation function
                    # check_assignment_validity returns (bool, reason)
                    is_valid, _ = check_assignment_validity(new_schedule, person_name, check_date, target_dept, staff_df, st.session_state.requests)
                    return is_valid
                # נרוץ על המתמודדים שנפסלו (Candidate A) וננסה למצוא פתרון שיאפשר לשבץ אותם
                for _, person_a in staff_df.iterrows():
                    name_a = person_a['name']
                    # אם הסיבה לפסילה היא "מכסה", אין טעם להציע החלפה (אלא אם כן מגדילים ראש, אבל נתמקד באילוצי לו"ז)
                    # אבל אם הוא תפוס במקום אחר, ננסה לשחרר אותו
                    
                    # תרחיש 1: החלפה ישירה (Direct Swap)
                    # A נמצא במקום אחר באותו יום. האם אפשר למצוא מישהו (B) שיחליף את A שם?
                    parallel_shift = next((s for s in new_schedule if s['date'] == d_str and s['employee'] == name_a), None)
                    if parallel_shift:
                        other_dept = parallel_shift['dept']
                        # מחפשים מחליף (B) לתפקיד השני
                        for _, person_b in staff_df.iterrows():
                            name_b = person_b['name']
                            if name_b == name_a: continue
                            
                            # בדיקה קפדנית: האם B יכול חוקית להיכנס ל-other_dept בתאריך d_str?
                            if is_valid_assignment_for_swap(name_b, d_str, other_dept):
                                suggestions.append(f"💡 החלפה: העבר את **{name_a}** מ-{other_dept} לפה, ושבץ את **{name_b}** שם.")
                                
                                # שמירת הצעה מובנית לביצוע בלחיצה
                                core_key = f"{d_str}_{dept}" # מפתח למשבצת הריקה הנוכחית
                                if core_key not in st.session_state.swap_suggestions: st.session_state.swap_suggestions[core_key] = []
                                st.session_state.swap_suggestions[core_key].append({
                                    'type': 'direct_swap',
                                    'target_date': d_str,
                                    'conflicted_emp': name_a, # מי שאנחנו רוצים לפה
                                    'source_dept': other_dept, # מאיפה הוא בא
                                    'replacement_emp': name_b, # מי יחליף אותו שם
                                    'desc': f"{name_a} ⬅️ {name_b} ({other_dept})"
                                })
                                break 

                    # תרחיש 2: הסטה (Shift/Move)
                    # A לא יכול לעבוד היום כי עבד אתמול (מנוחה). האם אפשר להזיז את המשמרת של אתמול למישהו אחר (B)?
                    prev_conflict = next((s for s in new_schedule if s['employee'] == name_a and s['date'] in [str(d - timedelta(days=i)) for i in [1, 2]]), None)
                    if prev_conflict:
                        conf_date = prev_conflict['date']
                        conf_dept = prev_conflict['dept']
                        
                        # מחפשים מחליף (B) לתאריך ההוא
                        for _, person_b in staff_df.iterrows():
                            name_b = person_b['name']
                            if name_b == name_a: continue
                            
                            # בדיקה קפדנית: האם B יכול להיכנס ל-conf_date?
                            if is_valid_assignment_for_swap(name_b, conf_date, conf_dept):
                                # Format conf_date for display
                                conf_date_obj = datetime.strptime(conf_date, '%Y-%m-%d')
                                conf_date_fmt = conf_date_obj.strftime('%d/%m/%Y')

                                suggestions.append(f"💡 הסטה: העבר את **{name_a}** מה-{conf_date_fmt} לפה, ושבץ שם את **{name_b}**.")
                                
                                core_key = f"{d_str}_{dept}"
                                if core_key not in st.session_state.swap_suggestions: st.session_state.swap_suggestions[core_key] = []
                                st.session_state.swap_suggestions[core_key].append({
                                    'type': 'move_shift',
                                    'conflict_date': conf_date,
                                    'conflicted_emp': name_a,
                                    'conflict_dept': conf_dept,
                                    'replacement_emp': name_b,
                                    'desc': f"הסטה: {name_a} (מ-{conf_date_fmt}) ⬅️ {name_b}"
                                })
                                break

                    # תרחיש 3: שרשור משולש (Triple Swap) - בקשת המשתמש
                    # אם תרחיש 1 נכשל (לא נמצא B פנוי), אולי B תפוס במקום אחר (C) אבל C פנוי?
                    # כלומר: A בא לפה -> B מחליף את A -> C מחליף את B
                    if parallel_shift: # A תפוס ב-other_dept
                         other_dept = parallel_shift['dept']
                         # רצים שוב על B פוטנציאליים (שאולי לא פנויים)
                         for _, person_b in staff_df.iterrows():
                            name_b = person_b['name']
                            if name_b == name_a: continue
                            
                            # אם B תפוס ב-other_dept_2 בתאריך d_str
                            parallel_shift_b = next((s for s in new_schedule if s['date'] == d_str and s['employee'] == name_b), None)
                            if parallel_shift_b:
                                other_dept_b = parallel_shift_b['dept']
                                # מחפשים C שיחליף את B
                                for _, person_c in staff_df.iterrows():
                                    name_c = person_c['name']
                                    if name_c in [name_a, name_b]: continue
                                    
                                    # האם C יכול להחליף את B ב-other_dept_b?
                                    if is_valid_assignment_for_swap(name_c, d_str, other_dept_b):
                                        # האם B (אחרי שהשתחרר) יכול להחליף את A ב-other_dept?
                                        # כאן ההנחה היא ש-B עובר מ-other_dept_b ל-other_dept באותו יום. האם זה חוקי?
                                        # בדרך כלל כן, כי זה אותו יום.
                                        
                                        suggestions.append(f"💡 שרשור: {name_a} לפה, {name_b} ל-{other_dept}, {name_c} ל-{other_dept_b}.")
                                        
                                        core_key = f"{d_str}_{dept}"
                                        if core_key not in st.session_state.swap_suggestions: st.session_state.swap_suggestions[core_key] = []
                                        st.session_state.swap_suggestions[core_key].append({
                                            'type': 'triple_swap',
                                            'target_date': d_str,
                                            'emp_a': name_a, 'dept_a_origin': other_dept,
                                            'emp_b': name_b, 'dept_b_origin': other_dept_b,
                                            'emp_c': name_c,
                                            'desc': f"שרשור: {name_a} ⬅️ {name_b} ⬅️ {name_c}"
                                        })
                                        break
                                if len(suggestions) > 3: break # הגבלה שלא נתפוצץ

                # סינון כפילויות בהצגה
                unique_suggestions = list(set([s.split(":")[0] + "..." for s in suggestions])) # תקציר
                if suggestions:
                    final_msg = "לא נמצא שיבוץ מתאים עבור אף עובד.\n" + "\n".join(suggestions[:3])
                else:
                    final_msg = "לא נמצא פתרון אוטומטי (סיבות לדוגמה: " + ", ".join(list(set(failure_reasons))[:3]) + ")"
                
                new_schedule.append({'date': d_str, 'dept': dept, 'employee': '---', 'is_manual': False, 'empty_reason': final_msg})

    # --- לוגיקה חדשה: שיבוץ משמרות כמו "שישי בוקר" (4 עובדים) ---
    # רצים על כל ימי השישי האמיתיים + ימי "כמו שישי" האקסטרה
    fridays = [d for d in all_dates if d.weekday() == 4 or get_functional_day_type(d, st.session_state.special_days) == 'כמו שישי (ערב חג)']
    for fri_date in fridays:
        fri_str = str(fri_date)
        sat_date = fri_date + timedelta(days=1)
        sat_str = str(sat_date)
        
        # 1. פנימית גריאטרית (2 עובדים) - מי שעושה שישי ושבת
        fri_worker_pnimia = next((s['employee'] for s in new_schedule if s['date'] == fri_str and s['dept'] == 'פנימית גריאטרית'), None)
        sat_worker_pnimia = next((s['employee'] for s in new_schedule if s['date'] == sat_str and s['dept'] == 'פנימית גריאטרית'), None)
        
        # בדיקה אם כבר יש שיבוץ (למשל ידני)
        has_pnimia_1 = any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == 'שישי בוקר - פנימית (1)')
        has_pnimia_2 = any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == 'שישי בוקר - פנימית (2)')

        if not has_pnimia_1 and fri_worker_pnimia and fri_worker_pnimia != '---':
                new_schedule.append({'date': fri_str, 'dept': 'שישי בוקר - פנימית (1)', 'employee': fri_worker_pnimia, 'is_manual': False, 'empty_reason': 'נגזר אוטומטית משישי'})
        if not has_pnimia_2 and sat_worker_pnimia and sat_worker_pnimia != '---':
                new_schedule.append({'date': fri_str, 'dept': 'שישי בוקר - פנימית (2)', 'employee': sat_worker_pnimia, 'is_manual': False, 'empty_reason': 'נגזר אוטומטית משבת'})

        # 2. שיקום (2 עובדים)
        fri_worker_rehab = next((s['employee'] for s in new_schedule if s['date'] == fri_str and s['dept'] == 'שיקום'), None)
        sat_worker_rehab = next((s['employee'] for s in new_schedule if s['date'] == sat_str and s['dept'] == 'שיקום'), None)
        
        def handle_rehab_morning(worker_name, source_day, slot_num):
            target_dept = f'שישי בוקר - שיקום ({slot_num})'
            # בדיקה למניעת כפילות עם שיבוץ ידני
            if any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == target_dept): return

            if not worker_name or worker_name == '---': return

            # בדיקת סוג העובד
            w_type = None
            if worker_name in staff_df['name'].values:
                w_type = staff_df[staff_df['name'] == worker_name]['type'].iloc[0]
            
            if w_type == 'מתמחה':
                # אם זה מתמחה - הוא עושה את הבוקר
                new_schedule.append({'date': fri_str, 'dept': target_dept, 'employee': worker_name, 'is_manual': False, 'empty_reason': f'נגזר אוטומטית מ{source_day}'})
            else:
                # אם זה תורן חוץ - מחפשים מחליף (מתמחה משיקום)
                # קריטריונים: מחלקת שיקום, פנוי בשישי, עומד בבדיקת מנוחה ±2 ימים
                candidates = []
                _fri_ym = fri_str[:7]
                for _, row in staff_df.iterrows():
                    if row['type'] != 'מתמחה' or row['name'] == worker_name:
                        continue
                    # Feature 5: pull dept from the Gantt for this month (not staff.dept).
                    _row_dept = _emp_night_dept(row['name'], _fri_ym) or str(row.get('dept', '') or '')
                    if _row_dept == 'שיקום':
                        emp = row['name']
                        
                        # האם פנוי ביום שישי (אילוץ - חסם)
                        is_blocked = not st.session_state.requests[(st.session_state.requests['employee'] == emp) & (st.session_state.requests['date'] == fri_str) & (st.session_state.requests['status'] == "אילוץ")].empty
                        if is_blocked: continue
                        
                        # בדיקת מנוחה: ±2 ימים מסביב ליום שישי
                        has_rest_conflict = any(
                            s['employee'] == emp and s['date'] == str(fri_date + timedelta(days=offset))
                            for s in new_schedule for offset in [-2, -1, 1, 2]
                        )
                        if has_rest_conflict: continue

                        # האם כבר משובץ בשישי במקום אחר (למשל תורנות רגילה במחלקת שיקום בצד השני?)
                        if any(s['employee'] == emp and s['date'] == fri_str for s in new_schedule): continue
                        
                        # HARD QUOTA CHECK: האם המכסה החודשית שלו חוסמת אותו?
                        monthly_quota = safe_int(row['monthly_quota'], 0)
                        if work_load.get(emp, 0) >= monthly_quota: continue

                        # חישוב ציון הוגנות: כמה שישי בוקר כבר יש לו?
                        fri_morning_count = len([s for s in new_schedule if s['employee'] == emp and 'שישי בוקר' in s['dept']])
                        candidates.append((emp, fri_morning_count))
                
                # מיון לפי הכמות הכי קטנה של שישי בוקר (איזון)
                if candidates:
                    candidates.sort(key=lambda x: x[1]) # מהקטן לגדול
                    best_candidate = candidates[0][0]
                    new_schedule.append({'date': fri_str, 'dept': target_dept, 'employee': best_candidate, 'is_manual': False, 'empty_reason': f'השלמה במקום {worker_name}'})
                    work_load[best_candidate] += 1 # עדכון מכסה כדי למנוע חריגה בחיפושים הבאים באותה ריצה
                else:
                    new_schedule.append({'date': fri_str, 'dept': target_dept, 'employee': '---', 'is_manual': False, 'empty_reason': 'לא נמצא מחליף לבוקר'})

        handle_rehab_morning(fri_worker_rehab, "שישי", "1")
        handle_rehab_morning(sat_worker_rehab, "שבת", "2")

    # יצירת DataFrame סופי וניקוי כפילויות (הגנה הרמטית)
    final_df = pd.DataFrame(new_schedule)
    if not final_df.empty:
        # אם יש כפילות של תאריך+מחלקה, משאירים את הראשון (שהוא הידני כי הוא הוסף ראשון)
        final_df = final_df.drop_duplicates(subset=['date', 'dept'], keep='first')

    st.session_state.schedule = final_df
    save_to_db("schedule", st.session_state.schedule)

    # הצגת דוח האיזון למשתמש
    st.info(balancing_msg)

def run_smart_scheduling_cp(year, month, only_weekends=False):
    """
    CP-SAT optimized scheduler. Replaces the greedy slot-by-slot approach with a
    global optimizer that sees all assignments simultaneously. Falls back to
    run_smart_scheduling() if ortools is unavailable or the solver fails.
    All hard/soft constraints are preserved exactly.
    """
    try:
        from ortools.sat.python import cp_model as _cp
    except ImportError:
        st.warning("⚠️ ortools לא מותקן — משתמש בשיבוץ חמדני")
        run_smart_scheduling(year, month, only_weekends)
        return

    # ── 1. Identical initialization to run_smart_scheduling() ──────────────
    num_days = calendar.monthrange(year, month)[1]
    staff_df = st.session_state.staff.copy()
    all_current_records = st.session_state.schedule.to_dict('records')

    new_schedule = []
    current_month_prefix = f"{year}-{month:02d}"
    for r in all_current_records:
        if not str(r['date']).startswith(current_month_prefix):
            new_schedule.append(r)
        else:
            if r['employee'] != '---':
                new_schedule.append(r)

    work_load       = {row['name']: 0   for _, row in staff_df.iterrows()}
    weekends_worked = {row['name']: set() for _, row in staff_df.iterrows()}
    last_assignment = {row['name']: -999 for _, row in staff_df.iterrows()}
    wed_counts      = {row['name']: 0   for _, row in staff_df.iterrows()}
    thu_counts      = {row['name']: 0   for _, row in staff_df.iterrows()}

    for s in new_schedule:
        if s['employee'] not in work_load or s['employee'] == '---': continue
        dt = datetime.strptime(s['date'], '%Y-%m-%d')
        if dt.weekday() == 2: wed_counts[s['employee']] += 1
        if dt.weekday() == 3: thu_counts[s['employee']] += 1
        if dt.toordinal() > last_assignment[s['employee']]:
            last_assignment[s['employee']] = dt.toordinal()
        if str(s['date']).startswith(current_month_prefix):
            work_load[s['employee']] += 1
            if is_functional_weekend(dt, st.session_state.special_days) and "שישי בוקר" not in s.get('dept', ''):
                weekends_worked[s['employee']].add(dt.isocalendar()[1])

    avg_wed = sum(wed_counts.values()) / len(wed_counts) if wed_counts else 0
    avg_thu = sum(thu_counts.values()) / len(thu_counts) if thu_counts else 0
    priority_wed = [k for k, v in wed_counts.items() if v < avg_wed]
    priority_thu = [k for k, v in thu_counts.items() if v < avg_thu]
    balancing_msg = (
        f"**דוח איזון הוגנות (רב-חודשי):**\n"
        f"- **רביעי:** ממוצע {avg_wed:.1f}. תועדפו: {len(priority_wed)} עובדים.\n"
        f"- **חמישי:** ממוצע {avg_thu:.1f}. תועדפו: {len(priority_thu)} עובדים.\n"
        f"- **שישי בוקר:** האיזון מבוצע אוטומטית על סמך כל ההיסטוריה."
    )

    all_dates = [date(year, month, d) for d in range(1, num_days + 1)]

    def safe_int(val, default=0):
        try:
            if pd.isna(val) or val == "": return default
            return int(float(val))
        except (ValueError, TypeError):
            return default

    # ── 2. Lookup structures ────────────────────────────────────────────────
    DEPTS = ["פנימית גריאטרית", "שיקום"]

    covered = {(str(s['date']), s['dept']) for s in new_schedule if s['employee'] not in ('---', '')}

    employees = []
    for _, row in staff_df.iterrows():
        name = str(row.get('name', '')).strip()
        if not name or name == '---' or name.upper() == 'ADMIN': continue
        if str(row.get('type', '')).strip() in ('מנהל/ת', 'מנהל מחלקה'): continue
        if safe_int(row.get('monthly_quota', 0), 0) == 0: continue
        employees.append(name)

    emp_row   = {}
    for e in employees:
        rows = staff_df[staff_df['name'] == e]
        if not rows.empty:
            emp_row[e] = rows.iloc[0]

    emp_type     = {e: str(emp_row[e].get('type', '')).strip()           for e in employees if e in emp_row}
    emp_quota    = {e: safe_int(emp_row[e].get('monthly_quota', 6), 6)   for e in employees if e in emp_row}
    emp_wq       = {e: safe_int(emp_row[e].get('weekend_quota', 0), 0)   for e in employees if e in emp_row}
    emp_only_home = {}
    for e in employees:
        if e not in emp_row: continue
        _oh = emp_row[e].get('only_home_dept', False)
        emp_only_home[e] = _oh if isinstance(_oh, bool) else str(_oh).strip().lower() == 'true'
    emp_home = {e: (_emp_night_dept(e, f"{year}-{month:02d}") or str(emp_row[e].get('dept', '') or ''))
                for e in employees if e in emp_row}
    employees = [e for e in employees if e in emp_row]  # keep only resolved rows

    blocked_set = set()
    wish_set    = set()
    for _, r in st.session_state.requests.iterrows():
        key = (str(r['employee']).strip(), str(r['date']))
        if str(r['status']) == 'אילוץ':  blocked_set.add(key)
        elif str(r['status']) == 'בקשה': wish_set.add(key)

    if only_weekends:
        target_dates = [d for d in all_dates if is_functional_weekend(d, st.session_state.special_days)]
    else:
        target_dates = all_dates

    optimizable_slots = [(d, k) for d in target_dates for k in DEPTS if (str(d), k) not in covered]

    if not optimizable_slots:
        st.info(f"✅ כל המשמרות כבר שובצו.\n\n{balancing_msg}")
        return

    # ── 3. Build CP-SAT model ───────────────────────────────────────────────
    model = _cp.CpModel()
    x = {}  # (emp, date_obj, dept_str) -> BoolVar

    for e in employees:
        for d, k in optimizable_slots:
            d_str = str(d)
            if (e, d_str) in blocked_set: continue
            if k == "פנימית גריאטרית" and emp_type.get(e) == 'תורן חוץ': continue
            if emp_only_home.get(e):
                home = emp_home.get(e, '')
                if home not in ('כללי', k): continue
            x[(e, d, k)] = model.new_bool_var(f"x|{e}|{d}|{k}")

    # HC-1: at most one employee per (day, dept)
    for d, k in optimizable_slots:
        slot_vars = [x[(e, d, k)] for e in employees if (e, d, k) in x]
        if slot_vars:
            model.add(sum(slot_vars) <= 1)

    # HC-2: monthly quota
    for e in employees:
        remaining = emp_quota.get(e, 6) - work_load.get(e, 0)
        all_e = [x[(e, d, k)] for d, k in optimizable_slots if (e, d, k) in x]
        if all_e:
            if remaining <= 0:
                for v in all_e: model.add(v == 0)
            else:
                model.add(sum(all_e) <= remaining)

    # HC-3: same-day uniqueness
    for e in employees:
        for d in target_dates:
            day_vars = [x[(e, d, k)] for k in DEPTS if (e, d, k) in x]
            if len(day_vars) > 1:
                model.add(sum(day_vars) <= 1)

    # HC-4: rest gap ±2 days (including cross-month boundary)
    day_list = sorted(set(d for d, k in optimizable_slots))
    for e in employees:
        last_ord = last_assignment.get(e, -999)
        for i, d1 in enumerate(day_list):
            vars_d1 = [x[(e, d1, k)] for k in DEPTS if (e, d1, k) in x]
            if not vars_d1: continue
            # cross-month: forbid if previous assignment is within 2 days
            if last_ord != -999 and (d1.toordinal() - last_ord) <= 2:
                for v in vars_d1: model.add(v == 0)
            # within-month pairs
            for offset in [1, 2]:
                if i + offset >= len(day_list): break
                d2 = day_list[i + offset]
                vars_d2 = [x[(e, d2, k)] for k in DEPTS if (e, d2, k) in x]
                if vars_d2:
                    model.add(sum(vars_d1) + sum(vars_d2) <= 1)

    # HC-5: Wed-Sat rule (מתמחה only)
    day_set_lookup = set(target_dates)
    for e in [e for e in employees if emp_type.get(e) == 'מתמחה']:
        for d in target_dates:
            vars_d = [x[(e, d, k)] for k in DEPTS if (e, d, k) in x]
            if not vars_d: continue
            if d.weekday() == 2:  # Wednesday → forbid Fri (+2) and Sat (+3)
                for gap in [2, 3]:
                    d2 = d + timedelta(days=gap)
                    if d2 in day_set_lookup:
                        vars_d2 = [x[(e, d2, k)] for k in DEPTS if (e, d2, k) in x]
                        if vars_d2:
                            model.add(sum(vars_d) + sum(vars_d2) <= 1)
            if d.weekday() == 5:  # Saturday → forbid Wed (-3)
                d2 = d - timedelta(days=3)
                if d2 in day_set_lookup:
                    vars_d2 = [x[(e, d2, k)] for k in DEPTS if (e, d2, k) in x]
                    if vars_d2:
                        model.add(sum(vars_d) + sum(vars_d2) <= 1)

    # HC-6: weekend quota
    wknd_days_by_week = {}
    for d in target_dates:
        if is_functional_weekend(d, st.session_state.special_days):
            w = d.isocalendar()[1]
            wknd_days_by_week.setdefault(w, []).append(d)
    for e in employees:
        pre_wknd = weekends_worked.get(e, set())
        new_wknd_bools = []
        for w, wdays in wknd_days_by_week.items():
            if w in pre_wknd: continue
            wvars = [x[(e, d, k)] for d in wdays for k in DEPTS if (e, d, k) in x]
            if not wvars: continue
            wb = model.new_bool_var(f"wknd|{e}|{w}")
            model.add(sum(wvars) >= 1).only_enforce_if(wb)
            model.add(sum(wvars) == 0).only_enforce_if(wb.negated())
            new_wknd_bools.append(wb)
        if new_wknd_bools:
            model.add(sum(new_wknd_bools) + len(pre_wknd) <= emp_wq.get(e, 0))

    # HC-7: wish priority — if eligible wishers exist for a slot, only they may fill it
    for d, k in optimizable_slots:
        d_str = str(d)
        wisher_vars    = [x[(e, d, k)] for e in employees if (e, d_str) in wish_set    and (e, d, k) in x]
        nonwisher_vars = [x[(e, d, k)] for e in employees if (e, d_str) not in wish_set and (e, d, k) in x]
        if wisher_vars and nonwisher_vars:
            model.add(sum(nonwisher_vars) == 0)

    # ── 4. Objective (same weights as greedy scoring) ───────────────────────
    FILL_BONUS = 50_000
    obj_vars   = []
    obj_coeffs = []

    for e in employees:
        q        = max(emp_quota.get(e, 1), 1)
        last_ord = last_assignment.get(e, -999)
        for d, k in optimizable_slots:
            v = x.get((e, d, k))
            if v is None: continue
            c  = FILL_BONUS
            c += int(-100 * work_load.get(e, 0) / q)
            days_diff = d.toordinal() - last_ord if last_ord != -999 else 30
            c += int(2 * days_diff)
            expected = q * (d.day / num_days)
            c += int(500 * (expected - work_load.get(e, 0)))
            if emp_type.get(e) == 'תורן חוץ' and k == "שיקום":
                if d.weekday() in [3, 4, 5] or is_functional_weekend(d, st.session_state.special_days):
                    c += 2000
            if emp_type.get(e) == 'מתמחה':
                if d.weekday() == 2: c += int(-200 * wed_counts.get(e, 0))
                if d.weekday() == 3: c += int(-200 * thu_counts.get(e, 0))
            home = emp_home.get(e, '')
            c += 500 if home in (k, 'כללי') else -5000
            if (e, str(d)) in wish_set:
                c += 1000
            obj_vars.append(v)
            obj_coeffs.append(int(c))

    if obj_vars:
        model.maximize(_cp.LinearExpr.weighted_sum(obj_vars, obj_coeffs))

    # ── 5. Solve Phase 1 (primary, 28 s) ───────────────────────────────────
    solver = _cp.CpSolver()
    solver.parameters.max_time_in_seconds = 28.0
    try:
        solver.parameters.num_workers = 4
    except Exception:
        pass

    status = solver.solve(model)

    if status in (_cp.INFEASIBLE, _cp.UNKNOWN):
        st.warning("⚠️ CP-SAT לא הצליח למצוא פתרון — עובר לשיבוץ חמדני")
        run_smart_scheduling(year, month, only_weekends)
        return

    # Extract Phase 1 assignments
    filled_slots = set()
    for e in employees:
        for d, k in optimizable_slots:
            v = x.get((e, d, k))
            if v is not None and solver.value(v) == 1:
                d_str = str(d)
                new_schedule.append({'date': d_str, 'dept': k, 'employee': e,
                                     'is_manual': False, 'empty_reason': ''})
                filled_slots.add((d_str, k))
                work_load[e]        = work_load.get(e, 0) + 1
                d_ord               = d.toordinal()
                if d_ord > last_assignment.get(e, -999): last_assignment[e] = d_ord
                if d.weekday() == 2: wed_counts[e] = wed_counts.get(e, 0) + 1
                if d.weekday() == 3: thu_counts[e] = thu_counts.get(e, 0) + 1
                if is_functional_weekend(d, st.session_state.special_days):
                    weekends_worked[e].add(d.isocalendar()[1])

    # ── 6. Phase 2: relax rest-gap + weekend-quota for unfilled slots ───────
    unfilled     = [(d, k) for d, k in optimizable_slots if (str(d), k) not in filled_slots]
    fallback_cnt = 0

    if unfilled:
        model2 = _cp.CpModel()
        x2     = {}
        for e in employees:
            for d, k in unfilled:
                d_str = str(d)
                if (e, d_str) in blocked_set: continue
                if k == "פנימית גריאטרית" and emp_type.get(e) == 'תורן חוץ': continue
                if work_load.get(e, 0) >= emp_quota.get(e, 6): continue
                # same-day conflict with phase-1 assignments
                if any(s['date'] == d_str and s['employee'] == e for s in new_schedule
                       if not str(s.get('empty_reason','')).startswith('לא נמצא')): continue
                x2[(e, d, k)] = model2.new_bool_var(f"x2|{e}|{d}|{k}")

        for d, k in unfilled:
            sv = [x2[(e, d, k)] for e in employees if (e, d, k) in x2]
            if sv: model2.add(sum(sv) <= 1)

        for e in employees:
            remaining = emp_quota.get(e, 6) - work_load.get(e, 0)
            all2 = [x2[(e, d, k)] for d, k in unfilled if (e, d, k) in x2]
            if all2:
                if remaining <= 0:
                    for v in all2: model2.add(v == 0)
                else:
                    model2.add(sum(all2) <= remaining)

        obj2_v, obj2_c = [], []
        for e in employees:
            for d, k in unfilled:
                v = x2.get((e, d, k))
                if v is not None:
                    obj2_v.append(v)
                    obj2_c.append(50000 - work_load.get(e, 0) * 100)
        if obj2_v:
            model2.maximize(_cp.LinearExpr.weighted_sum(obj2_v, obj2_c))

        solver2 = _cp.CpSolver()
        solver2.parameters.max_time_in_seconds = 10.0
        status2 = solver2.solve(model2)

        if status2 in (_cp.OPTIMAL, _cp.FEASIBLE):
            for e in employees:
                for d, k in unfilled:
                    v = x2.get((e, d, k))
                    if v is not None and solver2.value(v) == 1:
                        d_str = str(d)
                        new_schedule.append({'date': d_str, 'dept': k, 'employee': e,
                                             'is_manual': False,
                                             'empty_reason': 'הושלם מחוסר ברירה (הגמשת חוקים)'})
                        filled_slots.add((d_str, k))
                        work_load[e] = work_load.get(e, 0) + 1
                        fallback_cnt += 1

    # Mark truly empty slots
    still_empty = [(d, k) for d, k in optimizable_slots if (str(d), k) not in filled_slots]
    for d, k in still_empty:
        new_schedule.append({'date': str(d), 'dept': k, 'employee': '---',
                             'is_manual': False, 'empty_reason': 'לא נמצא פתרון (CP-SAT)'})

    # ── 7. Friday post-pass (verbatim from run_smart_scheduling) ───────────
    fridays = [d for d in all_dates if d.weekday() == 4
               or get_functional_day_type(d, st.session_state.special_days) == 'כמו שישי (ערב חג)']
    for fri_date in fridays:
        fri_str  = str(fri_date)
        sat_date = fri_date + timedelta(days=1)
        sat_str  = str(sat_date)

        fri_worker_pnimia = next((s['employee'] for s in new_schedule
                                   if s['date'] == fri_str and s['dept'] == 'פנימית גריאטרית'), None)
        sat_worker_pnimia = next((s['employee'] for s in new_schedule
                                   if s['date'] == sat_str and s['dept'] == 'פנימית גריאטרית'), None)
        has_pnimia_1 = any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == 'שישי בוקר - פנימית (1)')
        has_pnimia_2 = any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == 'שישי בוקר - פנימית (2)')
        if not has_pnimia_1 and fri_worker_pnimia and fri_worker_pnimia != '---':
            new_schedule.append({'date': fri_str, 'dept': 'שישי בוקר - פנימית (1)',
                                  'employee': fri_worker_pnimia, 'is_manual': False,
                                  'empty_reason': 'נגזר אוטומטית משישי'})
        if not has_pnimia_2 and sat_worker_pnimia and sat_worker_pnimia != '---':
            new_schedule.append({'date': fri_str, 'dept': 'שישי בוקר - פנימית (2)',
                                  'employee': sat_worker_pnimia, 'is_manual': False,
                                  'empty_reason': 'נגזר אוטומטית משבת'})

        fri_worker_rehab = next((s['employee'] for s in new_schedule
                                  if s['date'] == fri_str and s['dept'] == 'שיקום'), None)
        sat_worker_rehab = next((s['employee'] for s in new_schedule
                                  if s['date'] == sat_str and s['dept'] == 'שיקום'), None)

        def handle_rehab_morning(worker_name, source_day, slot_num):
            target_dept = f'שישי בוקר - שיקום ({slot_num})'
            if any(s for s in new_schedule if s['date'] == fri_str and s['dept'] == target_dept): return
            if not worker_name or worker_name == '---': return
            w_type = None
            if worker_name in staff_df['name'].values:
                w_type = staff_df[staff_df['name'] == worker_name]['type'].iloc[0]
            if w_type == 'מתמחה':
                new_schedule.append({'date': fri_str, 'dept': target_dept,
                                      'employee': worker_name, 'is_manual': False,
                                      'empty_reason': f'נגזר אוטומטית מ{source_day}'})
            else:
                cands = []
                _fri_ym = fri_str[:7]
                for _, row in staff_df.iterrows():
                    if row['type'] != 'מתמחה' or row['name'] == worker_name: continue
                    _row_dept = _emp_night_dept(row['name'], _fri_ym) or str(row.get('dept', '') or '')
                    if _row_dept == 'שיקום':
                        emp = row['name']
                        is_blocked = not st.session_state.requests[
                            (st.session_state.requests['employee'] == emp) &
                            (st.session_state.requests['date'] == fri_str) &
                            (st.session_state.requests['status'] == "אילוץ")
                        ].empty
                        if is_blocked: continue
                        has_rest = any(
                            s['employee'] == emp and s['date'] == str(fri_date + timedelta(days=off))
                            for s in new_schedule for off in [-2, -1, 1, 2]
                        )
                        if has_rest: continue
                        if any(s['employee'] == emp and s['date'] == fri_str for s in new_schedule): continue
                        if work_load.get(emp, 0) >= safe_int(row['monthly_quota'], 0): continue
                        fri_cnt = len([s for s in new_schedule if s['employee'] == emp and 'שישי בוקר' in s['dept']])
                        cands.append((emp, fri_cnt))
                if cands:
                    cands.sort(key=lambda z: z[1])
                    best = cands[0][0]
                    new_schedule.append({'date': fri_str, 'dept': target_dept,
                                          'employee': best, 'is_manual': False,
                                          'empty_reason': f'השלמה במקום {worker_name}'})
                    work_load[best] = work_load.get(best, 0) + 1
                else:
                    new_schedule.append({'date': fri_str, 'dept': target_dept,
                                          'employee': '---', 'is_manual': False,
                                          'empty_reason': 'לא נמצא מחליף לבוקר'})

        handle_rehab_morning(fri_worker_rehab, "שישי", "1")
        handle_rehab_morning(sat_worker_rehab, "שבת", "2")

    # ── 8. Finalize + save ──────────────────────────────────────────────────
    final_df = pd.DataFrame(new_schedule)
    if not final_df.empty:
        final_df = final_df.drop_duplicates(subset=['date', 'dept'], keep='first')
    st.session_state.schedule = final_df
    save_to_db("schedule", st.session_state.schedule)

    # ── 9. Summary banner ───────────────────────────────────────────────────
    n_total  = len(optimizable_slots)
    n_filled = n_total - len(still_empty)
    method   = "CP-SAT אופטימלי" if status == _cp.OPTIMAL else "CP-SAT פתרון ישים"
    if status == _cp.FEASIBLE:
        method += " (timeout)"
    fallback_note = f" ({fallback_cnt} בהגמשת חוקים)" if fallback_cnt else ""
    st.info(
        f"✅ שיבוץ הושלם ({method}) — "
        f"{n_filled}/{n_total} משמרות שובצו{fallback_note}.\n\n"
        f"{balancing_msg}"
    )

# --- כוננויות ---
def _render_konenut_tab(active_month_int):
    import calendar as _cal
    year = 2026
    _FREE_TEXT_OPT = "✏️ הקלד שם..."

    sel_month = st.selectbox(
        "חודש:", range(1, 13), index=active_month_int - 1,
        key="konenut_month_sel",
        format_func=lambda m: f"{m:02d}/{year}"
    )
    year_month = f"{year}-{sel_month:02d}"

    # Include ALL doctor types (רופא בכיר + מנהל מחלקה) — exclude interns and external
    try:
        _eligible_types = {'רופא בכיר', 'מנהל מחלקה', 'מנהל/ת', 'מנהל על'}
        _doc_names = sorted(
            st.session_state.staff[
                st.session_state.staff['type'].astype(str).str.strip().isin(_eligible_types)
            ]['name'].astype(str).str.strip().tolist()
        )
    except Exception:
        _doc_names = []
    # options: blank → known doctors → free-text marker
    seniors = [''] + _doc_names + [_FREE_TEXT_OPT]

    kdf = st.session_state.konenut
    month_kdf = kdf[kdf['date'].str.startswith(year_month)].set_index('date')

    # Calendar grid — same layout as draw_calendar_view (RTL: col0=Sat … col6=Sun)
    cal = [list(reversed(w)) for w in _cal.Calendar(firstweekday=6).monthdayscalendar(year, sel_month)]
    days_names = ["ש'", "ו'", "ה'", "ד'", "ג'", "ב'", "א'"]

    # Legend
    st.markdown(
        "<div style='font-size:0.78rem;color:#475569;margin-bottom:6px'>"
        "🏥 <b>פנימית</b> &nbsp;|&nbsp; 🦽 <b>שיקום 1</b> &nbsp;|&nbsp; 🦽 <b>שיקום 2</b>"
        "</div>",
        unsafe_allow_html=True,
    )

    # Header row
    hcols = st.columns(7)
    for i, name in enumerate(days_names):
        hcols[i].markdown(
            f"<div style='text-align:center;font-weight:bold;font-size:0.88rem;"
            f"padding:2px 0'>{name}</div>",
            unsafe_allow_html=True,
        )

    def _sel_idx(val):
        v = str(val).strip() if val else ''
        if v in seniors:
            return seniors.index(v)
        # saved value was free text — mark as free-text option
        return seniors.index(_FREE_TEXT_OPT) if _FREE_TEXT_OPT in seniors else 0

    def _slot(label, date_str, saved_val, sel_key, txt_key):
        """Render one on-call slot: selectbox + optional free-text input."""
        v = str(saved_val).strip() if saved_val else ''
        # If the saved value isn't in the dropdown it was free-typed → pre-select marker
        init_idx = _sel_idx(v)
        chosen = st.selectbox(label, seniors, index=init_idx,
                              key=sel_key, label_visibility="collapsed")
        if chosen == _FREE_TEXT_OPT:
            # pre-fill text box with the saved free-text if any
            init_txt = v if v not in seniors else ''
            return st.text_input("שם", value=init_txt, key=txt_key,
                                 placeholder="שם רופא/ה", label_visibility="collapsed")
        return chosen

    for week in cal:
        cols = st.columns(7)
        for i, day in enumerate(week):
            with cols[i]:
                if day == 0:
                    st.markdown("<div style='min-height:10px'></div>", unsafe_allow_html=True)
                    continue
                date_str = f"{year}-{sel_month:02d}-{day:02d}"
                existing = month_kdf.loc[date_str] if date_str in month_kdf.index else None
                is_wknd  = i <= 1
                bg       = "#fef9c3" if is_wknd else "#f1f5f9"

                st.markdown(
                    f"<div style='text-align:center;font-weight:700;font-size:0.82rem;"
                    f"background:{bg};border-radius:4px;padding:2px 0;margin-bottom:2px'>"
                    f"{day}</div>",
                    unsafe_allow_html=True,
                )

                pnim_v = existing['pnim_dr']   if existing is not None else ''
                r1_v   = existing['rehab_dr1'] if existing is not None else ''
                r2_v   = existing['rehab_dr2'] if existing is not None else ''

                st.markdown("<div style='font-size:0.6rem;color:#9a3412;font-weight:600;line-height:1;margin-top:2px'>🏥פנ׳</div>", unsafe_allow_html=True)
                _slot("פ",  date_str, pnim_v, f"kn_p_sel_{date_str}",  f"kn_p_txt_{date_str}")
                st.markdown("<div style='font-size:0.6rem;color:#1e3a8a;font-weight:600;line-height:1;margin-top:2px'>🦽ש׳1</div>", unsafe_allow_html=True)
                _slot("ש1", date_str, r1_v,   f"kn_r1_sel_{date_str}", f"kn_r1_txt_{date_str}")
                st.markdown("<div style='font-size:0.6rem;color:#1e3a8a;font-weight:600;line-height:1;margin-top:2px'>🦽ש׳2</div>", unsafe_allow_html=True)
                _slot("ש2", date_str, r2_v,   f"kn_r2_sel_{date_str}", f"kn_r2_txt_{date_str}")

    st.divider()
    if st.button("💾 שמור שינויים", key="konenut_save", type="primary", use_container_width=True):
        num_days = _cal.monthrange(year, sel_month)[1]

        def _get_val(sel_key, txt_key):
            chosen = st.session_state.get(sel_key, '')
            if chosen == _FREE_TEXT_OPT:
                return (st.session_state.get(txt_key, '') or '').strip()
            return (chosen or '').strip()

        new_rows = []
        for d in range(1, num_days + 1):
            date_str = f"{year}-{sel_month:02d}-{d:02d}"
            new_rows.append({
                'date':      date_str,
                'pnim_dr':   _get_val(f"kn_p_sel_{date_str}",  f"kn_p_txt_{date_str}"),
                'rehab_dr1': _get_val(f"kn_r1_sel_{date_str}", f"kn_r1_txt_{date_str}"),
                'rehab_dr2': _get_val(f"kn_r2_sel_{date_str}", f"kn_r2_txt_{date_str}"),
            })
        new_month_df = pd.DataFrame(new_rows)
        full = st.session_state.konenut
        full = full[~full['date'].str.startswith(year_month)]
        full = pd.concat([full, new_month_df], ignore_index=True)
        full = full.sort_values('date').reset_index(drop=True)
        st.session_state.konenut = full
        save_to_db("konenut", full)
        st.success("✅ הכוננויות נשמרו בהצלחה")
        st.rerun()


# --- 4. פונקציית ציור הלוח ---
def draw_calendar_view(year, month, role, user_name=None):
    # --- Pre-compute availability (admin only, only when ALL employees submitted) ---
    show_availability = False
    date_availability = {}  # {date_str: [available_names]}

    if role in ("מנהל/ת", "מנהל על"):
        month_prefix = f"{year}-{month:02d}"
        active_staff = st.session_state.staff[
            st.session_state.staff['type'].isin(['מתמחה', 'תורן חוץ']) &
            ~st.session_state.staff['name'].isin(['---', 'ADMIN'])
        ]
        reqs_month = st.session_state.requests[
            st.session_state.requests['date'].astype(str).str.startswith(month_prefix)
        ] if not st.session_state.requests.empty else pd.DataFrame()

        submitted_names  = set(reqs_month['employee'].astype(str).str.strip().unique()) if not reqs_month.empty else set()
        all_active_names = set(active_staff['name'].dropna().astype(str).str.strip().tolist())
        show_availability = bool(all_active_names) and all_active_names.issubset(submitted_names)

        if show_availability:
            # Build per-employee blocked/wished sets for this month
            blocked_dates = {}   # emp -> set of date strings
            wished_dates  = {}   # emp -> set of date strings
            if not reqs_month.empty:
                for _, r in reqs_month.iterrows():
                    emp = str(r['employee']).strip()
                    d   = str(r['date'])[:10]
                    if r['status'] == 'אילוץ':
                        blocked_dates.setdefault(emp, set()).add(d)
                    elif r['status'] == 'בקשה':
                        wished_dates.setdefault(emp, set()).add(d)

            num_days = calendar.monthrange(year, month)[1]
            sched_records = st.session_state.schedule.to_dict('records') if not st.session_state.schedule.empty else []

            date_wished = {}  # {date_str: [wished_names]}
            for day in range(1, num_days + 1):
                d_str = f"{year}-{month:02d}-{day:02d}"
                assigned_today = {s['employee'] for s in sched_records if str(s['date']) == d_str and s['employee'] != '---'}
                wished   = []
                available = []
                for _, emp in active_staff.iterrows():
                    name = str(emp['name']).strip()
                    if d_str in blocked_dates.get(name, set()):  continue  # blocked
                    if name in assigned_today:                    continue  # already scheduled
                    if d_str in wished_dates.get(name, set()):
                        wished.append(name)
                    else:
                        available.append(name)
                date_wished[d_str]       = wished
                date_availability[d_str] = available

    # Toggle for Mobile View (List vs Grid)
    # Mobile Detection
    # Mobile Detection
    if 'mobile_detected_persistent' not in st.session_state:
        st.session_state.mobile_detected_persistent = False
        
    try:
        from streamlit_javascript import st_javascript
        # Check User Agent and Width — run once per session, not every rerun
        if not st.session_state.get('analytics_device_captured'):
            ua_string = st_javascript("window.navigator.userAgent", key="ua_check_1")
            ui_width  = st_javascript("window.innerWidth",          key="width_check_1")
        else:
            ua_string = st.session_state.get('analytics_ua', '')
            ui_width  = st.session_state.get('analytics_vp_width', 0) or 0
        
        # 1. User Agent Check
        if ua_string and isinstance(ua_string, str):
             if any(x in ua_string for x in ["Android", "iPhone", "iPad", "Mobile", "webOS"]):
                 st.session_state.mobile_detected_persistent = True
                 
        # 2. Width Check (Backup)
        if ui_width and isinstance(ui_width, int) and 300 < ui_width < 768:
             st.session_state.mobile_detected_persistent = True
             
    except:
        pass

    is_mobile_view = st.toggle("📱 תצוגת רשימה", value=st.session_state.mobile_detected_persistent, key="mobile_list_view")

    cal = [list(reversed(w)) for w in calendar.Calendar(firstweekday=6).monthdayscalendar(year, month)]

    if is_mobile_view:
        # ── Night-shift mobile: one collapsible card per day ────────────────
        WD_FULL_N = ["ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת"]

        # Collect special days
        month_special_days = {}
        if "special_days" in st.session_state and not st.session_state.special_days.empty:
            for _, row in st.session_state.special_days.iterrows():
                month_special_days[str(row["date"])[:10]] = row["description"]

        month_sched = st.session_state.schedule
        num_days_m  = calendar.monthrange(year, month)[1]

        for day in range(1, num_days_m + 1):
            date_str = f"{year}-{month:02d}-{day:02d}"
            wd_idx   = (date(year, month, day).weekday() + 1) % 7
            wd_name  = WD_FULL_N[wd_idx]
            is_wknd  = wd_idx in (5, 6)

            day_rows = month_sched[month_sched["date"].astype(str) == date_str].copy()
            if role != "מנהל/ת":
                day_rows = day_rows[
                    (day_rows["employee"].astype(str).str.strip() == str(user_name).strip()) |
                    (day_rows["employee"] == "---")
                ]
            if day_rows.empty and role != "מנהל/ת":
                continue

            # Build display groups: main depts / friday morning / empty slots
            mains, friday_m, empty_slots = [], [], []
            for _, row in day_rows.iterrows():
                emp  = str(row["employee"]).replace("👤", "").replace("🛡️", "").replace("🍼", "").strip()
                dept = str(row["dept"])
                if emp == "---" or emp == "":
                    empty_slots.append(dept)
                elif "שישי בוקר" in dept:
                    friday_m.append((dept, emp))
                else:
                    mains.append((dept, emp))

            sd_suffix = f" — {month_special_days[date_str]}" if date_str in month_special_days else ""
            n_assigned = len(mains) + len(friday_m)
            n_empty    = len(empty_slots)

            with st.expander(
                f"📅 {day}/{month} {wd_name}{sd_suffix}  ✅ {n_assigned}  ⚠️ {n_empty}",
                expanded=False,
            ):
                if mains:
                    st.markdown(
                        "<div style='font-weight:700;font-size:0.85rem;margin-bottom:4px'>✅ משמרות לילה</div>",
                        unsafe_allow_html=True)
                    for dept, emp in mains:
                        dept_color = "#1e3a8a" if "שיקום" in dept else "#9a3412"
                        st.markdown(
                            f"<div style='margin-right:6px;font-size:0.92rem;padding:2px 0'>"
                            f"<span style='color:{dept_color};font-weight:600'>{dept}:</span> {emp}</div>",
                            unsafe_allow_html=True)
                if friday_m:
                    st.markdown(
                        "<div style='font-weight:700;font-size:0.85rem;margin:6px 0 4px'>☀️ שישי בוקר</div>",
                        unsafe_allow_html=True)
                    for dept, emp in friday_m:
                        st.markdown(
                            f"<div style='margin-right:6px;font-size:0.92rem;padding:2px 0'>"
                            f"<span style='color:#854d0e;font-weight:600'>{dept}:</span> {emp}</div>",
                            unsafe_allow_html=True)
                if empty_slots:
                    st.markdown(
                        "<div style='font-weight:700;color:#991b1b;font-size:0.85rem;margin:6px 0 4px'>"
                        "⚠️ ריקים</div>",
                        unsafe_allow_html=True)
                    for dept in empty_slots:
                        st.markdown(
                            f"<div style='color:#991b1b;font-size:0.85rem;margin-right:6px'>{dept}</div>",
                            unsafe_allow_html=True)
                if not mains and not friday_m and not empty_slots:
                    st.caption("אין שיבוצים")

                # Availability panel (admin only, after all submitted)
                if show_availability:
                    wished_here = date_wished.get(date_str, [])
                    avail       = date_availability.get(date_str, [])
                    if wished_here:
                        st.markdown(
                            f"<div style='font-size:11px;color:#854d0e;margin-top:4px'>⭐ ביקשו: {', '.join(wished_here)}</div>",
                            unsafe_allow_html=True)
                    if avail:
                        st.markdown(
                            f"<div style='font-size:11px;color:#166534'>✅ זמינים: {', '.join(avail)}</div>",
                            unsafe_allow_html=True)
                    if not wished_here and not avail:
                        st.markdown(
                            "<div style='font-size:11px;color:#991b1b'>⚠️ אין זמינים</div>",
                            unsafe_allow_html=True)
    else:
        # --- Standard Grid View ---
        # RTL: col 0=Saturday, col 6=Sunday (reads right→left: Sun … Fri Sat)
        days_names = ["ש'", "ו'", "ה'", "ד'", "ג'", "ב'", "א'"]
        st.markdown('<div class="calendar-grid-container">', unsafe_allow_html=True)
        header_cols = st.columns(7)
        for i, name in enumerate(days_names):
            header_cols[i].markdown(f"<div style='text-align: center; font-weight: bold;'>{name}</div>", unsafe_allow_html=True)

        for week in cal:
            cols = st.columns(7)
            for i, day in enumerate(week):
                if day == 0: continue
                with cols[i]:
                    date_str = f"{year}-{month:02d}-{day:02d}"
                    is_weekend = "weekend-day" if i <= 1 else ""  # col 0=Sat, col 1=Fri
                    day_sched = st.session_state.schedule[st.session_state.schedule['date'] == date_str]
                    
                    html = f'<div class="calendar-day {is_weekend}"><div class="day-number">{day}</div>'
                    
                    # Add special day note in grid
                    if 'special_days' in st.session_state and not st.session_state.special_days.empty:
                        sd_match = st.session_state.special_days[st.session_state.special_days['date'] == date_str]
                        if not sd_match.empty:
                            sd_desc = sd_match.iloc[0]['description']
                            html += f'<div style="font-size:10px; color:#b91c1c; font-weight:bold; text-align:center; padding-bottom:5px;">{sd_desc}</div>'
                            
                    for dept in ["שיקום", "פנימית גריאטרית", "שישי בוקר - שיקום (1)", "שישי בוקר - שיקום (2)", "שישי בוקר - פנימית (1)", "שישי בוקר - פנימית (2)"]:
                        rows = day_sched[day_sched['dept'] == dept]
                        # אם מדובר בשישי בוקר ואין שורה כזו (כי זה לא יום שישי), דלג
                        if "שישי בוקר" in dept and rows.empty: continue
                        
                        # שינוי: ריצה על כל השורות שנמצאו (כדי לתמוך בכפילויות, למשל 2 תורני בוקר)
                        for _, row in rows.iterrows():
                            val = row['employee']
                            reason = row['empty_reason'] if val == "---" else ""
                            
                            # פילטור עבור מתמחים - רואים רק את השיבוצים של עצמם
                            if role != "מנהל/ת" and val != user_name and val != "---":
                                continue
                                
                            css = "shikum-slot" if "שיקום" in dept else "pnimia-slot"
                            if val == "---": css = "empty-slot"
                            
                            label = "שיקום" if dept == "שיקום" else "פנימית"
                            if "שישי בוקר" in dept: label = "בוקר (" + ("שיקום" if "שיקום" in dept else "פנימית") + ")"
                            
                            html += f'<div class="slot {css}"><span class="dept-label">{label}</span> <span>{val}</span>'
                            if role in ("מנהל/ת", "מנהל על") and reason:
                                html += f'<span class="error-hint">❓ {reason}</span>'
                            html += '</div>'

                    # הצגת אילוצים ובקשות (למנהל בלבד או לעובד על עצמו)
                    if role in ("מנהל/ת", "מנהל על"):
                        reqs = st.session_state.requests[st.session_state.requests['date'] == date_str]
                        for _, r in reqs.iterrows():
                            icon = "❌" if r['status'] == "אילוץ" else "⭐"
                            html += f'<div style="font-size:10px; color:{"#991b1b" if r["status"] == "אילוץ" else "#eab308"};">{icon} {r["employee"]}</div>'

                    # Availability panel — shown only after all employees submitted
                    if show_availability:
                        wished_here = date_wished.get(date_str, [])
                        avail       = date_availability.get(date_str, [])
                        avail_html  = ''
                        if wished_here:
                            avail_html += (
                                f'<div style="font-size:9px; color:#854d0e; margin-top:4px;'
                                f' border-top:1px dashed #fde68a; padding-top:3px; line-height:1.5;">'
                                f'⭐ ביקשו:<br>{"<br>".join(wished_here)}</div>'
                            )
                        if avail:
                            avail_html += (
                                f'<div style="font-size:9px; color:#166534; margin-top:2px;'
                                f' line-height:1.5;">'
                                f'✅ זמינים:<br>{"<br>".join(avail)}</div>'
                            )
                        if not wished_here and not avail:
                            avail_html = '<div style="font-size:9px; color:#991b1b; margin-top:4px; border-top:1px dashed #fca5a5; padding-top:3px;">⚠️ אין זמינים</div>'
                        html += avail_html

                    st.markdown(html + "</div>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # --- Smart Swap Assistant (Manager Only) ---
    if role in ("מנהל/ת", "מנהל על"):
        st.divider()
        st.subheader("🤖 עוזר החלפות חכם")
        st.caption("כלי זה מציע החלפות משמרת לפי חוקיות: פנויים, החלפות הדדיות, והחלפות משולשות.")
        
        # On mobile, we avoid tight column wrapping by using stack layout logic (handled natively if grid block removed above)
        c1, c2 = st.columns(2)
        # Safe Date Input
        try:
            default_d = date(year, month, 1)
            max_d = date(year, month, calendar.monthrange(year, month)[1])
        except:
             default_d = date(2026, 1, 1)
             max_d = date(2026, 12, 31)

        target_date_swap = c1.date_input("תאריך לבדיקה", value=default_d, min_value=default_d, max_value=max_d)
        target_dept_swap = c2.selectbox("מחלקה/משמרת", ["שיקום", "פנימית גריאטרית", "שישי בוקר - שיקום (1)", "שישי בוקר - שיקום (2)", "שישי בוקר - פנימית (1)", "שישי בוקר - פנימית (2)"])
        
        # Analyze Current Slot
        t_date_str = str(target_date_swap)
        sche_df = st.session_state.schedule
        
        # Robust filtering with strip() to avoid whitespace mismatches
        current_assignment = sche_df[
            (sche_df['date'] == t_date_str) & 
            (sche_df['dept'].astype(str).str.strip() == target_dept_swap.strip())
        ]
        
        current_emp = "---"
        if not current_assignment.empty:
            current_emp = current_assignment.iloc[0]['employee']
        
        st.info(f"**מצב נוכחי:** {target_dept_swap} ב-{target_date_swap.strftime('%d/%m')}: **{current_emp}**")
        
        if st.button("🔍 מצא החלפות אפשריות"):
            candidates_direct = []
            candidates_exchange = []
            candidates_triple = []
            
            # Debug counters
            failure_reasons = {}
            debug_log = []
            
            # Pre-fetch data
            staff_df = st.session_state.staff
            requests_df = st.session_state.requests
            
            # CRITICAL FIX: Limit schedule analysis STRICTLY to the target month to prevent March swaps in February
            t_month_prefix = t_date_str[:7]
            schedule_records = [s for s in sche_df.to_dict('records') if str(s['date']).startswith(t_month_prefix)]
            
            debug_log.append(f"Target: {target_dept_swap} on {t_date_str}, Current: {current_emp}")
            
            # Count how many people work on the target date
            workers_on_date = [s for s in schedule_records if str(s['date']) == t_date_str]
            debug_log.append(f"Workers on {t_date_str}: {len(workers_on_date)}")
            for w in workers_on_date:
                debug_log.append(f"  - {w['employee']} @ {w['dept']}")
            
            # 1. Direct Replacements (Free & Valid)
            # 2. Exchanges (Busy but can swap)
            
            for _, person in staff_df.iterrows():
                p_name = person['name']
                # Filter out ADMIN (Case Insensitive), placeholder, and current employee
                if str(p_name).upper() == 'ADMIN' or p_name == '---' or p_name == current_emp: continue
                
                # Check validity for TARGET spot
                # Relaxed constraints for Manager Swap: Quota, Home Dept, Rest - but keep External/Internal check
                is_valid, reason = check_assignment_validity(
                    schedule_records, p_name, t_date_str, target_dept_swap, staff_df, requests_df,
                    ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                )

                if is_valid:
                    candidates_direct.append(p_name)
                else:
                    # Log failure reason
                    failure_reasons[reason] = failure_reasons.get(reason, 0) + 1
                    
                    # Logic: If failed due to a constraint we can relax for Swaps (because they are busy and we keep invariant)
                    ignorable_for_swap = ["Quota Exceeded", "Restricted to Home Dept", "Rest Violation", "Already working"]
                    is_ignorable = any(ign in reason for ign in ignorable_for_swap)
                    
                    if is_ignorable:
                        # Check if they are actually working today (Condition for Exchange)
                        other_spot = next((s for s in schedule_records if str(s['date']) == t_date_str and s['employee'] == p_name), None)
                        
                        if other_spot:
                            other_dept = other_spot['dept']
                            
                            # Re-validate B for Target with relaxed rules
                            is_valid_relaxed, reason_relaxed = check_assignment_validity(
                                schedule_records, p_name, t_date_str, target_dept_swap, staff_df, requests_df,
                                ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                            )
                            
                            # If valid (or blocked only by "Already working" which is expected)
                            if is_valid_relaxed or reason_relaxed == "Already working this day":
                                
                                # --- Exchange Check (A <-> B) ---
                                # Can Current Emp (A) take Other Spot (B's spot)?
                                if current_emp != "---":
                                    valid_for_a, reason_a = check_assignment_validity(
                                        schedule_records, current_emp, t_date_str, other_dept, staff_df, requests_df, 
                                        ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                                    )
                                    
                                    if valid_for_a or reason_a == "Already working this day": 
                                        # --- Strict Constraints Check ---
                                        # 1. Morning <-> Morning
                                        is_target_morning = 'שישי בוקר' in target_dept_swap or 'שיקום (1)' in target_dept_swap or 'שיקום (2)' in target_dept_swap
                                        is_other_morning = 'שישי בוקר' in other_dept or 'שיקום (1)' in other_dept or 'שיקום (2)' in other_dept
                                        
                                        # 2. Weekday <-> Weekday
                                        t_date_obj = datetime.strptime(t_date_str, '%Y-%m-%d')
                                        is_target_weekend = t_date_obj.weekday() in [4, 5]
                                        # Since exchange is same day, date is same. So this check is always passed logic-wise.
                                        
                                        if is_target_morning == is_other_morning:
                                             candidates_exchange.append({'name': p_name, 'dept': other_dept})

                            # --- Triple Swap Skipped (User Request) ---
            
            # Store results in Session State
            st.session_state['swap_results'] = {
                'direct': candidates_direct,
                'exchange': candidates_exchange,
                'triple': candidates_triple,
                'date': t_date_str,
                'dept': target_dept_swap,
                'fail_reasons': failure_reasons
            }

            # --- Advanced Swaps (Cross-Date) ---
            # Added as per user request for "Mutual Cross-Date" and "Circular Cross-Date"
            
            advanced_log = []
            
            # Helper for strict constraints
            def check_advanced_constraints(date1, dept1, date2, dept2):
                # 1. Morning Swap Rule: Morning only with Morning
                is_m1 = 'שישי בוקר' in dept1 or 'שיקום (1)' in dept1 or 'שיקום (2)' in dept1
                is_m2 = 'שישי בוקר' in dept2 or 'שיקום (1)' in dept2 or 'שיקום (2)' in dept2
                if is_m1 != is_m2: return False
                
                # 2. Weekend Swap Rule: Weekend only with Weekend
                # Weekday: Sun(6), Mon(0)-Thu(3). Weekend: Fri(4), Sat(5).
                d1 = datetime.strptime(date1, '%Y-%m-%d')
                d2 = datetime.strptime(date2, '%Y-%m-%d')
                is_w1 = d1.weekday() in [4, 5]
                is_w2 = d2.weekday() in [4, 5]
                if is_w1 != is_w2: return False
                
                return True

            # Helper for consecutive check (Zero Gap)
            def is_creating_consecutive_violation(emp, new_date, schedule_data):
                # Check if emp works on new_date +/- 1 day
                d_obj = datetime.strptime(new_date, '%Y-%m-%d').date()
                for offset in [-1, 1]:
                    check_s = str(d_obj + timedelta(days=offset))
                    # Ignore if the conflict is the slot we are moving OUT of? 
                    # No, we assume we move TO new_date.
                    if any(s for s in schedule_data if s['date'] == check_s and s['employee'] == emp):
                        return True
                return False
            
            # 4. Mutual Cross-Date Swaps
            # Scenario: A is on Date 1 (Target). B is on Date 2.
            # Proposal: A goes to Date 2 (replacing B), B goes to Date 1 (replacing A).
            
            candidates_mutual_cross = []
            
            # We iterate over all OTHER assignments in the schedule (same month)
            # To optimize, we focus on the displayed month or the relevant range
            
            if current_emp != "---":
                for s_rec in schedule_records:
                    # Skip same date (already covered by Exchange)
                    if s_rec['date'] == t_date_str: continue
                    
                    # Skip empty slots or Admin
                    other_emp = s_rec['employee']
                    if other_emp == "---" or other_emp == 'ADMIN': continue
                    if other_emp == current_emp: continue # Can't swap with self
                    
                    other_date = s_rec['date']
                    other_dept = s_rec['dept']
                    
                    # Check 1: Can 'Other Emp' (B) move to 'Target Date' (Date 1) in 'Target Dept'?
                    # We use relaxed constraints because B is technically busy on Date 2, but we are moving them.
                    # We must ensure B is NOT busy on Date 1 (unless they are, which blocks them).
                    
                    valid_b_to_target, reason_b = check_assignment_validity(
                        schedule_records, other_emp, t_date_str, target_dept_swap, staff_df, requests_df,
                        ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                    )
                    
                    # If B is working on target date, they can't move here (unless we do complex 3-way, but let's keep simple)
                    if not valid_b_to_target:
                         # unique constraint: invalid because "Already working" is acceptable IF we were swapping same day, 
                         # but here we are swapping CROSS date. So if B is working on Date 1, they can't take A's spot on Date 1.
                         continue
                         
                    # Check 2: Can 'Current Emp' (A) move to 'Other Date' (Date 2) in 'Other Dept'?
                    # We need to verify A is not busy on Date 2 (unless they are, which blocks them).
                    
                    valid_a_to_other, reason_a = check_assignment_validity(
                        schedule_records, current_emp, other_date, other_dept, staff_df, requests_df,
                        ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                    )
                    
                    if not valid_a_to_other:
                        continue
                        
                    # --- NEW STRICT CHECKS ---
                    if not check_advanced_constraints(t_date_str, target_dept_swap, other_date, other_dept):
                         continue
                    
                    if is_creating_consecutive_violation(current_emp, other_date, schedule_records):
                         continue
                    if is_creating_consecutive_violation(other_emp, t_date_str, schedule_records):
                         continue
                    # -------------------------

                    # If both valid, we have a match!
                    candidates_mutual_cross.append({
                        'b_name': other_emp,
                        'b_date': other_date,
                        'b_dept': other_dept
                    })
            
            st.session_state['swap_results']['mutual_cross'] = candidates_mutual_cross


            # 5. Circular Cross-Date Swaps (A -> B, B -> C, C -> A)
            # Date 1 (Target): A is here. We want C here.
            # Date 2: B is here. We want A here.
            # Date 3: C is here. We want B here.
            
            candidates_circular = []
            
            # This is O(N^2) or O(N^3) search space. We must limit it.
            # We already know who can accept A (from Mutual search step 2).
            # Let's refine:
            
            # Step 1: Find potential "Date 2" slots where A can go.
            # list of (Date 2, Dept 2, Person B) where A -> (Date 2, Dept 2) is valid.
            
            potential_destinations_for_a = []
            if current_emp != "---":
                for s_rec in schedule_records:
                    if s_rec['date'] == t_date_str: continue # distinct dates
                    person_b = s_rec['employee']
                    if person_b == "---" or person_b == 'ADMIN' or person_b == current_emp: continue
                    
                    valid_a, _ = check_assignment_validity(
                        schedule_records, current_emp, s_rec['date'], s_rec['dept'], staff_df, requests_df,
                        ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                    )
                    if valid_a:
                        potential_destinations_for_a.append(s_rec)
            
            # Step 2: For each such slot, Person B needs to go somewhere (Date 3).
            # We look for a "Date 3" slot where Person C exists, and B can go there.
            # AND matching that C can go to "Date 1" (Target).
            
            # Optimization: Pre-calculate who can go to Target (Date 1).
            # list of Person C who can Validly work at (Target Date, Target Dept)
            potential_replacements_for_a = []
            for _, person in staff_df.iterrows():
                c_name = person['name']
                if c_name == current_emp or c_name == 'ADMIN' or c_name == '---': continue
                
                # C must not be working on Date 1
                is_busy_on_target = any(s['date'] == t_date_str and s['employee'] == c_name for s in schedule_records)
                if is_busy_on_target: continue
                
                val_c, _ = check_assignment_validity(
                    schedule_records, c_name, t_date_str, target_dept_swap, staff_df, requests_df,
                    ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                )
                if val_c:
                    potential_replacements_for_a.append(c_name)
                    
            # Now Step 2 loop
            import random
            # Limit loop if too huge
            for b_rec in potential_destinations_for_a[:50]: 
                person_b = b_rec['employee']
                date_2 = b_rec['date']
                dept_2 = b_rec['dept']
                
                # Find where B can go (Date 3) where Person C is one of 'potential_replacements_for_a'
                # We iterate over slots occupied by potential C's
                
                for c_name in potential_replacements_for_a:
                    if c_name == person_b: continue 
                    
                    # Find slots occupied by C (Date 3)
                    c_slots = [s for s in schedule_records if s['employee'] == c_name and s['date'] != t_date_str and s['date'] != date_2]
                    
                    for c_rec in c_slots:
                        date_3 = c_rec['date']
                        dept_3 = c_rec['dept']
                        
                        # Check: Can B go to (Date 3, Dept 3)?
                        valid_b, _ = check_assignment_validity(
                            schedule_records, person_b, date_3, dept_3, staff_df, requests_df,
                            ignore_quota=True, ignore_home_restrict=True, ignore_rest=True
                        )
                        
                        if valid_b:
                            # --- NEW STRICT CHECKS for Circular Swap ---
                            # A -> (Date 2, Dept 2)
                            # B -> (Date 3, Dept 3)
                            # C -> (Target Date, Target Dept)

                            # Check A's move
                            if not check_advanced_constraints(t_date_str, target_dept_swap, date_2, dept_2): continue
                            if is_creating_consecutive_violation(current_emp, date_2, schedule_records): continue
                            
                            # Check B's move
                            if not check_advanced_constraints(date_2, dept_2, date_3, dept_3): continue
                            if is_creating_consecutive_violation(person_b, date_3, schedule_records): continue

                            # Check C's move
                            if not check_advanced_constraints(date_3, dept_3, t_date_str, target_dept_swap): continue
                            if is_creating_consecutive_violation(c_name, t_date_str, schedule_records): continue
                            # -------------------------------------------

                            # Found a chain!
                            # A -> (Date 2, Dept 2) [Replacing B]
                            # B -> (Date 3, Dept 3) [Replacing C]
                            # C -> (Target Date, Target Dept) [Replacing A]
                            
                            candidates_circular.append({
                                'b_name': person_b, 'b_date': date_2, 'b_dept': dept_2,
                                'c_name': c_name, 'c_date': date_3, 'c_dept': dept_3
                            })
                            if len(candidates_circular) > 10: break # Limit results
                    if len(candidates_circular) > 10: break
                if len(candidates_circular) > 10: break
                
            st.session_state['swap_results']['circular'] = candidates_circular


        # --- Display Results from Session State ---
        if 'swap_results' in st.session_state and \
           st.session_state['swap_results']['date'] == t_date_str and \
           st.session_state['swap_results']['dept'] == target_dept_swap:
            
            res = st.session_state['swap_results']
            
            st.write("---")
            
            # 1. Direct - Available Replacements (main swap suggestions)
            st.markdown("##### ✅ מחליפים זמינים להחלפה")
            if res['direct']:
                st.success(f"נמצאו **{len(res['direct'])}** מחליפים זמינים עבור {target_dept_swap} ב-{t_date_str}:")
                
                # Selectbox to pick a replacement
                selected_replacement = st.selectbox(
                    "בחר מחליף/ה:", 
                    res['direct'], 
                    key="swap_select_direct"
                )
                
                if st.button("✅ בצע החלפה", key="do_selected_swap"):
                    # Remove old assignment
                    st.session_state.schedule = st.session_state.schedule[
                        ~((st.session_state.schedule['date'] == t_date_str) & (st.session_state.schedule['dept'] == target_dept_swap))
                    ]
                    # Add new assignment
                    new_row = {'date': t_date_str, 'dept': target_dept_swap, 'employee': selected_replacement, 'is_manual': True, 'empty_reason': ''}
                    st.session_state.schedule = pd.concat([st.session_state.schedule, pd.DataFrame([new_row])], ignore_index=True)
                    save_to_db("schedule", st.session_state.schedule)
                    
                    del st.session_state['swap_results']
                    st.success(f"בוצע! {selected_replacement} שובץ/ה במקום {current_emp}")
                    st.rerun()
            else:
                st.caption("לא נמצאו מחליפים זמינים.")

            # 2. Exchanges
            if current_emp != "---":
                st.markdown("##### 🔄 החלפות הדדיות (ראש בראש)")
                if res['exchange']:
                    for item in res['exchange']:
                        b = item['name']
                        b_dept = item['dept']
                        # Format date for display
                        d_disp = datetime.strptime(t_date_str, '%Y-%m-%d').strftime('%d/%m')
                        
                        c1, c2 = st.columns([3, 1])
                        c1.write(f"**{b}** ({b_dept} ב-**{d_disp}**) ↔️ **{current_emp}** ({target_dept_swap} ב-**{d_disp}**)")
                        if c2.button("החלף", key=f"do_swap_{b}"):
                            # Update DB - Swap Depts
                            mask_a = (st.session_state.schedule['date'] == t_date_str) & (st.session_state.schedule['dept'] == target_dept_swap)
                            mask_b = (st.session_state.schedule['date'] == t_date_str) & (st.session_state.schedule['dept'] == b_dept)
                            
                            st.session_state.schedule.loc[mask_a, 'employee'] = b
                            st.session_state.schedule.loc[mask_b, 'employee'] = current_emp
                            st.session_state.schedule.loc[mask_a | mask_b, 'is_manual'] = True # Mark as manual
                            
                            save_to_db("schedule", st.session_state.schedule)
                            
                            del st.session_state['swap_results']
                            st.success("ההחלפה בוצעה בהצלחה!")
                            st.rerun()
                else:
                    st.caption("לא נמצאו החלפות הדדיות מתאימות.")
            
            # 4. Mutual Cross-Date (New)
            st.markdown("##### 📅⚡ החלפות הדדיות בין תאריכים")
            st.caption(f"תרחיש: {current_emp} מחליף עם B בתאריך אחר.")
            
            candidates_mutual_cross = res.get('mutual_cross', [])
            if candidates_mutual_cross:
                for i, item in enumerate(candidates_mutual_cross[:5]):
                     b_name = item['b_name']
                     b_date = item['b_date']
                     b_dept = item['b_dept']
                     
                     # Displays
                     d_disp_current = datetime.strptime(t_date_str, '%Y-%m-%d').strftime('%d/%m')
                     d_disp_b = datetime.strptime(b_date, '%Y-%m-%d').strftime('%d/%m')
                     
                     c1, c2 = st.columns([3, 1])
                     c1.write(f"**{b_name}** ({b_dept} ב-**{d_disp_b}**) ↔️ **{current_emp}** ({target_dept_swap} ב-**{d_disp_current}**)")
                     
                     if c2.button("החלף", key=f"do_cross_mutual_{i}"):
                         # A goes to B's spot (Date 2)
                         # B goes to A's spot (Date 1)
                         
                         sched = st.session_state.schedule
                         
                         # Mask for A's spot (Date 1)
                         mask_a = (sched['date'] == t_date_str) & (sched['dept'] == target_dept_swap)
                         # Mask for B's spot (Date 2)
                         mask_b = (sched['date'] == b_date) & (sched['dept'] == b_dept)
                         
                         st.session_state.schedule.loc[mask_a, 'employee'] = b_name
                         st.session_state.schedule.loc[mask_a, 'is_manual'] = True
                         
                         st.session_state.schedule.loc[mask_b, 'employee'] = current_emp
                         st.session_state.schedule.loc[mask_b, 'is_manual'] = True
                         
                         save_to_db("schedule", st.session_state.schedule)
                         del st.session_state['swap_results']
                         st.success("החלפה הדדית בין תאריכים בוצעה!")
                         st.rerun()
            else:
                st.caption("לא נמצאו החלפות הדדיות בין תאריכים.")

            # 5. Circular Cross-Date (New)
            st.markdown("##### 📅🔄 מעגל החלפות (3 תאריכים)")
            st.caption(f"תרחיש: A (פה) ⬅️ B (תאריך 2) ⬅️ C (תאריך 3) ⬅️ A.")
            
            candidates_circular = res.get('circular', [])
            if candidates_circular:
                for i, item in enumerate(candidates_circular[:5]):
                    b_name = item['b_name']
                    b_date = item['b_date']
                    b_dept = item['b_dept']
                    
                    c_name = item['c_name']
                    c_date = item['c_date']
                    c_dept = item['c_dept']
                    
                    d_disp_1 = datetime.strptime(t_date_str, '%Y-%m-%d').strftime('%d/%m') # Target (A is here)
                    d_disp_2 = datetime.strptime(b_date, '%Y-%m-%d').strftime('%d/%m') # B is here
                    d_disp_3 = datetime.strptime(c_date, '%Y-%m-%d').strftime('%d/%m') # C is here
                    
                    c1, c2 = st.columns([3, 1])
                    c1.markdown(f"""
                    1. **{current_emp}** עובר אל {b_dept} (**{d_disp_2}**)
                    2. **{b_name}** עובר אל {c_dept} (**{d_disp_3}**)
                    3. **{c_name}** עובר אל {target_dept_swap} (**{d_disp_1}**)
                    """)
                    
                    if c2.button("בצע מעגל", key=f"do_circular_{i}"):
                        sched = st.session_state.schedule
                        
                        # A's Spot (Date 1)
                        mask_1 = (sched['date'] == t_date_str) & (sched['dept'] == target_dept_swap)
                        # B's Spot (Date 2)
                        mask_2 = (sched['date'] == b_date) & (sched['dept'] == b_dept)
                        # C's Spot (Date 3)
                        mask_3 = (sched['date'] == c_date) & (sched['dept'] == c_dept)
                        
                        # Apply Changes
                        # 1. C -> Spot 1
                        st.session_state.schedule.loc[mask_1, 'employee'] = c_name
                        st.session_state.schedule.loc[mask_1, 'is_manual'] = True
                        
                        # 2. A -> Spot 2
                        st.session_state.schedule.loc[mask_2, 'employee'] = current_emp
                        st.session_state.schedule.loc[mask_2, 'is_manual'] = True
                        
                        # 3. B -> Spot 3
                        st.session_state.schedule.loc[mask_3, 'employee'] = b_name
                        st.session_state.schedule.loc[mask_3, 'is_manual'] = True
                        
                        save_to_db("schedule", st.session_state.schedule)
                        del st.session_state['swap_results']
                        st.success("מעגל החלפות בוצע בהצלחה!")
                        st.rerun()
            else:
                 st.caption("לא נמצאו מעגלי החלפות.")


            # Debug Info Removed


# --- 5. ממשק המנהל והעובד ---
# Header Area
st.markdown('<div class="main-header">', unsafe_allow_html=True)

# FORCE INDIGO CHECKBOX CSS (Hardcoded Fix for Streamlit Cloud)
st.markdown("""
<style>
    /* Force Indigo Checkbox Color - Broad Selector */
    div[data-testid="stCheckbox"] *[aria-checked="true"] {
        background-color: #6366f1 !important;
        border-color: #6366f1 !important;
    }
    
    /* Force SVG Color */
    div[data-testid="stCheckbox"] *[aria-checked="true"] svg {
        fill: white !important;
        stroke: white !important;
    }
    
    /* Fallback for specific structure if broad fails */
    span[data-baseweb="checkbox"] div[aria-checked="true"] {
         background-color: #6366f1 !important;
    }

    /* ── Day-absence calendar ── */
    /* free / clickable day */
    div[class*="st-key-dayabs_free_"] > div[data-testid="stButton"] > button {
        background: white !important; color: #334155 !important;
        border: 1px solid #e2e8f0 !important;
        font-size: 0.78rem !important; min-height: 34px !important;
    }
    div[class*="st-key-dayabs_free_"] > div[data-testid="stButton"] > button:hover {
        background: #f0fdf4 !important; border-color: #86efac !important;
        color: #166534 !important; transform: none !important;
    }
    /* light green — pending חופש request */
    div[class*="st-key-dayabs_vac_"] > div[data-testid="stButton"] > button {
        background: #bbf7d0 !important; color: #166534 !important;
        border: 1px solid #86efac !important;
        font-size: 0.78rem !important; min-height: 34px !important; font-weight: 600 !important;
    }
    div[class*="st-key-dayabs_vac_"] > div[data-testid="stButton"] > button:hover {
        background: #fef9c3 !important; border-color: #fde68a !important;
        color: #854d0e !important; transform: none !important;
    }
    /* light yellow — pending 202 request */
    div[class*="st-key-dayabs_202_"] > div[data-testid="stButton"] > button {
        background: #fef9c3 !important; color: #854d0e !important;
        border: 1px solid #fde68a !important;
        font-size: 0.78rem !important; min-height: 34px !important; font-weight: 600 !important;
    }
    div[class*="st-key-dayabs_202_"] > div[data-testid="stButton"] > button:hover {
        background: white !important; border-color: #e2e8f0 !important;
        color: #334155 !important; transform: none !important;
    }
    /* non-clickable fixed states */
    .dayabs-vac   { background:#16a34a; color:white; border-radius:8px; padding:5px 2px;
                    text-align:center; font-size:0.78rem; min-height:34px;
                    display:flex; align-items:center; justify-content:center; }
    .dayabs-202   { background:#eab308; color:white; border-radius:8px; padding:5px 2px;
                    text-align:center; font-size:0.78rem; min-height:34px;
                    display:flex; align-items:center; justify-content:center; }
    .dayabs-night { background:#1e3a5f; color:white; border-radius:8px; padding:5px 2px;
                    text-align:center; font-size:0.78rem; min-height:34px;
                    display:flex; align-items:center; justify-content:center; }
    .dayabs-post  { background:#f97316; color:white; border-radius:8px; padding:5px 2px;
                    text-align:center; font-size:0.78rem; min-height:34px;
                    display:flex; align-items:center; justify-content:center; }
    .dayabs-empty { padding:5px 2px; min-height:34px; }
    .dayabs-hdr   { text-align:center; font-weight:700; color:#475569;
                    font-size:0.82rem; padding-bottom:4px; }

    /* ── WSD dept-grid cell coloring (Phase 6) ── */
    div[class*="st-key-wsdcell_"] > div[data-testid="stButton"] > button {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        font-size:0.7rem !important; min-height:30px !important; padding:2px 1px !important;
    }
    div[class*="st-key-wsdcell_w_"] > div[data-testid="stButton"] > button {
        background:#16a34a !important; color:white !important;
    }
    div[class*="st-key-wsdcell_h_"] > div[data-testid="stButton"] > button {
        background:#2563eb !important; color:white !important;
    }
    div[class*="st-key-wsdcell_2_"] > div[data-testid="stButton"] > button {
        background:#ca8a04 !important; color:white !important;
    }
    div[class*="st-key-wsdcell_p_"] > div[data-testid="stButton"] > button {
        background:#f97316 !important; color:white !important;
    }
    div[class*="st-key-wsdcell_a_"] > div[data-testid="stButton"] > button {
        background:#94a3b8 !important; color:white !important;
    }
    div[class*="st-key-wsdcell_t_"] > div[data-testid="stButton"] > button {
        background:#7c3aed !important; color:white !important;
    }
    /* Note icon buttons — minimal appearance */
    div[class*="st-key-notepop_"] button {
        background:transparent !important; color:#94a3b8 !important;
        border:1px solid #e2e8f0 !important;
        font-size:0.65rem !important; min-height:30px !important; padding:1px !important;
        border-radius:4px !important;
    }
    div[class*="st-key-notepop_"] button:hover {
        color:#4f46e5 !important; border-color:#4f46e5 !important;
    }
</style>
""", unsafe_allow_html=True)
st.title("מערכת סידור עבודה המערך הגריאטרי")

# Logout button centered under title for mobile robustness
if st.button("התנתק", key="logout_top", use_container_width=False):
    _login_time = st.session_state.get('analytics_login_time', datetime.now())
    _session_dur = int((datetime.now() - _login_time).total_seconds())
    _log_async('logout', str(_session_dur), 'explicit')
    st.session_state.logged_in = False
    st.rerun()
st.markdown('</div>', unsafe_allow_html=True)

role = st.session_state.user_role
user_name = st.session_state.user_name

# שליפת החודש הפעיל — cached with 30 s TTL so admin changes propagate quickly
# but don't trigger a new API call on every widget interaction.
try:
    if (time.time() - st.session_state.get('_settings_fetched_at', 0) > 30):
        st.session_state.settings = get_db_data("settings")
        st.session_state['_settings_fetched_at'] = time.time()
except:
    st.session_state.settings = pd.DataFrame([{'key': 'active_month', 'value': str(date.today().month + 1)}])

if 'key' not in st.session_state.settings.columns:
    st.session_state.settings = pd.DataFrame([{'key': 'active_month', 'value': str(date.today().month + 1)}])
    save_to_db("settings", st.session_state.settings)

try:
    active_month_setting = st.session_state.settings[st.session_state.settings['key'] == 'active_month']['value'].iloc[0]
    active_month_int = int(active_month_setting)
except:
    active_month_int = date.today().month

# ── Daily-schedule settings (Phase 2+) ──
def _parse_manage_depts(raw):
    """
    Split a comma-separated manage_depts string into a clean list of dept names.
    Normalises Hebrew Geresh ׳ (U+05F3) → ASCII apostrophe ' so manually-typed
    values like 'שיקום גריאטרי א׳' match DAILY_DEPTS_ALL entries exactly.
    """
    normalised = str(raw).strip().replace('׳', "'").replace('״', '״')
    return [d.strip() for d in normalised.split(',') if d.strip()]


# Managers who stay in `staff` (for approvals, emails, etc.) but must NOT appear
# as default rows in the admin schedule grid. They can still be added ad-hoc via
# the "העברה זמנית" widget when needed.
_MANUAL_PLANT_ONLY_MANAGERS = {"רון צליק", "רתם תלם"}


def _get_dept_managers(dept_name: str) -> list[str]:
    """Return all מנהל מחלקה staff names whose manage_depts includes dept_name.
    Managers listed in _MANUAL_PLANT_ONLY_MANAGERS are excluded — they must be
    planted manually (via cell click or the temporary-add widget)."""
    try:
        sf = st.session_state.staff.copy()
        sf['name'] = sf['name'].astype(str).str.strip()
        sf['type'] = sf['type'].astype(str).str.strip()
        out = []
        for _, r in sf[sf['type'] == 'מנהל מחלקה'].iterrows():
            _nm = str(r['name']).strip()
            if _nm in _MANUAL_PLANT_ONLY_MANAGERS:
                continue
            if dept_name in _parse_manage_depts(r.get('manage_depts', '')):
                out.append(_nm)
        return out
    except Exception:
        return []

_ROLE_ORDER = {'מנהל מחלקה': 0, 'רופא בכיר': 1, 'מתמחה': 2}

def _sort_employees_by_role(emp_list: list[str]) -> list[str]:
    """
    Sort a list of employee names by role:
      מנהל/ת מחלקה → רופא/ה בכיר/ה → מתמחה → everything else.
    Looks up each name in st.session_state.staff; unknown names go last.
    Preserves original order within the same role (stable sort).
    """
    try:
        sf = st.session_state.staff
        if sf.empty or 'name' not in sf.columns:
            return emp_list
        _name_to_role: dict[str, str] = {}
        for _, r in sf.iterrows():
            _name_to_role[str(r.get('name', '')).strip()] = str(r.get('type', '')).strip()
        return sorted(
            emp_list,
            key=lambda n: _ROLE_ORDER.get(_name_to_role.get(str(n).strip(), ''), 99)
        )
    except Exception:
        return emp_list


def _get_setting(key, default):
    try:
        rows = st.session_state.settings[st.session_state.settings['key'] == key]
        if not rows.empty:
            return rows['value'].iloc[0]
    except Exception:
        pass
    return default

def _set_setting(key, value):
    """Upsert a setting key+value, persist to sheet."""
    s = st.session_state.settings
    if (s['key'] == key).any():
        s.loc[s['key'] == key, 'value'] = str(value)
    else:
        s = pd.concat([s, pd.DataFrame([{'key': key, 'value': str(value)}])], ignore_index=True)
    st.session_state.settings = s
    save_to_db("settings", s)

def _sw_month_selector_12(key_ns: str) -> tuple[int, int]:
    """
    Auto-advancing 12-month forward selector for the 'סידור עבודה' tab.
    Returns (year, month). Recomputes the window from `datetime.now()` on every
    render so it rolls forward automatically when a new month begins.

    Example: in June 2026 → 'יוני 26 / יולי 26 / ... / מאי 27'.
             In July 2026 (one render later) → 'יולי 26 / ... / יוני 27'.
    """
    now = datetime.now()
    months = []
    for i in range(12):
        m_idx = (now.month - 1 + i) % 12
        y     = now.year + (now.month - 1 + i) // 12
        months.append((y, m_idx + 1))
    labels = [f"{_HEB_MONTHS[m - 1]} {str(y)[-2:]}" for (y, m) in months]
    sk = f"{key_ns}_vym"
    # Stale-selection guard: if a previously-stored (y, m) dropped out of the
    # rolling window, fall back to the current (now) month.
    cur_key = (now.year, now.month)
    cur = st.session_state.get(sk)
    if cur not in months:
        st.session_state[sk] = cur_key
        cur = cur_key
    cur_idx = months.index(cur)
    c1, _ = st.columns([3, 9])
    with c1:
        sel_label = st.selectbox(
            "חודש:", labels, index=cur_idx,
            key=f"{key_ns}_sel12", label_visibility="collapsed")
    new_idx = labels.index(sel_label)
    if months[new_idx] != cur:
        st.session_state[sk] = months[new_idx]
        st.rerun()
    return months[new_idx]


try:
    daily_active_month_int = int(_get_setting('daily_active_month', active_month_int))
except Exception:
    daily_active_month_int = active_month_int

daily_requests_open = str(_get_setting('daily_requests_open', 'True')).strip().lower() == 'true'

# Rebuild approved-absence map on every render (Streamlit reruns full script on every interaction).
# This ensures _derive_auto_status always sees current approved absences with O(1) per cell.
_build_approved_map()

# Render Navigation Bar
selected_nav = ui_components.render_navbar(role)

# --- אנליטיקת טאב: רישום כניסה ויציאה מטאבים ---
_prev_tab = st.session_state.get('analytics_last_tab', None)
if selected_nav != _prev_tab:
    _now_tab = datetime.now()
    if _prev_tab is not None:
        _secs_on_prev = int((_now_tab - st.session_state.get('analytics_tab_enter', _now_tab)).total_seconds())
        _log_async('tab_view', _prev_tab, str(_secs_on_prev))
    _log_async('tab_view', selected_nav, 'enter')
    st.session_state.analytics_last_tab = selected_nav
    st.session_state.analytics_tab_enter = _now_tab

# Month Selection Logic - MOVED TO SCHEDULE TAB ONLY (as requested)
# sel_month defaults to active month, but admins can override it LOCALLY in the schedule tab
sel_month = active_month_int 

# Admin Global control (Hidden here, moved logic down)
# We keep the "active_month_setter" logic but it needs to be where the selector is.
# So we skip the global expander here.



if selected_nav == 'הגדרות':
    st.subheader("⚙️ הגדרות משתמש")
    
    # --- Change Password Section ---
    with st.expander("🔑 שינוי סיסמה", expanded=True):
        with st.form("change_password_form"):
            st.caption("כאן ניתן לשנות את סיסמת ההתחברות למערכת.")
            
            curr_pass = st.text_input("סיסמה נוכחית", type="password")
            new_pass = st.text_input("סיסמה חדשה", type="password")
            conf_pass = st.text_input("אימות סיסמה חדשה", type="password")
            
            if st.form_submit_button("עדכן סיסמה"):
                # 1. Verify Current Password
                # Get user record
                user_record = st.session_state.staff[st.session_state.staff['name'] == user_name]
                if user_record.empty:
                    st.error("שגיאה: משתמש לא נמצא.")
                else:
                    stored_hash = user_record.iloc[0]['password']
                    input_hash = hashlib.sha256(curr_pass.encode()).hexdigest()
                    
                    if input_hash != stored_hash:
                        st.error("סיסמה נוכחית שגויה.")
                    elif not new_pass:
                        st.error("לא ניתן להגדיר סיסמה ריקה.")
                    elif new_pass != conf_pass:
                        st.error("הסיסמאות החדשות אינן תואמות.")
                    else:
                        # 2. Update Password
                        new_hash = hashlib.sha256(new_pass.encode()).hexdigest()
                        st.session_state.staff.loc[st.session_state.staff['name'] == user_name, 'password'] = new_hash
                        
                        # 3. Save to DB
                        save_to_db("staff", st.session_state.staff)
                        
                        # 4. Clear Cache to force reload on next login
                        st.cache_data.clear()
                        
                        st.success("הסיסמה עודכנה בהצלחה!")

    # --- Swap Search Section ---
    if role != "מנהל/ת":
        with st.expander("🔄 חיפוש החלפות", expanded=False):


            sched_df = st.session_state.schedule
            today_str = date.today().strftime('%Y-%m-%d')
            user_shifts = sched_df[
                (sched_df['employee'].astype(str).str.strip() == str(user_name).strip()) &
                (sched_df['date'].astype(str) >= today_str)
            ].sort_values('date').reset_index(drop=True)

            if user_shifts.empty:
                st.info("אין לך משמרות עתידיות משובצות.")
            else:
                DAY_NAMES_HE = {0: 'שני', 1: 'שלישי', 2: 'רביעי', 3: 'חמישי', 4: 'שישי', 5: 'שבת', 6: 'ראשון'}

                shift_options = []
                for _, s in user_shifts.iterrows():
                    d_obj = datetime.strptime(str(s['date']), '%Y-%m-%d').date()
                    label = f"{d_obj.strftime('%d/%m/%Y')} ({DAY_NAMES_HE[d_obj.weekday()]}) — {s['dept']}"
                    shift_options.append((label, str(s['date']), s['dept']))

                labels = ["— בחר/י משמרת —"] + [opt[0] for opt in shift_options]
                sel_label = st.selectbox("המשמרות הקרובות שלי:", labels, key=f"swap_search_select_upcoming")

                if sel_label != "— בחר/י משמרת —":
                    chosen = next(o for o in shift_options if o[0] == sel_label)
                    swap_date, swap_dept = chosen[1], chosen[2]
                    swap_month_int = int(swap_date.split('-')[1])
                    _log_async('swap_search', swap_date, swap_dept)

                    with st.spinner("מחפש החלפות..."):
                        results = find_swap_candidates(
                            st.session_state.schedule,
                            st.session_state.requests,
                            st.session_state.staff,
                            user_name, swap_date, swap_dept, swap_month_int
                        )

                    full = results['full']
                    partial = results['partial']
                    chain = results.get('chain', [])

                    if not full and not partial and not chain:
                        st.warning("לא נמצאו אפשרויות החלפה תקפות למשמרת זו.")
                    else:
                        if full:
                            st.markdown("##### ✅ החלפה מלאה")
                            st.caption("יכולים לכסות את המשמרת שלך, ויש להם משמרת שתוכל/י לקחת בתמורה")
                            for cand in full:
                                ts_d_obj = datetime.strptime(cand['their_shift']['date'], '%Y-%m-%d').date()
                                their_label = f"{ts_d_obj.strftime('%d/%m')} ({DAY_NAMES_HE[ts_d_obj.weekday()]}) — {cand['their_shift']['dept']}"
                                wish_pill = " ⭐" if cand['wished'] else ""

                                c1, c2 = st.columns([3, 1])
                                with c1:
                                    st.markdown(f"**{cand['name']}**{wish_pill}  \n_{cand['type']} · {cand['dept']}_  \n⇄ מציע/ה: **{their_label}**")
                                with c2:
                                    btn_key = f"swap_send_full_{cand['name']}_{swap_date}"
                                    if st.button("🔄 בקש החלפה", key=btn_key, use_container_width=True):
                                        new_req = pd.DataFrame([{
                                            'requester': user_name,
                                            'requester_date': swap_date,
                                            'requester_dept': swap_dept,
                                            'candidate': cand['name'],
                                            'candidate_date': cand['their_shift']['date'],
                                            'candidate_dept': cand['their_shift']['dept'],
                                            'swap_type': 'full',
                                            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                                            'status': 'pending',
                                        }])
                                        existing = st.session_state.swap_requests
                                        combined = new_req if existing.empty else pd.concat([existing, new_req], ignore_index=True)
                                        save_to_db("swap_requests", combined)
                                        _log_async('swap_request_sent', swap_date, 'full')
                                        st.session_state.swap_requests = combined
                                        st.success("הבקשה נשלחה למנהל/ת לאישור.")
                                st.divider()

                        if partial:
                            st.markdown("##### ⚠️ כיסוי חד-צדדי")
                            st.caption("יכולים לכסות את המשמרת שלך, אך לא נמצאה משמרת הדדית מתאימה")
                            for cand in partial:
                                wish_pill = " ⭐" if cand['wished'] else ""
                                c1, c2 = st.columns([3, 1])
                                with c1:
                                    st.markdown(f"**{cand['name']}**{wish_pill}  \n_{cand['type']} · {cand['dept']}_")
                                with c2:
                                    btn_key = f"swap_send_partial_{cand['name']}_{swap_date}"
                                    if st.button("🔄 בקש החלפה", key=btn_key, use_container_width=True):
                                        new_req = pd.DataFrame([{
                                            'requester': user_name,
                                            'requester_date': swap_date,
                                            'requester_dept': swap_dept,
                                            'candidate': cand['name'],
                                            'candidate_date': '',
                                            'candidate_dept': '',
                                            'swap_type': 'partial',
                                            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                                            'status': 'pending',
                                        }])
                                        existing = st.session_state.swap_requests
                                        combined = new_req if existing.empty else pd.concat([existing, new_req], ignore_index=True)
                                        save_to_db("swap_requests", combined)
                                        _log_async('swap_request_sent', swap_date, 'partial')
                                        st.session_state.swap_requests = combined
                                        st.success("הבקשה נשלחה למנהל/ת לאישור.")
                                st.divider()

                        if chain:
                            st.markdown("##### 🔗 החלפה משולשת")
                            st.caption("מתמחה עובר/ת מ-שיקום ל-פנימית, תורן חוץ מכסה שיקום")
                            for ch in chain:
                                fac_pill = " ⭐" if ch['facilitator_wished'] else ""
                                ext_pill = " ⭐" if ch['external_wished'] else ""
                                c1, c2 = st.columns([3, 1])
                                with c1:
                                    st.markdown(
                                        f"**{ch['facilitator_name']}**{fac_pill} ⇄ פנימית  +  "
                                        f"**{ch['external_name']}**{ext_pill} → שיקום"
                                    )
                                with c2:
                                    btn_key = f"swap_send_chain_{ch['facilitator_name']}_{ch['external_name']}_{swap_date}"
                                    if st.button("🔄 בקש החלפה", key=btn_key, use_container_width=True):
                                        new_req = pd.DataFrame([{
                                            'requester': user_name,
                                            'requester_date': swap_date,
                                            'requester_dept': swap_dept,
                                            'candidate': ch['facilitator_name'],
                                            'candidate_date': swap_date,
                                            'candidate_dept': 'פנימית גריאטרית',
                                            'swap_type': 'chain',
                                            'chain_ext': ch['external_name'],
                                            'chain_ext_dept': 'שיקום',
                                            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                                            'status': 'pending',
                                        }])
                                        existing = st.session_state.swap_requests
                                        combined = new_req if existing.empty else pd.concat([existing, new_req], ignore_index=True)
                                        save_to_db("swap_requests", combined)
                                        _log_async('swap_request_sent', swap_date, 'chain')
                                        st.session_state.swap_requests = combined
                                        st.success("הבקשה נשלחה למנהל/ת לאישור.")
                                st.divider()

    # --- Manage Special Days Section (מנהל על only) ---
    if role == "מנהל על":
        with st.expander("📅 ניהול ימים מיוחדים וחגים", expanded=False):
            st.caption("הוסף תאריכים מיוחדים כדי שמתמחים יראו אותם לפני הגשת אילוצים.")
            
            c_sd1, c_sd2, c_sd_type, c_sd3 = st.columns([1, 1.5, 1, 1])
            sd_date = c_sd1.date_input("תאריך ליום מיוחד:", format="DD/MM/YYYY")
            sd_desc = c_sd2.text_input("תיאור (למשל: ערב פסח):")
            sd_type = c_sd_type.selectbox("סוג יום:", ["לידיעה בלבד", "כמו שישי (ערב חג)", "כמו שבת (חג)"])
            
            if c_sd3.button("➕ הוסף יום", use_container_width=True):
                if sd_desc:
                    new_sd = pd.DataFrame([{'date': str(sd_date), 'description': sd_desc, 'day_type': sd_type}])
                    st.session_state.special_days = pd.concat([st.session_state.special_days, new_sd], ignore_index=True)
                    save_to_db("special_days", st.session_state.special_days)
                    st.success("היום המיוחד התווסף בהצלחה!")
                    st.rerun()
                else:
                    st.error("חובה להזין תיאור.")
            
            if not st.session_state.special_days.empty:
                st.write("ימים מיוחדים קיימים:")
                for idx, row in st.session_state.special_days.iterrows():
                    colA, colB, colC, colD = st.columns([1, 1.5, 1, 0.5])
                    colA.write(f"**{row['date']}**")
                    colB.write(row['description'])
                    colC.write(row.get('day_type', 'רגיל'))
                    if colD.button("🗑️", key=f"del_sd_{idx}"):
                        st.session_state.special_days = st.session_state.special_days.drop(idx).reset_index(drop=True)
                        save_to_db("special_days", st.session_state.special_days)
                        st.rerun()

elif role in ("מנהל/ת", "מנהל על"):
    if selected_nav == 'סידור כוננויות':
        _render_konenut_tab(active_month_int)

    elif selected_nav == 'סידור תורנויות':
        
        # --- Month Selector for Schedule Tab (Admin Only) ---
        with st.expander("הגדרות תצוגה", expanded=False):
            c_m1, c_m2 = st.columns(2)
            # Use a specialized key to avoid conflicts if previously used
            sel_month = c_m1.selectbox("חודש לצפייה/עריכה", range(1, 13), index=active_month_int - 1, key="sched_month_select")
            
            # Admin Global Control
            new_active_month = c_m2.selectbox(
                "חודש פתוח להגשת אילוצים:", 
                range(1, 13), 
                index=active_month_int - 1,
                key="admin_active_month_setter_sched"
            )
            if new_active_month != active_month_int:
                st.session_state.settings['value'] = st.session_state.settings['value'].astype(object)
                st.session_state.settings.loc[st.session_state.settings['key'] == 'active_month', 'value'] = str(new_active_month)
                save_to_db("settings", st.session_state.settings)
                st.success(f"החודש הפעיל עודכן ל-{new_active_month}")
                st.rerun()
        # ----------------------------------------------------

        # --- ייצוא ל-Schedule_Export (הועבר מטאב "דוחות וניהול") ---
        col_exp, _ = st.columns([1, 3])
        with col_exp:
            if st.button("📤 עדכן את הגיליון Schedule_Export",
                         key="export_schedule_top", use_container_width=True):
                ok, err = _export_schedule_wide(sel_month)
                if ok:
                    st.success("✅ הנתונים יוצאו לגיליון 'Schedule_Export'!")
                else:
                    st.error(f"שגיאה בייצוא: {err}")

        # --- הוספת כלי שיבוץ ידני ---
        with st.expander("כלי שיבוץ ידני (דריסה)", expanded=True):
            c_date, c_dept, c_emp, c_btn_add, c_btn_del = st.columns([1, 1, 1, 0.7, 0.7])
            
            # עדכון ערך ברירת מחדל לתאריך רק אם החודש השתנה בסרגל הצד
            # זה מאפשר לשמור על בחירת המשתמש בתוך אותו חודש
            default_date = date(2026, sel_month, 1)
            if 'manual_date' in st.session_state:
                if st.session_state.manual_date.month != sel_month:
                     st.session_state.manual_date = default_date

            d_man = c_date.date_input("תאריך:", key="manual_date", format="DD/MM/YYYY")
            dept_man = c_dept.selectbox("מחלקה:", ["שיקום", "פנימית גריאטרית", "שישי בוקר - שיקום (1)", "שישי בוקר - שיקום (2)", "שישי בוקר - פנימית (1)", "שישי בוקר - פנימית (2)"], key="manual_dept")
            
            # סינון רשימת העובדים - הצגת מי שמשובץ כרגע למעלה או סימון מיוחד? לא קריטי כרגע.
            emp_man = c_emp.selectbox("עובד:", st.session_state.staff['name'].tolist(), key="manual_emp")
            
            # כפתור שיבוץ
            if c_btn_add.button("✅ שיבוץ"):
                d_man_str = str(d_man)
                # Check for constraints
                is_blocked = False
                if not st.session_state.requests[(st.session_state.requests['employee'] == emp_man) & (st.session_state.requests['date'] == d_man_str) & (st.session_state.requests['status'] == "אילוץ")].empty:
                    is_blocked = True
                
                if is_blocked:
                    st.error(f"לא ניתן לשבץ את {emp_man} ב-{d_man_str}. יש לו אילוץ על יום זה.")
                else:
                    # הסרת שיבוץ קיים לתאריך ולמחלקה הזו (אם יש)
                    st.session_state.schedule = st.session_state.schedule[
                        ~((st.session_state.schedule['date'] == d_man_str) & 
                          (st.session_state.schedule['dept'] == dept_man))
                    ]
                    # הוספת השיבוץ הידני עם סימון ידני=True
                    new_entry = pd.DataFrame([{
                        'date': d_man_str, 
                        'dept': dept_man, 
                        'employee': emp_man, 
                        'is_manual': True, 
                        'empty_reason': ''
                    }])
                    st.session_state.schedule = pd.concat([st.session_state.schedule, new_entry], ignore_index=True)
                    save_to_db("schedule", st.session_state.schedule)
                    st.success(f"שובץ: {emp_man}")
                    st.rerun()
 
            # כפתור ביטול שיבוץ
            if c_btn_del.button("❌ בטל"):
                # מחיקת השיבוץ הספציפי הזה
                st.session_state.schedule = st.session_state.schedule[
                    ~((st.session_state.schedule['date'] == str(d_man)) & 
                      (st.session_state.schedule['dept'] == dept_man))
                ]
                save_to_db("schedule", st.session_state.schedule)
                st.success("השיבוץ בוטל")
                st.rerun()
        # ---------------------------
        c1, c2, c3 = st.columns(3)
        if c1.button("🪄 שיבוץ אוטומטי מלא"): run_smart_scheduling_cp(2026, sel_month, only_weekends=False); st.rerun()
        if c2.button("☕ שיבוץ סופ\"שים בלבד"): run_smart_scheduling_cp(2026, sel_month, only_weekends=True); st.rerun()
        with c3:
            with st.expander("🗑️ ניקוי"):
                if st.button("🧹 נקה אוטומטי (חודש זה)", help="מוחק שיבוצים אוטומטיים בחודש הנוכחי בלבד"):
                    current_prefix = f"2026-{sel_month:02d}"
                    
                    # Ensure is_manual column exists
                    if 'is_manual' not in st.session_state.schedule.columns:
                        st.session_state.schedule['is_manual'] = False
                    
                    # Logic: Delete if (Date matches prefix) AND (is_manual is NOT True)
                    mask_current_month = st.session_state.schedule['date'].astype(str).str.startswith(current_prefix)
                    
                    # Handle both boolean True and string "TRUE" from Google Sheets serialization
                    is_manual_s = st.session_state.schedule['is_manual']
                    mask_manual = (is_manual_s == True) | (is_manual_s.astype(str).str.upper() == 'TRUE')
                    mask_auto = ~mask_manual
                    mask_to_delete = mask_current_month & mask_auto
                    
                    st.session_state.schedule = st.session_state.schedule[~mask_to_delete]
                    
                    save_to_db("schedule", st.session_state.schedule)
                    st.success(f"שיבוצים אוטומטיים לחודש {sel_month}/2026 נמחקו (חודשים אחרים וידניים נשמרו).")
                    st.rerun()
                
                if st.button("💥 נקה הכל (חודש זה)", help="מוחק את כל הלוח לחודש זה"): 
                    current_prefix = f"2026-{sel_month:02d}"
                    
                    # Logic: Delete if Date matches prefix
                    mask_to_delete = st.session_state.schedule['date'].astype(str).str.startswith(current_prefix)
                    
                    st.session_state.schedule = st.session_state.schedule[~mask_to_delete]
                    save_to_db("schedule", st.session_state.schedule)
                    st.success(f"כל השיבוצים לחודש {sel_month}/2026 נמחקו.")
                    st.rerun()
        
        # --- התראה על משמרות שלא שובצו ---
        if not st.session_state.schedule.empty:
            failures = st.session_state.schedule[st.session_state.schedule['employee'] == '---']
            if not failures.empty:
                st.error(f"⚠️ שימו לב: נמצאו {len(failures)} משמרות שלא ניתן היה לשבץ!")
                with st.expander("🔻 לחץ כאן לפירוט השגיאות והסיבות", expanded=False):
                    for _, row in failures.iterrows():
                        # Format date for display
                        d_obj = datetime.strptime(row['date'], '%Y-%m-%d')
                        fmt_date = d_obj.strftime('%d/%m/%Y')
                        st.markdown(f"❌ **{fmt_date}** ({row['dept']}): {row['empty_reason']}")
                        
                        # כפתורי ביצוע החלפה (Swap Actions)
                        actions_found = False
                        if 'swap_suggestions' in st.session_state:
                            core_key = f"{row['date']}_{row['dept']}"
                            if core_key in st.session_state.swap_suggestions:
                                actions_found = True
                                for i, sugg in enumerate(st.session_state.swap_suggestions[core_key]):
                                    btn_label = f"✨ בצע: {sugg['desc']}"
                                    if st.button(btn_label, key=f"swap_btn_{core_key}_{i}"):
                                        # --- בדיקת תקינות נוספת לפני ביצוע (פותר באג של הצעות ישנות/לא תקינות) ---
                                        # וידוא שתורן חוץ לא נכנס לפנימית גריאטרית
                                        current_schedule = st.session_state.schedule
                                        current_staff = st.session_state.staff
                                        
                                        # פונקציה לבדיקת חוקיות עובד-מחלקה
                                        def validate_emp_dept(emp_name, dept_name):
                                            emp_row = current_staff[current_staff['name'] == emp_name]
                                            if not emp_row.empty:
                                                e_type = emp_row['type'].iloc[0]
                                                if e_type == 'תורן חוץ' and 'פנימית' in dept_name:
                                                    return False, f"שגיאה: לא ניתן לשבץ את {emp_name} (תורן חוץ) לפנימית גריאטרית!"
                                            return True, ""

                                        # בדיקה לפי סוג ההחלפה
                                        validation_passed = True
                                        fail_reason = ""
                                        
                                        if sugg['type'] == 'direct_swap':
                                            # נכנס ל-row['dept'] -> conflicted_emp (A)
                                            # נכנס ל-other_dept -> replacement_emp (B)
                                            ok1, msg1 = validate_emp_dept(sugg['conflicted_emp'], row['dept'])
                                            if not ok1: validation_passed, fail_reason = False, msg1
                                            ok2, msg2 = validate_emp_dept(sugg['replacement_emp'], sugg['source_dept'])
                                            if not ok2: validation_passed, fail_reason = False, msg2
                                            
                                        elif sugg['type'] == 'move_shift':
                                            # נכנס ל-row['dept'] -> conflicted_emp (A)
                                            # נכנס ל-conf_dept -> replacement_emp (B)
                                            ok1, msg1 = validate_emp_dept(sugg['conflicted_emp'], row['dept'])
                                            if not ok1: validation_passed, fail_reason = False, msg1
                                            ok2, msg2 = validate_emp_dept(sugg['replacement_emp'], sugg['conflict_dept'])
                                            if not ok2: validation_passed, fail_reason = False, msg2

                                        if not validation_passed:
                                            st.error(f"❌ לא ניתן לבצע את ההחלפה: {fail_reason}")
                                            st.info("ייתכן שהנתונים השתנו מאז ההרצה האחרונה. מומלץ להריץ שיבוץ אוטומטי מחדש.")
                                        else:
                                            # ביצוע ההחלפה בפועל!
                                            sched = st.session_state.schedule
                                            
                                            if sugg['type'] == 'direct_swap':
                                                mask_other = (sched['date'] == sugg['target_date']) & (sched['dept'] == sugg['source_dept'])
                                                st.session_state.schedule.loc[mask_other, 'employee'] = sugg['replacement_emp']
                                                
                                                mask_here = (sched['date'] == sugg['target_date']) & (sched['dept'] == row['dept'])
                                                st.session_state.schedule.loc[mask_here, 'employee'] = sugg['conflicted_emp']
                                                st.session_state.schedule.loc[mask_here, 'empty_reason'] = '' 
                                                
                                            elif sugg['type'] == 'move_shift':
                                                mask_conflict = (sched['date'] == sugg['conflict_date']) & (sched['dept'] == sugg['conflict_dept'])
                                                st.session_state.schedule.loc[mask_conflict, 'employee'] = sugg['replacement_emp']
                                                
                                                mask_here = (sched['date'] == row['date']) & (sched['dept'] == row['dept'])
                                                st.session_state.schedule.loc[mask_here, 'employee'] = sugg['conflicted_emp']
                                                st.session_state.schedule.loc[mask_here, 'empty_reason'] = ''
                                                
                                            elif sugg['type'] == 'triple_swap':
                                                mask_b_origin = (sched['date'] == sugg['target_date']) & (sched['dept'] == sugg['dept_b_origin'])
                                                st.session_state.schedule.loc[mask_b_origin, 'employee'] = sugg['emp_c']
                                                
                                                mask_a_origin = (sched['date'] == sugg['target_date']) & (sched['dept'] == sugg['dept_a_origin'])
                                                st.session_state.schedule.loc[mask_a_origin, 'employee'] = sugg['emp_b']
                                                
                                                mask_here = (sched['date'] == sugg['target_date']) & (sched['dept'] == row['dept'])
                                                st.session_state.schedule.loc[mask_here, 'employee'] = sugg['emp_a']
                                                st.session_state.schedule.loc[mask_here, 'empty_reason'] = ''
                                                
                                            save_to_db("schedule", st.session_state.schedule)
                                            st.success("ההחלפה בוצע בהצלחה! מרענן...")
                                            st.rerun()
                        
                        if not actions_found:
                             st.caption("כדי לראות כפתורי החלפה, יש להריץ 'שיבוץ אוטומטי' מחדש.")
        # ---------------------------------

        draw_calendar_view(2026, sel_month, "מנהל/ת")

        # --- Swap Requests Panel ---
        if st.session_state.pop('show_swap_approved', False):
            st.success("✅ ההחלפה אושרה ובוצעה בלוח השיבוץ!")
        if st.session_state.pop('show_swap_rejected', False):
            st.info("הבקשה נדחתה.")

        _swap_reqs = st.session_state.swap_requests
        if not _swap_reqs.empty and 'status' in _swap_reqs.columns:
            _pending = _swap_reqs[_swap_reqs['status'].astype(str).str.strip() == 'pending']
            if not _pending.empty:
                st.divider()
                st.markdown(f"### 🔄 בקשות החלפה ממתינות לאישור ({len(_pending)})")
                for _idx, _req in _pending.iterrows():
                    _requester  = str(_req.get('requester', ''))
                    _r_date     = str(_req.get('requester_date', ''))
                    _r_dept     = str(_req.get('requester_dept', ''))
                    _candidate  = str(_req.get('candidate', ''))
                    _c_date     = str(_req.get('candidate_date', ''))
                    _c_dept     = str(_req.get('candidate_dept', ''))
                    _stype      = str(_req.get('swap_type', 'partial'))
                    _created    = str(_req.get('created_at', ''))

                    _chain_ext      = str(_req.get('chain_ext', ''))
                    _chain_ext_dept = str(_req.get('chain_ext_dept', ''))

                    with st.container(border=True):
                        _ca, _cb = st.columns([4, 2])
                        with _ca:
                            if _stype == 'full':
                                st.markdown(f"**✅ החלפה מלאה** | נשלח: {_created}")
                                st.write(f"**{_requester}** מוותר/ת על: **{_r_date}** ({_r_dept})")
                                st.write(f"← **{_candidate}** ייקח/תיקח אותה ויעביר/ת: **{_c_date}** ({_c_dept})")
                            elif _stype == 'chain':
                                st.markdown(f"**🔗 החלפה משולשת** | נשלח: {_created}")
                                st.write(f"**{_requester}** מסיר/ה: **{_r_date}** ({_r_dept})")
                                st.write(f"← **{_candidate}** עובר/ת מ-שיקום ל-פנימית")
                                st.write(f"← **{_chain_ext}** מכסה שיקום")
                            else:
                                st.markdown(f"**⚠️ כיסוי חד-צדדי** | נשלח: {_created}")
                                st.write(f"**{_requester}** מבקש/ת כיסוי: **{_r_date}** ({_r_dept})")
                                st.write(f"← **{_candidate}** מוצע/ת כמחליף/ה (ללא משמרת הדדית)")

                        with _cb:
                            _col_a, _col_r = st.columns(2)
                            if _col_a.button("✅ אשר", key=f"approve_swap_{_idx}", use_container_width=True):
                                _sched = st.session_state.schedule
                                # Mutation 1: requester's shift → candidate (or remove requester from פנימית for chain)
                                _mask_req = (_sched['date'].astype(str) == _r_date) & (_sched['dept'] == _r_dept)
                                if _stype == 'chain':
                                    # Remove requester from פנימית
                                    st.session_state.schedule.loc[_mask_req, 'employee'] = '---'
                                    st.session_state.schedule.loc[_mask_req, 'is_manual'] = True
                                    # Mutation 2: move facilitator (מתמחה) from שיקום → פנימית
                                    _mask_fac = (_sched['date'].astype(str) == _r_date) & (_sched['dept'] == 'שיקום') & \
                                                (_sched['employee'].astype(str).str.strip() == _candidate)
                                    st.session_state.schedule.loc[_mask_fac, 'employee'] = '---'
                                    st.session_state.schedule.loc[_mask_fac, 'is_manual'] = True
                                    st.session_state.schedule.loc[_mask_req, 'employee'] = _candidate
                                    # Mutation 3: assign תורן חוץ to שיקום (add new row if needed)
                                    _mask_sha = (_sched['date'].astype(str) == _r_date) & (_sched['dept'] == 'שיקום') & \
                                                (_sched['employee'].astype(str).str.strip().isin(['---', '']))
                                    if _mask_sha.any():
                                        st.session_state.schedule.loc[_mask_sha, 'employee'] = _chain_ext
                                        st.session_state.schedule.loc[_mask_sha, 'is_manual'] = True
                                    else:
                                        _new_row = pd.DataFrame([{
                                            'date': _r_date, 'dept': 'שיקום',
                                            'employee': _chain_ext, 'is_manual': True, 'empty_reason': ''
                                        }])
                                        st.session_state.schedule = pd.concat(
                                            [st.session_state.schedule, _new_row], ignore_index=True
                                        )
                                else:
                                    st.session_state.schedule.loc[_mask_req, 'employee'] = _candidate
                                    st.session_state.schedule.loc[_mask_req, 'is_manual'] = True
                                    if _stype == 'full' and _c_date:
                                        _mask_cand = (_sched['date'].astype(str) == _c_date) & (_sched['dept'] == _c_dept)
                                        st.session_state.schedule.loc[_mask_cand, 'employee'] = _requester
                                        st.session_state.schedule.loc[_mask_cand, 'is_manual'] = True
                                save_to_db("schedule", st.session_state.schedule)
                                _swap_reqs.loc[_idx, 'status'] = 'approved'
                                save_to_db("swap_requests", _swap_reqs)
                                st.session_state.swap_requests = _swap_reqs.copy()
                                st.session_state['show_swap_approved'] = True
                                st.rerun()
                            if _col_r.button("❌ דחה", key=f"reject_swap_{_idx}", use_container_width=True):
                                _swap_reqs.loc[_idx, 'status'] = 'rejected'
                                save_to_db("swap_requests", _swap_reqs)
                                st.session_state.swap_requests = _swap_reqs.copy()
                                st.session_state['show_swap_rejected'] = True
                                st.rerun()

    elif selected_nav == 'צוות':
        st.subheader("ניהול צוות עובדים")
        
        # --- טופס הוספת עובד ---
        with st.expander("➕ הוספת עובד חדש", expanded=False):
            with st.form("add_emp_form"):
                col_new1, col_new2 = st.columns(2)
                with col_new1:
                    new_name = st.text_input("שם מלא:")
                    new_type = st.selectbox(
                        "תפקיד:",
                        ["מתמחה", "תורן חוץ", "מנהל/ת", "רופא בכיר", "מנהל מחלקה"],
                        format_func=lambda x: _TYPE_DISPLAY.get(x, x),
                    )
                    new_email = st.text_input("אימייל (להתראות):", placeholder="optional@example.com")
                with col_new2:
                    new_dept = st.selectbox("מחלקה:", ["שיקום", "פנימית גריאטרית", "כללי", "הנהלה", "זה״ב"])
                    new_quota = st.number_input("מכסה חודשית:", min_value=0, value=6)
                    new_weekend_quota = st.number_input("מכסת סופ\"ש:", min_value=0, value=1)
                    new_only_home = st.checkbox("מוגבל למחלקה זו בלבד?", value=False)

                # שדה manage_depts — רק למנהל מחלקה
                new_manage_depts = ""
                if new_type == "מנהל מחלקה":
                    selected_depts = st.multiselect(
                        "מחלקות בניהולו:",
                        options=DAILY_DEPTS_ALL,
                        default=[],
                        key="new_emp_manage_depts"
                    )
                    new_manage_depts = ",".join(selected_depts)

                if st.form_submit_button("הוסף עובד"):
                    if new_name.strip():
                        if new_name not in st.session_state.staff['name'].values:
                            # סיסמת ברירת מחדל מוצפנת
                            def_pass_hash = hashlib.sha256("1234".encode()).hexdigest()

                            new_emp_row = pd.DataFrame([{
                                'name': new_name,
                                'type': new_type,
                                'dept': new_dept,
                                'monthly_quota': new_quota,
                                'weekend_quota': new_weekend_quota,
                                'only_home_dept': new_only_home,
                                'password': def_pass_hash,
                                'email': new_email.strip(),
                                'manage_depts': new_manage_depts.strip(),
                                'recurring_absent_days': '',
                            }])

                            st.session_state.staff = pd.concat([st.session_state.staff, new_emp_row], ignore_index=True)
                            save_to_db("staff", st.session_state.staff)
                            st.session_state['_staff_editor_ver'] = st.session_state.get('_staff_editor_ver', 0) + 1
                            st.success(f"העובד/ת {new_name} נוספ/ה בהצלחה! (סיסמה: 1234)")
                            st.rerun()
        
        st.divider()
        st.caption("שינויים בטבלה נשמרים רק בלחיצה על כפתור השמירה")
        
        # עטיפה בטופס (Form) כדי למנוע טעינה מחדש בכל שינוי תא
        with st.form(key=f"staff_batch_edit_form_{st.session_state.get('_staff_editor_ver', 0)}"):
            # Ensure only_home_dept exists
            if 'only_home_dept' not in st.session_state.staff.columns:
                st.session_state.staff['only_home_dept'] = False

            # Fix True/False serialization from Google Sheets
            def _is_true(v):
                if pd.isna(v): return False
                if isinstance(v, bool): return v
                return str(v).strip().lower() == 'true'

            st.session_state.staff['only_home_dept'] = st.session_state.staff['only_home_dept'].apply(_is_true)

            # Multi-select for Home Dept only
            valid_names = [n for n in st.session_state.staff['name'].tolist() if str(n).strip() and n != '---']
            default_homes = st.session_state.staff[st.session_state.staff['only_home_dept']]['name'].dropna().tolist()
            default_homes = [n for n in default_homes if n in valid_names]

            selected_home_depts = st.multiselect(
                "עובדים המוגבלים למחלקת האם בלבד (ללא חציות למחלקות אחרות):",
                options=valid_names,
                default=default_homes
            )
            st.divider()

            # Build editor view: exclude password & only_home_dept from the editable table
            # (only_home_dept handled via the multiselect above)
            preferred_order = ['name', 'type', 'dept', 'monthly_quota', 'weekend_quota',
                               'email', 'manage_depts', 'recurring_absent_days',
                               'manual_schedule_only']
            available = [c for c in preferred_order if c in st.session_state.staff.columns]
            # tack on any unexpected extra cols (except hidden)
            extras = [c for c in st.session_state.staff.columns
                      if c not in available and c not in ('password', 'only_home_dept', 'whatsapp_link')]
            cols_to_show = available + extras
            staff_view = st.session_state.staff[cols_to_show].copy()

            # Reverse for RTL display
            reversed_cols = cols_to_show[::-1]
            reversed_staff_view = staff_view[reversed_cols]

            staff_editor = st.data_editor(
                reversed_staff_view,
                use_container_width=True,
                num_rows="dynamic",
                hide_index=True,
                column_config={
                    "name": st.column_config.TextColumn(
                        "👤 שם", help="שם מלא של העובד", width="medium", required=True),
                    "type": st.column_config.SelectboxColumn(
                        "תפקיד",
                        options=["מתמחה", "תורן חוץ", "מנהל/ת", "רופא בכיר", "מנהל מחלקה"],
                        width="small", required=True),
                    "dept": st.column_config.SelectboxColumn(
                        "מחלקה",
                        options=["שיקום", "פנימית גריאטרית", "כללי", "הנהלה"],
                        width="small"),
                    "monthly_quota": st.column_config.NumberColumn(
                        "מכסה חודשית", help="מספר תורנויות חודשיות נדרשות",
                        min_value=0, max_value=20, step=1, width="small"),
                    "weekend_quota": st.column_config.NumberColumn(
                        "מכסת סופ\"ש", help="תורנויות סוף שבוע נדרשות",
                        min_value=0, max_value=10, step=1, width="small"),
                    "email": st.column_config.TextColumn(
                        "📧 אימייל", help="לקבלת התראות (אופציונלי)", width="medium"),
                    "manage_depts": st.column_config.TextColumn(
                        "מחלקות בניהול",
                        help="רק למנהל מחלקה — שמות מופרדים בפסיקים, לדוגמה: שיקום גריאטרי א',שיקום גריאטרי ב'",
                        width="medium"),
                    "recurring_absent_days": st.column_config.TextColumn(
                        "🔁 ימי היעדרות קבועים",
                        help="ימי שבוע שבהם העובד נעדר באופן קבוע (יומי). אותיות עברית מופרדות בפסיקים: א,ב,ג,ד,ה,ו,ש. דוגמה: 'ד,ה' = רביעי וחמישי. ייושם אוטומטית בכל סידור.",
                        width="medium"),
                    "manual_schedule_only": st.column_config.CheckboxColumn(
                        "✋ שיבוץ ידני בלבד",
                        help="אם מסומן — העובד לא ישובץ אוטומטית לסידור העבודה. יופיע רק אם הוספת ידנית.",
                        width="medium"),
                }
            )
            submit_changes = st.form_submit_button("💾 שמור שינויים בצוות", use_container_width=False)
        
        if submit_changes:
            # Merge logic:
            # 1. Take the editor result (staff_editor)
            # 2. Get the original passwords from st.session_state.staff
            # WE MUST BE CAREFUL: "dynamic" rows means users can add/delete rows.
            # If a row is added here, it won't have a password. We should assign default.
            
            # Because we reversed columns for display, we reverse back to normal for processing
            edited_df = staff_editor[staff_editor.columns[::-1]]

            # Reattach only_home_dept dynamically from the multiselect
            edited_df['only_home_dept'] = edited_df['name'].isin(selected_home_depts)

            # We need to preserve passwords for existing users.
            # Simple approach: Join with original on 'name' IF name is unique and didn't change.
            # But users might change names. 
            # Best effort: 
            # If the index matches, keep the password.
            # If new row (index not in old), set default password.
            
            final_df_list = []
            # משיכת נתונים עדכניים כדי למנוע דריסת נתונים
            latest_staff = get_db_data("staff")
            if not latest_staff.empty:
                old_df = latest_staff
            else:
                old_df = st.session_state.staff
            
            for index, row in edited_df.iterrows():
                # Check if this index existed in old_df
                pass_val = hashlib.sha256("1234".encode()).hexdigest() # Default
                
                if index in old_df.index:
                    # Check if name matched (sanity check, though index usually persists in editor unless sorted)
                     pass_val = old_df.loc[index, 'password']
                
                row['password'] = pass_val
                final_df_list.append(row)
            
            # Reconstruct DataFrame including password
            final_new_staff = pd.DataFrame(final_df_list)
            
            st.session_state.staff = final_new_staff
            save_to_db("staff", st.session_state.staff)
            st.success("הנתונים נשמרו בהצלחה!")
            st.rerun()
        
        st.markdown("---")
        col_sync, col_warn = st.columns([1, 3])
        with col_sync:
            if st.button("🔄 סנכרן נתונים מהענן"):
                 # Force reload of all data
                 st.cache_data.clear()
                 for key in ['staff', 'schedule', 'requests']:
                     if key in st.session_state:
                         del st.session_state[key]
                 st.rerun()
        with col_warn:
             st.warning("⚠️ שים לב: עריכה בטבלה זו דורסת שינויים שנעשו ישירות ב-Google Sheets. אם ערכת שם, לחץ קודם על 'סנכרן נתונים'.")

        st.divider()
        st.subheader("ניהול אילוצים ומשמרות")
        
        # בחירת עובד לניהול אילוצים
        selected_emp_mgr = st.selectbox("בחר עובד לניהול אילוצים:", st.session_state.staff['name'].tolist())
        
        if selected_emp_mgr:
            st.write(f"עריכת אילוצים עבור: **{selected_emp_mgr}**")
            
            # --- טעינת נתונים קיימים ---
            current_month_prefix = f"2026-{sel_month:02d}"
            
            # נרמול עמודת התאריך למחרוזת
            requests_df = st.session_state.requests.copy()
            requests_df['date'] = requests_df['date'].astype(str)
            
            # סינון רשומות רלוונטיות לעובד ולחודש
            emp_reqs = requests_df[
                (requests_df['employee'] == selected_emp_mgr) & 
                (requests_df['date'].str.startswith(current_month_prefix))
            ]
            
            existing_constraints = emp_reqs[emp_reqs['status'] == 'אילוץ']['date'].tolist()
            existing_wishes = emp_reqs[emp_reqs['status'] == 'בקשה']['date'].tolist()
            
            # --- ממשק ויזואלי (מודרני) ---
            # Use employee index for key_prefix — Hebrew names break CSS class selectors
            emp_names_list = st.session_state.staff['name'].tolist()
            emp_idx        = emp_names_list.index(selected_emp_mgr) if selected_emp_mgr in emp_names_list else 0
            mgr_key_prefix = f"mgr_e{emp_idx}"
            mgr_const_days = [int(d.split('-')[2]) for d in existing_constraints]
            mgr_wish_days  = [int(d.split('-')[2]) for d in existing_wishes]
            mgr_cal_sd     = st.session_state.get('special_days')
            if isinstance(mgr_cal_sd, pd.DataFrame) and mgr_cal_sd.empty:
                mgr_cal_sd = None

            st.divider()
            new_const_day_nums, new_wish_day_nums = render_modern_calendar(
                2026, sel_month,
                mgr_const_days, mgr_wish_days,
                special_days_df=mgr_cal_sd,
                key_prefix=mgr_key_prefix,
                show_validation=False
            )
            new_constraints = [f"2026-{sel_month:02d}-{d:02d}" for d in new_const_day_nums]
            new_wishes       = [f"2026-{sel_month:02d}-{d:02d}" for d in new_wish_day_nums]
            
            st.divider()
            
            if st.button("💾 שמור שינויים לעובד זה"):
                # בדיקת חפיפה (סתירה)
                overlap = set(new_constraints).intersection(set(new_wishes))
                if overlap:
                    formatted_overlap = [datetime.strptime(d, '%Y-%m-%d').strftime('%d/%m') for d in overlap]
                    st.error(f"שגיאה: התאריכים הבאים מסומנים גם כחסימה וגם כבקשה: {', '.join(formatted_overlap)}")
                else:
                    # הכל תקין - שמירה
                    # משיכת נתונים חיים מהגיליון (עוקף cache למניעת דריסה הדדית)
                    live = _fetch_live("requests")
                    base = live if not live.empty else st.session_state.requests

                    # 1. מחיקת הישן לחודש זה
                    mask_keep = ~((base['employee'].astype(str).str.strip() == str(selected_emp_mgr).strip()) &
                                  (base['date'].astype(str).str.startswith(current_month_prefix)))
                    base = base[mask_keep]

                    # 2. הוספת החדש
                    new_records = []
                    for d in new_constraints:
                        new_records.append({'employee': selected_emp_mgr, 'date': d, 'status': 'אילוץ'})
                    for d in new_wishes:
                        new_records.append({'employee': selected_emp_mgr, 'date': d, 'status': 'בקשה'})

                    merged = pd.concat([base, pd.DataFrame(new_records)], ignore_index=True) if new_records else base
                    save_to_db("requests", merged)
                    st.session_state.requests = merged
                    _fetch_sheet_data_silently.clear()
                    st.success(f"האילוצים של {selected_emp_mgr} עודכנו בהצלחה!")
                    st.session_state.pop(f"{mgr_key_prefix}_init_{sel_month}", None)
                    st.rerun()

            # ── ☀️ היעדרויות יומיות לעובד הנבחר ──────────────────────
            st.divider()
            st.markdown("### ☀️ היעדרויות יומיות (עבודת בוקר)")
            st.caption("בקשות ההיעדרות של העובד — ניתן לאשר/לדחות בקשות ממתינות.")

            ar_tzevet = st.session_state.absence_requests.copy()
            if ar_tzevet.empty or 'employee' not in ar_tzevet.columns:
                st.info("אין בקשות במערכת.")
            else:
                ar_tzevet['employee'] = ar_tzevet['employee'].astype(str).str.strip()
                ar_tzevet['status']   = ar_tzevet['status'].astype(str).str.lower()
                emp_n = str(selected_emp_mgr).strip()
                emp_reqs_da = ar_tzevet[ar_tzevet['employee'] == emp_n]

                if emp_reqs_da.empty:
                    st.info(f"אין בקשות היעדרות לעובד {selected_emp_mgr}.")
                else:
                    pending_da = emp_reqs_da[emp_reqs_da['status'] == 'pending']
                    if not pending_da.empty:
                        st.markdown(f"**🟡 ממתינות לאישור ({len(pending_da)})**")
                        for idx_da, row_da in pending_da.iterrows():
                            with st.container(border=True):
                                ca, cb, cc, cd, ce = st.columns([2, 2, 2, 1, 1])
                                ca.write(f"{row_da['start_date']} – {row_da['end_date']}")
                                cb.write(row_da.get('type', '—'))
                                cc.write(row_da.get('dept_at_request', '—') or '—')
                                req_id_da = str(row_da.get('id', idx_da))
                                if cd.button("✅ אשר", key=f"tzv_ap_{req_id_da}",
                                             use_container_width=True):
                                    _approve_request(req_id_da, str(user_name).strip())
                                    st.rerun()
                                if ce.button("❌ דחה", key=f"tzv_rj_{req_id_da}",
                                             use_container_width=True):
                                    _reject_request(req_id_da, str(user_name).strip())
                                    st.rerun()
                                if row_da.get('notes'):
                                    st.caption(f"💬 {row_da['notes']}")
                    # All requests (including approved/rejected) as a compact table
                    st.markdown("**📋 היסטוריה מלאה**")
                    show_da = emp_reqs_da[['start_date','end_date','type','status','approved_by','notes']].copy()
                    show_da['status'] = show_da['status'].map({
                        'pending':'⏳ ממתין', 'approved':'✅ אושר', 'rejected':'❌ נדחה'
                    }).fillna(show_da['status'])
                    show_da = show_da.rename(columns={
                        'start_date':'תאריך התחלה', 'end_date':'תאריך סיום',
                        'type':'סוג', 'status':'סטטוס', 'approved_by':'מאשר', 'notes':'הערה',
                    })
                    st.dataframe(show_da, use_container_width=True, hide_index=True)

    elif selected_nav == 'דוחות וניהול':
        # st.header("דוח סטטוס ומסכמים") - Removed by user request
        
        # --- חלק חדש: לוח בקרה מודרני ---
        st.markdown("### 📊 לוח בקרה - סיכום חודשי")
        
        # --- 1. חישוב נתונים למדדים ---
        current_month_prefix = f"2026-{sel_month:02d}"
        relevant_staff = st.session_state.staff[st.session_state.staff['type'].isin(['מתמחה', 'תורן חוץ'])]
        n_total = len(relevant_staff)
        submitted_count = 0
        status_list = []
        
        reqs_df = st.session_state.requests.copy()
        if not reqs_df.empty:
            reqs_df['date'] = reqs_df['date'].astype(str)
            reqs_df['employee'] = reqs_df['employee'].astype(str).str.strip()

        for _, emp in relevant_staff.iterrows():
            name = str(emp['name']).strip()

            n_c, n_w = 0, 0
            if not reqs_df.empty:
                user_reqs = reqs_df[(reqs_df['employee'] == name) & (reqs_df['date'].str.startswith(current_month_prefix))]
                n_c = len(user_reqs[user_reqs['status'] == 'אילוץ'])
                n_w = len(user_reqs[user_reqs['status'] == 'בקשה'])
            
            has_submitted = (n_c + n_w) > 0
            if has_submitted: submitted_count += 1
            
            status_list.append({
                "שם העובד": name.strip(),
                "תפקיד": emp['type'],
                "סטטוס": "✅ הוגש" if has_submitted else "❌ טרם הוגש",
                "חסימות": n_c,
                "בקשות": n_w
            })

        # --- 2. הצגת כרטיסי מדדים מותאמים אישית (Custom Metric Cards) ---
        m_cols = st.columns(3)
        
        # Helper function for generating HTML cards
        def metric_html(title, value, subtitle, icon, color, gradient="from-[#ffffff] to-[#f9fafb]"):
            return f"""
            <div style="
                background: linear-gradient(135deg, {gradient.split('to-')[0].replace('from-[', '').replace(']', '').strip()} 0%, {gradient.split('to-')[1].replace('[', '').replace(']', '').strip()} 100%);
                border-radius: 12px;
                padding: 20px;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                text-align: right;
                direction: rtl;
                border-top: 4px solid {color};
                transition: transform 0.2s ease-in-out;
                margin-bottom: 20px;
            " onmouseover="this.style.transform='translateY(-5px)'" onmouseout="this.style.transform='translateY(0)'">
                <h4 style="color: #6b7280; font-size: 14px; margin: 0; padding-bottom: 8px;">{title}</h4>
                <div style="display: flex; justify-content: space-between; align-items: baseline;">
                    <h2 style="color: #111827; font-size: 32px; font-weight: 700; margin: 0;">{value}</h2>
                    <span style="font-size: 24px;">{icon}</span>
                </div>
                <p style="color: #10b981; font-size: 12px; margin-top: 8px; margin-bottom:0;">{subtitle}</p>
            </div>
            """

        with m_cols[0]:
            st.markdown(metric_html("סה״כ צוות", f"{n_total}", "מתמחים ותורני חוץ", "👥", "#4f46e5"), unsafe_allow_html=True)
        with m_cols[1]:
            st.markdown(metric_html("הגישו אילוצים", f"{submitted_count}", f"עבור חודש {sel_month}", "✅", "#10b981"), unsafe_allow_html=True)
        with m_cols[2]:
            pending = n_total - submitted_count
            p_color = "#ef4444" if pending > 0 else "#10b981"
            st.markdown(metric_html("ממתינים להגשה", f"{pending}", "נדרש תזכורת", "⚠️", p_color), unsafe_allow_html=True)

        st.divider()

        # --- 3. גרף משמרות וסטטוס הגשה (שני טורים) ---
        col_chart, col_table = st.columns([1.2, 1])

        with col_chart:
            st.markdown("##### 📈 ספירת תורנויות")
            sched = st.session_state.schedule
            
            # סינון ספירת המשמרות לחודש הנוכחי בלבד
            if not sched.empty:
                sched_month = sched[sched['date'].astype(str).str.startswith(current_month_prefix)]
            else:
                sched_month = sched

            if not sched_month.empty:
                # ספירת משמרות לפי עובד וסוג
                reg_counts = sched_month[~sched_month['dept'].astype(str).str.contains("שישי בוקר", na=False)]['employee'].value_counts()
                morn_counts = sched_month[sched_month['dept'].astype(str).str.contains("שישי בוקר", na=False)]['employee'].value_counts()
                combined_df = pd.DataFrame({'תורנויות': reg_counts, 'שישי בוקר': morn_counts}).fillna(0)
                st.bar_chart(combined_df, color=["#4f46e5", "#fb923c"]) # Indigo and Orange
            else:
                st.info("אין נתוני שיבוץ להצגה בגרף")

        with col_table:
            st.markdown("##### 📋 סטטוס הגשות")
            df_status = pd.DataFrame(status_list)
            
            if not df_status.empty:
                df_status = df_status[["שם העובד", "תפקיד", "סטטוס", "חסימות", "בקשות"]] # Ensure order
                # עיצוב טבלה מותאם
                def style_status(val):
                    try:
                        color = '#dcfce7' if '✅' in str(val) else '#fee2e2'
                        return f'background-color: {color}; color: #1e293b; font-weight: 500; border-radius: 4px;'
                    except:
                        return ''

                # rtl support
                reversed_df_status = df_status[df_status.columns[::-1]]
                
                # Make mobile text unbreak using a dedicated mobile-responsive CSS injection inside ui_components.py
                try:
                     styled_df = reversed_df_status.style.applymap(style_status, subset=['סטטוס']).set_properties(**{'text-align': 'right', 'direction': 'rtl'})
                     st.dataframe(styled_df, use_container_width=True, hide_index=True)
                except AttributeError: # For newer pandas without applymap handling
                     styled_df = reversed_df_status.style.map(style_status, subset=['סטטוס']).set_properties(**{'text-align': 'right', 'direction': 'rtl'})
                     st.dataframe(styled_df, use_container_width=True, hide_index=True)
                except Exception as e:
                     st.dataframe(reversed_df_status, use_container_width=True, hide_index=True)
            else:
                 st.info("אין נתונים להצגה.")

        # ── "ייצוא נתונים" הוסר מכאן והועבר לטאב "סידור תורנויות" (ראה כפתור 📤 שם) ──


    if selected_nav == 'דוחות וניהול': # Fairness merged into Reports
        st.subheader("מעקב הוגנות - ימי רביעי, חמישי ושישי (מתמחים בלבד)")
        
        # טעינת כל הנתונים
        full_schedule = st.session_state.schedule.copy()
        if not full_schedule.empty:
            # המרת תאריך ל-datetime
            full_schedule['date_dt'] = pd.to_datetime(full_schedule['date'])
            full_schedule['weekday'] = full_schedule['date_dt'].dt.weekday
            full_schedule['month_year'] = full_schedule['date_dt'].dt.strftime('%m/%Y')
            
            # סינון רק למתמחים
            staff_types = st.session_state.staff.set_index('name')['type'].to_dict()
            full_schedule['staff_type'] = full_schedule['employee'].map(staff_types)
            intern_schedule = full_schedule[full_schedule['staff_type'] == 'מתמחה']
            
            # יצירת טבלת סיכום
            # ימי רביעי = 2, ימי חמישי = 3
            tracker = []
            
            # קבלת רשימת כל המתמחים (כולל אלו שלא שובצו כלל)
            all_interns = st.session_state.staff[st.session_state.staff['type'] == 'מתמחה']['name'].tolist()
            
            for name in all_interns:
                emp_sched = intern_schedule[intern_schedule['employee'] == name]
                wed_count = len(emp_sched[emp_sched['weekday'] == 2])
                thu_count = len(emp_sched[emp_sched['weekday'] == 3])
                fri_morn_count = len(emp_sched[emp_sched['dept'].str.contains('שישי בוקר')])
                
                tracker.append({
                    'שם המתמחה': name,
                    "סה\"כ ימי ד' (מבוקש)": wed_count,
                    "סה\"כ ימי ה' (קשה)": thu_count,
                    "שישי בוקר": fri_morn_count,
                    'ציון הוגנות (נטו)': wed_count - (thu_count + fri_morn_count) # חיובי = קיבל יותר טובים, שלילי = קיבל יותר קשים
                })
            
            df_fairness = pd.DataFrame(tracker).sort_values('ציון הוגנות (נטו)', ascending=False)
            
            # עיצוב הטבלה
            st.dataframe(
                df_fairness[df_fairness.columns[::-1]].style.background_gradient(subset=['ציון הוגנות (נטו)'], cmap="RdYlGn"),
                use_container_width=True
            )
            
            # פירוט לפי חודשים הוסר לבקשת המשתמש.
        else:
            st.info("הלוח עדיין ריק.")

    # --- דוח יומי אוטומטי ---
    if selected_nav == 'דוחות וניהול':
        st.divider()
        st.subheader("🤖 דוח בעיות צפויות")

        # Headline: heavily blocked days from last report
        _report_preview = get_db_data("daily_report")
        if not _report_preview.empty and 'problem_type' in _report_preview.columns:
            _peak_days = _report_preview[_report_preview['problem_type'] == 'ימי שיא חסימה']
            if not _peak_days.empty:
                _critical_peaks = _peak_days[_peak_days['severity'] == 'קריטי']
                _warn_peaks = _peak_days[_peak_days['severity'] == 'אזהרה']
                _headline_parts = []
                if not _critical_peaks.empty:
                    _headline_parts.append(f"🔴 {len(_critical_peaks)} ימים קריטיים (50%+ חסמו)")
                if not _warn_peaks.empty:
                    _headline_parts.append(f"🟡 {len(_warn_peaks)} ימים בסיכון (30%+ חסמו)")
                if _headline_parts:
                    st.info("**ימים קשים לאיוש החודש:** " + " · ".join(_headline_parts) +
                            " — ראה פירוט בדוח למטה")

        col_rep1, col_rep2 = st.columns([3, 1])
        with col_rep2:
            if st.button("🔄 הרץ סריקה עכשיו", use_container_width=True):
                with st.spinner("סורק..."):
                    try:
                        # Import the analysis functions directly from daily_report.py
                        import importlib.util, sys as _sys
                        _spec = importlib.util.spec_from_file_location(
                            "daily_report",
                            os.path.join(os.path.dirname(__file__), "daily_report.py")
                            if hasattr(__file__, '__file__') else "daily_report.py"
                        )
                        _dr = importlib.util.module_from_spec(_spec)
                        _spec.loader.exec_module(_dr)

                        _staff    = st.session_state.staff.copy()
                        _requests = st.session_state.requests.copy()
                        _month    = sel_month

                        _problems = _dr.analyze_month(2026, _month, _staff, _requests)
                        if not _problems:
                            _problems = [{'severity': 'תקין', 'problem_type': 'ללא בעיות',
                                          'description': f'לא נמצאו בעיות צפויות לחודש {_month}/2026', 'day': 0}]

                        # Save to Google Sheets via gspread
                        _now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
                        _month_str = f"2026-{_month:02d}"
                        _gc = get_gspread_client()
                        _sh = _gc.open_by_url(st.secrets["connections"]["gsheets"]["spreadsheet"])
                        _header = [['generated_at', 'month', 'severity', 'problem_type', 'description']]
                        _rows = [[_now_str, _month_str, p['severity'], p['problem_type'], p['description']]
                                 for p in _problems]
                        try:
                            _ws = _sh.worksheet('daily_report')
                            _ws.clear()
                        except Exception:
                            _ws = _sh.add_worksheet('daily_report', rows=500, cols=6)
                        _ws.update('A1', _header + _rows)

                        _fetch_sheet_data_silently.clear()
                        st.toast(f"סריקה הושלמה — {len(_problems)} ממצאים")
                    except Exception as e:
                        st.error(f"שגיאה בהרצת הסריקה: {e}")
                st.rerun()

        report_df = get_db_data("daily_report")

        if report_df.empty:
            st.info("טרם הופק דוח. לחץ 'הרץ סריקה עכשיו' או המתן להרצה האוטומטית היומית.")
        else:
            generated_at = report_df['generated_at'].iloc[0] if 'generated_at' in report_df.columns else "לא ידוע"
            report_month = report_df['month'].iloc[0] if 'month' in report_df.columns else ""
            with col_rep1:
                st.caption(f"עודכן לאחרונה: {generated_at} | חודש: {report_month}")

            severity_icon = {'קריטי': '🔴', 'אזהרה': '🟡', 'מידע': 'ℹ️', 'תקין': '✅'}
            severity_order = {'קריטי': 0, 'אזהרה': 1, 'מידע': 2, 'תקין': 3}

            report_df['_order'] = report_df['severity'].map(severity_order).fillna(9)
            report_df = report_df.sort_values('_order')

            current_type = None
            for _, row in report_df.iterrows():
                ptype = row.get('problem_type', '')
                sev   = row.get('severity', '')
                desc  = row.get('description', '')
                icon  = severity_icon.get(sev, '•')

                if ptype != current_type:
                    current_type = ptype
                    st.markdown(f"**{ptype}**")

                if sev == 'קריטי':
                    st.error(f"{icon} {desc}")
                elif sev == 'אזהרה':
                    st.warning(f"{icon} {desc}")
                elif sev == 'תקין':
                    st.success(f"{icon} {desc}")
                else:
                    st.info(f"{icon} {desc}")

    # --- ניתוח שימוש במערכת ---
    if selected_nav == 'דוחות וניהול':
        st.divider()
        with st.expander("📈 ניתוח שימוש במערכת", expanded=False):
            _alog = get_db_data("analytics_log")
            if _alog.empty or 'event_type' not in _alog.columns:
                st.info("אין נתוני שימוש עדיין. הנתונים יצטברו אוטומטית ככל שמשתמשים נכנסים.")
            else:
                # ─── A. סיכום כללי ───────────────────────────────────────────────
                st.markdown("#### A. סיכום כללי")
                _logins = _alog[_alog['event_type'] == 'login_success']
                _submits = _alog[_alog['event_type'] == 'constraint_submit']
                _unique_users = _alog['user_name'].nunique()
                _mobile_pct = (
                    int((_logins['device_type'] == 'mobile').sum() / len(_logins) * 100)
                    if not _logins.empty else 0
                )
                _avg_submits = (
                    _submits.groupby('active_month').size().mean()
                    if not _submits.empty else 0
                )
                _mc1, _mc2, _mc3, _mc4 = st.columns(4)
                _mc1.metric("סה\"כ התחברויות", len(_logins))
                _mc2.metric("משתמשים ייחודיים", _unique_users)
                _mc3.metric("שיעור מובייל", f"{_mobile_pct}%")
                _mc4.metric("ממוצע הגשות/חודש", f"{_avg_submits:.1f}")

                st.divider()

                # ─── B. פעילות לפי משתמש ─────────────────────────────────────────
                st.markdown("#### B. פעילות לפי משתמש")
                _by_user = pd.DataFrame()
                if not _alog.empty:
                    _grp = _alog.groupby('user_name')
                    _b_logins  = _grp.apply(lambda g: (g['event_type'] == 'login_success').sum())
                    _b_submits = _grp.apply(lambda g: (g['event_type'] == 'constraint_submit').sum())
                    _b_searches= _grp.apply(lambda g: (g['event_type'] == 'swap_search').sum())
                    _b_swaps   = _grp.apply(lambda g: (g['event_type'] == 'swap_request_sent').sum())
                    _b_device  = _alog[_alog['event_type'] == 'login_success'].groupby('user_name')['device_type'].agg(
                        lambda x: x.value_counts().index[0] if len(x) > 0 else 'unknown'
                    )
                    _by_user = pd.DataFrame({
                        'שם': _b_logins.index,
                        'התחברויות': _b_logins.values,
                        'הגשות': _b_submits.reindex(_b_logins.index, fill_value=0).values,
                        'חיפושי החלפה': _b_searches.reindex(_b_logins.index, fill_value=0).values,
                        'בקשות החלפה': _b_swaps.reindex(_b_logins.index, fill_value=0).values,
                        'מכשיר נפוץ': _b_device.reindex(_b_logins.index, fill_value='unknown').values,
                    })
                    st.dataframe(_by_user[_by_user.columns[::-1]], use_container_width=True, hide_index=True)

                st.divider()

                # ─── C. שיעור חסימה לעובד ────────────────────────────────────────
                st.markdown("#### C. שיעור חסימה לעובד (חודש פעיל)")
                _reqs_df = get_db_data("requests")
                if not _reqs_df.empty:
                    _month_prefix_a = f"2026-{sel_month:02d}"
                    _days_in_month = calendar.monthrange(2026, sel_month)[1]
                    _blocks = _reqs_df[
                        (_reqs_df['status'] == 'אילוץ') &
                        (_reqs_df['date'].astype(str).str.startswith(_month_prefix_a))
                    ].groupby('employee').size().reset_index(name='חסימות')
                    _blocks['חסימות'] = _blocks['חסימות'].astype(int)
                    _blocks['% חסימה'] = (_blocks['חסימות'] / _days_in_month * 100).round(1)
                    _blocks = _blocks.sort_values('% חסימה', ascending=False)

                    def _block_color(pct):
                        if pct > 60: return '🔴'
                        if pct > 40: return '🟡'
                        return '🟢'

                    _blocks['סטטוס'] = _blocks['% חסימה'].apply(_block_color)
                    _risk = _blocks[_blocks['% חסימה'] > 60]['employee'].tolist()
                    if _risk:
                        st.warning(f"⚠️ סיכון שיבוץ — חסימה מעל 60%: {', '.join(_risk)}")
                    st.dataframe(_blocks[['employee', 'חסימות', '% חסימה', 'סטטוס']].rename(
                        columns={'employee': 'עובד/ת'}
                    )[['סטטוס', '% חסימה', 'חסימות', 'עובד/ת']], use_container_width=True, hide_index=True)
                else:
                    st.info("אין נתוני בקשות לחודש זה.")

                st.divider()

                # ─── D. שימוש בטאבים ─────────────────────────────────────────────
                st.markdown("#### D. שימוש בטאבים")
                _tab_entries = _alog[(_alog['event_type'] == 'tab_view') & (_alog['detail_2'] == 'enter')]
                if not _tab_entries.empty:
                    _tab_counts = _tab_entries.groupby('detail_1').size().reset_index(name='כניסות')
                    _tab_counts = _tab_counts.rename(columns={'detail_1': 'טאב'})
                    st.bar_chart(_tab_counts.set_index('טאב')['כניסות'])
                else:
                    st.info("אין נתוני טאב עדיין.")

                st.divider()

                # ─── E. זמן ממוצע בטאב ────────────────────────────────────────────
                st.markdown("#### E. זמן ממוצע בטאב (שניות)")
                _tab_exit = _alog[(_alog['event_type'] == 'tab_view') & (_alog['detail_2'] != 'enter')].copy()
                if not _tab_exit.empty:
                    _tab_exit['שניות'] = pd.to_numeric(_tab_exit['detail_2'], errors='coerce')
                    _tab_exit = _tab_exit.dropna(subset=['שניות'])
                    if not _tab_exit.empty:
                        _avg_time = _tab_exit.groupby('detail_1')['שניות'].mean().round(1).reset_index()
                        _avg_time = _avg_time.rename(columns={'detail_1': 'טאב', 'שניות': 'ממוצע שניות'})
                        st.bar_chart(_avg_time.set_index('טאב')['ממוצע שניות'])
                    else:
                        st.info("אין נתוני זמן בטאב עדיין.")
                else:
                    st.info("אין נתוני זמן בטאב עדיין.")

                st.divider()

                # ─── F. התחברויות לפי שעה ─────────────────────────────────────────
                st.markdown("#### F. התחברויות לפי שעה ביום")
                if not _logins.empty:
                    _logins_ts = _logins.copy()
                    _logins_ts['שעה'] = pd.to_datetime(_logins_ts['timestamp'], errors='coerce').dt.hour
                    _hour_counts = _logins_ts.groupby('שעה').size().reindex(range(24), fill_value=0).reset_index(name='כניסות')
                    st.bar_chart(_hour_counts.set_index('שעה')['כניסות'])
                else:
                    st.info("אין נתוני כניסה עדיין.")

                st.divider()

                # ─── G. סוג מכשיר ─────────────────────────────────────────────────
                st.markdown("#### G. סוג מכשיר")
                if not _logins.empty:
                    _dev_counts = _logins['device_type'].value_counts().reset_index()
                    _dev_counts.columns = ['מכשיר', 'כניסות']
                    _gc1, _gc2 = st.columns(2)
                    with _gc1:
                        st.dataframe(_dev_counts[['כניסות', 'מכשיר']], use_container_width=True, hide_index=True)
                    with _gc2:
                        _user_dev = _logins.groupby('user_name')['device_type'].agg(
                            lambda x: x.value_counts().index[0] if len(x) > 0 else 'unknown'
                        ).reset_index().rename(columns={'user_name': 'שם', 'device_type': 'מכשיר נפוץ'})
                        st.dataframe(_user_dev[['מכשיר נפוץ', 'שם']], use_container_width=True, hide_index=True)
                else:
                    st.info("אין נתוני מכשיר עדיין.")

                st.divider()

                # ─── H. זיהוי רענון דף ────────────────────────────────────────────
                st.markdown("#### H. זיהוי רענון דף")
                if not _logins.empty:
                    _logins_ts2 = _logins.copy()
                    _logins_ts2['ts'] = pd.to_datetime(_logins_ts2['timestamp'], errors='coerce')
                    _logins_ts2 = _logins_ts2.sort_values(['user_name', 'ts'])
                    _logins_ts2['prev_ts'] = _logins_ts2.groupby('user_name')['ts'].shift(1)
                    _logins_ts2['gap_min'] = (_logins_ts2['ts'] - _logins_ts2['prev_ts']).dt.total_seconds() / 60
                    _reloads = _logins_ts2[_logins_ts2['gap_min'] < 5]['user_name'].unique().tolist()
                    if _reloads:
                        st.warning(f"⚠️ משתמשים עם כניסות כפולות תוך 5 דקות (ייתכן רענון דף): {', '.join(_reloads)}")
                    else:
                        st.success("לא זוהו רענונים חשודים.")
                else:
                    st.info("אין נתונים.")

                st.divider()

                # ─── I. תזמון הגשות ───────────────────────────────────────────────
                st.markdown("#### I. תזמון הגשות (יום בחודש)")
                if not _submits.empty:
                    _sub_day = _submits['day_of_month'].astype(str).str.extract(r'(\d+)')[0]
                    _sub_day = pd.to_numeric(_sub_day, errors='coerce').dropna().astype(int)
                    _day_counts = _sub_day.value_counts().reindex(range(1, 32), fill_value=0).reset_index()
                    _day_counts.columns = ['יום בחודש', 'הגשות']
                    st.bar_chart(_day_counts.set_index('יום בחודש')['הגשות'])
                else:
                    st.info("אין נתוני הגשות עדיין.")

    # ── סידור חודשי (אדמין) ────────────────────────────────────
    if selected_nav == 'גאנט חודשי':
        st.subheader("🗓️ גאנט חודשי — ניהול")

        # ── Top: month selector + active-month action ──────────
        hebrew_months = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני","יולי",
                         "אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

        # Show deferred success message from "הפוך לחודש פעיל" button
        if st.session_state.get('show_active_month_success'):
            st.success(f"✅ סידור לחודש {st.session_state.pop('show_active_month_success')} נוצר. הגשות נפתחו.")

        # 12 consecutive months starting from the active month, each carrying its
        # correct year so the window rolls into the next year past December
        # (e.g. active=December → ... דצמבר 2026 / ינואר 2027 / ... נובמבר 2027).
        # base_year follows the real clock: if the active month already passed
        # this calendar year, it refers to next year's occurrence.
        _now = datetime.now()
        base_year = _now.year if daily_active_month_int >= _now.month else _now.year + 1
        view_pairs = [
            (base_year + (daily_active_month_int - 1 + i) // 12,
             ((daily_active_month_int - 1 + i) % 12) + 1)
            for i in range(12)
        ]
        view_opts = [m for (_y, m) in view_pairs]          # months in window order
        _month_to_year = {m: y for (y, m) in view_pairs}   # month → its window year

        # Single session-state key — the selectbox IS the state; no double-assignment
        if 'daily_view_month' not in st.session_state or \
                st.session_state.daily_view_month not in view_opts:
            st.session_state.daily_view_month = daily_active_month_int

        col_top1, col_top2, col_top3 = st.columns([2, 2, 2])
        with col_top1:
            st.selectbox(
                "🗓️ חודש לתכנון/עריכה:",
                view_opts,
                format_func=lambda m: f"{hebrew_months[m-1]} {_month_to_year[m]}",
                key="daily_view_month",   # session state IS the widget state
            )
        view_month = st.session_state.daily_view_month
        view_year  = _month_to_year[view_month]
        view_year_month = f"{view_year}-{view_month:02d}"

        with col_top2:
            is_active = (view_month == daily_active_month_int)
            if is_active:
                st.markdown(
                    f"<div style='padding:8px;background:#dcfce7;border-radius:8px;"
                    f"text-align:center;color:#166534;font-weight:600'>"
                    f"✅ זהו החודש הפעיל</div>",
                    unsafe_allow_html=True)
            else:
                if st.button(f"⚡ הפוך את {hebrew_months[view_month-1]} לחודש פעיל",
                             key="set_active_month", use_container_width=True):
                    _set_setting('daily_active_month', view_month)
                    _set_setting('daily_requests_open', 'True')
                    # _set_setting already updated st.session_state.settings in memory.
                    # Refresh TTL timestamp so the TTL block does NOT overwrite the
                    # in-memory settings with a potentially-stale DB read on the next render.
                    st.session_state['_settings_fetched_at'] = time.time()
                    # Sync month selector so both tabs open on the new month
                    st.session_state.daily_view_month = view_month
                    st.session_state['show_active_month_success'] = hebrew_months[view_month-1]
                    st.rerun()
        with col_top3:
            req_open_now = daily_requests_open if is_active else None
            if is_active:
                if req_open_now:
                    if st.button(f"🔓 סגור הגשות לחודש {hebrew_months[view_month-1]}",
                                 key="toggle_req_close", use_container_width=True):
                        _set_setting('daily_requests_open', 'False')
                        st.success("הגשות נסגרו"); st.rerun()
                else:
                    if st.button(f"🔒 פתח הגשות לחודש {hebrew_months[view_month-1]}",
                                 key="toggle_req_open", use_container_width=True):
                        _set_setting('daily_requests_open', 'True')
                        st.success("הגשות נפתחו"); st.rerun()
            else:
                st.caption("פתיחה/סגירה זמינה רק לחודש הפעיל")

        st.divider()

        # ── שיבוץ חודשי — flat layout, no sub-tabs ───────────────────
        st.markdown(f"#### שיוך עובדים למחלקות — {hebrew_months[view_month-1]} {view_year}")
        st.caption("בחר לכל עובד את המחלקה היומית שלו לחודש זה. תורן חוץ ומנהלים אינם בטבלה.")

        DAILY_DEPTS = list(DAILY_DEPTS_ALL) + ["— לא שובץ —"]

        staff = st.session_state.staff.copy()
        staff['name'] = staff['name'].astype(str).str.strip()
        staff['type'] = staff['type'].astype(str).str.strip()
        eligible = staff[staff['type'].isin(['מתמחה', 'רופא בכיר'])]
        eligible = eligible[eligible['name'].str.len() > 0]
        eligible = eligible[eligible['name'] != '---']

        if eligible.empty:
            st.warning("אין עובדים פעילים מסוג מתמחה או רופא/ה בכיר/ה.")
        else:
            dr = st.session_state.dept_rotation.copy()
            if 'employee' in dr.columns:
                dr['employee']   = dr['employee'].astype(str).str.strip()
                dr['year_month'] = dr['year_month'].astype(str)
                dr['daily_dept'] = dr['daily_dept'].astype(str)
                if 'side' not in dr.columns: dr['side'] = ''
                dr['side'] = dr['side'].fillna('').astype(str).str.strip()
                _cur_month_dr = dr[dr['year_month'] == view_year_month]
                existing = _cur_month_dr.set_index('employee')['daily_dept'].to_dict()
                existing_sides = _cur_month_dr.set_index('employee')['side'].to_dict()
            else:
                existing = {}
                existing_sides = {}

            _gantt_saved = st.session_state.pop('show_gantt_save_success', None)
            if _gantt_saved is not None:
                st.success(f"✅ נשמרו {_gantt_saved} שיבוצים לחודש {hebrew_months[view_month-1]}")
            if _mig := st.session_state.pop('show_gantt_migrate_info', 0):
                st.info(f"↪️ הועברו {_mig} בקשות היעדרות למחלקה החדשה.")

            with st.form(f"rotation_form_{view_month}"):
                new_assignments = {}
                new_sides = {}
                # Header row
                _hc1, _hc2, _hc3 = st.columns([2, 2, 1])
                _hc2.caption("מחלקה")
                _hc3.caption("צד (פנימית)")
                for _, emp_row in eligible.iterrows():
                    emp_name = emp_row['name']
                    emp_type = emp_row['type']
                    current = existing.get(emp_name, "— לא שובץ —")
                    if current not in DAILY_DEPTS:
                        current = "— לא שובץ —"
                    cur_side = existing_sides.get(emp_name, "")
                    if cur_side not in PNIM_SIDES: cur_side = ""
                    c1, c2, c3 = st.columns([2, 2, 1])
                    c1.markdown(
                        f"<div style='padding:6px 0;font-weight:500'>"
                        f"{emp_name} <span style='color:#64748b;font-size:0.8rem'>({emp_type})</span></div>",
                        unsafe_allow_html=True)
                    new_assignments[emp_name] = c2.selectbox(
                        "", DAILY_DEPTS,
                        index=DAILY_DEPTS.index(current),
                        key=f"rot_{emp_name}_{view_month}",
                        label_visibility="collapsed"
                    )
                    # Side selector: show for all; only relevant when dept=פנימית
                    # (current DB value drives whether side options are highlighted)
                    _side_opts = [""] + PNIM_SIDES
                    _side_idx = _side_opts.index(cur_side) if cur_side in _side_opts else 0
                    new_sides[emp_name] = c3.selectbox(
                        "", _side_opts,
                        index=_side_idx,
                        key=f"rot_side_{emp_name}_{view_month}",
                        label_visibility="collapsed",
                        format_func=lambda s: ("🌸" if s == "ורוד" else "🔵" if s == "כחול" else "—"),
                    )
                submit = st.form_submit_button(f"💾 שמור שיבוץ לחודש {hebrew_months[view_month-1]}")

            if submit:
                # Detect all dept changes for this month
                dept_changes = [
                    (emp_n, existing[emp_n], new_dept)
                    for emp_n, new_dept in new_assignments.items()
                    if (emp_n in existing
                        and existing[emp_n]
                        and new_dept != "— לא שובץ —"
                        and existing[emp_n] != new_dept)
                ]

                other_months = dr[dr['year_month'] != view_year_month] if not dr.empty else pd.DataFrame(
                    columns=['employee','year_month','daily_dept','side'])
                new_rows = []
                for emp_n, dept_n in new_assignments.items():
                    if dept_n != "— לא שובץ —":
                        # Only carry the side value for פנימית assignments
                        side_val = new_sides.get(emp_n, '') if dept_n == PNIM_DEPT else ''
                        new_rows.append({
                            'employee': emp_n,
                            'year_month': view_year_month,
                            'daily_dept': dept_n,
                            'side': side_val,
                        })
                new_dr = _norm_dr(pd.concat([other_months, pd.DataFrame(new_rows)], ignore_index=True))
                st.session_state.dept_rotation = new_dr
                save_to_db("dept_rotation", new_dr)

                # Migrate absence requests + WSD rows for each dept change
                total_migrated = sum(
                    _migrate_requests_on_dept_change(emp_n, old_d, new_d, view_year_month)
                    for emp_n, old_d, new_d in dept_changes
                )
                if total_migrated:
                    st.session_state['show_gantt_migrate_info'] = total_migrated

                st.session_state['show_gantt_save_success'] = len(new_rows)
                st.rerun()

            # Summary
            st.divider()
            if not dr.empty:
                cur = dr[dr['year_month'] == view_year_month]
                if not cur.empty:
                    st.caption(f"📊 שיבוצים נוכחיים בחודש: {len(cur)} עובדים")
                    for d in DAILY_DEPTS[:-1]:
                        cnt = (cur['daily_dept'] == d).sum()
                        if cnt > 0:
                            st.caption(f"   • {d}: {cnt} עובדים")

        # ── ייצוא כל המחלקות ─────────────────────────────────────
        st.divider()
        st.markdown("##### 📥 ייצוא סידור — כל המחלקות")
        _exp_col1, _exp_col2, _ = st.columns([2, 2, 3])
        with _exp_col1:
            _all_bytes, _all_fname = _export_all_depts_batched(view_year_month, view_month)
            if _all_bytes:
                st.download_button(
                    f"📥 Excel — כל המחלקות ({hebrew_months[view_month-1]})",
                    data=_all_bytes,
                    file_name=_all_fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"xl_all_depts_{view_month}",
                    use_container_width=True,
                )
            else:
                st.caption("אין שיבוצים לייצוא לחודש זה.")


    # ── סידור עבודה (אדמין) — same view as מנהל מחלקה but for any dept of choice ──
    if selected_nav == 'סידור עבודה':
        st.subheader("🗓️ סידור עבודה — תצוגת מנהל/ת מחלקה (אדמין)")
        st.caption("בחר מחלקה לראות אותה כפי שמנהל המחלקה רואה.")

        sel_dept_admin = st.selectbox("מחלקה לתצוגה:", DAILY_DEPTS_ALL, key="adm_sy_dept")
        # Feature 4: 12-month rolling forward window (auto-advances on month change).
        adm_year, adm_view_m = _sw_month_selector_12("sw_adm")
        adm_y_m = f"{adm_year}-{adm_view_m:02d}"

        st.markdown(f"#### לוח {sel_dept_admin} — {_HEB_MONTHS[adm_view_m-1]} {adm_year}")

        # Build employees: dept_rotation + ALL managers of this dept.
        # Managers always appear as rows; their cells default to empty
        # (via _derive_auto_status) — click any cell to "plant" them.
        _adm_dr = st.session_state.dept_rotation
        _adm_emps = []
        if not _adm_dr.empty and 'employee' in _adm_dr.columns:
            _adm_mask = ((_adm_dr['year_month'].astype(str) == adm_y_m) &
                         (_adm_dr['daily_dept'].astype(str) == sel_dept_admin))
            _adm_emps = _adm_dr[_adm_mask]['employee'].astype(str).str.strip().tolist()
        _dept_managers = _get_dept_managers(sel_dept_admin)
        _all_adm_emps = _sort_employees_by_role(
            _dept_managers + [e for e in _adm_emps if e not in _dept_managers]
        )

        if sel_dept_admin == PNIM_DEPT:
            _render_pnim_sided(adm_y_m, adm_view_m, "adm_mgrview",
                               employees=_all_adm_emps, allow_temp_add=True,
                               year=adm_year)
        else:
            _render_dept_grid(sel_dept_admin, adm_y_m, adm_view_m,
                              "adm_mgrview", employees=_all_adm_emps,
                              allow_temp_add=True, year=adm_year)

        st.divider()
        _render_export_buttons(sel_dept_admin, adm_y_m, adm_view_m,
                               "adm_mgrview", user_name)

    # ── ניהול בקשות (אדמין) ───────────────────────────────────────
    if selected_nav == 'ניהול בקשות':
        st.subheader("📋 ניהול בקשות היעדרות")

        # ── Section 1: בקשות ממתינות ──────────────────────────────
        st.markdown("#### ⏳ בקשות ממתינות")
        nb_dept_filter = st.selectbox(
            "סנן לפי מחלקה:",
            ["הכל"] + DAILY_DEPTS_ALL,
            key="nb_adm_dept"
        )
        ar_nb = st.session_state.absence_requests.copy()
        if ar_nb.empty or 'status' not in ar_nb.columns:
            st.info("אין בקשות במערכת.")
        else:
            ar_nb['status']          = ar_nb['status'].astype(str).str.lower()
            ar_nb['dept_at_request'] = ar_nb['dept_at_request'].astype(str)
            pen_nb = ar_nb[ar_nb['status'] == 'pending']
            if nb_dept_filter != "הכל":
                pen_nb = pen_nb[pen_nb['dept_at_request'] == nb_dept_filter]
            if 'start_date' in pen_nb.columns:
                pen_nb = pen_nb.sort_values('start_date')
            if pen_nb.empty:
                st.success("אין בקשות ממתינות ✅")
            else:
                st.caption(f"📋 {len(pen_nb)} בקשות ממתינות")
                # Pre-compute approved rows for overlap check (once, outside loop)
                _ar_appr_nb = ar_nb.copy()
                _ar_appr_nb['_sd'] = pd.to_datetime(_ar_appr_nb['start_date'], errors='coerce')
                _ar_appr_nb['_ed'] = pd.to_datetime(_ar_appr_nb['end_date'], errors='coerce')
                _ar_appr_nb = _ar_appr_nb[_ar_appr_nb['status'] == 'approved']
                for idx, row in pen_nb.iterrows():
                    with st.container(border=True):
                        # Compute conflicts up front — they drive the approve UX.
                        req_id = str(row.get('id', idx))
                        _rs_d  = pd.to_datetime(row.get('start_date', ''), errors='coerce')
                        _re_d  = pd.to_datetime(row.get('end_date', ''), errors='coerce')
                        _remp  = str(row.get('employee', '')).strip()
                        _rdept_canon = _emp_dept_for_date(_remp, _rs_d)
                        _row_conflicts = _absence_conflicts(
                            _remp, _rdept_canon, _rs_d, _re_d,
                            exclude_id=row.get('id'))

                        cc1, cc2, cc3, cc4, cc5, cc6 = st.columns([2, 1.5, 1.5, 1.5, 1, 1])
                        cc1.markdown(f"**{row.get('employee', '—')}**")
                        cc2.write(_emp_dept_for_date(row.get('employee', ''),
                                                     row.get('start_date', '')) or '—')
                        cc3.write(f"{row['start_date']} – {row['end_date']}")
                        cc4.write(row.get('type', '—'))
                        # Conflict → disable the inline ✅ button; an explicit
                        # confirm button appears below the warning instead.
                        if cc5.button("✅ אשר", key=f"nb_adm_ap_{req_id}",
                                      use_container_width=True,
                                      disabled=bool(_row_conflicts)):
                            _approve_request(req_id, str(user_name).strip())
                            st.rerun()
                        if cc6.button("❌ דחה", key=f"nb_adm_rj_{req_id}",
                                      use_container_width=True):
                            _reject_request(req_id, str(user_name).strip())
                            st.rerun()
                        if _row_conflicts:
                            st.warning(_format_absence_conflict_question(_row_conflicts))
                            if st.button("✅ כן, אשר למרות החפיפה",
                                         key=f"nb_adm_ap_force_{req_id}",
                                         type="primary"):
                                _approve_request(req_id, str(user_name).strip())
                                st.rerun()
                        if row.get('notes'):
                            st.caption(f"💬 {row['notes']}")

        st.divider()

        # ── Section 2: Gantt-style approved-absence view (Feature 3) ────
        st.markdown("#### ✅ בקשות שאושרו — תצוגת גאנט")
        # 12-month selector for which month to visualise (auto-advancing).
        _gantt_year, _gantt_month = _sw_month_selector_12("nb_adm_gantt")
        _render_absence_gantt(_gantt_year, _gantt_month)

        if st.session_state.pop('show_nb_adm_del_ok', False):
            st.success("🗑️ הבקשה נמחקה.")

        # Hard-delete an approved request (admin power-tool).
        with st.expander("🗑️ מחיקת בקשה מאושרת", expanded=False):
            _del_df = st.session_state.absence_requests.copy()
            if _del_df.empty or 'status' not in _del_df.columns:
                st.info("אין בקשות במערכת.")
            else:
                _del_df['_status'] = _del_df['status'].astype(str).str.lower()
                _del_df = _del_df[_del_df['_status'] == 'approved']
                if _del_df.empty:
                    st.info("אין בקשות מאושרות למחיקה.")
                else:
                    _del_df['_label'] = _del_df.apply(
                        lambda r: (f"{str(r.get('employee', '')).strip()} · "
                                   f"{_emp_dept_for_date(r.get('employee', ''), r.get('start_date', '')) or '—'} · "
                                   f"{str(r.get('start_date', ''))[:10]} – {str(r.get('end_date', ''))[:10]} · "
                                   f"{str(r.get('type', '') or '').strip()}"),
                        axis=1)
                    _del_options = _del_df[['id', '_label']].values.tolist()
                    _label_to_id = {lbl: rid for rid, lbl in _del_options}
                    _sel_del = st.selectbox(
                        "בחר/י בקשה למחיקה:",
                        ["—"] + [lbl for _, lbl in _del_options],
                        key="nb_adm_del_sel")
                    _confirm_del = st.checkbox(
                        "✓ אני בטוח/ה — המחיקה אינה הפיכה",
                        key="nb_adm_del_confirm",
                        value=False)
                    if st.button("🗑️ מחק בקשה", key="nb_adm_del_btn",
                                 disabled=not (_sel_del and _sel_del != "—"
                                               and _confirm_del),
                                 type="primary"):
                        _rid = _label_to_id.get(_sel_del)
                        if _rid and _delete_absence_request(_rid):
                            st.session_state['show_nb_adm_del_ok'] = True
                            st.rerun()
                        else:
                            st.error("שגיאה במחיקה.")

        # Original tabular view kept as fallback for raw inspection.
        with st.expander("📋 תצוגת טבלה (כל הבקשות העתידיות שאושרו)", expanded=False):
            ar_nb2 = st.session_state.absence_requests.copy()
            if not ar_nb2.empty and 'status' in ar_nb2.columns:
                ar_nb2['status']    = ar_nb2['status'].astype(str).str.lower()
                ar_nb2['end_date']  = pd.to_datetime(ar_nb2['end_date'], errors='coerce')
                ar_nb2['start_date'] = pd.to_datetime(ar_nb2['start_date'], errors='coerce')
                _today_ts = pd.Timestamp(date.today())
                ap_nb2 = ar_nb2[
                    (ar_nb2['status'] == 'approved') &
                    (ar_nb2['end_date'] >= _today_ts)
                ].copy()
                if ap_nb2.empty:
                    st.info("אין בקשות שאושרו לתאריכים עתידיים.")
                else:
                    # Normalize dept_at_request via Gantt for display (Feature 1).
                    ap_nb2['_dept_canon'] = ap_nb2.apply(
                        lambda r: _emp_dept_for_date(r['employee'], r['start_date']) or '—',
                        axis=1)
                    ap_nb2 = ap_nb2.sort_values(['_dept_canon', 'start_date'])
                    _show_nb2 = ap_nb2[['employee', '_dept_canon', 'start_date', 'end_date',
                                        'type', 'approved_by']].copy()
                    _show_nb2.columns = ['עובד/ת', 'מחלקה', 'מתאריך', 'עד תאריך', 'סוג', 'אושר ע"י']
                    _show_nb2['מתאריך']  = _show_nb2['מתאריך'].dt.strftime('%Y-%m-%d')
                    _show_nb2['עד תאריך'] = _show_nb2['עד תאריך'].dt.strftime('%Y-%m-%d')
                    st.dataframe(_show_nb2, use_container_width=True, hide_index=True)
            else:
                st.info("אין בקשות שאושרו.")

        st.divider()

        # ── Section 3: הוספת היעדרות עתידית לעובד ─────────────────
        st.markdown("#### ➕ הוסף היעדרות עתידית לעובד")
        sf_all_nb = st.session_state.staff
        active_emps_nb = sorted(
            sf_all_nb[sf_all_nb['type'].isin(['מתמחה', 'רופא בכיר', 'מנהל מחלקה', 'מנהל/ת'])]
            ['name'].astype(str).str.strip().tolist()
        )
        adm_nb_emp = st.selectbox("עובד/ת:", active_emps_nb, key="nb_adm_emp")
        nb_adm_col1, nb_adm_col2 = st.columns(2)
        with nb_adm_col1:
            adm_nb_start = st.date_input("מתאריך:", key="nb_adm_start",
                                         value=date.today(),
                                         min_value=date(2026, 1, 1),
                                         max_value=date(2026, 12, 31))
        with nb_adm_col2:
            adm_nb_end = st.date_input("עד תאריך:", key="nb_adm_end",
                                       value=date.today(),
                                       min_value=date(2026, 1, 1),
                                       max_value=date(2026, 12, 31))
        adm_nb_type = st.selectbox("סוג היעדרות:", ["חופש", "202", "היעדרות אחרת"], key="nb_adm_type")
        adm_nb_note = st.text_input("הערה (אופציונלי):", key="nb_adm_note")

        # Pre-write conflict warning (Feature 2): show overlap with already-approved
        # absences in the same (Gantt-canonical) dept BEFORE the user clicks submit.
        _adm_nb_conflicts = []
        if adm_nb_emp and adm_nb_end >= adm_nb_start:
            _adm_nb_dept_preview = _emp_dept_for_date(adm_nb_emp, adm_nb_start)
            _adm_nb_conflicts = _absence_conflicts(
                adm_nb_emp, _adm_nb_dept_preview, adm_nb_start, adm_nb_end)
        # Conflict-confirmation flow: when there's an overlap, the regular save
        # button is replaced by an explicit "yes, save despite overlap" button.
        _adm_nb_force = False
        if _adm_nb_conflicts:
            st.warning(_format_absence_conflict_question(_adm_nb_conflicts))
            _adm_nb_force = st.button(
                "✅ כן, הוסף למרות החפיפה",
                key="nb_adm_submit_force", type="primary")

        if (not _adm_nb_conflicts and
            st.button("✅ הוסף היעדרות מאושרת", key="nb_adm_submit")) or _adm_nb_force:
            if adm_nb_end < adm_nb_start:
                st.error("תאריך סיום לפני תאריך התחלה.")
            else:
                # Gantt-canonical dept (Feature 1) — falls back to staff.dept only
                # when no dept_rotation row exists for the requested year-month.
                emp_dept_nb = _emp_dept_for_date(adm_nb_emp, adm_nb_start)
                new_nb_row = {
                    'id':              str(uuid.uuid4()),
                    'employee':        adm_nb_emp,
                    'start_date':      adm_nb_start.strftime('%Y-%m-%d'),
                    'end_date':        adm_nb_end.strftime('%Y-%m-%d'),
                    'type':            adm_nb_type,
                    'status':          'approved',
                    'dept_at_request': emp_dept_nb,
                    'manager_email':   '',
                    'approved_by':     str(user_name).strip(),
                    'notes':           adm_nb_note.strip(),
                    'created_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'responded_at':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }
                new_nb_df = pd.concat(
                    [st.session_state.absence_requests, pd.DataFrame([new_nb_row])],
                    ignore_index=True
                )
                st.session_state.absence_requests = new_nb_df
                save_to_db("absence_requests", new_nb_df)
                _materialize_absence_to_wsd(
                    adm_nb_emp, adm_nb_type,
                    new_nb_row['start_date'], new_nb_row['end_date'])
                _build_approved_map()
                st.session_state['show_nb_adm_success'] = True
                st.rerun()

        if st.session_state.pop('show_nb_adm_success', False):
            st.success("✅ היעדרות נוספה ואושרה — תופיע בלוח מיד.")


else:
    user_name = st.session_state.user_name

    # ── Night-shift constraints (existing UI) — only for roles that do night shifts ──
    if selected_nav == 'הגשת בקשות' and role in ('מתמחה', 'תורן חוץ'):
        # Display the active month name clearly
        hebrew_months = ["ינואר", "פברואר", "מרץ", "אפריל", "מאי", "יוני", "יולי", "אוגוסט", "ספטמבר", "אוקטובר", "נובמבר", "דצמבר"]
        month_name = hebrew_months[sel_month - 1]
        st.subheader(f"📋 הגשת בקשות לחודש: {month_name}")
        st.markdown("### 🌙 אילוצים לתורנויות לילה")
        
        # --- הצגת ימים מיוחדים / חגים לחודש זה ---
        if 'special_days' in st.session_state and not st.session_state.special_days.empty:
            special_days_month = []
            for _, row in st.session_state.special_days.iterrows():
                try:
                    d_obj = datetime.strptime(row['date'], '%Y-%m-%d').date()
                    if d_obj.month == sel_month and d_obj.year == 2026:
                        special_days_month.append(f"{d_obj.strftime('%d/%m/%Y')} - {row['description']}")
                except: pass
            
            if special_days_month:
                st.warning("⚠️ **שים לב לימים מיוחדים בחודש זה:**\n\n" + "\n".join([f"- {sd}" for sd in special_days_month]))
        
        # --- Pre-initialize Chip States for the entire month ---
        cal = calendar.monthcalendar(2026, sel_month)
        
        # Function to safely parse day from various date formats
        def get_day_nums(df, status_type):
            if df.empty: return []
            # Create a copy to avoid SettingWithCopy warnings
            user_rehab = df[(df['employee'] == user_name) & (df['status'] == status_type)].copy()
            
            if user_rehab.empty: return []
            
            # Robust conversion to datetime
            user_rehab['date_dt'] = pd.to_datetime(user_rehab['date'], errors='coerce')
            
            # Filter for valid dates in the selected month/year
            # sel_month is integer (e.g., 2)
            valid_dates = user_rehab[
                (user_rehab['date_dt'].dt.month == sel_month) & 
                (user_rehab['date_dt'].dt.year == 2026)
            ]
            
            return valid_dates['date_dt'].dt.day.tolist()

        default_day_nums = get_day_nums(st.session_state.requests, "אילוץ")
        default_wish_nums = get_day_nums(st.session_state.requests, "בקשה")
        
        default_dates  = [date(2026, sel_month, d) for d in default_day_nums]
        default_wishes = [date(2026, sel_month, d) for d in default_wish_nums]

        # --- הצגת אילוצים ובקשות קיימים ---
        existing_constraints = st.session_state.requests[(st.session_state.requests['employee'] == user_name) & (st.session_state.requests['status'] == "אילוץ")].copy()
        existing_wishes_all = st.session_state.requests[(st.session_state.requests['employee'] == user_name) & (st.session_state.requests['status'] == "בקשה")].copy()
        
        # סינון להצגה רק של החודש הפעיל הנבחר ב-UI (sel_month)
        if not existing_constraints.empty:
            existing_constraints['date_dt'] = pd.to_datetime(existing_constraints['date'], errors='coerce')
            existing_constraints = existing_constraints[(existing_constraints['date_dt'].dt.month == sel_month) & (existing_constraints['date_dt'].dt.year == 2026)]

        if not existing_wishes_all.empty:
            existing_wishes_all['date_dt'] = pd.to_datetime(existing_wishes_all['date'], errors='coerce')
            existing_wishes_all = existing_wishes_all[(existing_wishes_all['date_dt'].dt.month == sel_month) & (existing_wishes_all['date_dt'].dt.year == 2026)]
        
        if not existing_constraints.empty or not existing_wishes_all.empty:
            msg = ""
            if not existing_constraints.empty:
                msg += f"📅 **חסימות ({len(existing_constraints)}):** " + ", ".join([r['date'] for _, r in existing_constraints.iterrows()]) + "\n\n"
            if not existing_wishes_all.empty:
                msg += f"⭐ **בקשות ({len(existing_wishes_all)}):** " + ", ".join([r['date'] for _, r in existing_wishes_all.iterrows()])
            st.info(msg)
        else:
            st.info("עדיין לא הגשת אילוצים או בקשות לחודש זה.")
        # ----------------------------

        st.divider()
        cal_sd = st.session_state.get('special_days')
        if isinstance(cal_sd, pd.DataFrame) and cal_sd.empty:
            cal_sd = None
        constraint_days, wish_days = render_modern_calendar(
            2026, sel_month,
            default_day_nums, default_wish_nums,
            special_days_df=cal_sd,
            key_prefix='user_cal',
            show_validation=(st.session_state.user_role == 'מתמחה')
        )
        selected_from_grid = [date(2026, sel_month, d) for d in constraint_days]
        selected_wishes    = [date(2026, sel_month, d) for d in wish_days]
        
        # -----------------------------------
        st.divider()
        
        # ── Live "what will be saved" preview ─────────────────────────
        # (formerly hidden behind a two-step confirm flow — employees were
        # missing the second button and losing their submissions; collapsed
        # to a single atomic save here.)
        _added         = set(selected_from_grid) - set(default_dates)
        _removed       = set(default_dates)      - set(selected_from_grid)
        _added_wishes  = set(selected_wishes)    - set(default_wishes)
        _removed_wishes = set(default_wishes)    - set(selected_wishes)
        _has_changes = bool(_added or _removed or _added_wishes or _removed_wishes)
        _is_empty_save = (not selected_from_grid and not selected_wishes
                          and not default_dates and not default_wishes)

        if _has_changes:
            _preview_lines = []
            if _added:
                _preview_lines.append(f"➕ **נוספו לחסימה:** {', '.join([d.strftime('%d/%m/%Y') for d in sorted(_added)])}")
            if _removed:
                _preview_lines.append(f"➖ **הוסרו מחסימה:** {', '.join([d.strftime('%d/%m/%Y') for d in sorted(_removed)])}")
            if _added_wishes:
                _preview_lines.append(f"⭐ **נוספו לבקשה:** {', '.join([d.strftime('%d/%m/%Y') for d in sorted(_added_wishes)])}")
            if _removed_wishes:
                _preview_lines.append(f"⭐❌ **הוסרו מבקשה:** {', '.join([d.strftime('%d/%m/%Y') for d in sorted(_removed_wishes)])}")
            if not selected_from_grid and default_dates:
                st.warning("⚠️ אתה עומד להסיר את **כל** החסימות שלך לחודש זה.")
            st.info("📋 **השינויים שיישמרו:**\n\n" + "\n\n".join(_preview_lines))

        _save_label = ("💾 שמור אילוצים ובקשות"
                       if (_has_changes or _is_empty_save)
                       else "אין שינויים לשמירה")
        _save_disabled = not (_has_changes or _is_empty_save)

        if st.button(_save_label, type="primary",
                     use_container_width=True,
                     disabled=_save_disabled,
                     key="save_constraints_single"):
            # --- ולידציה (חוקים) ---
            errors = []
            if len(selected_wishes) > 2:
                errors.append("שגיאה: ניתן לבחור עד 2 בקשות חיוביות (⭐) בלבד.")
            overlap = set(selected_from_grid).intersection(set(selected_wishes))
            if overlap:
                errors.append(f"שגיאה: בחרת באותו יום ({list(overlap)[0].strftime('%d/%m')}) גם אילוץ וגם בקשה. נא בחר רק אחד.")
            if st.session_state.user_role == 'מתמחה':
                num_days = calendar.monthrange(2026, sel_month)[1]
                month_days = [date(2026, sel_month, d) for d in range(1, num_days+1)]
                total_thursdays = len([d for d in month_days if d.weekday() == 3])
                total_weekends  = len([d for d in month_days if d.weekday() in [4, 5]])
                blocked_thursdays = len([d for d in selected_from_grid if d.weekday() == 3])
                blocked_weekends  = len([d for d in selected_from_grid if d.weekday() in [4, 5]])
                avail_thursdays = total_thursdays - blocked_thursdays
                avail_weekends  = total_weekends  - blocked_weekends
                if avail_thursdays < 2:
                    errors.append("נותר רק יום חמישי אחד פנוי (או פחות). חובה להשאיר לפחות 2 ימי חמישי פנויים.")
                if avail_weekends < 4:
                    errors.append(f"נותרו רק {avail_weekends} ימי סוף שבוע פנויים. חובה להשאיר לפחות 4 (שישי/שבת).")

            _is_admin = st.session_state.user_role == 'מנהל/ת'
            if errors and not _is_admin:
                for e in errors:
                    st.error(e)
            else:
                # משיכת נתונים חיים מהגיליון (עוקף cache למניעת דריסה הדדית בין משתמשים)
                live = _fetch_live("requests")
                base = live if not live.empty else st.session_state.requests

                # הסרת כל האילוצים והבקשות הקודמים של המשתמש לחודש זה
                current_month_prefix = f"2026-{sel_month:02d}"
                mask_keep = ~((base['employee'].astype(str).str.strip() == str(user_name).strip()) &
                              (base['date'].astype(str).str.startswith(current_month_prefix)))
                base = base[mask_keep]

                # הוספת הרשימה החדשה והמעודכנת
                new_rows = []
                if selected_from_grid:
                    new_rows += [{'employee': user_name, 'date': str(d), 'status': "אילוץ"}
                                 for d in selected_from_grid]
                if selected_wishes:
                    new_rows += [{'employee': user_name, 'date': str(d), 'status': "בקשה"}
                                 for d in selected_wishes]

                merged = pd.concat([base, pd.DataFrame(new_rows)], ignore_index=True) if new_rows else base
                save_to_db("requests", merged)
                _log_async('constraint_submit', str(len(selected_from_grid)), str(len(selected_wishes)))
                st.session_state.requests = merged
                _fetch_sheet_data_silently.clear()
                st.session_state.pop(f"user_cal_init_{sel_month}", None)
                st.session_state.pop('confirm_request_save', None)   # legacy flag cleanup
                st.session_state['show_update_success'] = True
                st.rerun()

    if st.session_state.get('show_update_success'):
        st.success("✅ האילוצים עודכנו בהצלחה!")
        st.session_state['show_update_success'] = False

    # ── Day-absence section (Phase 3) — for מתמחה / רופא בכיר ─────────────────────────
    if selected_nav == 'הגשת בקשות' and role in ('מתמחה', 'רופא בכיר', 'מנהל מחלקה'):
        hebrew_months_da = ["ינואר","פברואר","מרץ","אפריל","מאי","יוני","יולי",
                            "אוגוסט","ספטמבר","אוקטובר","נובמבר","דצמבר"]

        # מנהל/ת מחלקה / רופא/ה בכיר/ה — no night UI above, so render page header
        if role in ('רופא בכיר', 'מנהל מחלקה'):
            st.subheader(f"📋 הגשת בקשות לחודש: {hebrew_months_da[daily_active_month_int - 1]}")
            if role == 'מנהל מחלקה':
                st.info("💡 כמנהל/ת מחלקה בקשותיך מאושרות אוטומטית ומשתקפות מיד בלוח.")
        else:
            st.divider()

        st.markdown("### ☀️ היעדרות מעבודה יומית")

        # Status banner: open/closed
        if not daily_requests_open:
            st.warning(f"🔒 הגשות לחודש {hebrew_months_da[daily_active_month_int - 1]} נסגרו על ידי המנהל.")

        # Use the daily_active_month for this section (separate from sel_month/active_month)
        da_year_month = f"2026-{daily_active_month_int:02d}"
        da_num_days = calendar.monthrange(2026, daily_active_month_int)[1]
        da_first_wd = (date(2026, daily_active_month_int, 1).weekday() + 1) % 7  # Sun=0

        # Selection state — namespaced per month
        sel_start_key = f"dayabs_sel_start_{daily_active_month_int}"
        sel_end_key   = f"dayabs_sel_end_{daily_active_month_int}"
        if sel_start_key not in st.session_state: st.session_state[sel_start_key] = None
        if sel_end_key   not in st.session_state: st.session_state[sel_end_key]   = None

        # Pre-color sets from absence_requests for this user × this month
        ar_df = st.session_state.absence_requests.copy()
        approved_vac_days, approved_202_days, approved_future_days, pending_days = set(), set(), set(), set()
        if not ar_df.empty and 'employee' in ar_df.columns:
            ar_df['employee'] = ar_df['employee'].astype(str).str.strip()
            user_n = str(user_name).strip()
            mine = ar_df[ar_df['employee'] == user_n]
            for _, r in mine.iterrows():
                try:
                    sd = datetime.strptime(str(r['start_date']), '%Y-%m-%d').date()
                    ed = datetime.strptime(str(r['end_date']),   '%Y-%m-%d').date()
                except Exception:
                    continue
                # Only include days that fall within the displayed month
                d_iter = sd
                while d_iter <= ed:
                    if d_iter.month == daily_active_month_int and d_iter.year == 2026:
                        status = str(r.get('status', '')).strip().lower()
                        atype = str(r.get('type', '')).strip()
                        if status == 'approved':
                            if atype == '202':
                                approved_202_days.add(d_iter.day)
                            elif atype == 'חופש עתידי':
                                approved_future_days.add(d_iter.day)
                            else:  # חופש / היעדרות אחרת → 🔵
                                approved_vac_days.add(d_iter.day)
                        elif status == 'pending':
                            pending_days.add(d_iter.day)
                    d_iter = d_iter + timedelta(days=1)

        # ── Night-shift + post-shift days for this user × this month ──
        night_shift_days, post_shift_days = set(), set()
        sched_df_da = st.session_state.schedule.copy()
        if not sched_df_da.empty and 'employee' in sched_df_da.columns:
            sched_df_da['employee'] = sched_df_da['employee'].astype(str).str.strip()
            sched_df_da['date']     = sched_df_da['date'].astype(str)
            sched_df_da['dept']     = sched_df_da['dept'].astype(str)
            user_shifts = sched_df_da[
                (sched_df_da['employee'] == str(user_name).strip()) &
                (~sched_df_da['dept'].str.contains('שישי בוקר', na=False))
            ]
            for _, sr in user_shifts.iterrows():
                try:
                    sd = datetime.strptime(sr['date'], '%Y-%m-%d').date()
                except Exception:
                    continue
                if sd.month == daily_active_month_int and sd.year == 2026:
                    night_shift_days.add(sd.day)
                # Post-shift = day AFTER the shift
                next_d = sd + timedelta(days=1)
                if next_d.month == daily_active_month_int and next_d.year == 2026:
                    post_shift_days.add(next_d.day)

        # ── Click-cycle state: 0=free  1=חופש request (light green)  2=202 request (light yellow)
        _cycle_ns = f"dayabs_cycle_{daily_active_month_int}"
        _init_key = f"dayabs_cyc_init_{daily_active_month_int}"

        # Pre-populate cycle from existing pending requests (once per session/month)
        if not st.session_state.get(_init_key, False):
            _cyc_init: dict[int, int] = {}
            _ar_init = st.session_state.absence_requests
            if not _ar_init.empty and 'employee' in _ar_init.columns:
                _user_n = str(user_name).strip()
                _mine_init = _ar_init[_ar_init['employee'].astype(str).str.strip() == _user_n]
                for _, _r in _mine_init.iterrows():
                    if str(_r.get('status', '')).strip().lower() != 'pending':
                        continue
                    try:
                        _sd = datetime.strptime(str(_r['start_date']), '%Y-%m-%d').date()
                        _ed = datetime.strptime(str(_r['end_date']),   '%Y-%m-%d').date()
                    except Exception:
                        continue
                    _atype = str(_r.get('type', '')).strip()
                    _val   = 2 if _atype == '202' else 1
                    _di    = _sd
                    while _di <= _ed:
                        if _di.month == daily_active_month_int and _di.year == 2026:
                            _cyc_init[_di.day] = _val
                        _di += timedelta(days=1)
            st.session_state[_cycle_ns] = _cyc_init
            st.session_state[_init_key] = True

        _cycle: dict[int, int] = st.session_state.get(_cycle_ns, {})

        # ── Calendar grid — RTL: col 0=Saturday (ש), col 6=Sunday (א) ──
        HEB_WEEK = ["ש", "ו", "ה", "ד", "ג", "ב", "א"]
        _cal_da = [list(reversed(w)) for w in
                   calendar.Calendar(firstweekday=6).monthdayscalendar(2026, daily_active_month_int)]

        with st.container(border=True):
            # Header row
            cols_h = st.columns(7)
            for i, h in enumerate(HEB_WEEK):
                cols_h[i].markdown(f"<div class='dayabs-hdr'>{h}</div>", unsafe_allow_html=True)

            for week in _cal_da:
                cols = st.columns(7)
                for ci, day in enumerate(week):
                    with cols[ci]:
                        if day == 0:
                            st.markdown("<div class='dayabs-empty'></div>", unsafe_allow_html=True)
                            continue
                        # ── Fixed non-clickable states ───────────────
                        if day in night_shift_days:
                            st.markdown(f"<div class='dayabs-night'>{day}</div>",
                                        unsafe_allow_html=True)
                        elif day in post_shift_days:
                            st.markdown(f"<div class='dayabs-post'>{day}</div>",
                                        unsafe_allow_html=True)
                        elif day in approved_future_days or day in approved_vac_days:
                            st.markdown(f"<div class='dayabs-vac'>{day}</div>",
                                        unsafe_allow_html=True)
                        elif day in approved_202_days:
                            st.markdown(f"<div class='dayabs-202'>{day}</div>",
                                        unsafe_allow_html=True)
                        # ── Clickable: 3-state cycle ─────────────────
                        else:
                            _state = _cycle.get(day, 0)
                            if _state == 1:
                                # Light green — בקשה לחופש; click → 202 (or → free for רופא בכיר)
                                if st.button(str(day),
                                             key=f"dayabs_vac_{daily_active_month_int}_{day}",
                                             use_container_width=True):
                                    _cycle[day] = 0 if role == 'רופא בכיר' else 2
                                    st.session_state[_cycle_ns] = _cycle
                                    st.rerun()
                            elif _state == 2:
                                # Light yellow — בקשה ל-202; click → free
                                if st.button(str(day),
                                             key=f"dayabs_202_{daily_active_month_int}_{day}",
                                             use_container_width=True):
                                    _cycle[day] = 0
                                    st.session_state[_cycle_ns] = _cycle
                                    st.rerun()
                            else:
                                # White — free; click → חופש request
                                if st.button(str(day),
                                             key=f"dayabs_free_{daily_active_month_int}_{day}",
                                             use_container_width=True):
                                    _cycle[day] = 1
                                    st.session_state[_cycle_ns] = _cycle
                                    st.rerun()

            # Legend — clean colored squares, no emoji inside cells
            # רופא בכיר never sees the 202 entries (they can't request 202)
            _legend_html = (
                "<div style='direction:rtl;font-size:0.77rem;color:#475569;"
                "margin:8px 0 2px 0;line-height:2.2'>"
                "<span style='background:#1e3a5f;color:#1e3a5f;border-radius:4px;"
                "padding:1px 10px'>&nbsp;</span> תורנות &nbsp;|&nbsp; "
                "<span style='background:#f97316;color:#f97316;border-radius:4px;"
                "padding:1px 10px'>&nbsp;</span> אחרי תורנות &nbsp;|&nbsp; "
                "<span style='background:#bbf7d0;border-radius:4px;"
                "padding:1px 10px;border:1px solid #86efac'>&nbsp;</span> בקשה לחופש &nbsp;|&nbsp; "
                "<span style='background:#16a34a;color:#16a34a;border-radius:4px;"
                "padding:1px 10px'>&nbsp;</span> חופש מאושר &nbsp;|&nbsp; "
            )
            if role != 'רופא בכיר':
                _legend_html += (
                    "<span style='background:#fef9c3;border-radius:4px;"
                    "padding:1px 10px;border:1px solid #fde68a'>&nbsp;</span> בקשה ל-202 &nbsp;|&nbsp; "
                    "<span style='background:#eab308;color:#eab308;border-radius:4px;"
                    "padding:1px 10px'>&nbsp;</span> 202 מאושר &nbsp;|&nbsp; "
                )
            _legend_html += (
                "<span style='border:1px solid #e2e8f0;padding:1px 10px;border-radius:4px;"
                "background:white'>&nbsp;</span> פנוי"
                "</div>"
            )
            st.markdown(_legend_html, unsafe_allow_html=True)

        # ── Summary message + שמור button ──────────────────────────────
        _vac_sel = sorted(d for d, s in _cycle.items() if s == 1)
        _202_sel = sorted(d for d, s in _cycle.items() if s == 2)

        if _vac_sel or _202_sel:
            st.divider()
            # ── Conflict warning (Feature 2) — show same-dept overlap with already-approved
            # absences BEFORE the user clicks "שמור". Soft warning, doesn't block submission.
            _ea_first_day = (_vac_sel + _202_sel)[0]
            _ea_first_date = date(2026, daily_active_month_int, _ea_first_day)
            _ea_last_day = (_vac_sel + _202_sel)[-1]
            _ea_last_date = date(2026, daily_active_month_int, _ea_last_day)
            _ea_dept_preview = _emp_dept_for_date(user_name, _ea_first_date)
            _ea_conflicts = _absence_conflicts(
                user_name, _ea_dept_preview, _ea_first_date, _ea_last_date)
            if _ea_conflicts:
                st.warning(_format_absence_conflict_warning(_ea_conflicts))
            _btn_col, _msg_col = st.columns([1, 4])
            with _msg_col:
                _lines = []
                for _d in _vac_sel:
                    _lines.append(
                        f"<span style='background:#bbf7d0;color:#166534;border-radius:4px;"
                        f"padding:1px 7px;margin:2px;display:inline-block'>"
                        f"בקשה לחופש — {_d:02d}/{daily_active_month_int:02d}</span>")
                for _d in _202_sel:
                    _lines.append(
                        f"<span style='background:#fef9c3;color:#854d0e;border-radius:4px;"
                        f"padding:1px 7px;margin:2px;display:inline-block'>"
                        f"בקשה ל-202 — {_d:02d}/{daily_active_month_int:02d}</span>")
                st.markdown(
                    "<div style='direction:rtl;line-height:2;font-size:0.85rem'>"
                    + " ".join(_lines) + "</div>",
                    unsafe_allow_html=True)
            with _btn_col:
                can_submit = daily_requests_open or role == 'מנהל מחלקה'
                if st.button("💾 שמור", use_container_width=True,
                             key=f"dayabs_save_{daily_active_month_int}",
                             type="primary", disabled=not can_submit):
                    # ── Lookup dept + manager email (Feature 1 — Gantt-canonical) ───────────────────
                    # _emp_dept_for_date returns dept_rotation.daily_dept for the request's
                    # year-month, with a staff.dept fallback if the user hasn't yet been
                    # placed on the Gantt for that month.
                    _ea_first_d = (_vac_sel + _202_sel)[0]
                    _dept_at_req = _emp_dept_for_date(
                        user_name, date(2026, daily_active_month_int, _ea_first_d))
                    _mgr_email   = ""
                    if _dept_at_req:
                        for _, _sr in st.session_state.staff.iterrows():
                            if str(_sr.get('type', '')).strip() != 'מנהל מחלקה':
                                continue
                            if _dept_at_req in _parse_manage_depts(_sr.get('manage_depts', '')):
                                _mgr_email = str(_sr.get('email', '')).strip()
                                break

                    _is_mgr = (role == 'מנהל מחלקה')
                    _req_st = 'approved' if _is_mgr else 'pending'
                    _appr   = str(user_name).strip() if _is_mgr else ''
                    _resp   = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if _is_mgr else ''
                    _now    = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                    # ── Remove existing pending rows for this user × month ─
                    _ar_base = st.session_state.absence_requests.copy()
                    if not _ar_base.empty and 'employee' in _ar_base.columns:
                        _keep = ~(
                            (_ar_base['employee'].astype(str).str.strip() == str(user_name).strip()) &
                            (_ar_base['status'].astype(str).str.lower() == 'pending') &
                            (_ar_base['start_date'].astype(str).str.startswith(
                                f"2026-{daily_active_month_int:02d}"))
                        )
                        _ar_base = _ar_base[_keep]

                    # ── Add one row per selected day ──────────────────
                    _new_rows = []
                    for _d in _vac_sel:
                        _ds = f"2026-{daily_active_month_int:02d}-{_d:02d}"
                        _new_rows.append({
                            'id': str(uuid.uuid4()), 'employee': str(user_name).strip(),
                            'start_date': _ds, 'end_date': _ds, 'type': 'חופש',
                            'status': _req_st, 'dept_at_request': _dept_at_req,
                            'manager_email': _mgr_email, 'approved_by': _appr,
                            'notes': '', 'created_at': _now, 'responded_at': _resp,
                        })
                    for _d in _202_sel:
                        _ds = f"2026-{daily_active_month_int:02d}-{_d:02d}"
                        _new_rows.append({
                            'id': str(uuid.uuid4()), 'employee': str(user_name).strip(),
                            'start_date': _ds, 'end_date': _ds, 'type': '202',
                            'status': _req_st, 'dept_at_request': _dept_at_req,
                            'manager_email': _mgr_email, 'approved_by': _appr,
                            'notes': '', 'created_at': _now, 'responded_at': _resp,
                        })

                    _ar_final = pd.concat(
                        [_ar_base, pd.DataFrame(_new_rows)], ignore_index=True)
                    st.session_state.absence_requests = _ar_final
                    save_to_db("absence_requests", _ar_final)

                    if _is_mgr:
                        # Auto-approved: write directly to WSD
                        _mgr_dept2 = _dept_at_req
                        if not _mgr_dept2:
                            _srf2 = st.session_state.staff
                            _srf2 = _srf2[_srf2['name'].astype(str).str.strip() == str(user_name).strip()]
                            if not _srf2.empty:
                                _mdl2 = _parse_manage_depts(_srf2.iloc[0].get('manage_depts', ''))
                                _mgr_dept2 = _mdl2[0] if _mdl2 else ''
                        if _mgr_dept2:
                            for _d in _vac_sel:
                                _wsd_upsert(f"2026-{daily_active_month_int:02d}-{_d:02d}",
                                            str(user_name).strip(), _mgr_dept2,
                                            'חופש', is_manual=True, note='')
                            for _d in _202_sel:
                                _wsd_upsert(f"2026-{daily_active_month_int:02d}-{_d:02d}",
                                            str(user_name).strip(), _mgr_dept2,
                                            '202', is_manual=True, note='')
                        _build_approved_map()
                    else:
                        # Email manager
                        if _mgr_email:
                            _all_dates = (
                                [f"חופש — {_d:02d}/{daily_active_month_int:02d}" for _d in _vac_sel] +
                                [f"202 — {_d:02d}/{daily_active_month_int:02d}" for _d in _202_sel]
                            )
                            send_notification_email(
                                _mgr_email,
                                f"בקשת היעדרות חדשה: {user_name}",
                                f"<div dir='rtl'><p>שלום,</p>"
                                f"<p><b>{user_name}</b> ({_dept_at_req}) הגיש/ה בקשות היעדרות:</p>"
                                f"<ul>{''.join(f'<li>{x}</li>' for x in _all_dates)}</ul>"
                                f"<p>נא להיכנס למערכת לאישור.</p></div>"
                            )

                    # Reset init so calendar re-reads from DB on next render
                    st.session_state[_init_key] = False
                    st.session_state['show_dayabs_saved'] = True
                    st.rerun()

        if st.session_state.pop('show_dayabs_saved', False):
            st.success("✅ הבקשות נשמרו!")
        elif not (_vac_sel or _202_sel):
            st.caption("לחץ על יום כדי לבקש חופש או 202")

        # ── History table ──
        st.divider()
        st.markdown("#### 📋 היסטוריית בקשות")
        if ar_df.empty or 'employee' not in ar_df.columns:
            st.caption("אין בקשות עדיין.")
        else:
            mine = ar_df[ar_df['employee'].astype(str).str.strip() == str(user_name).strip()]
            if mine.empty:
                st.caption("אין בקשות עדיין.")
            else:
                show_cols = ['start_date', 'end_date', 'type', 'status', 'approved_by', 'notes']
                disp = mine[[c for c in show_cols if c in mine.columns]].copy()
                # Translate status for display
                disp['status'] = disp['status'].astype(str).str.lower().map({
                    'pending':  '⏳ ממתין',
                    'approved': '✅ אושר',
                    'rejected': '❌ נדחה',
                }).fillna(disp['status'])
                disp = disp.rename(columns={
                    'start_date': 'תאריך התחלה',
                    'end_date':   'תאריך סיום',
                    'type':       'סוג',
                    'status':     'סטטוס',
                    'approved_by': 'מאשר',
                    'notes':      'הערה',
                })
                st.dataframe(disp, use_container_width=True, hide_index=True)

        # ── הגשת חופש עתידי (תאריך מרוחק) ─────────────────────────
        st.divider()
        with st.expander("📅 הגשת חופש עתידי (לתאריך מרוחק / חודש אחר)", expanded=False):
            st.caption("לבקשת היעדרות בחודש שאינו פתוח להגשה. דורש אישור מנהל מראש. לא תלוי במצב פתח/סגור של החודש הנוכחי.")
            fc1, fc2, fc3 = st.columns([2, 2, 2])
            with fc1:
                fut_start = st.date_input("מתאריך:",
                                          value=date.today() + timedelta(days=30),
                                          min_value=date.today(),
                                          key="fut_abs_start")
            with fc2:
                fut_end = st.date_input("עד תאריך:",
                                        value=date.today() + timedelta(days=30),
                                        min_value=date.today(),
                                        key="fut_abs_end")
            with fc3:
                fut_type = st.selectbox("סוג:",
                                        ["חופש עתידי", "חופש", "202", "היעדרות אחרת"],
                                        key="fut_abs_type")
            fut_note = st.text_input("הערה (רשות):", key="fut_abs_note")

            # Persist request types using the display-normalized form
            # ("חופש עתידי" → "חופש") so downstream label lookups never miss.
            fut_type = _ABSENCE_TYPE_TO_STATUS.get(fut_type, fut_type)

            if st.button("✅ שלח בקשה לחופש עתידי", key="fut_abs_submit"):
                if fut_end < fut_start:
                    st.error("תאריך הסיום חייב להיות שווה או אחרי תאריך ההתחלה.")
                else:
                    fut_year_month = fut_start.strftime("%Y-%m")
                    # Lookup user's daily dept for that future month
                    dr = st.session_state.dept_rotation
                    fut_dept = ""
                    fut_mgr_email = ""
                    if not dr.empty and 'employee' in dr.columns:
                        my_rot = dr[
                            (dr['employee'].astype(str).str.strip() == str(user_name).strip()) &
                            (dr['year_month'].astype(str) == fut_year_month)
                        ]
                        if not my_rot.empty:
                            fut_dept = str(my_rot.iloc[0].get('daily_dept', '')).strip()
                    if fut_dept:
                        for _, sr in st.session_state.staff.iterrows():
                            if str(sr.get('type', '')).strip() != 'מנהל מחלקה':
                                continue
                            mdepts = _parse_manage_depts(sr.get('manage_depts', ''))
                            if fut_dept in mdepts:
                                fut_mgr_email = str(sr.get('email', '')).strip()
                                break
                    new_fut_row = {
                        'id': str(uuid.uuid4()),
                        'employee': str(user_name).strip(),
                        'start_date': fut_start.strftime('%Y-%m-%d'),
                        'end_date':   fut_end.strftime('%Y-%m-%d'),
                        'type': fut_type,
                        'status': 'pending',
                        'dept_at_request': fut_dept,
                        'manager_email': fut_mgr_email,
                        'approved_by': '',
                        'notes': fut_note.strip(),
                        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'responded_at': '',
                    }
                    new_fut_df = pd.concat(
                        [st.session_state.absence_requests, pd.DataFrame([new_fut_row])],
                        ignore_index=True
                    )
                    st.session_state.absence_requests = new_fut_df
                    save_to_db("absence_requests", new_fut_df)

                    if fut_mgr_email:
                        send_notification_email(
                            fut_mgr_email,
                            f"בקשת חופש עתידי: {user_name}",
                            f"<div dir='rtl'><p>שלום,</p>"
                            f"<p>עובד <b>{user_name}</b> ({fut_dept or '—'}) הגיש בקשת חופש עתידי:</p>"
                            f"<ul><li>סוג: {fut_type}</li>"
                            f"<li>תאריכים: {new_fut_row['start_date']} – {new_fut_row['end_date']}</li>"
                            f"<li>הערה: {fut_note or '—'}</li></ul>"
                            f"<p>נא להיכנס למערכת לאישור או דחיה.</p></div>"
                        )
                    if not fut_dept:
                        st.warning("⚠️ לא נמצא שיבוץ מחלקה לחודש זה — הבקשה תועבר לכל מנהלי המחלקות.")
                    st.success(f"✅ בקשת חופש עתידי ל-{fut_start} – {fut_end} נשלחה!")
                    st.rerun()

    if selected_nav == 'סידור תורנויות':
        draw_calendar_view(2026, sel_month, "עובד/ת", user_name)

    # ── סידור עבודה (עובדים / מנהלי מחלקה) ────────────────────────
    if selected_nav == 'סידור עבודה':
        if role == "מנהל מחלקה":
            st.subheader("🗓️ סידור עבודה — מנהל/ת מחלקה")
            # Look up the depts this manager controls
            staff_row = st.session_state.staff[
                st.session_state.staff['name'].astype(str).str.strip() == str(user_name).strip()
            ]
            managed_depts = []
            if not staff_row.empty:
                managed_depts = _parse_manage_depts(staff_row.iloc[0].get('manage_depts', ''))
            if not managed_depts:
                st.warning("⚠️ לא הוגדרו מחלקות בניהולך. אנא פנה למנהל המערכת.")
            else:
                st.caption(f"מחלקות בניהולך: {' | '.join(managed_depts)}")

                # Feature 4: 12-month rolling forward window (auto-advances on month change).
                mgr_year, mgr_view_m = _sw_month_selector_12("sw_mgr")
                da_y_m = f"{mgr_year}-{mgr_view_m:02d}"
                st.markdown(f"#### לוח מחלקה — {_HEB_MONTHS[mgr_view_m-1]} {mgr_year}")
                st.caption("עורך/ת מחלקה: לחץ/י על תא לשינוי סטטוס. שורתך מצורפת לכל מחלקה — עריכה ישירה.")
                def _compose_mgr_emps(d_name):
                    dr = st.session_state.dept_rotation
                    emps = []
                    if not dr.empty and 'employee' in dr.columns:
                        mask = ((dr['year_month'].astype(str) == da_y_m) &
                                (dr['daily_dept'].astype(str) == d_name))
                        emps = dr[mask]['employee'].astype(str).str.strip().tolist()
                    mgrs = _get_dept_managers(d_name)
                    # ensure current user is included even if not in manage_depts list
                    if user_name not in mgrs:
                        mgrs = [user_name] + mgrs
                    return _sort_employees_by_role(
                        mgrs + [e for e in emps if e not in mgrs]
                    )

                def _mgr_render_dept(d_name, emps, ns):
                    if d_name == PNIM_DEPT:
                        _render_pnim_sided(da_y_m, mgr_view_m, ns,
                                           employees=emps, allow_temp_add=True,
                                           year=mgr_year)
                    else:
                        _render_dept_grid(d_name, da_y_m, mgr_view_m,
                                          ns, employees=emps, allow_temp_add=True,
                                          year=mgr_year)

                if len(managed_depts) > 1:
                    sub_dept_tabs = st.tabs(managed_depts)
                    for di, d_name in enumerate(managed_depts):
                        with sub_dept_tabs[di]:
                            emps = _compose_mgr_emps(d_name)
                            _mgr_render_dept(d_name, emps, f"mgr_{di}")
                            st.divider()
                            _render_export_buttons(d_name, da_y_m, mgr_view_m,
                                                   f"mgr_{di}", user_name)
                else:
                    d_name = managed_depts[0]
                    emps = _compose_mgr_emps(d_name)
                    _mgr_render_dept(d_name, emps, "mgr_solo")
                    st.divider()
                    _render_export_buttons(d_name, da_y_m, mgr_view_m,
                                           "mgr_solo", user_name)

    # ── ניהול בקשות (מנהל מחלקה) ─────────────────────────────────
    if selected_nav == 'ניהול בקשות' and role == "מנהל מחלקה":
        st.subheader("📋 ניהול בקשות היעדרות")

        # Lookup managed departments
        _mgr_staff_row = st.session_state.staff[
            st.session_state.staff['name'].astype(str).str.strip() == str(user_name).strip()
        ]
        _mgr_managed_depts = []
        if not _mgr_staff_row.empty:
            _mgr_managed_depts = _parse_manage_depts(_mgr_staff_row.iloc[0].get('manage_depts', ''))

        if not _mgr_managed_depts:
            st.warning("⚠️ לא הוגדרו מחלקות בניהולך.")
        else:
            # Build dept/emp match helpers (uses global _norm_dept)
            _mgr_norm_depts = [_norm_dept(d) for d in _mgr_managed_depts]

            _ar_mgr = st.session_state.absence_requests.copy()
            if not _ar_mgr.empty and 'status' in _ar_mgr.columns:
                _ar_mgr['status']          = _ar_mgr['status'].astype(str).str.lower()
                _ar_mgr['dept_at_request'] = _ar_mgr['dept_at_request'].astype(str).str.strip()
                _ar_mgr['employee']        = _ar_mgr['employee'].astype(str).str.strip()
                _dept_match_mgr = _ar_mgr['dept_at_request'].apply(_norm_dept).isin(_mgr_norm_depts)

                _dr_all_mgr = st.session_state.dept_rotation
                _mgr_emps_set = set()
                if not _dr_all_mgr.empty and 'daily_dept' in _dr_all_mgr.columns:
                    for _md in _mgr_norm_depts:
                        _emask = _dr_all_mgr['daily_dept'].apply(_norm_dept) == _md
                        _mgr_emps_set.update(
                            _dr_all_mgr[_emask]['employee'].astype(str).str.strip().tolist()
                        )
                _emp_match_mgr = _ar_mgr['employee'].isin(_mgr_emps_set)
            else:
                _ar_mgr = pd.DataFrame()
                _dept_match_mgr = pd.Series(dtype=bool)
                _emp_match_mgr  = pd.Series(dtype=bool)

            # ── Section 1: בקשות ממתינות ──────────────────────────
            st.markdown("#### ⏳ בקשות ממתינות")
            if _ar_mgr.empty:
                st.info("אין בקשות במערכת.")
            else:
                _pen_mgr = _ar_mgr[
                    (_ar_mgr['status'] == 'pending') &
                    (_dept_match_mgr | _emp_match_mgr)
                ]
                if 'start_date' in _pen_mgr.columns:
                    _pen_mgr = _pen_mgr.sort_values('start_date')
                if _pen_mgr.empty:
                    st.success("אין בקשות ממתינות במחלקותיך ✅")
                else:
                    st.caption(f"📋 {len(_pen_mgr)} בקשות ממתינות")
                    # Pre-compute approved rows for overlap check (once, outside loop)
                    _ar_appr_mgr = _ar_mgr.copy()
                    _ar_appr_mgr['_sd'] = pd.to_datetime(_ar_appr_mgr['start_date'], errors='coerce')
                    _ar_appr_mgr['_ed'] = pd.to_datetime(_ar_appr_mgr['end_date'], errors='coerce')
                    _ar_appr_mgr = _ar_appr_mgr[_ar_appr_mgr['status'] == 'approved']
                    for _idx, _row in _pen_mgr.iterrows():
                        with st.container(border=True):
                            # Conflict-check via the unified helper (normalises dept on both
                            # sides) — drives the approve UX below.
                            _rid = str(_row.get('id', _idx))
                            _rs2 = pd.to_datetime(_row.get('start_date', ''), errors='coerce')
                            _re2 = pd.to_datetime(_row.get('end_date', ''), errors='coerce')
                            _remp2 = str(_row.get('employee', '')).strip()
                            _rdept2_canon = _emp_dept_for_date(_remp2, _rs2)
                            _row_conflicts2 = _absence_conflicts(
                                _remp2, _rdept2_canon, _rs2, _re2,
                                exclude_id=_row.get('id'))

                            _c1, _c2, _c3, _c4, _c5 = st.columns([2, 2, 2, 1, 1])
                            _c1.markdown(f"**{_row.get('employee', '—')}**")
                            _c2.write(f"{_row['start_date']} – {_row['end_date']}")
                            _c3.write(_row.get('type', '—'))
                            if _c4.button("✅ אשר", key=f"nb_mgr_ap_{_rid}",
                                          use_container_width=True,
                                          disabled=bool(_row_conflicts2)):
                                _approve_request(_rid, str(user_name).strip())
                                st.rerun()
                            if _c5.button("❌ דחה", key=f"nb_mgr_rj_{_rid}",
                                          use_container_width=True):
                                _reject_request(_rid, str(user_name).strip())
                                st.rerun()
                            if _row_conflicts2:
                                st.warning(_format_absence_conflict_question(_row_conflicts2))
                                if st.button("✅ כן, אשר למרות החפיפה",
                                             key=f"nb_mgr_ap_force_{_rid}",
                                             type="primary"):
                                    _approve_request(_rid, str(user_name).strip())
                                    st.rerun()
                            if _row.get('notes'):
                                st.caption(f"💬 {_row['notes']}")

            st.divider()

            # ── Section 2: Gantt-style approved-absence view (Feature 3) ────
            st.markdown("#### ✅ בקשות שאושרו — תצוגת גאנט")
            _mgr_gantt_year, _mgr_gantt_month = _sw_month_selector_12("nb_mgr_gantt")
            _render_absence_gantt(_mgr_gantt_year, _mgr_gantt_month,
                                  dept_filter=set(_mgr_managed_depts))

            if st.session_state.pop('show_nb_mgr_del_ok', False):
                st.success("🗑️ הבקשה נמחקה.")

            # Hard-delete an approved request, scoped to managed depts.
            with st.expander("🗑️ מחיקת בקשה מאושרת", expanded=False):
                _mgr_del_df = st.session_state.absence_requests.copy()
                if _mgr_del_df.empty or 'status' not in _mgr_del_df.columns:
                    st.info("אין בקשות במערכת.")
                else:
                    _mgr_del_df['_status'] = _mgr_del_df['status'].astype(str).str.lower()
                    _mgr_del_df = _mgr_del_df[_mgr_del_df['_status'] == 'approved'].copy()
                    # Only show requests whose canonical dept is in this manager's scope
                    _mgr_del_df['_dept_canon'] = _mgr_del_df.apply(
                        lambda r: _emp_dept_for_date(r.get('employee', ''),
                                                     r.get('start_date', '')) or '',
                        axis=1)
                    _mgr_del_df = _mgr_del_df[
                        _mgr_del_df['_dept_canon'].isin(_mgr_managed_depts)]
                    if _mgr_del_df.empty:
                        st.info("אין בקשות מאושרות במחלקות בניהולך.")
                    else:
                        _mgr_del_df['_label'] = _mgr_del_df.apply(
                            lambda r: (f"{str(r.get('employee', '')).strip()} · "
                                       f"{r.get('_dept_canon', '') or '—'} · "
                                       f"{str(r.get('start_date', ''))[:10]} – {str(r.get('end_date', ''))[:10]} · "
                                       f"{str(r.get('type', '') or '').strip()}"),
                            axis=1)
                        _mgr_del_opts = _mgr_del_df[['id', '_label']].values.tolist()
                        _mgr_label_to_id = {lbl: rid for rid, lbl in _mgr_del_opts}
                        _mgr_sel_del = st.selectbox(
                            "בחר/י בקשה למחיקה:",
                            ["—"] + [lbl for _, lbl in _mgr_del_opts],
                            key="nb_mgr_del_sel")
                        _mgr_confirm_del = st.checkbox(
                            "✓ אני בטוח/ה — המחיקה אינה הפיכה",
                            key="nb_mgr_del_confirm",
                            value=False)
                        if st.button("🗑️ מחק בקשה", key="nb_mgr_del_btn",
                                     disabled=not (_mgr_sel_del and _mgr_sel_del != "—"
                                                   and _mgr_confirm_del),
                                     type="primary"):
                            _mgr_rid = _mgr_label_to_id.get(_mgr_sel_del)
                            if _mgr_rid and _delete_absence_request(_mgr_rid):
                                st.session_state['show_nb_mgr_del_ok'] = True
                                st.rerun()
                            else:
                                st.error("שגיאה במחיקה.")

            with st.expander("📋 תצוגת טבלה (כל הבקשות העתידיות שאושרו)", expanded=False):
                if _ar_mgr.empty:
                    st.info("אין בקשות שאושרו.")
                else:
                    _ar_mgr2 = _ar_mgr.copy()
                    _ar_mgr2['end_date']   = pd.to_datetime(_ar_mgr2['end_date'], errors='coerce')
                    _ar_mgr2['start_date'] = pd.to_datetime(_ar_mgr2['start_date'], errors='coerce')
                    _today_mgr = pd.Timestamp(date.today())
                    _ap_mgr = _ar_mgr2[
                        (_ar_mgr2['status'] == 'approved') &
                        (_dept_match_mgr | _emp_match_mgr) &
                        (_ar_mgr2['end_date'] >= _today_mgr)
                    ].copy()
                    if _ap_mgr.empty:
                        st.info("אין בקשות שאושרו לתאריכים עתידיים.")
                    else:
                        # Normalize dept via Gantt for display (Feature 1).
                        _ap_mgr['_dept_canon'] = _ap_mgr.apply(
                            lambda r: _emp_dept_for_date(r['employee'], r['start_date']) or '—',
                            axis=1)
                        _ap_mgr = _ap_mgr.sort_values(['_dept_canon', 'start_date'])
                        _show_mgr = _ap_mgr[['employee', '_dept_canon', 'start_date',
                                              'end_date', 'type', 'approved_by']].copy()
                        _show_mgr.columns = ['עובד/ת', 'מחלקה', 'מתאריך', 'עד תאריך', 'סוג', 'אושר ע"י']
                        _show_mgr['מתאריך']  = _show_mgr['מתאריך'].dt.strftime('%Y-%m-%d')
                        _show_mgr['עד תאריך'] = _show_mgr['עד תאריך'].dt.strftime('%Y-%m-%d')
                        st.dataframe(_show_mgr, use_container_width=True, hide_index=True)

            st.divider()

            # ── Section 3: הוספת היעדרות לעובד ────────────────────
            st.markdown("#### ➕ הוסף היעדרות לעובד")
            st.caption("בחר עובד/ת מהמחלקה לרישום היעדרות מאושרת ישירות.")

            # Build employee list: managed dept rotation members + manager themselves
            _nb_mgr_da_ym = f"2026-{daily_active_month_int:02d}"
            _nb_mgr_emps_list = list(_mgr_emps_set)
            if str(user_name).strip() not in _nb_mgr_emps_list:
                _nb_mgr_emps_list.append(str(user_name).strip())
            _nb_mgr_emps_list = sorted(_nb_mgr_emps_list)

            _nb_mgr_sel_emp = st.selectbox("עובד/ת:", _nb_mgr_emps_list, key="nb_mgr_emp")
            _nb_mgr_c1, _nb_mgr_c2 = st.columns(2)
            with _nb_mgr_c1:
                _nb_mgr_start = st.date_input("מתאריך:", key="nb_mgr_start",
                                               value=date.today(),
                                               min_value=date(2026, 1, 1),
                                               max_value=date(2026, 12, 31))
            with _nb_mgr_c2:
                _nb_mgr_end = st.date_input("עד תאריך:", key="nb_mgr_end",
                                             value=date.today(),
                                             min_value=date(2026, 1, 1),
                                             max_value=date(2026, 12, 31))
            _nb_mgr_type = st.selectbox("סוג היעדרות:", ["חופש", "202", "היעדרות אחרת"],
                                         key="nb_mgr_type")
            _nb_mgr_note = st.text_input("הערה (אופציונלי):", key="nb_mgr_note")

            # Pre-write conflict warning (Feature 2): show same-dept overlap
            # against approved absences BEFORE the manager submits.
            _nb_mgr_conflicts = []
            if _nb_mgr_sel_emp and _nb_mgr_end >= _nb_mgr_start:
                _nb_mgr_dept_preview = _emp_dept_for_date(_nb_mgr_sel_emp, _nb_mgr_start)
                _nb_mgr_conflicts = _absence_conflicts(
                    _nb_mgr_sel_emp, _nb_mgr_dept_preview, _nb_mgr_start, _nb_mgr_end)
            # Conflict-confirmation flow: replace the regular save button with an
            # explicit "yes, save despite overlap" button when a conflict exists.
            _nb_mgr_force = False
            if _nb_mgr_conflicts:
                st.warning(_format_absence_conflict_question(_nb_mgr_conflicts))
                _nb_mgr_force = st.button(
                    "✅ כן, הוסף למרות החפיפה",
                    key="nb_mgr_submit_force", type="primary")

            if (not _nb_mgr_conflicts and
                st.button("✅ הוסף היעדרות מאושרת", key="nb_mgr_submit")) or _nb_mgr_force:
                if _nb_mgr_end < _nb_mgr_start:
                    st.error("תאריך סיום לפני תאריך התחלה.")
                else:
                    # Gantt-canonical dept (Feature 1) — replaces the old
                    # _mgr_managed_depts[0] shortcut which captured the *manager's*
                    # dept rather than the *employee's* actual Gantt dept.
                    _nb_emp_dept = _emp_dept_for_date(_nb_mgr_sel_emp, _nb_mgr_start)
                    _nb_new_row = {
                        'id':              str(uuid.uuid4()),
                        'employee':        _nb_mgr_sel_emp,
                        'start_date':      _nb_mgr_start.strftime('%Y-%m-%d'),
                        'end_date':        _nb_mgr_end.strftime('%Y-%m-%d'),
                        'type':            _nb_mgr_type,
                        'status':          'approved',
                        'dept_at_request': _nb_emp_dept,
                        'manager_email':   '',
                        'approved_by':     str(user_name).strip(),
                        'notes':           _nb_mgr_note.strip(),
                        'created_at':      datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'responded_at':    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    }
                    _nb_new_df = pd.concat(
                        [st.session_state.absence_requests, pd.DataFrame([_nb_new_row])],
                        ignore_index=True
                    )
                    st.session_state.absence_requests = _nb_new_df
                    save_to_db("absence_requests", _nb_new_df)
                    _materialize_absence_to_wsd(
                        _nb_mgr_sel_emp, _nb_mgr_type,
                        _nb_new_row['start_date'], _nb_new_row['end_date'])
                    _build_approved_map()
                    st.session_state['show_nb_mgr_success'] = True
                    st.rerun()

            if st.session_state.pop('show_nb_mgr_success', False):
                st.success("✅ היעדרות נוספה ואושרה — תופיע בלוח מיד.")

    if selected_nav == 'סידור עבודה' and role not in ("מנהל מחלקה",):
        # מתמחה / רופא בכיר — read-only full department calendar.
        # Feature 4: 12-month rolling forward window (auto-advances on month change).
        emp_year, view_m = _sw_month_selector_12("sw_emp")
        es_year_month = f"{emp_year}-{view_m:02d}"

        # Resolve the user's daily dept for this month
        dr = st.session_state.dept_rotation
        my_dept = "—"
        if not dr.empty and 'employee' in dr.columns:
            my_row = dr[
                (dr['employee'].astype(str).str.strip() == str(user_name).strip()) &
                (dr['year_month'].astype(str) == es_year_month)
            ]
            if not my_row.empty:
                my_dept = str(my_row.iloc[0].get('daily_dept', '—'))

        st.subheader(f"🗓️ לוח מחלקה — {_HEB_MONTHS[view_m-1]} {emp_year}")
        if my_dept == "—":
            st.info("טרם שובצת למחלקה בחודש זה. פנה/י למנהל המערכת.")
        else:
            st.caption(f"מחלקתך: **{my_dept}**")
            # Build employees: dept_rotation + all managers of this dept
            _u_dr = st.session_state.dept_rotation
            _u_emps = []
            if not _u_dr.empty and 'employee' in _u_dr.columns:
                _u_mask = ((_u_dr['year_month'].astype(str) == es_year_month) &
                           (_u_dr['daily_dept'].astype(str) == my_dept))
                _u_emps = _u_dr[_u_mask]['employee'].astype(str).str.strip().tolist()
            _u_mgrs = _get_dept_managers(my_dept)
            _u_all  = _sort_employees_by_role(
                _u_mgrs + [e for e in _u_emps if e not in _u_mgrs]
            )
            if my_dept == PNIM_DEPT:
                _render_pnim_sided(es_year_month, view_m, "user_view",
                                   employees=_u_all,
                                   readonly=True, highlight_user=user_name,
                                   year=emp_year)
            else:
                _render_dept_grid(my_dept, es_year_month, view_m,
                                  "user_view", employees=_u_all,
                                  readonly=True, highlight_user=user_name,
                                  year=emp_year)

        # ── Personal schedule Excel download ──────────────────────────
        st.markdown("---")
        _pers_bytes, _pers_fname = _export_personal_schedule_excel(
            user_name, view_m,
            daily_dept=my_dept if my_dept != "—" else None)
        if _pers_bytes:
            st.download_button(
                "📥 הורד לוח עבודה שלי (Excel)",
                data=_pers_bytes,
                file_name=_pers_fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="personal_xl_dl",
            )